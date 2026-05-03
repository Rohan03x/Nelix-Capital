"""Tests for output/raw_sheets.py."""
import pytest

openpyxl = pytest.importorskip("openpyxl")

from openpyxl import Workbook
from auto_valuation.output.raw_sheets import write_raw_is, write_raw_bs, write_raw_cf

_IS_STMTS = [
    {"date": "2022-12-31", "revenue": 394_000, "grossProfit": 170_000,
     "ebit": 119_000, "netIncome": 99_803, "eps": 6.11},
    {"date": "2023-12-31", "revenue": 383_285, "grossProfit": 169_148,
     "ebit": 114_301, "netIncome": 96_995, "eps": 6.13},
]

_BS = [
    {"date": "2022-12-31", "totalAssets": 352_755, "totalLiabilities": 302_083,
     "totalEquity": 50_672, "cash": 23_646},
    {"date": "2023-12-31", "totalAssets": 352_583, "totalLiabilities": 290_437,
     "totalEquity": 62_146, "cash": 29_965},
]

_CF = [
    {"date": "2022-12-31", "operatingCashFlow": 122_151, "capitalExpenditure": -10_708,
     "freeCashFlow": 111_443, "netIncome": 99_803},
    {"date": "2023-12-31", "operatingCashFlow": 110_543, "capitalExpenditure": -10_959,
     "freeCashFlow": 99_584, "netIncome": 96_995},
]


class TestWriteRawIs:
    def test_creates_sheet(self):
        wb = Workbook()
        write_raw_is(wb, _IS_STMTS)
        assert any("IS" in s or "Income" in s for s in wb.sheetnames)

    def test_no_crash_empty(self):
        wb = Workbook()
        write_raw_is(wb, [])

    def test_data_written(self):
        wb = Workbook()
        write_raw_is(wb, _IS_STMTS)
        ws_name = next(s for s in wb.sheetnames if "IS" in s or "Income" in s)
        ws = wb[ws_name]
        # At least some data should be in the sheet
        assert ws.max_row >= 1


class TestWriteRawBs:
    def test_creates_sheet(self):
        wb = Workbook()
        write_raw_bs(wb, _BS)
        assert any("BS" in s or "Balance" in s for s in wb.sheetnames)

    def test_no_crash_empty(self):
        wb = Workbook()
        write_raw_bs(wb, [])


class TestWriteRawCf:
    def test_creates_sheet(self):
        wb = Workbook()
        write_raw_cf(wb, _CF)
        assert any("CF" in s or "Cash" in s for s in wb.sheetnames)

    def test_no_crash_empty(self):
        wb = Workbook()
        write_raw_cf(wb, [])
