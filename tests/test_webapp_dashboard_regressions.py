from __future__ import annotations

import copy
import json
from io import BytesIO

import pytest
from openpyxl import load_workbook

from auto_valuation.data import macro as macro_module
from auto_valuation.learning.confidence import build_ranked_confidence_model
from auto_valuation.learning import deployment_seed
from webapp.app import app as web_app
from webapp.data.ai_commentary import generate_commentary
from webapp.data import eodhd_client
from webapp.data import fmp_client
from webapp.data import knowledge_model as knowledge_model_module
from webapp.data import samples as samples_module
from webapp.data import yfinance_client
from webapp.data.eodhd_client import _derive_ebit_margin_target
from webapp.data.excel_export import build_excel_bytes
from webapp.data.samples import REGISTRY, _apply_overrides


def _midyear_annuity(rate_pct: float, years: int) -> float:
    rate = rate_pct / 100
    return sum(1 / (1 + rate) ** (period - 0.5) for period in range(1, years + 1))


def test_apply_overrides_matches_corrected_dcf_math() -> None:
    base = copy.deepcopy(REGISTRY["NKE"])
    overridden = copy.deepcopy(REGISTRY["NKE"])

    _apply_overrides(overridden, {"wacc": 9.4, "g": 2.5})

    forecast_years = max(len(base.get("forecast") or []), 7)
    expected_pv_tv = (
        base["terminal_ufcf"]
        * (1 + 2.5 / 100)
        / ((9.4 - 2.5) / 100)
        / (1 + 9.4 / 100) ** forecast_years
    )
    expected_pv_uf = base["pv_ufcfs"] * (
        _midyear_annuity(9.4, forecast_years) / _midyear_annuity(base["wacc"], forecast_years)
    )
    expected_iv = (expected_pv_tv + expected_pv_uf - base["net_debt"]) / base["diluted_shares"]

    assert overridden["intrinsic_value"] == pytest.approx(expected_iv, abs=0.01)

    sensitivity = overridden["sensitivity"]
    # Correct orientation: rows = g values, columns = wacc values
    # iv_grid[g_idx][wacc_idx]
    center = sensitivity["iv_grid"][sensitivity["base_g_idx"]][sensitivity["base_wacc_idx"]]
    assert center == pytest.approx(overridden["intrinsic_value"], abs=0.1)

    # Along a fixed g row: higher wacc (rightmost) → lower IV
    mid_row = sensitivity["base_g_idx"]
    assert sensitivity["iv_grid"][mid_row][0] > sensitivity["iv_grid"][mid_row][-1]

    # Along a fixed wacc column: higher g (bottom row) → higher IV
    mid_col = sensitivity["base_wacc_idx"]
    assert sensitivity["iv_grid"][0][mid_col] < sensitivity["iv_grid"][-1][mid_col]


def test_excel_export_forecast_rows_follow_updated_forecast_inputs() -> None:
    data = copy.deepcopy(REGISTRY["NKE"])
    _apply_overrides(data, {"wacc": 9.4, "g": 2.5})

    workbook = load_workbook(BytesIO(build_excel_bytes(data)), data_only=False)
    sheet = workbook["valuation"]
    forecast_cols = list(range(12, 19))

    interest_expense = [sheet.cell(348, col).value for col in forecast_cols]
    diluted_shares = [sheet.cell(371, col).value for col in forecast_cols]
    supplemental_shares = [sheet.cell(543, col).value for col in forecast_cols]
    residual_nwc = [sheet.cell(444, col).value for col in forecast_cols]
    change_in_nwc = [sheet.cell(475, col).value for col in forecast_cols]

    assert all(value < 0 for value in interest_expense)
    assert diluted_shares[0] > diluted_shares[-1]
    assert supplemental_shares == diluted_shares
    assert any(abs(value) > 1e-6 for value in residual_nwc)

    component_sums = [
        sum(sheet.cell(row, col).value for row in (441, 442, 443, 444))
        for col in forecast_cols
    ]
    for total, nowc in zip(component_sums, change_in_nwc):
        assert total == pytest.approx(-nowc, abs=1e-6)

    first_forecast_col = forecast_cols[0]
    assert sheet.cell(321, first_forecast_col).value == pytest.approx(
        sheet.cell(339, first_forecast_col).value / sheet.cell(337, first_forecast_col).value,
        abs=1e-9,
    )
    assert "Levered Hist." in sheet.cell(474, 1).value


def test_excel_export_uses_display_currency_and_profile_metadata() -> None:
    data = copy.deepcopy(REGISTRY["NKE"])
    data.update(
        {
            "display_currency": "EUR",
            "display_currency_symbol": "€",
            "currency": "EUR",
            "model_currency": "USD",
            "quote_currency": "USD",
            "exchange": "EPA",
            "sector": "Industrials",
            "industry": "Electrical Equipment",
            "price_date": "2025-05-01",
            "confidence_breakdown": {
                "suitability_note": "Use DCF with caution for capital-intensive cyclicals.",
            },
        }
    )

    workbook = load_workbook(BytesIO(build_excel_bytes(data)), data_only=False)
    sheet = workbook["valuation"]

    assert "EUR M" in sheet["B5"].value
    assert "model currency USD" in sheet["B5"].value
    assert sheet["B6"].value == "EPA | Industrials | Electrical Equipment"
    assert "capital-intensive cyclicals" in sheet["B7"].value
    assert sheet["A11"].value == "Current Market Price (EUR/share)"
    assert sheet["A35"].value == "DCF Intrinsic Value (EUR/share)"


def test_fmp_cached_reads_disk_cache(tmp_path, monkeypatch) -> None:
    monkeypatch.setattr(fmp_client, "_CACHE_DIR", tmp_path)
    fmp_client._CACHE.clear()

    first = fmp_client._cached("fred:dgs10", lambda: {"rate": 4.2})
    assert first == {"rate": 4.2}
    assert (tmp_path / "fmp_fred_dgs10.json").exists()

    fmp_client._CACHE.clear()
    second = fmp_client._cached("fred:dgs10", lambda: {"rate": 9.9})
    assert second == {"rate": 4.2}


def test_dashboard_keeps_knowledge_model_when_macro_cache_disabled(monkeypatch) -> None:
    monkeypatch.setattr(macro_module, "_CACHE_DIR", None)

    data = samples_module.get_dashboard_data("NKE")

    assert data["knowledge_model"] is not None
    assert data["knowledge_model"]["summary"].startswith("Weighted knowledge model active")


def test_learning_loaders_fallback_to_seed_snapshot(tmp_path, monkeypatch) -> None:
    seed_path = tmp_path / "dashboard_learning_seed.json"
    seed_path.write_text(
        json.dumps(
            {
                "cohort_observations": [
                    {
                        "ticker": "NKE",
                        "sector": "Consumer Cyclical",
                        "industry": "Footwear & Accessories",
                        "data_vintage_years": 5,
                        "market_cap_regime": "large",
                        "macro_regime": "neutral",
                        "predicted_revenue_growth": 0.05,
                        "actual_revenue_growth": 0.06,
                        "predicted_ebit_margin": 0.10,
                        "actual_ebit_margin": 0.11,
                        "predicted_wacc": 0.09,
                        "actual_wacc": 0.09,
                        "predicted_terminal_growth": 0.025,
                        "actual_terminal_growth": 0.025,
                        "predicted_beta": 1.1,
                        "actual_beta": 1.1,
                        "predicted_ufcf_margin": 0.08,
                        "actual_ufcf_margin": 0.09,
                        "predicted_reinvestment_rate": 0.01,
                        "actual_reinvestment_rate": 0.015,
                        "feature_vector": [0.1] * 10,
                        "structural_break_flag": False,
                        "structural_break_hints": [],
                    }
                ],
                "analog_observations": [
                    {
                        "ticker": "ADDYY",
                        "sector": "Consumer Cyclical",
                        "industry": "Footwear & Accessories",
                        "vintage_year": 5,
                        "feature_vector": [0.1] * 10,
                        "feature_map": {},
                        "outcome_revenue_cagr_5y": 0.02,
                        "outcome_margin_change_bps": 50.0,
                        "outcome_ev_multiple_change": 0.1,
                        "pattern_label": "MATURE_COMPOUNDER",
                        "market_cap_regime": "large",
                        "macro_regime": "neutral",
                        "maturity_stage": "mid",
                        "valuation_regime": "fair",
                        "volatility_regime": "stable",
                        "data_quality_score": 0.8,
                        "sample_size": 5,
                        "predictive_usefulness": 0.7,
                        "as_of_year": 2026,
                    }
                ],
            }
        ),
        encoding="utf-8",
    )
    monkeypatch.setattr(deployment_seed, "SEED_PATH", seed_path)
    deployment_seed.reset_seed_cache()

    class _BrokenLedgerReader:
        def __init__(self, *args, **kwargs):
            raise RuntimeError("ledger unavailable")

    monkeypatch.setattr(knowledge_model_module, "LedgerReader", _BrokenLedgerReader)

    cohort = knowledge_model_module._load_learning_cohort(limit=10)
    analogs = knowledge_model_module._load_analog_candidates(limit=10)

    assert len(cohort) == 1
    assert cohort[0]["ticker"] == "NKE"
    assert len(analogs) == 1
    assert analogs[0].ticker == "ADDYY"


def test_watchlist_api_falls_back_to_seed_when_discovery_unavailable(monkeypatch) -> None:
    monkeypatch.setattr("webapp.app._safe_discovery_store", lambda: None)
    monkeypatch.setattr("webapp.app._seed_watchlist", lambda limit=30: [{"ticker": "AAPL", "company_name": "Apple Inc."}])

    client = web_app.test_client()
    response = client.get("/api/watchlist")

    assert response.status_code == 200
    payload = response.get_json()
    assert payload["ok"] is True
    assert payload["seeded"] is True
    assert payload["items"][0]["ticker"] == "AAPL"


def test_ai_commentary_uses_correct_premium_discount_wording() -> None:
    undervalued = generate_commentary(
        {
            "company_name": "Shell plc",
            "price": 45.30,
            "intrinsic_value": 129.79,
            "upside_pct": 186.5,
            "wacc": 6.5,
            "terminal_growth": 2.5,
            "tv_pct": 78.0,
            "confidence_score": 62,
        }
    )["valuation_summary"]
    overvalued = generate_commentary(
        {
            "company_name": "Tesla Inc",
            "price": 381.63,
            "intrinsic_value": 20.65,
            "upside_pct": -94.6,
            "wacc": 12.8,
            "terminal_growth": 2.5,
            "tv_pct": 57.8,
            "confidence_score": 61,
        }
    )["valuation_summary"]

    assert "premium to the current market price" in undervalued
    assert "discount to the current market price" in overvalued


def test_get_dashboard_data_falls_through_when_provider_raises(monkeypatch) -> None:
    fallback = copy.deepcopy(REGISTRY["NKE"])
    fallback["data_source"] = "yfinance"

    monkeypatch.setattr(eodhd_client, "is_available", lambda: True)
    monkeypatch.setattr(
        eodhd_client,
        "build_dashboard_data",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(RuntimeError("boom")),
    )
    monkeypatch.setattr(yfinance_client, "is_available", lambda: True)
    monkeypatch.setattr(yfinance_client, "build_dashboard_data", lambda _ticker: copy.deepcopy(fallback))
    monkeypatch.setattr(fmp_client, "is_available", lambda: False)

    data = samples_module.get_dashboard_data("CDI.PA")

    assert data["data_source"] == "yfinance"
    assert data["company_name"] == fallback["company_name"]


def test_get_dashboard_data_preserves_knowledge_model_confidence_breakdown(monkeypatch) -> None:
    live_payload = copy.deepcopy(REGISTRY["NKE"])
    live_payload.update(
        {
            "data_source": "eodhd",
            "is_demo": False,
            "confidence_score": 53,
            "confidence_breakdown": {"total": 53, "label": "Guarded Confidence"},
            "knowledge_model": {
                "confidence_model": {
                    "dashboard_breakdown": {
                        "total": 53,
                        "grade": "C",
                        "label": "Guarded Confidence",
                        "color": "amber",
                        "dcf_suitable": True,
                        "suitability_note": "",
                        "warnings": ["No close same-industry peer basket is available to sanity-check the live valuation."],
                        "dimensions": [],
                    }
                }
            },
        }
    )

    monkeypatch.setattr(eodhd_client, "is_available", lambda: True)
    monkeypatch.setattr(eodhd_client, "build_dashboard_data", lambda *_args, **_kwargs: copy.deepcopy(live_payload))
    monkeypatch.setattr(samples_module, "score_confidence", lambda _data: (_ for _ in ()).throw(AssertionError("legacy confidence should not run")))

    data = samples_module.get_dashboard_data("NKE")

    assert data["confidence_score"] == 53
    assert data["confidence_breakdown"]["label"] == "Guarded Confidence"


def test_yfinance_safe_last_price_ignores_broken_fast_info() -> None:
    class _BrokenFastInfo:
        @property
        def last_price(self) -> float:
            raise AttributeError("PriceHistory object has no attribute '_dividends'")

    class _FakeTicker:
        fast_info = _BrokenFastInfo()

    assert yfinance_client._safe_last_price(_FakeTicker(), {"currentPrice": 12.34}) == pytest.approx(12.34)


def test_yfinance_safe_fast_info_value_ignores_broken_attributes() -> None:
    class _BrokenFastInfo:
        @property
        def market_cap(self) -> float:
            raise AttributeError("broken fast_info market cap")

    assert yfinance_client._safe_fast_info_value(_BrokenFastInfo(), "market_cap", 123.0) == pytest.approx(123.0)


def test_live_confidence_marks_reit_as_not_dcf_suitable() -> None:
    model = build_ranked_confidence_model({"sector": "Real Estate", "industry": "Retail REIT"})

    assert model["dashboard_breakdown"]["dcf_suitable"] is False
    assert "FFO/AFFO" in model["dashboard_breakdown"]["suitability_note"]


def test_live_confidence_marks_blank_metadata_as_provisional() -> None:
    model = build_ranked_confidence_model({"sector": "", "industry": ""})

    assert model["dashboard_breakdown"]["dcf_suitable"] is False
    assert "metadata are incomplete" in model["dashboard_breakdown"]["suitability_note"]


def test_get_dashboard_data_survives_enrichment_failures(monkeypatch) -> None:
    monkeypatch.setattr(eodhd_client, "is_available", lambda: False)
    monkeypatch.setattr(yfinance_client, "is_available", lambda: False)
    monkeypatch.setattr(fmp_client, "is_available", lambda: False)
    monkeypatch.setattr(samples_module, "score_confidence", lambda _data: (_ for _ in ()).throw(RuntimeError("conf")))
    monkeypatch.setattr(samples_module, "generate_commentary", lambda _data: (_ for _ in ()).throw(RuntimeError("commentary")))
    monkeypatch.setattr(samples_module, "_build_investment_memo", lambda _data: (_ for _ in ()).throw(RuntimeError("memo")))
    monkeypatch.setattr(samples_module, "_build_market_expectations", lambda _data: (_ for _ in ()).throw(RuntimeError("expectations")))

    data = samples_module.get_dashboard_data("NKE")

    assert data["is_demo"] is True
    assert data["confidence_score"] == 50
    assert data["confidence_breakdown"]["label"] == "Low Confidence"
    assert data["ai_commentary"] == {}
    assert data["investment_memo"] is None
    assert data["market_expectations"] is None


def test_ebit_margin_target_respects_recent_profitable_regime_shift() -> None:
    target, source = _derive_ebit_margin_target(
        5.9,
        [-7.8, -14.8, -1.6, 0.1, 6.0, 12.4, 17.5, 10.5, 9.6, 5.9],
        24.6,
        "Auto Manufacturers",
    )

    assert target == pytest.approx(12.2, abs=0.1)
    assert source == "Recent profitable regime + historical anchor"


def test_ebit_margin_target_restrains_shrinking_mature_business() -> None:
    target, source = _derive_ebit_margin_target(
        8.1,
        [5.0, 6.3, 6.5, 6.4, 6.5, 7.8, 9.8, 5.2, 8.4, 8.1],
        40.1,
        "Electrical Equipment & Parts",
        revenues=[7115, 6965, 6357, 6247, 6502, 6860, 7514, 6704, 6143, 5765],
    )

    assert target == pytest.approx(8.9, abs=0.1)
    assert source == "Historical peak margin + restrained expansion"


def test_global_overlay_prefers_matching_structural_break_regime() -> None:
    stable_records = [
        {
            "sector": f"Stable {idx % 3}",
            "market_cap_regime": "mid",
            "macro_regime": "neutral",
            "data_vintage_years": 8,
            "predicted_ebit_margin": 0.08,
            "actual_ebit_margin": 0.10,
            "predicted_revenue_growth": 0.01,
            "actual_revenue_growth": 0.03,
            "predicted_wacc": 0.09,
            "actual_wacc": 0.088,
            "predicted_terminal_growth": 0.025,
            "actual_terminal_growth": 0.026,
            "predicted_beta": 1.0,
            "actual_beta": 0.98,
            "structural_break_flag": False,
        }
        for idx in range(6)
    ]
    break_records = [
        {
            "sector": f"Break {idx % 3}",
            "market_cap_regime": "mid",
            "macro_regime": "neutral",
            "data_vintage_years": 8,
            "predicted_ebit_margin": 0.08,
            "actual_ebit_margin": 0.03,
            "predicted_revenue_growth": 0.01,
            "actual_revenue_growth": -0.04,
            "predicted_wacc": 0.09,
            "actual_wacc": 0.102,
            "predicted_terminal_growth": 0.025,
            "actual_terminal_growth": 0.018,
            "predicted_beta": 1.0,
            "actual_beta": 1.12,
            "structural_break_flag": True,
        }
        for idx in range(6)
    ]

    stable_overlay = knowledge_model_module._global_cross_symbol_overlay(
        stable_records + break_records,
        data_vintage_years=8,
        market_cap_regime="mid",
        macro_regime="neutral",
        subject_structural_break_like=False,
    )
    break_overlay = knowledge_model_module._global_cross_symbol_overlay(
        stable_records + break_records,
        data_vintage_years=8,
        market_cap_regime="mid",
        macro_regime="neutral",
        subject_structural_break_like=True,
    )

    assert stable_overlay["regime_filter"] == "matched"
    assert break_overlay["regime_filter"] == "matched"
    assert stable_overlay["ebit_margin_adj_pp"] > 0
    assert break_overlay["ebit_margin_adj_pp"] < 0
    assert stable_overlay["wacc_adj_pp"] <= 0
    assert break_overlay["wacc_adj_pp"] > 0


def test_refine_live_assumptions_caps_bullish_memory_for_declining_structural_break(monkeypatch) -> None:
    class _Diagnostics:
        def to_dict(self):
            return {}

    class _Calibrated:
        revenue_growth_adj = 0.04
        ebit_margin_adj = 0.14
        beta_adj = -0.05
        wacc_adj = -0.01
        terminal_growth_adj = 0.01
        calibration_confidence = 0.9
        calibration_cohort_size = 20
        scenario_width_multiplier = 1.6
        calibration_diagnostics = _Diagnostics()

    monkeypatch.setattr(knowledge_model_module, "calibrate", lambda *args, **kwargs: _Calibrated())
    monkeypatch.setattr(
        knowledge_model_module,
        "_global_cross_symbol_overlay",
        lambda *args, **kwargs: {
            "enabled": True,
            "scope": "regime",
            "cohort_size": 12,
            "sector_span": 3,
            "confidence": 0.8,
            "regime_filter": "matched",
            "revenue_growth_adj_pp": 2.0,
            "ebit_margin_adj_pp": 3.0,
            "wacc_adj_pp": -0.7,
            "terminal_growth_adj_pp": 0.4,
            "beta_adj": -0.08,
            "note": "Global cross-symbol learning active.",
        },
    )
    monkeypatch.setattr(
        knowledge_model_module,
        "build_relationship_graph",
        lambda **kwargs: {
            "enabled": True,
            "overlay": {
                "revenue_growth_adj_pp": 2.0,
                "ebit_margin_adj_pp": 2.0,
                "wacc_adj_pp": -0.6,
                "terminal_growth_adj_pp": 0.3,
                "beta_adj": -0.06,
            },
            "node_count": 4,
            "note": "Relationship graph learning active.",
        },
    )
    monkeypatch.setattr(knowledge_model_module, "_load_analog_candidates", lambda limit=None: [])
    monkeypatch.setattr(
        knowledge_model_module,
        "_build_layered_learning_snapshot",
        lambda **kwargs: {
            "learned_metrics": {"reinvestment_confidence": 0.0},
            "uncertainty": {"scenario_width_multiplier": 2.0, "effective_confidence": 0.6, "weak_evidence": False},
            "structural_break": {"detected": True, "score": 0.88, "note": "Break detected."},
            "layer_mix": {},
        },
    )
    monkeypatch.setattr(
        knowledge_model_module,
        "match_pattern_library",
        lambda features: ("MATURE_COMPOUNDER", 0.9, {"revenue_growth_adj": 0.02, "ebit_margin_adj": 0.02}),
    )
    monkeypatch.setattr(knowledge_model_module, "_build_learning_explainability", lambda **kwargs: {"summary": "ok"})
    monkeypatch.setattr(
        knowledge_model_module,
        "build_ranked_confidence_model",
        lambda payload: {
            "summary": "ok",
            "dominant_risk": "regime",
            "assumption_confidence": {"score": 0.62},
            "valuation_confidence": {"score": 0.58, "expected_error_pct": {"p50": 18.0}},
            "components": [],
            "ranking_signal": 0.4,
        },
    )
    monkeypatch.setattr(knowledge_model_module, "_build_memory_hierarchy", lambda **kwargs: {"items": []})

    result = knowledge_model_module.refine_live_assumptions(
        ticker="TEST",
        sector="Industrials",
        industry="Electrical Equipment & Parts",
        market_cap=9_000.0,
        revenues=[7115, 6965, 6357, 6247, 6502, 6860, 7514, 6704, 6143, 5765],
        ebit_margins=[5.0, 6.3, 6.5, 6.4, 6.5, 7.8, 9.8, 5.2, 8.4, 8.1],
        gross_margin_base_pct=40.1,
        revenue_growth_near=-2.2,
        terminal_growth=2.5,
        ebit_margin_base_pct=8.1,
        ebit_margin_target=8.9,
        beta=1.05,
        wacc=7.1,
        rf_rate=4.2,
        erp=4.8,
        kd_post=3.6,
        e_wt=70.0,
        d_wt=30.0,
        total_assets=8_500.0,
        total_debt=1_700.0,
        revenue_base=5765.0,
        operating_cf=620.0,
        fcf=399.0,
        capex_pct=2.4,
        capexes=[150, 155, 160, 165, 170, 175, 180, 185, 190, 195],
        da_pct=2.0,
        das=[130, 132, 135, 138, 140, 142, 145, 147, 148, 150],
        sbc_pct=0.0,
        sbcs=[0.0] * 10,
        tax_rate_pct=24.0,
        pretax_incomes=[300, 320, 310, 290, 275, 260, 245, 230, 220, 210],
        tax_provisions=[72, 77, 74, 69, 66, 63, 59, 56, 53, 50],
        dso=48.0,
        dio=62.0,
        dpo=44.0,
        observations=[],
    )

    assert result["regime_guardrail"]["applied"] is True
    assert result["ebit_margin_target"] <= 9.1
    assert result["revenue_growth_near"] <= 0.0
    assert result["wacc"] >= 7.1
    assert result["terminal_growth"] <= 2.5
    assert result["beta"] >= 1.05
    assert "safeguard applied" in result["assumption_weights"]["ebit_margin_target"]["source"]
    assert "positive learned memory is capped" in result["assumption_weights"]["wacc"]["warn"]