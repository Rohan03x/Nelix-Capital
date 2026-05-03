"""Tests for output/model_sheet.py."""
import pytest

openpyxl = pytest.importorskip("openpyxl")

from openpyxl import Workbook
from auto_valuation.output.model_sheet import write_model_sheet

_HIST = [
    {"date": "2022-12-31", "revenue": 394_000, "ebit": 119_000,
     "netIncome": 99_803, "da": 11_284, "sbc": 9_038, "capex": 10_708,
     "nowc": -18_000},
    {"date": "2023-12-31", "revenue": 383_285, "ebit": 114_301,
     "netIncome": 96_995, "da": 11_519, "sbc": 10_833, "capex": 10_959,
     "nowc": -17_000},
]


def _make_forecast_year(yr):
    """Return a minimal namespace mimicking a ForecastYear dataclass."""
    class FY:
        pass
    fy = FY()
    fy.year = yr
    fy.revenue = 400_000 + yr * 10_000
    fy.ebit = 120_000
    fy.ebit_margin = 0.30
    fy.nopat = 94_800
    fy.da = 12_000
    fy.sbc = 10_000
    fy.capex = 11_000
    fy.delta_nowc = -500
    fy.ufcf = 95_300
    fy.ocf = 107_000
    fy.net_income = 97_000
    fy.tax_rate = 0.21
    fy.ibd = 100_000
    fy.interest_expense = 5_000
    return fy


_FORECAST = [_make_forecast_year(yr) for yr in range(1, 8)]

_VALUATION = {
    "enterprise_value_mm": 2_500_000,
    "equity_value_mm": 2_400_000,
    "implied_share_price": 150.0,
    "current_price": 130.0,
    "upside_pct": 0.154,
    "wacc": 0.09,
    "terminal_g": 0.025,
    "tv_pct_of_ev": 0.72,
    "net_debt_mm": 100_000,
    "diluted_shares_mm": 15_552,
}


class TestWriteModelSheet:
    def test_creates_model_sheet(self):
        wb = Workbook()
        write_model_sheet(wb, "AAPL", _HIST, _FORECAST, _VALUATION)
        assert "Model" in wb.sheetnames

    def test_replaces_existing_sheet(self):
        wb = Workbook()
        wb.create_sheet("Model")
        write_model_sheet(wb, "AAPL", _HIST, _FORECAST, _VALUATION)
        assert "Model" in wb.sheetnames
        assert wb.sheetnames.count("Model") == 1

    def test_no_crash_empty_hist(self):
        wb = Workbook()
        write_model_sheet(wb, "AAPL", [], _FORECAST, _VALUATION)
        assert "Model" in wb.sheetnames

    def test_data_written(self):
        wb = Workbook()
        write_model_sheet(wb, "AAPL", _HIST, _FORECAST, _VALUATION)
        ws = wb["Model"]
        assert ws.max_row >= 2
