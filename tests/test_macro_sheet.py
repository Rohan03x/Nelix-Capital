"""Tests for output/macro_sheet.py."""
import pytest
from datetime import date

openpyxl = pytest.importorskip("openpyxl")

from openpyxl import Workbook
from auto_valuation.output.macro_sheet import write_macro_sheet

_MACRO = {
    "risk_free_rate": 0.043,
    "risk_free_series": "GS10",
    "risk_free_currency": "USD",
    "rf_as_of_date": date(2024, 1, 15),
    "equity_risk_premium": 0.056,
    "erp_source": "Damodaran",
    "country_risk_premium": 0.0,
    "country_name": "United States",
    "size_premium": 0.0,
    "size_decile": "Large Cap",
    "beta_source": "Bloomberg",
    "beta_value": 1.29,
    "beta_date": date(2024, 1, 15),
    "fx_rate": 1.0,
    "fx_pair": "USD/USD",
    "reporting_currency": "USD",
}


class TestWriteMacroSheet:
    def test_creates_macro_sheet(self):
        wb = Workbook()
        write_macro_sheet(wb, "AAPL", _MACRO)
        assert "Macro" in wb.sheetnames

    def test_no_crash_empty_macro(self):
        wb = Workbook()
        write_macro_sheet(wb, "AAPL", {})
        assert "Macro" in wb.sheetnames

    def test_replaces_existing_sheet(self):
        wb = Workbook()
        wb.create_sheet("Macro")
        write_macro_sheet(wb, "AAPL", _MACRO)
        assert wb.sheetnames.count("Macro") == 1

    def test_custom_run_date(self):
        wb = Workbook()
        write_macro_sheet(wb, "AAPL", _MACRO, run_date=date(2024, 6, 30))
        assert "Macro" in wb.sheetnames

    def test_data_written(self):
        wb = Workbook()
        write_macro_sheet(wb, "AAPL", _MACRO)
        ws = wb["Macro"]
        assert ws.max_row >= 2
