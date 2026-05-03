"""
tests/test_session12_gaps.py — Tests for all 48 functions added in Session 12.

Covers:
- Canonical aliases (checklist-canonical names → existing implementations)
- Genuine new implementations (error_recovery, compute_segment_forecast,
  check_wc_seasonality_flag, compute_da_forecast, capex_convergence_to_da,
  compute_interest_expense, compute_pv_ufcfs, compute_enterprise_value,
  write_sensitivity_sheet, write_validation_sheet, write_tornado_chart,
  compute_ebitda, validate_terminal_roic, check_ufcf_sign, etc.)
"""
from __future__ import annotations

import pytest


# ─────────────────────────────────────────────────────────────────────────────
# Phase 0 — error_recovery
# ─────────────────────────────────────────────────────────────────────────────

def test_error_recovery_returns_fallback():
    from auto_valuation.utils.error import error_recovery
    result = error_recovery(ValueError("boom"), context="test", fallback=42)
    assert result == 42


def test_error_recovery_none_fallback():
    from auto_valuation.utils.error import error_recovery
    result = error_recovery(RuntimeError("oops"))
    assert result is None


# ─────────────────────────────────────────────────────────────────────────────
# Phase 1 — fetcher quarterly aliases
# ─────────────────────────────────────────────────────────────────────────────

def test_quarterly_aliases_exist():
    from auto_valuation.data.fetcher import (
        fetch_quarterly_income_statement,
        fetch_quarterly_balance_sheet,
        fetch_quarterly_cash_flow,
        fetch_income_quarterly,
        fetch_balance_quarterly,
        fetch_cashflow_quarterly,
    )
    assert fetch_quarterly_income_statement is fetch_income_quarterly
    assert fetch_quarterly_balance_sheet is fetch_balance_quarterly
    assert fetch_quarterly_cash_flow is fetch_cashflow_quarterly


def test_align_fiscal_year_alias():
    from auto_valuation.data.fiscal_year import align_fiscal_year, align_to_calendar_year
    assert align_fiscal_year is align_to_calendar_year


# ─────────────────────────────────────────────────────────────────────────────
# Phase 2 — income statement / working capital / balance sheet
# ─────────────────────────────────────────────────────────────────────────────

def test_compute_revenue_forecast_alias():
    from auto_valuation.model.income_statement import (
        compute_revenue_forecast, build_revenue_forecast,
    )
    assert compute_revenue_forecast is build_revenue_forecast


def test_normalize_effective_tax_rate_alias():
    from auto_valuation.model.income_statement import (
        normalize_effective_tax_rate, normalise_tax_rate,
    )
    assert normalize_effective_tax_rate is normalise_tax_rate


def test_normalize_effective_tax_rate_clamps():
    from auto_valuation.model.income_statement import normalize_effective_tax_rate
    stmts = [
        {"incomeTaxExpense": 500, "incomeBeforeTax": 1000},
        {"incomeTaxExpense": 400, "incomeBeforeTax": 1000},
        {"incomeTaxExpense": 300, "incomeBeforeTax": 1000},
    ]
    rate = normalize_effective_tax_rate(stmts)
    assert 0.05 <= rate <= 1.0


def test_compute_revenue_cagr_alias():
    from auto_valuation.model.income_statement import (
        compute_revenue_cagr, historical_revenue_cagr,
    )
    assert compute_revenue_cagr is historical_revenue_cagr


def test_compute_segment_forecast_basic():
    from auto_valuation.model.income_statement import compute_segment_forecast
    segments = [
        {"name": "North America", "revenue": 100, "growth_rate": 0.10},
        {"name": "Europe",        "revenue": 50,  "growth_rate": 0.05},
    ]
    result = compute_segment_forecast(segments, forecast_years=3)
    assert set(result.keys()) == {"North America", "Europe"}
    assert len(result["North America"]) == 3
    # Year 1: 100 * 1.10 = 110
    assert abs(result["North America"][0] - 110.0) < 0.01
    # Year 3: 100 * 1.10^3 = 133.1
    assert abs(result["North America"][2] - 133.1) < 0.1


def test_compute_segment_forecast_empty():
    from auto_valuation.model.income_statement import compute_segment_forecast
    result = compute_segment_forecast([], forecast_years=5)
    assert result == {}


def test_compute_segment_forecast_default_growth():
    from auto_valuation.model.income_statement import compute_segment_forecast
    segs = [{"name": "A", "revenue": 200}]
    result = compute_segment_forecast(segs, forecast_years=1, base_growth=0.08)
    assert abs(result["A"][0] - 216.0) < 0.01


def test_compute_working_capital_days_alias():
    from auto_valuation.model.working_capital import (
        compute_working_capital_days, compute_cwc_days,
    )
    assert compute_working_capital_days is compute_cwc_days


def test_forecast_nowc_alias():
    from auto_valuation.model.working_capital import forecast_nowc, build_nowc_forecast
    assert forecast_nowc is build_nowc_forecast


def test_check_wc_seasonality_flag_no_data():
    from auto_valuation.model.working_capital import check_wc_seasonality_flag
    assert check_wc_seasonality_flag([]) is False
    assert check_wc_seasonality_flag(None) is False  # type: ignore[arg-type]


def test_check_wc_seasonality_flag_high_swing():
    from auto_valuation.model.working_capital import check_wc_seasonality_flag
    # Q1 = high NOWC, Q3 = low NOWC → big swing (need at least 3 items)
    bs = [
        {"period": "Q1", "net_receivables": 1000, "inventory": 500, "accountPayables": 200},
        {"period": "Q2", "net_receivables": 600,  "inventory": 300, "accountPayables": 200},
        {"period": "Q3", "net_receivables": 100,  "inventory": 50,  "accountPayables": 200},
    ]
    result = check_wc_seasonality_flag(bs, revenue_threshold=0.30)
    assert result is True


def test_check_wc_seasonality_flag_low_swing():
    from auto_valuation.model.working_capital import check_wc_seasonality_flag
    bs = [
        {"period": "Q1", "net_receivables": 500, "inventory": 200, "account_payables": 100},
        {"period": "Q3", "net_receivables": 520, "inventory": 200, "account_payables": 100},
    ]
    result = check_wc_seasonality_flag(bs, revenue_threshold=0.30)
    assert result is False


def test_compute_capex_forecast_alias():
    from auto_valuation.model.balance_sheet import compute_capex_forecast, build_capex_forecast
    assert compute_capex_forecast is build_capex_forecast


def test_rollforward_ppe_alias():
    from auto_valuation.model.balance_sheet import rollforward_ppe, build_ppe_rollforward
    assert rollforward_ppe is build_ppe_rollforward


def test_deferred_tax_rollforward_alias():
    from auto_valuation.model.balance_sheet import (
        deferred_tax_rollforward, rollforward_deferred_tax,
    )
    assert deferred_tax_rollforward is rollforward_deferred_tax


def test_goodwill_rollforward_alias():
    from auto_valuation.model.balance_sheet import goodwill_rollforward, rollforward_goodwill
    assert goodwill_rollforward is rollforward_goodwill


def test_intangibles_amortization_rollforward_alias():
    from auto_valuation.model.balance_sheet import (
        intangibles_amortization_rollforward, rollforward_intangibles,
    )
    assert intangibles_amortization_rollforward is rollforward_intangibles


def test_compute_da_forecast_basic():
    from auto_valuation.model.balance_sheet import compute_da_forecast
    result = compute_da_forecast(
        opening_ppe=1000.0,
        opening_intangibles=200.0,
        da_pct_ppe=0.05,
        amort_pct_intangibles=0.10,
    )
    assert result["depreciation"] == pytest.approx(50.0)
    assert result["amortization"] == pytest.approx(20.0)
    assert result["total_da"] == pytest.approx(70.0)


def test_compute_da_forecast_zero_inputs():
    from auto_valuation.model.balance_sheet import compute_da_forecast
    result = compute_da_forecast(0.0, 0.0)
    assert result["total_da"] == 0.0


def test_capex_convergence_to_da_year1():
    from auto_valuation.model.balance_sheet import capex_convergence_to_da
    # At year 5/5, should equal da
    result = capex_convergence_to_da(base_capex=200, base_da=100, convergence_years=5, year=5)
    assert result == pytest.approx(100.0)


def test_capex_convergence_to_da_midpoint():
    from auto_valuation.model.balance_sheet import capex_convergence_to_da
    # At year 2.5/5, midpoint between 200 and 100 = 150
    result = capex_convergence_to_da(base_capex=200, base_da=100, convergence_years=5, year=2)
    # progress = 2/5 = 0.4; result = 200 + 0.4*(100-200) = 200 - 40 = 160
    assert result == pytest.approx(160.0)


def test_capex_convergence_to_da_zero_years():
    from auto_valuation.model.balance_sheet import capex_convergence_to_da
    result = capex_convergence_to_da(200, 100, convergence_years=0, year=1)
    assert result == pytest.approx(100.0)


# ─────────────────────────────────────────────────────────────────────────────
# Phase 2G — debt / interest
# ─────────────────────────────────────────────────────────────────────────────

def test_compute_cost_of_debt_alias():
    from auto_valuation.model.debt import compute_cost_of_debt, historical_cost_of_debt
    assert compute_cost_of_debt is historical_cost_of_debt


def test_compute_interest_expense_basic():
    from auto_valuation.model.debt import compute_interest_expense
    result = compute_interest_expense(avg_ibd=1000.0, cost_of_debt=0.05)
    assert result == pytest.approx(50.0)


def test_compute_interest_expense_zero():
    from auto_valuation.model.debt import compute_interest_expense
    assert compute_interest_expense(0.0, 0.05) == pytest.approx(0.0)


# ─────────────────────────────────────────────────────────────────────────────
# Phase 2H — shares
# ─────────────────────────────────────────────────────────────────────────────

def test_compute_diluted_shares_tsm_alias():
    from auto_valuation.model.shares import compute_diluted_shares_tsm, compute_diluted_shares
    assert compute_diluted_shares_tsm is compute_diluted_shares


def test_compute_diluted_shares_tsm_works():
    from auto_valuation.model.shares import compute_diluted_shares_tsm
    result = compute_diluted_shares_tsm(basic_shares_mm=100.0)
    assert result == pytest.approx(100.0)


# ─────────────────────────────────────────────────────────────────────────────
# Phase 2I — ratios
# ─────────────────────────────────────────────────────────────────────────────

def test_coverage_ratios_alias():
    from auto_valuation.model.ratios import coverage_ratios, compute_coverage_ratios
    assert coverage_ratios is compute_coverage_ratios


def test_compute_ebitda_basic():
    from auto_valuation.model.ratios import compute_ebitda
    assert compute_ebitda(ebit=100.0, da=30.0) == pytest.approx(130.0)


def test_compute_ebitda_zero():
    from auto_valuation.model.ratios import compute_ebitda
    assert compute_ebitda(0.0, 0.0) == 0.0


# ─────────────────────────────────────────────────────────────────────────────
# Phase 3 — WACC aliases
# ─────────────────────────────────────────────────────────────────────────────

def test_compute_unlevered_beta_alias():
    from auto_valuation.assumptions.wacc import compute_unlevered_beta, unlever_beta
    assert compute_unlevered_beta is unlever_beta


def test_compute_relevered_beta_alias():
    from auto_valuation.assumptions.wacc import compute_relevered_beta, relever_beta
    assert compute_relevered_beta is relever_beta


def test_compute_cost_of_equity_alias():
    from auto_valuation.assumptions.wacc import compute_cost_of_equity, cost_of_equity_capm
    assert compute_cost_of_equity is cost_of_equity_capm


def test_wacc_mean_reversion_schedule_alias():
    from auto_valuation.assumptions.wacc import wacc_mean_reversion_schedule, apply_wacc_step_down
    assert wacc_mean_reversion_schedule is apply_wacc_step_down


def test_fetch_size_premium_alias():
    from auto_valuation.data.macro import fetch_size_premium, compute_size_premium
    assert fetch_size_premium is compute_size_premium


def test_fetch_crp_alias():
    from auto_valuation.data.macro import fetch_crp, compute_crp
    assert fetch_crp is compute_crp


def test_fetch_size_premium_returns_float():
    from auto_valuation.data.macro import fetch_size_premium
    # mega-cap: 0% premium
    assert fetch_size_premium(500_000) == pytest.approx(0.0)
    # micro-cap: positive premium
    assert fetch_size_premium(100) > 0.0


# ─────────────────────────────────────────────────────────────────────────────
# Phase 4 — DCF engine
# ─────────────────────────────────────────────────────────────────────────────

def test_compute_terminal_value_gordon_alias():
    from auto_valuation.forecast.terminal_value import (
        compute_terminal_value_gordon, gordon_growth_tv,
    )
    assert compute_terminal_value_gordon is gordon_growth_tv


def test_compute_terminal_value_exit_multiple_alias():
    from auto_valuation.forecast.terminal_value import (
        compute_terminal_value_exit_multiple, exit_multiple_tv,
    )
    assert compute_terminal_value_exit_multiple is exit_multiple_tv


def test_compute_tv_nopat_reinvestment_alias():
    from auto_valuation.forecast.terminal_value import (
        compute_tv_nopat_reinvestment, gordon_growth_tv_from_nopat,
    )
    assert compute_tv_nopat_reinvestment is gordon_growth_tv_from_nopat


def test_compute_pv_terminal_value_alias():
    from auto_valuation.forecast.terminal_value import (
        compute_pv_terminal_value, pv_terminal_value,
    )
    assert compute_pv_terminal_value is pv_terminal_value


def test_compute_pv_ufcfs_mid_year():
    from auto_valuation.forecast.dcf import compute_pv_ufcfs
    ufcfs = [100.0, 100.0, 100.0]
    wacc  = 0.10
    pv    = compute_pv_ufcfs(ufcfs, wacc, mid_year_convention=True)
    # PV = 100/1.10^0.5 + 100/1.10^1.5 + 100/1.10^2.5
    expected = sum(100 / (1.10 ** (t - 0.5)) for t in range(1, 4))
    assert pv == pytest.approx(expected, rel=1e-6)


def test_compute_pv_ufcfs_end_year():
    from auto_valuation.forecast.dcf import compute_pv_ufcfs
    ufcfs = [100.0, 100.0]
    wacc  = 0.10
    pv    = compute_pv_ufcfs(ufcfs, wacc, mid_year_convention=False)
    expected = 100 / 1.10 + 100 / 1.10 ** 2
    assert pv == pytest.approx(expected, rel=1e-6)


def test_compute_pv_ufcfs_empty():
    from auto_valuation.forecast.dcf import compute_pv_ufcfs
    assert compute_pv_ufcfs([], 0.10) == 0.0


def test_compute_enterprise_value_basic():
    from auto_valuation.forecast.dcf import compute_enterprise_value
    ev = compute_enterprise_value(pv_ufcfs=500.0, pv_terminal_value_=1000.0)
    assert ev == pytest.approx(1500.0)


def test_compute_enterprise_value_zero_tv():
    from auto_valuation.forecast.dcf import compute_enterprise_value
    assert compute_enterprise_value(300.0, 0.0) == pytest.approx(300.0)


def test_validate_terminal_roic_alias():
    from auto_valuation.validation.checks import validate_terminal_roic, check_terminal_roic_vs_wacc
    assert validate_terminal_roic is check_terminal_roic_vs_wacc


# ─────────────────────────────────────────────────────────────────────────────
# Phase 5 — comps aliases
# ─────────────────────────────────────────────────────────────────────────────

def test_select_peer_group_alias():
    from auto_valuation.data.peers import select_peer_group, find_peer_group
    assert select_peer_group is find_peer_group


def test_compute_comps_summary_stats_alias():
    from auto_valuation.data.comps import compute_comps_summary_stats, compute_peer_set_stats
    assert compute_comps_summary_stats is compute_peer_set_stats


def test_apply_multiples_to_subject_alias():
    from auto_valuation.data.comps import apply_multiples_to_subject, apply_comps_to_subject
    assert apply_multiples_to_subject is apply_comps_to_subject


# ─────────────────────────────────────────────────────────────────────────────
# Phase 6 — sensitivity aliases
# ─────────────────────────────────────────────────────────────────────────────

def test_build_sensitivity_grid_alias():
    from auto_valuation.sensitivity.analysis import build_sensitivity_grid, wacc_growth_sensitivity
    assert build_sensitivity_grid is wacc_growth_sensitivity


def test_compute_wacc_sensitivity_alias():
    from auto_valuation.sensitivity.analysis import compute_wacc_sensitivity, wacc_growth_sensitivity
    assert compute_wacc_sensitivity is wacc_growth_sensitivity


def test_build_tornado_chart_data_alias():
    from auto_valuation.sensitivity.analysis import build_tornado_chart_data, build_tornado_chart
    assert build_tornado_chart_data is build_tornado_chart


# ─────────────────────────────────────────────────────────────────────────────
# Phase 7 — sector aliases
# ─────────────────────────────────────────────────────────────────────────────

def test_ebitdar_adjustment_alias():
    from auto_valuation.model.sector import ebitdar_adjustment, apply_ebitdar_adjustment
    assert ebitdar_adjustment is apply_ebitdar_adjustment


def test_mining_nav_unsupported_alias():
    from auto_valuation.model.sector import mining_nav_unsupported, mining_company_gate
    assert mining_nav_unsupported is mining_company_gate


# ─────────────────────────────────────────────────────────────────────────────
# Phase 8 — validation additions
# ─────────────────────────────────────────────────────────────────────────────

def test_check_tv_pct_ev_alias():
    from auto_valuation.validation.checks import check_tv_pct_ev, check_tv_pct_of_ev
    assert check_tv_pct_ev is check_tv_pct_of_ev


def test_check_ufcf_sign_all_positive():
    from auto_valuation.validation.checks import check_ufcf_sign
    result = check_ufcf_sign([100.0, 150.0, 200.0])
    assert result.status == "PASS"


def test_check_ufcf_sign_negative_present():
    from auto_valuation.validation.checks import check_ufcf_sign
    result = check_ufcf_sign([-50.0, 100.0, 200.0])
    assert result.status == "PASS"   # negative is valid; always PASS


def test_check_ufcf_sign_empty():
    from auto_valuation.validation.checks import check_ufcf_sign
    result = check_ufcf_sign([])
    assert result.status == "PASS"


def test_validate_terminal_roic_pass():
    from auto_valuation.validation.checks import validate_terminal_roic
    result = validate_terminal_roic(terminal_roic=0.20, wacc=0.10)
    assert result.status == "PASS"


def test_validate_terminal_roic_warn():
    from auto_valuation.validation.checks import validate_terminal_roic
    result = validate_terminal_roic(terminal_roic=0.08, wacc=0.10)
    assert result.status == "WARN"


# ─────────────────────────────────────────────────────────────────────────────
# Phase 9 — output writers
# ─────────────────────────────────────────────────────────────────────────────

def test_write_sensitivity_sheet_creates_sheet():
    import openpyxl
    from auto_valuation.output.excel_writer import write_sensitivity_sheet
    wb = openpyxl.Workbook()
    grid = {
        "wacc_steps":   [0.08, 0.09, 0.10],
        "growth_steps": [0.02, 0.025, 0.03],
        "grid": [[100, 110, 120], [90, 100, 110], [80, 90, 100]],
    }
    write_sensitivity_sheet(wb, grid)
    assert "Sensitivity" in wb.sheetnames
    ws = wb["Sensitivity"]
    assert ws.cell(row=1, column=1).value == "WACC \\ g"
    assert ws.cell(row=1, column=2).value == "2.0%"


def test_write_sensitivity_sheet_no_data():
    import openpyxl
    from auto_valuation.output.excel_writer import write_sensitivity_sheet
    wb = openpyxl.Workbook()
    write_sensitivity_sheet(wb, {})
    assert "Sensitivity" in wb.sheetnames


def test_write_validation_sheet_creates_sheet():
    import openpyxl
    from auto_valuation.output.excel_writer import write_validation_sheet
    from auto_valuation.validation.checks import ValidationResult
    wb = openpyxl.Workbook()
    vrs = [
        ValidationResult(name="check_wacc", status="PASS", message="OK"),
        ValidationResult(name="check_tv",   status="WARN", message="High TV"),
    ]
    write_validation_sheet(wb, vrs)
    assert "Validation" in wb.sheetnames
    ws = wb["Validation"]
    assert ws.cell(row=1, column=1).value == "Check"
    assert ws.cell(row=2, column=1).value == "check_wacc"
    assert ws.cell(row=2, column=2).value == "PASS"


def test_write_validation_sheet_dict_input():
    import openpyxl
    from auto_valuation.output.excel_writer import write_validation_sheet
    wb = openpyxl.Workbook()
    vrs = [{"name": "x", "status": "PASS", "message": "msg"}]
    write_validation_sheet(wb, vrs)
    ws = wb["Validation"]
    assert ws.cell(row=2, column=1).value == "x"


def test_write_validation_sheet_empty():
    import openpyxl
    from auto_valuation.output.excel_writer import write_validation_sheet
    wb = openpyxl.Workbook()
    write_validation_sheet(wb, [])
    assert "Validation" in wb.sheetnames


def test_write_tornado_chart_creates_sheet():
    import openpyxl
    from auto_valuation.output.tornado_chart import write_tornado_chart
    wb = openpyxl.Workbook()
    bars = [
        {"driver": "WACC",       "impact_low": -10.0, "impact_high": 12.0},
        {"driver": "Rev Growth", "impact_low": -8.0,  "impact_high": 9.0},
    ]
    write_tornado_chart(wb, bars)
    assert "Tornado" in wb.sheetnames
    ws = wb["Tornado"]
    assert ws.cell(row=1, column=1).value == "Driver"
    assert ws.cell(row=2, column=1).value == "WACC"


def test_write_tornado_chart_empty():
    import openpyxl
    from auto_valuation.output.tornado_chart import write_tornado_chart
    wb = openpyxl.Workbook()
    write_tornado_chart(wb, [])
    assert "Tornado" in wb.sheetnames
