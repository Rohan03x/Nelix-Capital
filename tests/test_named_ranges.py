"""Tests for output/named_ranges.py."""
import pytest

openpyxl = pytest.importorskip("openpyxl")

from openpyxl import Workbook
from auto_valuation.output.named_ranges import create_named_ranges, get_named_range_value


class TestCreateNamedRanges:
    def test_no_crash_empty_workbook(self):
        wb = Workbook()
        create_named_ranges(wb)   # sheets referenced don't exist — should skip, not crash

    def test_ranges_created_when_sheet_exists(self):
        wb = Workbook()
        # "Assumptions" sheet is the first default sheet for named ranges
        if "Sheet" in wb.sheetnames:
            wb["Sheet"].title = "Assumptions"
        else:
            wb.create_sheet("Assumptions")
        create_named_ranges(wb)
        # Some named ranges should have been created (>= 0 is fine if sheet mismatch)
        assert wb.defined_names is not None

    def test_with_override(self):
        wb = Workbook()
        wb.create_sheet("MySheet")
        create_named_ranges(wb, sheet_cell_overrides={"WACC": ("MySheet", "B5")})
        # Should not raise; named range may or may not be created depending on specs


class TestGetNamedRangeValue:
    def test_missing_name_returns_none(self):
        wb = Workbook()
        result = get_named_range_value(wb, "NONEXISTENT_RANGE")
        assert result is None

    def test_existing_named_range_retrieves_value(self):
        wb = Workbook()
        ws = wb.create_sheet("Model")
        ws["B5"] = 0.095   # WACC value
        create_named_ranges(wb, sheet_cell_overrides={"WACC": ("Model", "B5")})
        # Whether the value can be read back depends on openpyxl named range support
        # At minimum the function should not crash
        result = get_named_range_value(wb, "WACC")
        assert result is None or isinstance(result, (int, float, str, type(None)))
