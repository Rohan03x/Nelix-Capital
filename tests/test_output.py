"""
tests/test_output.py — Unit tests for the output layer

Phase 10 — Output Layer:
  output/report.py  : format_valuation_summary, print_valuation_summary (no-crash),
                      write_json_output (valid JSON, tuple-key handling)
  output/excel.py   : write_excel_output smoke test (writes file, readable,
                      correct sheet names, numeric values present)

No live API calls.  All fixtures are built in-memory.
"""

from __future__ import annotations

import io
import json
import os
import tempfile
from pathlib import Path

import openpyxl
import pytest

from auto_valuation.forecast.dcf import DCFResult, ForecastYear
from auto_valuation.output.excel import write_excel_output
from auto_valuation.output.report import (
    format_valuation_summary,
    print_valuation_summary,
    write_json_output,
)


# ─────────────────────────────────────────────────────────────────────────────
# Shared fixtures
# ─────────────────────────────────────────────────────────────────────────────

def _make_forecast_year(year: int, revenue: float = 55_000) -> ForecastYear:
    return ForecastYear(
        year=year,
        revenue=revenue,
        ebit_margin=0.15,
        ebit=revenue * 0.15,
        tax_rate=0.21,
        nopat=revenue * 0.15 * 0.79,
        da=revenue * 0.04,
        capex=revenue * 0.04,
        nowc=revenue * 0.06,
        delta_nowc=200.0,
        ufcf=revenue * 0.15 * 0.79 - 200.0,
        discount_factor=1 / (1.09 ** (year - 0.5)),
        pv_ufcf=(revenue * 0.15 * 0.79 - 200.0) / (1.09 ** (year - 0.5)),
    )


def _make_dcf_result(warnings: list[str] | None = None) -> DCFResult:
    years = [_make_forecast_year(i, 50_000 + i * 2_000) for i in range(1, 11)]
    pv_ufcfs = sum(y.pv_ufcf for y in years)
    return DCFResult(
        ticker="ACME",
        scenario="base",
        forecast_years_data=years,
        terminal_ufcf=8_000.0,
        terminal_value_ggm=114_285.0,
        terminal_value_em=110_000.0,
        pv_terminal_value=48_000.0,
        tv_pct_of_ev=0.62,
        pv_ufcfs=pv_ufcfs,
        enterprise_value=pv_ufcfs + 48_000.0,
        wacc=0.09,
        terminal_growth=0.03,
        tax_rate=0.21,
        forecast_years=10,
        warnings=warnings or [],
    )


_FOOTBALL_FIELD = [
    {"method": "DCF", "ev_low_mm": 80_000, "ev_high_mm": 100_000,
     "price_low": 145.0, "price_high": 185.0},
    {"method": "Comps EV/EBITDA", "ev_low_mm": 75_000, "ev_high_mm": 95_000,
     "price_low": 135.0, "price_high": 175.0},
]

_SCENARIO_TABLE = [
    {"scenario": "bull",  "enterprise_value": 120_000, "equity_value": 115_000,
     "price_per_share": 210.0, "wacc": 0.08, "terminal_growth": 0.035},
    {"scenario": "base",  "enterprise_value": 90_000,  "equity_value": 85_000,
     "price_per_share": 155.0, "wacc": 0.09, "terminal_growth": 0.030},
    {"scenario": "bear",  "enterprise_value": 65_000,  "equity_value": 60_000,
     "price_per_share": 110.0, "wacc": 0.10, "terminal_growth": 0.025},
]

_ASSUMPTIONS = {
    "wacc": 0.09,
    "terminal_growth": 0.03,
    "near_term_growth": 0.10,
    "target_ebit_margin": 0.18,
    "forecast_years": 10,
    "ticker": "ACME",
}


# ─────────────────────────────────────────────────────────────────────────────
# 1 — format_valuation_summary
# ─────────────────────────────────────────────────────────────────────────────

class TestFormatValuationSummary:
    @pytest.fixture
    def dcf(self):
        return _make_dcf_result()

    def test_returns_dict(self, dcf):
        result = format_valuation_summary("ACME", dcf, net_debt=5_000, shares_mm=500)
        assert isinstance(result, dict)

    def test_ticker_in_summary(self, dcf):
        result = format_valuation_summary("ACME", dcf, net_debt=5_000, shares_mm=500)
        assert result["ticker"] == "ACME"

    def test_enterprise_value_preserved(self, dcf):
        result = format_valuation_summary("ACME", dcf, net_debt=5_000, shares_mm=500)
        assert result["enterprise_value"] == pytest.approx(dcf.enterprise_value)

    def test_equity_value_computed(self, dcf):
        result = format_valuation_summary("ACME", dcf, net_debt=5_000, shares_mm=500)
        assert result["equity_value"] == pytest.approx(dcf.enterprise_value - 5_000)

    def test_intrinsic_price_per_share(self, dcf):
        result = format_valuation_summary("ACME", dcf, net_debt=5_000, shares_mm=500)
        expected = (dcf.enterprise_value - 5_000) / 500
        assert result["intrinsic_price"] == pytest.approx(expected)

    def test_implied_upside_with_current_price(self, dcf):
        result = format_valuation_summary(
            "ACME", dcf, net_debt=5_000, shares_mm=500, current_price=100.0
        )
        assert "implied_upside" in result
        assert isinstance(result["implied_upside"], float)

    def test_no_current_price_upside_zero(self, dcf):
        result = format_valuation_summary("ACME", dcf, net_debt=5_000, shares_mm=500)
        assert result["implied_upside"] == pytest.approx(0.0)

    def test_football_field_included(self, dcf):
        result = format_valuation_summary(
            "ACME", dcf, net_debt=5_000, shares_mm=500,
            football_field=_FOOTBALL_FIELD,
        )
        assert result["football_field"] == _FOOTBALL_FIELD

    def test_scenario_table_included(self, dcf):
        result = format_valuation_summary(
            "ACME", dcf, net_debt=5_000, shares_mm=500,
            scenario_table=_SCENARIO_TABLE,
        )
        assert result["scenarios"] == _SCENARIO_TABLE

    def test_warnings_from_dcf_result(self, dcf):
        dcf_with_warnings = _make_dcf_result(["TV > 80% of EV", "WACC < 7%"])
        result = format_valuation_summary("ACME", dcf_with_warnings, 5_000, 500)
        assert len(result["warnings"]) == 2

    def test_all_required_keys_present(self, dcf):
        result = format_valuation_summary("ACME", dcf, 5_000, 500)
        for key in ("ticker", "enterprise_value", "net_debt", "equity_value",
                    "shares_mm", "intrinsic_price", "wacc", "terminal_growth",
                    "pv_ufcfs", "pv_terminal_value", "tv_pct_of_ev",
                    "terminal_ufcf", "warnings", "football_field", "scenarios"):
            assert key in result, f"Missing key: {key}"

    def test_wacc_from_dcf_result(self, dcf):
        result = format_valuation_summary("ACME", dcf, 5_000, 500)
        assert result["wacc"] == pytest.approx(0.09)

    def test_assumptions_included(self, dcf):
        result = format_valuation_summary(
            "ACME", dcf, 5_000, 500, assumptions=_ASSUMPTIONS
        )
        assert result["assumptions"]["ticker"] == "ACME"


# ─────────────────────────────────────────────────────────────────────────────
# 2 — print_valuation_summary (no-crash tests)
# ─────────────────────────────────────────────────────────────────────────────

class TestPrintValuationSummary:
    @pytest.fixture
    def summary(self):
        dcf = _make_dcf_result()
        return format_valuation_summary(
            "ACME", dcf, net_debt=5_000, shares_mm=500,
            current_price=120.0,
            football_field=_FOOTBALL_FIELD,
            scenario_table=_SCENARIO_TABLE,
        )

    def test_does_not_raise(self, summary, capsys):
        """Printing must not raise any exceptions."""
        print_valuation_summary(summary)

    def test_outputs_to_stdout(self, summary, capsys):
        print_valuation_summary(summary)
        out = capsys.readouterr().out
        assert len(out) > 0

    def test_ticker_in_output(self, summary, capsys):
        print_valuation_summary(summary)
        out = capsys.readouterr().out
        assert "ACME" in out

    def test_minimal_summary_no_crash(self, capsys):
        """Minimal summary with only required keys should not crash."""
        minimal = {"ticker": "TEST"}
        print_valuation_summary(minimal)

    def test_with_warnings_no_crash(self, capsys):
        dcf = _make_dcf_result(["TV > 80% of EV — review terminal assumptions"])
        summary = format_valuation_summary("ACME", dcf, 5_000, 500)
        print_valuation_summary(summary)

    def test_with_football_field_no_crash(self, summary, capsys):
        """Football field section renders without error."""
        print_valuation_summary(summary)

    def test_with_scenario_table_no_crash(self, summary, capsys):
        """Scenario section renders without error."""
        print_valuation_summary(summary)

    def test_no_current_price_no_crash(self, capsys):
        dcf = _make_dcf_result()
        summary = format_valuation_summary("ACME", dcf, 5_000, 500)
        print_valuation_summary(summary)

    def test_empty_warnings_list_no_crash(self, capsys):
        summary = {"ticker": "X", "warnings": []}
        print_valuation_summary(summary)


# ─────────────────────────────────────────────────────────────────────────────
# 3 — write_json_output
# ─────────────────────────────────────────────────────────────────────────────

class TestWriteJsonOutput:
    @pytest.fixture
    def summary(self):
        dcf = _make_dcf_result()
        return format_valuation_summary("ACME", dcf, 5_000, 500,
                                        current_price=120.0,
                                        football_field=_FOOTBALL_FIELD,
                                        scenario_table=_SCENARIO_TABLE,
                                        assumptions=_ASSUMPTIONS)

    def test_creates_file(self, summary, tmp_path):
        out_file = str(tmp_path / "output.json")
        returned = write_json_output(summary, out_file)
        assert Path(returned).exists()

    def test_returns_file_path(self, summary, tmp_path):
        out_file = str(tmp_path / "output.json")
        returned = write_json_output(summary, out_file)
        assert str(returned) == out_file

    def test_valid_json(self, summary, tmp_path):
        out_file = str(tmp_path / "output.json")
        write_json_output(summary, out_file)
        with open(out_file, "r", encoding="utf-8") as fh:
            data = json.load(fh)
        assert isinstance(data, dict)

    def test_ticker_preserved(self, summary, tmp_path):
        out_file = str(tmp_path / "output.json")
        write_json_output(summary, out_file)
        with open(out_file) as fh:
            data = json.load(fh)
        assert data["ticker"] == "ACME"

    def test_numeric_values_preserved(self, summary, tmp_path):
        out_file = str(tmp_path / "output.json")
        write_json_output(summary, out_file)
        with open(out_file) as fh:
            data = json.load(fh)
        assert isinstance(data["enterprise_value"], (int, float))
        assert data["enterprise_value"] > 0

    def test_tuple_values_serialised_as_lists(self, summary, tmp_path):
        """Tuple *values* in a nested list/tuple are converted to lists by _default."""
        summary_with_tuples = dict(summary)
        summary_with_tuples["ranges"] = [(0.08, 0.09, 0.10), (0.02, 0.03, 0.04)]
        out_file = str(tmp_path / "output_tuples.json")
        write_json_output(summary_with_tuples, out_file)
        with open(out_file) as fh:
            data = json.load(fh)
        assert "ranges" in data
        assert isinstance(data["ranges"], list)

    def test_creates_parent_directories(self, summary, tmp_path):
        out_file = str(tmp_path / "subdir" / "nested" / "output.json")
        write_json_output(summary, out_file)
        assert Path(out_file).exists()

    def test_output_is_indented(self, summary, tmp_path):
        out_file = str(tmp_path / "output.json")
        write_json_output(summary, out_file)
        with open(out_file) as fh:
            raw = fh.read()
        assert "\n" in raw   # indented → multiline


# ─────────────────────────────────────────────────────────────────────────────
# 4 — write_excel_output  (smoke tests)
# ─────────────────────────────────────────────────────────────────────────────

class TestWriteExcelOutput:
    @pytest.fixture
    def dcf(self):
        return _make_dcf_result()

    def test_creates_excel_file(self, dcf, tmp_path):
        out = str(tmp_path / "valuation.xlsx")
        write_excel_output(out, "ACME", dcf, net_debt=5_000, shares_mm=500)
        assert Path(out).exists()

    def test_returns_output_path(self, dcf, tmp_path):
        out = str(tmp_path / "valuation.xlsx")
        returned = write_excel_output(out, "ACME", dcf, net_debt=5_000, shares_mm=500)
        assert returned == out

    def test_file_is_readable_xlsx(self, dcf, tmp_path):
        out = str(tmp_path / "valuation.xlsx")
        write_excel_output(out, "ACME", dcf, net_debt=5_000, shares_mm=500)
        wb = openpyxl.load_workbook(out)
        assert wb is not None
        wb.close()

    def test_summary_sheet_exists(self, dcf, tmp_path):
        out = str(tmp_path / "valuation.xlsx")
        write_excel_output(out, "ACME", dcf, net_debt=5_000, shares_mm=500)
        wb = openpyxl.load_workbook(out)
        assert "Summary" in wb.sheetnames
        wb.close()

    def test_dcf_sheet_exists(self, dcf, tmp_path):
        out = str(tmp_path / "valuation.xlsx")
        write_excel_output(out, "ACME", dcf, net_debt=5_000, shares_mm=500)
        wb = openpyxl.load_workbook(out)
        assert "DCF" in wb.sheetnames
        wb.close()

    def test_sensitivity_sheet_exists(self, dcf, tmp_path):
        out = str(tmp_path / "valuation.xlsx")
        write_excel_output(out, "ACME", dcf, net_debt=5_000, shares_mm=500)
        wb = openpyxl.load_workbook(out)
        assert "Sensitivity" in wb.sheetnames
        wb.close()

    def test_comps_sheet_created_when_data_provided(self, dcf, tmp_path):
        out = str(tmp_path / "valuation.xlsx")
        peer_data = [{"ticker": "PEER1", "ev": 80_000, "ev_ebitda_ltm": 10.5}]
        write_excel_output(out, "ACME", dcf, net_debt=5_000, shares_mm=500,
                           peer_multiples=peer_data)
        wb = openpyxl.load_workbook(out)
        assert "Comps" in wb.sheetnames
        wb.close()

    def test_no_comps_sheet_when_no_peer_data(self, dcf, tmp_path):
        out = str(tmp_path / "valuation.xlsx")
        write_excel_output(out, "ACME", dcf, net_debt=5_000, shares_mm=500)
        wb = openpyxl.load_workbook(out)
        assert "Comps" not in wb.sheetnames
        wb.close()

    def test_transactions_sheet_created_when_data_provided(self, dcf, tmp_path):
        out = str(tmp_path / "valuation.xlsx")
        txn_data = [{"target": "TargetCo", "acquirer": "BigCo",
                     "date": "2022-01-01", "ev_mm": 5_000}]
        write_excel_output(out, "ACME", dcf, net_debt=5_000, shares_mm=500,
                           transactions=txn_data)
        wb = openpyxl.load_workbook(out)
        assert "Transactions" in wb.sheetnames
        wb.close()

    def test_assumptions_sheet_created_when_provided(self, dcf, tmp_path):
        out = str(tmp_path / "valuation.xlsx")
        write_excel_output(out, "ACME", dcf, net_debt=5_000, shares_mm=500,
                           assumptions=_ASSUMPTIONS)
        wb = openpyxl.load_workbook(out)
        assert "Assumptions" in wb.sheetnames
        wb.close()

    def test_summary_sheet_contains_ev_value(self, dcf, tmp_path):
        out = str(tmp_path / "valuation.xlsx")
        write_excel_output(out, "ACME", dcf, net_debt=5_000, shares_mm=500)
        wb = openpyxl.load_workbook(out, data_only=True)
        ws = wb["Summary"]
        all_values = [ws.cell(row=r, column=c).value
                      for r in range(1, 40) for c in range(1, 5)]
        numeric_values = [v for v in all_values if isinstance(v, (int, float)) and v > 0]
        assert len(numeric_values) >= 3, "Expected at least 3 positive numeric values in Summary"
        wb.close()

    def test_dcf_sheet_has_10_year_columns(self, dcf, tmp_path):
        out = str(tmp_path / "valuation.xlsx")
        write_excel_output(out, "ACME", dcf, net_debt=5_000, shares_mm=500)
        wb = openpyxl.load_workbook(out, data_only=True)
        ws = wb["DCF"]
        header_row = [ws.cell(row=2, column=c).value for c in range(1, 15)]
        year_headers = [v for v in header_row if isinstance(v, str) and "Year" in v]
        assert len(year_headers) == 10
        wb.close()

    def test_with_football_field_no_crash(self, dcf, tmp_path):
        out = str(tmp_path / "valuation.xlsx")
        write_excel_output(out, "ACME", dcf, net_debt=5_000, shares_mm=500,
                           football_field=_FOOTBALL_FIELD)
        assert Path(out).exists()

    def test_with_scenario_table_no_crash(self, dcf, tmp_path):
        out = str(tmp_path / "valuation.xlsx")
        write_excel_output(out, "ACME", dcf, net_debt=5_000, shares_mm=500,
                           scenario_table=_SCENARIO_TABLE)
        assert Path(out).exists()

    def test_creates_parent_directory(self, dcf, tmp_path):
        out = str(tmp_path / "subdir" / "output.xlsx")
        write_excel_output(out, "ACME", dcf, net_debt=5_000, shares_mm=500)
        assert Path(out).exists()

    def test_with_warnings_in_dcf_result_no_crash(self, tmp_path):
        dcf_warn = _make_dcf_result(["TV > 75% of EV"])
        out = str(tmp_path / "valuation.xlsx")
        write_excel_output(out, "ACME", dcf_warn, net_debt=5_000, shares_mm=500)
        assert Path(out).exists()
