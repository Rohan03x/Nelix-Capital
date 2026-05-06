from __future__ import annotations

import copy
import html
import io
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

import auto_valuation.learning.background_runner as background_runner_module
from auto_valuation.learning._layered_calibrator import CalibrationObservation
from auto_valuation.learning.cross_industry import AnalogObservation
from auto_valuation.learning.discovery import DiscoveryStore
from auto_valuation.learning.feature_space import build_symbol_features
from auto_valuation.learning.universe import SymbolUniverseStore
from webapp.data import eodhd_client
from webapp.data import knowledge_model as knowledge_model_module
from webapp.data.excel_export import build_excel_bytes
from webapp.data.reverse_dcf import compute_reverse_dcf
from webapp.data.samples import _NKE, _apply_overrides


def _sample_dashboard_data() -> dict:
    data = copy.deepcopy(_NKE)
    hist = data["historical"]
    revenues = hist["revenue"]
    gross_margin = hist["gross_margin"]
    gross_profit = [round(revenue * margin / 100) for revenue, margin in zip(revenues, gross_margin)]
    cogs = [revenue - profit for revenue, profit in zip(revenues, gross_profit)]

    cash = [3500, 3800, 4200, 4300, 4500, 5200, 6900, 6200, 7800, 7464]
    total_assets = [12800, 14000, 13900, 15300, 16200, 17100, 19600, 20700, 22700, 23000]
    equity = [4700, 5600, 6100, 1900, 2000, 2100, 5600, 5500, 5600, 5400]
    receivables = [round(revenue * 32.4 / 365) for revenue in revenues]
    inventory = [round(cost * 92.8 / 365) for cost in cogs]
    payables = [round(cost * 45.2 / 365) for cost in cogs]
    total_current_assets = [cash[idx] + receivables[idx] + inventory[idx] + 500 for idx in range(len(revenues))]
    total_current_liabilities = [payables[idx] + 3000 for idx in range(len(revenues))]
    net_ppe = [2500, 2600, 2700, 2800, 2900, 3000, 3150, 3250, 3350, 3450]
    retained_earnings = [2200, 2600, 3000, 3200, 3600, 4200, 5200, 6100, 7100, 8100]

    hist.update({
        "gross_profit": gross_profit,
        "cash": cash,
        "total_assets": total_assets,
        "equity": equity,
        "accounts_receivable": receivables,
        "inventory_bs": inventory,
        "accounts_payable": payables,
        "total_current_assets": total_current_assets,
        "total_current_liabilities": total_current_liabilities,
        "net_ppe": net_ppe,
        "retained_earnings": retained_earnings,
        "dividends_paid": [700] * len(revenues),
        "buybacks": [2500] * len(revenues),
        "stock_issued": [0] * len(revenues),
    })
    return data


def _load_sample_workbook():
    return load_workbook(io.BytesIO(build_excel_bytes(_sample_dashboard_data())))["valuation"]


def _total_assets(ws, col: int) -> float:
    return sum(ws.cell(row, col).value for row in (500, 501, 502, 504, 505, 507, 508, 509, 511, 512, 514, 515, 516, 517))


def _total_liabilities(ws, col: int) -> float:
    return sum(ws.cell(row, col).value for row in (520, 521, 522, 523, 524, 525, 526, 528, 529, 530, 531, 532))


def _total_equity(ws, col: int) -> float:
    return sum(ws.cell(row, col).value for row in (535, 536, 537, 538))


def _mock_fundamentals() -> dict:
    return {
        "General": {
            "Name": "Test Co",
            "Exchange": "NYSE",
            "CurrencyCode": "USD",
            "Sector": "Industrials",
            "Industry": "Manufacturing",
            "Description": "Synthetic fundamentals for regression testing.",
            "FiscalYearEnd": "December",
            "CountryName": "United States",
        },
        "Highlights": {
            "EPSEstimateNextYear": 5.0,
            "DividendYield": 0.01,
            "WallStreetTargetPrice": 120.0,
        },
        "Technicals": {
            "Beta": 1.1,
            "52WeekHigh": 120.0,
            "52WeekLow": 80.0,
        },
        "SharesStats": {
            "SharesOutstanding": 1_100_000_000,
            "SharesFloat": 1_050_000_000,
        },
        "AnalystRatings": {
            "StrongBuy": 1,
            "Buy": 1,
            "Hold": 1,
            "Sell": 0,
            "StrongSell": 0,
            "TargetPrice": 110.0,
        },
        "Financials": {
            "Income_Statement": {
                "yearly": {
                    "2024-12-31": {
                        "date": "2024-12-31",
                        "totalRevenue": 1_000_000_000,
                        "grossProfit": 400_000_000,
                        "ebit": 100_000_000,
                        "netIncome": 70_000_000,
                        "incomeBeforeTax": 90_000_000,
                        "taxProvision": 20_000_000,
                        "interestExpense": 10_000_000,
                        "ebitda": 120_000_000,
                    },
                    "2023-12-31": {
                        "date": "2023-12-31",
                        "totalRevenue": 900_000_000,
                        "grossProfit": 360_000_000,
                        "ebit": 95_000_000,
                        "netIncome": 75_000_000,
                        "incomeBeforeTax": 100_000_000,
                        "taxProvision": 25_000_000,
                        "interestExpense": 5_000_000,
                        "ebitda": 112_000_000,
                    },
                }
            },
            "Balance_Sheet": {
                "yearly": {
                    "2024-12-31": {
                        "date": "2024-12-31",
                        "totalAssets": 800_000_000,
                        "totalStockholderEquity": 350_000_000,
                        "shortLongTermDebtTotal": 120_000_000,
                        "cashAndEquivalents": 50_000_000,
                        "inventory": 40_000_000,
                        "netReceivables": 120_000_000,
                        "accountsPayable": 60_000_000,
                        "totalCurrentAssets": 130_000_000,
                        "totalCurrentLiabilities": 100_000_000,
                        "retainedEarnings": 150_000_000,
                        "totalLiab": 450_000_000,
                        "longTermDebtTotal": 100_000_000,
                        "commonStockSharesOutstanding": 1_100_000_000,
                        "propertyPlantEquipment": 250_000_000,
                    },
                    "2023-12-31": {
                        "date": "2023-12-31",
                        "totalAssets": 780_000_000,
                        "totalStockholderEquity": 360_000_000,
                        "shortLongTermDebtTotal": 130_000_000,
                        "cashAndEquivalents": 55_000_000,
                        "inventory": 35_000_000,
                        "netReceivables": 110_000_000,
                        "accountsPayable": 50_000_000,
                        "totalCurrentAssets": 140_000_000,
                        "totalCurrentLiabilities": 100_000_000,
                        "retainedEarnings": 140_000_000,
                        "totalLiab": 420_000_000,
                        "longTermDebtTotal": 110_000_000,
                        "commonStockSharesOutstanding": 1_000_000_000,
                        "propertyPlantEquipment": 240_000_000,
                    },
                }
            },
            "Cash_Flow": {
                "yearly": {
                    "2024-12-31": {
                        "date": "2024-12-31",
                        "totalCashFromOperatingActivities": 110_000_000,
                        "capitalExpenditures": -20_000_000,
                        "freeCashFlow": 90_000_000,
                        "depreciation": 15_000_000,
                        "stockBasedCompensation": 5_000_000,
                        "salePurchaseOfStock": -10_000_000,
                        "dividendsPaid": -5_000_000,
                        "netBorrowings": -10_000_000,
                    },
                    "2023-12-31": {
                        "date": "2023-12-31",
                        "totalCashFromOperatingActivities": 120_000_000,
                        "capitalExpenditures": -18_000_000,
                        "freeCashFlow": 102_000_000,
                        "depreciation": 14_000_000,
                        "stockBasedCompensation": 4_000_000,
                        "salePurchaseOfStock": -8_000_000,
                        "dividendsPaid": -4_000_000,
                        "netBorrowings": -5_000_000,
                    },
                }
            },
        },
    }


def _mock_high_margin_fundamentals() -> dict:
    data = copy.deepcopy(_mock_fundamentals())
    data["General"]["Industry"] = "Semiconductors"
    data["Financials"]["Income_Statement"]["yearly"]["2024-12-31"].update({
        "grossProfit": 850_000_000,
        "ebit": 650_000_000,
        "netIncome": 520_000_000,
        "incomeBeforeTax": 600_000_000,
        "taxProvision": 80_000_000,
        "interestExpense": 50_000_000,
        "ebitda": 680_000_000,
    })
    data["Financials"]["Income_Statement"]["yearly"]["2023-12-31"].update({
        "grossProfit": 760_000_000,
        "ebit": 620_000_000,
        "netIncome": 500_000_000,
        "incomeBeforeTax": 580_000_000,
        "taxProvision": 80_000_000,
        "interestExpense": 40_000_000,
        "ebitda": 650_000_000,
    })
    data["Financials"]["Cash_Flow"]["yearly"]["2024-12-31"].update({
        "totalCashFromOperatingActivities": 560_000_000,
        "freeCashFlow": 530_000_000,
        "capitalExpenditures": -30_000_000,
    })
    data["Financials"]["Cash_Flow"]["yearly"]["2023-12-31"].update({
        "totalCashFromOperatingActivities": 540_000_000,
        "freeCashFlow": 515_000_000,
        "capitalExpenditures": -25_000_000,
    })
    return data


def _mock_cross_currency_fundamentals() -> dict:
    data = copy.deepcopy(_mock_fundamentals())
    data["General"]["Exchange"] = "LSE"
    data["General"]["CurrencyCode"] = "GBX"
    for section in data["Financials"].values():
        for period in section["yearly"].values():
            period["currency_symbol"] = "USD"
    return data


def _mock_holdco_fundamentals() -> dict:
    data = copy.deepcopy(_mock_fundamentals())
    yearly_is = data["Financials"]["Income_Statement"]["yearly"]
    yearly_cf = data["Financials"]["Cash_Flow"]["yearly"]
    yearly_bs = data["Financials"]["Balance_Sheet"]["yearly"]

    yearly_is["2024-12-31"].update({
        "totalRevenue": 80_000_000_000,
        "grossProfit": 53_000_000_000,
        "ebit": 17_200_000_000,
        "netIncome": 4_400_000_000,
        "incomeBeforeTax": 6_000_000_000,
        "taxProvision": 1_600_000_000,
        "interestExpense": 500_000_000,
        "ebitda": 20_000_000_000,
    })
    yearly_is["2023-12-31"].update({
        "totalRevenue": 76_000_000_000,
        "grossProfit": 50_000_000_000,
        "ebit": 16_400_000_000,
        "netIncome": 4_300_000_000,
        "incomeBeforeTax": 5_900_000_000,
        "taxProvision": 1_600_000_000,
        "interestExpense": 500_000_000,
        "ebitda": 19_000_000_000,
    })
    yearly_is["2022-12-31"] = {
        "date": "2022-12-31",
        "totalRevenue": 72_000_000_000,
        "grossProfit": 47_000_000_000,
        "ebit": 15_100_000_000,
        "netIncome": 4_100_000_000,
        "incomeBeforeTax": 5_500_000_000,
        "taxProvision": 1_400_000_000,
        "interestExpense": 450_000_000,
        "ebitda": 17_500_000_000,
    }
    yearly_is["2021-12-31"] = {
        "date": "2021-12-31",
        "totalRevenue": 67_000_000_000,
        "grossProfit": 43_000_000_000,
        "ebit": 13_800_000_000,
        "netIncome": 3_800_000_000,
        "incomeBeforeTax": 5_100_000_000,
        "taxProvision": 1_300_000_000,
        "interestExpense": 420_000_000,
        "ebitda": 16_200_000_000,
    }

    yearly_cf["2024-12-31"].update({
        "totalCashFromOperatingActivities": 11_000_000_000,
        "capitalExpenditures": -2_700_000_000,
        "freeCashFlow": 8_300_000_000,
        "depreciation": 6_300_000_000,
        "stockBasedCompensation": 0,
        "salePurchaseOfStock": 0,
        "dividendsPaid": -5_500_000_000,
    })
    yearly_cf["2023-12-31"].update({
        "totalCashFromOperatingActivities": 10_600_000_000,
        "capitalExpenditures": -2_500_000_000,
        "freeCashFlow": 8_100_000_000,
        "depreciation": 6_000_000_000,
        "stockBasedCompensation": 0,
        "salePurchaseOfStock": 0,
        "dividendsPaid": -5_300_000_000,
    })
    yearly_cf["2022-12-31"] = {
        "date": "2022-12-31",
        "totalCashFromOperatingActivities": 10_100_000_000,
        "capitalExpenditures": -2_400_000_000,
        "freeCashFlow": 7_700_000_000,
        "depreciation": 5_700_000_000,
        "stockBasedCompensation": 0,
        "salePurchaseOfStock": 0,
        "dividendsPaid": -5_100_000_000,
        "netBorrowings": 0,
    }
    yearly_cf["2021-12-31"] = {
        "date": "2021-12-31",
        "totalCashFromOperatingActivities": 9_700_000_000,
        "capitalExpenditures": -2_300_000_000,
        "freeCashFlow": 7_400_000_000,
        "depreciation": 5_400_000_000,
        "stockBasedCompensation": 0,
        "salePurchaseOfStock": 0,
        "dividendsPaid": -4_900_000_000,
        "netBorrowings": 0,
    }

    yearly_bs["2024-12-31"].update({
        "totalAssets": 139_000_000_000,
        "totalStockholderEquity": 24_500_000_000,
        "shortLongTermDebtTotal": 49_700_000_000,
        "cashAndEquivalents": 8_900_000_000,
        "cashAndShortTermInvestments": 13_700_000_000,
        "shortTermInvestments": 4_800_000_000,
        "inventory": 22_600_000_000,
        "netReceivables": 5_100_000_000,
        "accountsPayable": 8_200_000_000,
        "totalCurrentAssets": 48_300_000_000,
        "totalCurrentLiabilities": 30_400_000_000,
        "retainedEarnings": 4_500_000_000,
        "totalLiab": 72_700_000_000,
        "longTermDebtTotal": 25_800_000_000,
        "commonStockSharesOutstanding": 180_410_580,
        "propertyPlantEquipment": 43_900_000_000,
    })
    yearly_bs["2023-12-31"].update({
        "totalAssets": 141_000_000_000,
        "totalStockholderEquity": 21_500_000_000,
        "shortLongTermDebtTotal": 38_500_000_000,
        "cashAndEquivalents": 7_900_000_000,
        "cashAndShortTermInvestments": 11_500_000_000,
        "shortTermInvestments": 3_600_000_000,
        "inventory": 23_300_000_000,
        "netReceivables": 5_600_000_000,
        "accountsPayable": 9_000_000_000,
        "totalCurrentAssets": 44_000_000_000,
        "totalCurrentLiabilities": 33_200_000_000,
        "retainedEarnings": 6_300_000_000,
        "totalLiab": 80_600_000_000,
        "longTermDebtTotal": 23_100_000_000,
        "commonStockSharesOutstanding": 180_410_580,
        "propertyPlantEquipment": 42_000_000_000,
    })
    yearly_bs["2022-12-31"] = {
        "date": "2022-12-31",
        "totalAssets": 132_000_000_000,
        "totalStockholderEquity": 6_100_000_000,
        "shortLongTermDebtTotal": 35_200_000_000,
        "cashAndEquivalents": 7_600_000_000,
        "cashAndShortTermInvestments": 11_200_000_000,
        "shortTermInvestments": 3_600_000_000,
        "inventory": 20_300_000_000,
        "netReceivables": 8_500_000_000,
        "accountsPayable": 8_800_000_000,
        "totalCurrentAssets": 40_100_000_000,
        "totalCurrentLiabilities": 31_600_000_000,
        "retainedEarnings": 5_800_000_000,
        "totalLiab": 77_600_000_000,
        "longTermDebtTotal": 23_200_000_000,
        "commonStockSharesOutstanding": 180_410_580,
        "propertyPlantEquipment": 37_000_000_000,
    }
    yearly_bs["2021-12-31"] = {
        "date": "2021-12-31",
        "totalAssets": 128_000_000_000,
        "totalStockholderEquity": 3_000_000_000,
        "shortLongTermDebtTotal": 36_800_000_000,
        "cashAndEquivalents": 8_100_000_000,
        "cashAndShortTermInvestments": 10_700_000_000,
        "shortTermInvestments": 2_600_000_000,
        "inventory": 18_900_000_000,
        "netReceivables": 7_300_000_000,
        "accountsPayable": 7_600_000_000,
        "totalCurrentAssets": 37_500_000_000,
        "totalCurrentLiabilities": 32_000_000_000,
        "retainedEarnings": 2_700_000_000,
        "totalLiab": 78_200_000_000,
        "longTermDebtTotal": 23_100_000_000,
        "commonStockSharesOutstanding": 180_410_580,
        "propertyPlantEquipment": 35_000_000_000,
    }
    return data


def test_apply_overrides_recenters_sensitivity_table():
    data = copy.deepcopy(_NKE)
    _apply_overrides(data, {"wacc": 9.9, "g": 2.0})
    sens = data["sensitivity"]
    center = sens["iv_grid"][sens["base_wacc_idx"]][sens["base_g_idx"]]
    assert abs(data["intrinsic_value"] - center) < 1.0


def test_reverse_dcf_matches_base_case_value():
    reverse = compute_reverse_dcf(copy.deepcopy(_NKE))
    assert abs(reverse["iv_at_model_g"] - _NKE["intrinsic_value"]) < 1.0


def test_excel_forecast_rows_stay_consistent():
    ws = _load_sample_workbook()

    for col in range(12, 19):
        revenue = ws.cell(337, col).value
        cogs = ws.cell(339, col).value
        sga = ws.cell(342, col).value
        da = ws.cell(344, col).value
        ebit = ws.cell(380, col).value
        operating_income = revenue - cogs - sga - da
        assert operating_income == pytest.approx(ebit)
        assert _total_assets(ws, col) == pytest.approx(_total_liabilities(ws, col) + _total_equity(ws, col), abs=1.0)

    assert ws.cell(22, 1).value == "Free Cash Flow (Levered Hist.) / UFCF (Fcst) ($M)"
    assert ws.cell(33, 11).value.startswith("=IFERROR(")
    assert ws.cell(348, 12).value < 0
    assert any(ws.cell(row, 12).value != 0 for row in (441, 442, 443))
    assert ws.cell(537, 12).value != 0


def test_eodhd_build_uses_prior_year_liquidity_and_share_inputs(monkeypatch):
    monkeypatch.setattr(eodhd_client, "_fetch_price", lambda _code: {"close": 100.0})
    monkeypatch.setattr(eodhd_client, "_fetch_fundamentals", lambda _code: _mock_fundamentals())
    monkeypatch.setattr(eodhd_client, "_get_risk_free_rate", lambda: 4.0)
    monkeypatch.setattr(eodhd_client, "_PEERS_AVAILABLE", False)

    data = eodhd_client.build_dashboard_data("TEST")
    tests = {test["name"]: test for test in data["financial_scores"]["piotroski_f"]["tests"]}

    assert tests["Improving Current Ratio"]["pass"] is False
    assert tests["No Share Dilution"]["pass"] is False
    assert data["historical"]["shares"] == [1000.0, 1100.0]


def test_eodhd_build_falls_back_to_balance_sheet_shares_when_share_stats_missing(monkeypatch):
    fundamentals = copy.deepcopy(_mock_fundamentals())
    fundamentals["SharesStats"] = {}

    monkeypatch.setattr(eodhd_client, "_fetch_price", lambda _code: {"close": 100.0})
    monkeypatch.setattr(eodhd_client, "_fetch_fundamentals", lambda _code: fundamentals)
    monkeypatch.setattr(eodhd_client, "_get_risk_free_rate", lambda: 4.0)
    monkeypatch.setattr(eodhd_client, "_PEERS_AVAILABLE", False)

    data = eodhd_client.build_dashboard_data("TEST")

    assert data is not None
    assert data["diluted_shares"] == pytest.approx(1100.0)
    assert data["historical"]["shares"] == [1000.0, 1100.0]


def test_eodhd_build_surfaces_knowledge_model_payload(monkeypatch):
    monkeypatch.setattr(eodhd_client, "_fetch_price", lambda _code: {"close": 100.0})
    monkeypatch.setattr(eodhd_client, "_fetch_fundamentals", lambda _code: _mock_fundamentals())
    monkeypatch.setattr(eodhd_client, "_get_risk_free_rate", lambda: 4.0)
    monkeypatch.setattr(eodhd_client, "_PEERS_AVAILABLE", False)
    monkeypatch.setattr(knowledge_model_module, "fetch_damodaran_industry_beta", lambda _sector: 0.9)

    observations = [
        CalibrationObservation(
            sector="Industrials",
            industry="Manufacturing",
            data_vintage_years=2,
            market_cap_regime="large",
            macro_regime="neutral",
            predicted_revenue_growth=0.05,
            actual_revenue_growth=0.07,
            predicted_ebit_margin=0.11,
            actual_ebit_margin=0.12,
            predicted_wacc=0.09,
            actual_wacc=0.088,
            predicted_terminal_growth=0.025,
            actual_terminal_growth=0.026,
            predicted_beta=0.95,
            actual_beta=0.9,
        )
        for _ in range(5)
    ] + [
        CalibrationObservation(
            sector="Technology",
            industry="Software",
            data_vintage_years=2,
            market_cap_regime="large",
            macro_regime="neutral",
            predicted_revenue_growth=0.06,
            actual_revenue_growth=0.065,
            predicted_ebit_margin=0.16,
            actual_ebit_margin=0.17,
            predicted_wacc=0.095,
            actual_wacc=0.09,
            predicted_terminal_growth=0.03,
            actual_terminal_growth=0.03,
            predicted_beta=1.05,
            actual_beta=1.0,
        ),
        CalibrationObservation(
            sector="Consumer Staples",
            industry="Beverages",
            data_vintage_years=2,
            market_cap_regime="large",
            macro_regime="neutral",
            predicted_revenue_growth=0.04,
            actual_revenue_growth=0.045,
            predicted_ebit_margin=0.14,
            actual_ebit_margin=0.145,
            predicted_wacc=0.085,
            actual_wacc=0.082,
            predicted_terminal_growth=0.022,
            actual_terminal_growth=0.023,
            predicted_beta=0.8,
            actual_beta=0.78,
        ),
    ]
    monkeypatch.setattr(knowledge_model_module, "_load_learning_cohort", lambda limit=500: observations)
    analog_features = build_symbol_features(
        ticker="GRAPH",
        sector="Industrials",
        industry="Manufacturing",
        revenues=[720.0, 780.0, 845.0, 915.0, 980.0, 1_040.0],
        ebit_margins=[9.2, 9.8, 10.4, 11.1, 11.6, 12.0],
        gross_margin_base_pct=41.0,
        capex_pct=3.1,
        total_assets=1_260.0,
        total_debt=180.0,
        revenue_base=1_040.0,
        operating_cf=210.0,
        fcf=150.0,
        da_pct=1.8,
        tax_rate_pct=22.0,
        market_cap=9_500.0,
        observation_year=2024,
    )
    monkeypatch.setattr(
        knowledge_model_module,
        "_load_analog_candidates",
        lambda limit=500: [
            AnalogObservation(
                ticker="GRAPH",
                sector="Industrials",
                industry="Manufacturing",
                vintage_year=6,
                feature_map=dict(analog_features.feature_map),
                outcome_revenue_cagr_5y=0.08,
                outcome_margin_change_bps=90.0,
                outcome_ev_multiple_change=0.5,
                market_cap_regime=analog_features.market_cap_regime,
                data_quality_score=0.88,
                sample_size=analog_features.sample_size,
                predictive_usefulness=0.84,
                as_of_year=2024,
            )
        ],
    )

    data = eodhd_client.build_dashboard_data("TEST")

    assert data["knowledge_model"] is not None
    assert data["knowledge_model"]["summary"].startswith("Weighted knowledge model active")
    assert data["knowledge_model"]["global_learning"]["enabled"] is True
    assert data["revenue_growth_near"] == data["knowledge_model"]["revenue_growth_near"]
    explainability = data["knowledge_model"]["explainability"]
    layered_learning = data["knowledge_model"]["layered_learning"]
    assert explainability["company_memory"]["weight_pct"] > 0
    assert explainability["sector_memory"]["weight_pct"] >= 0
    assert explainability["cohort_memory"]["records"] == data["knowledge_model"]["calibration_cohort_size"]
    assert explainability["global_brain"]["enabled"] is True
    assert "structural_break" in explainability
    assert "uncertainty" in explainability
    assert "learned_metrics" in explainability
    assert "confidence_model" in data["knowledge_model"]
    assert data["knowledge_model"]["confidence_model"]["valuation_confidence"]["score"] > 0
    assert data["knowledge_model"]["confidence_model"]["assumption_confidence"]["score"] > 0
    assert data["knowledge_model"]["confidence_model"]["valuation_confidence"]["expected_error_pct"]["p90"] >= data["knowledge_model"]["confidence_model"]["valuation_confidence"]["expected_error_pct"]["p50"]
    assert explainability["confidence_decomposition"]["valuation_confidence"]["score_100"] == data["knowledge_model"]["confidence_model"]["valuation_confidence"]["score_100"]
    assert layered_learning["layer_mix"]["global_memory"]["enabled"] is True
    assert layered_learning["uncertainty"]["scenario_width_multiplier"] == data["knowledge_model"]["scenario_width_multiplier"]
    assert explainability["realized_evidence"]["matured_records"] == 0
    assert explainability["maintenance"]["reason"] == "disabled"
    assert explainability["current_snapshot"]["reason"] == "disabled"
    assert data["knowledge_model"]["relationship_graph"]["enabled"] is True
    assert data["knowledge_model"]["relationship_graph"]["node_count"] >= 2
    assert explainability["relationship_graph"]["enabled"] is True
    assert data["knowledge_model"]["memory_hierarchy"]["relational"]["score"] > 0
    assert explainability["memory_hierarchy"]["procedural"]["score"] > 0
    assert data["confidence_score"] == data["knowledge_model"]["confidence_model"]["valuation_confidence"]["score_100"]
    assert data["confidence_breakdown"]["total"] == data["confidence_score"]
    assert any(layer["driver"] == "Revenue Growth" for layer in explainability["forecast_layers"])


def test_eodhd_build_scales_holdco_style_cashflows_to_attributable_earnings(monkeypatch):
    monkeypatch.setattr(eodhd_client, "_fetch_price", lambda _code: {"close": 425.2})
    monkeypatch.setattr(eodhd_client, "_fetch_fundamentals", lambda _code: _mock_holdco_fundamentals())
    monkeypatch.setattr(eodhd_client, "_get_risk_free_rate", lambda: 4.0)
    monkeypatch.setattr(eodhd_client, "_PEERS_AVAILABLE", False)
    monkeypatch.setattr(eodhd_client, "fetch_historical_price_series", lambda *args, **kwargs: [])

    data = eodhd_client.build_dashboard_data("HOLDCO")
    monkeypatch.setattr(eodhd_client, "_attributable_earnings_adjustment", lambda _net, _ebit: (1.0, None))
    unadjusted = eodhd_client.build_dashboard_data("HOLDCO")

    assert data["attributable_earnings_adjustment_applied"] is True
    assert data["attributable_earnings_ratio"] == pytest.approx(0.2665, abs=0.02)
    assert data["intrinsic_value"] < unadjusted["intrinsic_value"] * 0.2
    assert data["enterprise_value"] < unadjusted["enterprise_value"] * 0.35
    flags = {flag["name"]: flag for flag in data["flags"]}
    assert "Attributable Earnings" in flags

    assumptions = {row["driver"]: row for row in data["assumptions"]}
    assert assumptions["Revenue Growth (Near-Term)"]["source"].startswith("Knowledge model:")
    assert assumptions["WACC"]["source"].startswith("Knowledge model:")


def test_eodhd_build_surfaces_symbol_brain_analogs(monkeypatch):
    monkeypatch.setattr(eodhd_client, "_fetch_price", lambda _code: {"close": 100.0})
    monkeypatch.setattr(eodhd_client, "_fetch_fundamentals", lambda _code: _mock_fundamentals())
    monkeypatch.setattr(eodhd_client, "_get_risk_free_rate", lambda: 4.0)
    monkeypatch.setattr(eodhd_client, "_PEERS_AVAILABLE", False)
    monkeypatch.setattr(knowledge_model_module, "_load_learning_cohort", lambda limit=500: [])

    analog_features = build_symbol_features(
        ticker="ANLG",
        sector="Consumer Discretionary",
        industry="Internet Retail",
        revenues=[700.0, 810.0, 930.0, 1_070.0, 1_230.0, 1_410.0],
        ebit_margins=[5.0, 6.5, 8.0, 9.5, 11.0, 12.5],
        gross_margin_base_pct=61.0,
        capex_pct=3.6,
        total_assets=980.0,
        total_debt=120.0,
        revenue_base=1_410.0,
        operating_cf=300.0,
        fcf=215.0,
        da_pct=1.8,
        tax_rate_pct=22.0,
        market_cap=13_800.0,
        observation_year=2024,
    )
    monkeypatch.setattr(
        knowledge_model_module,
        "_load_analog_candidates",
        lambda limit=500: [
            AnalogObservation(
                ticker="ANLG",
                sector="Consumer Discretionary",
                industry="Internet Retail",
                vintage_year=6,
                feature_map=dict(analog_features.feature_map),
                outcome_revenue_cagr_5y=0.12,
                outcome_margin_change_bps=125.0,
                outcome_ev_multiple_change=1.0,
                market_cap_regime=analog_features.market_cap_regime,
                data_quality_score=0.92,
                sample_size=analog_features.sample_size,
                predictive_usefulness=0.90,
                as_of_year=2024,
            )
        ],
    )

    data = eodhd_client.build_dashboard_data("TEST")
    knowledge = data["knowledge_model"]

    assert knowledge["symbol_brain"]["summary"]
    assert knowledge["analogs"]["enabled"] is True
    assert knowledge["analogs"]["items"][0]["ticker"] == "ANLG"
    assert knowledge["analogs"]["items"][0]["evidence"]
    assert knowledge["analogs"]["overlay"]["enabled"] is True
    analog_explain = knowledge["explainability"]["analog_evidence"]
    assert analog_explain["enabled"] is True
    assert analog_explain["match_count"] >= 1
    assert analog_explain["top_matches"][0]["ticker"] == "ANLG"
    assert analog_explain["top_matches"][0]["evidence"]
    assert "ANLG" in analog_explain["note"]


def test_health_check_payload_validator_accepts_learning_contract():
    from check import evaluate_dashboard_payload

    payload = copy.deepcopy(_NKE)
    payload["data_source"] = "eodhd"
    payload["data_quality"] = {"source": "EODHD synthetic regression"}
    payload["knowledge_model"] = {
        "summary": "Weighted knowledge model active",
        "global_learning": {"enabled": True, "scope": "regime", "cohort_size": 8, "sector_span": 4, "confidence": 0.72},
        "assumption_weights": {"revenue_growth_near": {"source": "Knowledge model: test"}},
        "layered_learning": {"uncertainty": {"scenario_width_multiplier": 1.4}},
        "calibration_cohort_size": 8,
    }

    report = evaluate_dashboard_payload("NKE", payload, response_ms=240, strict_learning=True)

    assert report["ok"] is True
    assert report["issues"] == []


def test_eodhd_scenarios_expand_with_learning_uncertainty(monkeypatch):
    monkeypatch.setattr(eodhd_client, "_fetch_price", lambda _code: {"close": 100.0})
    monkeypatch.setattr(eodhd_client, "_fetch_fundamentals", lambda _code: _mock_fundamentals())
    monkeypatch.setattr(eodhd_client, "_get_risk_free_rate", lambda: 4.0)
    monkeypatch.setattr(eodhd_client, "_PEERS_AVAILABLE", False)

    def _knowledge_payload(width: float) -> dict:
        return {
            "summary": "Weighted knowledge model active.",
            "global_learning": {"enabled": False, "scope": None, "cohort_size": 0, "sector_span": 0, "confidence": 0.0},
            "assumption_weights": {"revenue_growth_near": {"source": "Knowledge model: test"}},
            "layered_learning": {"uncertainty": {"scenario_width_multiplier": width}},
            "calibration_cohort_size": 5,
            "scenario_width_multiplier": width,
            "revenue_growth_near": 7.0,
            "terminal_growth": 2.5,
            "ebit_margin_target": 12.0,
            "beta": 1.0,
            "wacc": 9.0,
            "tax_rate_pct": 20.0,
            "da_pct": 3.0,
            "capex_pct": 3.5,
            "sbc_pct": 0.5,
            "dso": 35.0,
            "dio": 50.0,
            "dpo": 40.0,
        }

    monkeypatch.setattr(knowledge_model_module, "refine_live_assumptions", lambda **_: _knowledge_payload(1.0))
    narrow = eodhd_client.build_dashboard_data("TEST")

    monkeypatch.setattr(knowledge_model_module, "refine_live_assumptions", lambda **_: _knowledge_payload(2.0))
    wide = eodhd_client.build_dashboard_data("TEST")

    assert wide["scenarios"]["bull"]["rev_growth"] - wide["scenarios"]["base"]["rev_growth"] > narrow["scenarios"]["bull"]["rev_growth"] - narrow["scenarios"]["base"]["rev_growth"]
    assert wide["scenarios"]["bear"]["wacc"] - wide["scenarios"]["base"]["wacc"] > narrow["scenarios"]["bear"]["wacc"] - narrow["scenarios"]["base"]["wacc"]
    assert wide["scenarios"]["bull"]["margin_target"] - wide["scenarios"]["base"]["margin_target"] > narrow["scenarios"]["bull"]["margin_target"] - narrow["scenarios"]["base"]["margin_target"]
    assert wide["recommendation_basis"]["method"] == "learned-scenario-weighted-expected-upside"
    assert wide["learned_expected_upside_pct"] is not None
    assert all("learning_basis" in scenario for scenario in wide["scenarios"].values())
    assert sum(scenario["probability"] for scenario in wide["scenarios"].values()) == pytest.approx(1.0, abs=0.01)


def test_learning_cohort_pins_subject_historical_replay_before_global_cap(monkeypatch):
    subject_record = SimpleNamespace(
        record_id="subject-ledger",
        ticker="LIGHT.AS",
        sector="Industrials",
        industry="Electrical Components",
        data_vintage_years=9,
        market_cap_regime="mid",
        macro_regime="neutral",
        predicted_revenue_mm=100.0,
        actual_revenue_mm=110.0,
        near_term_revenue_growth=0.03,
        predicted_ebit_margin=0.10,
        actual_ebit_margin=0.11,
        predicted_wacc=0.08,
        predicted_terminal_growth=0.025,
        beta=1.0,
        predicted_ufcf_mm=8.0,
        actual_ufcf_mm=9.0,
        capex_pct_revenue=0.03,
        da_pct_revenue=0.02,
        predicted_ev_mm=1000.0,
        predicted_equity_value_mm=900.0,
        predicted_price_per_share=20.0,
        actual_price_at_prediction=18.0,
        actual_price_at_horizon=22.0,
        macro_backdrop={},
        feature_vector=None,
        structural_break_hints=[],
        prediction_context={"source": "historical_replay_bootstrap"},
    )
    subject_historical = [
        CalibrationObservation(
            ticker="LIGHT.AS",
            sector="Industrials",
            industry="Electrical Components",
            data_vintage_years=9,
            market_cap_regime="mid",
            macro_regime="neutral",
            predicted_revenue_growth=0.0,
            actual_revenue_growth=0.02 + idx / 100,
            predicted_ebit_margin=0.10,
            actual_ebit_margin=0.11,
            predicted_wacc=0.08,
            actual_wacc=0.08,
            predicted_terminal_growth=0.025,
            actual_terminal_growth=0.025,
        )
        for idx in range(3)
    ]
    other_historical = [
        CalibrationObservation(
            ticker=f"OTHER{idx}",
            sector="Industrials",
            industry="Electrical Components",
            data_vintage_years=9,
            market_cap_regime="mid",
            macro_regime="neutral",
            predicted_revenue_growth=0.0,
            actual_revenue_growth=0.01,
            predicted_ebit_margin=0.10,
            actual_ebit_margin=0.10,
            predicted_wacc=0.08,
            actual_wacc=0.08,
            predicted_terminal_growth=0.025,
            actual_terminal_growth=0.025,
        )
        for idx in range(6)
    ]

    monkeypatch.setattr(knowledge_model_module, "_cached_ledger_records", lambda limit=None: [subject_record])
    monkeypatch.setattr(knowledge_model_module, "_cached_stratified_sample", lambda records, max_records, target: list(records))
    monkeypatch.setattr(knowledge_model_module.LedgerReader, "query", lambda *_args, **_kwargs: [subject_record])
    monkeypatch.setattr(
        knowledge_model_module,
        "_get_historical_observations",
        lambda: other_historical[:4] + subject_historical + other_historical[4:],
    )
    monkeypatch.setitem(knowledge_model_module.LEARNING_CONFIG, "historical_replay_limit", 2)

    observations = knowledge_model_module._load_learning_cohort(limit=1, subject_ticker="LIGHT.AS")
    subject_count = sum(
        1
        for observation in observations
        if str(knowledge_model_module._obs_value(observation, "ticker", "")).upper() == "LIGHT.AS"
    )

    assert subject_count == 4


def test_learning_cohort_uses_seeded_replay_summary_when_raw_cache_empty(monkeypatch):
    seeded_summary_observation = {
        "ticker": "LIGHT.AS",
        "sector": "Industrials",
        "industry": "Electrical Components",
        "data_vintage_years": 13,
        "market_cap_regime": "mid",
        "macro_regime": "neutral",
        "predicted_revenue_growth": 0.0,
        "actual_revenue_growth": -0.024,
        "predicted_ebit_margin": 0.0,
        "actual_ebit_margin": 0.001,
        "predicted_wacc": 0.084,
        "actual_wacc": 0.084,
        "predicted_terminal_growth": 0.025,
        "actual_terminal_growth": 0.025,
        "evidence_count": 51,
        "observation_type": "deployment_historical_replay_summary",
    }
    monkeypatch.setattr(knowledge_model_module, "_cached_ledger_records", lambda limit=None: [])
    monkeypatch.setattr(knowledge_model_module, "_cached_stratified_sample", lambda records, max_records, target: [])
    monkeypatch.setattr(knowledge_model_module.LedgerReader, "query", lambda *_args, **_kwargs: [])
    monkeypatch.setattr(knowledge_model_module, "_get_historical_observations", lambda: [])
    monkeypatch.setattr(knowledge_model_module, "seeded_cohort_observations", lambda limit=None: [])
    monkeypatch.setattr(
        knowledge_model_module,
        "seeded_replay_summary_observations",
        lambda ticker=None: [seeded_summary_observation] if ticker == "LIGHT.AS" else [],
    )

    observations = knowledge_model_module._load_learning_cohort(limit=1, subject_ticker="LIGHT.AS")

    assert observations == [seeded_summary_observation]
    assert knowledge_model_module._observation_evidence_count(observations) == 51


@pytest.mark.parametrize(
    ("subject_ticker", "broad_ticker"),
    [
        ("SUBJECT.ONE", "BROAD.ONE"),
        ("SUBJECT.TWO", "BROAD.TWO"),
        (None, "BROAD.THREE"),
    ],
)
def test_learning_cohort_keeps_seeded_broad_cohort_with_any_ledger_rows(monkeypatch, subject_ticker, broad_ticker):
    record_ticker = subject_ticker or "LEDGER.ONLY"
    subject_record = SimpleNamespace(
        record_id="subject-ledger",
        ticker=record_ticker,
        sector="Consumer Cyclical",
        industry="Footwear & Accessories",
        data_vintage_years=12,
        market_cap_regime="large",
        macro_regime="neutral",
        predicted_revenue_mm=100.0,
        actual_revenue_mm=108.0,
        near_term_revenue_growth=0.02,
        predicted_ebit_margin=0.10,
        actual_ebit_margin=0.11,
        predicted_wacc=0.08,
        predicted_terminal_growth=0.025,
        beta=1.0,
        predicted_ufcf_mm=8.0,
        actual_ufcf_mm=9.0,
        capex_pct_revenue=0.03,
        da_pct_revenue=0.02,
        predicted_ev_mm=1000.0,
        predicted_equity_value_mm=900.0,
        predicted_price_per_share=20.0,
        actual_price_at_prediction=18.0,
        actual_price_at_horizon=22.0,
        macro_backdrop={},
        feature_vector=None,
        structural_break_hints=[],
        prediction_context={"source": "historical_replay_bootstrap"},
    )
    seeded_summary_observation = {
        "ticker": record_ticker,
        "sector": "Consumer Cyclical",
        "industry": "Footwear & Accessories",
        "data_vintage_years": 12,
        "market_cap_regime": "large",
        "macro_regime": "neutral",
        "predicted_revenue_growth": 0.0,
        "actual_revenue_growth": 0.03,
        "predicted_ebit_margin": 0.0,
        "actual_ebit_margin": 0.01,
        "predicted_wacc": 0.08,
        "actual_wacc": 0.08,
        "predicted_terminal_growth": 0.025,
        "actual_terminal_growth": 0.025,
        "evidence_count": 51,
        "observation_type": "deployment_historical_replay_summary",
    }
    duplicate_subject_seed = {
        "ticker": record_ticker,
        "sector": "Consumer Cyclical",
        "industry": "Footwear & Accessories",
        "data_vintage_years": 12,
        "market_cap_regime": "large",
        "macro_regime": "neutral",
    }
    broad_seed = {
        "ticker": broad_ticker,
        "sector": "Consumer Cyclical",
        "industry": "Apparel Retail",
        "data_vintage_years": 12,
        "market_cap_regime": "large",
        "macro_regime": "neutral",
        "predicted_revenue_growth": 0.0,
        "actual_revenue_growth": 0.04,
        "predicted_ebit_margin": 0.12,
        "actual_ebit_margin": 0.13,
        "predicted_wacc": 0.08,
        "actual_wacc": 0.08,
        "predicted_terminal_growth": 0.025,
        "actual_terminal_growth": 0.025,
    }
    monkeypatch.setattr(knowledge_model_module, "_cached_ledger_records", lambda limit=None: [subject_record])
    monkeypatch.setattr(knowledge_model_module, "_cached_stratified_sample", lambda records, max_records, target: list(records))
    monkeypatch.setattr(knowledge_model_module.LedgerReader, "query", lambda *_args, **_kwargs: [subject_record])
    monkeypatch.setattr(knowledge_model_module, "_get_historical_observations", lambda: [])
    monkeypatch.setattr(knowledge_model_module, "seeded_cohort_observations", lambda limit=None: [duplicate_subject_seed, broad_seed])
    monkeypatch.setattr(
        knowledge_model_module,
        "seeded_replay_summary_observations",
        lambda ticker=None: [seeded_summary_observation] if ticker == subject_ticker else [],
    )

    observations = knowledge_model_module._load_learning_cohort(limit=20, subject_ticker=subject_ticker)

    tickers = [str(knowledge_model_module._obs_value(observation, "ticker", "")).upper() for observation in observations]
    assert broad_ticker in tickers
    assert tickers.count(record_ticker) == 2


def test_historical_replay_evidence_summary_uses_seed_when_raw_cache_empty(monkeypatch):
    import auto_valuation.learning.deployment_seed as deployment_seed_module
    import auto_valuation.learning.historical_replay as historical_replay_module

    monkeypatch.setattr(historical_replay_module, "get_all_observations", lambda: [])
    monkeypatch.setattr(
        deployment_seed_module,
        "historical_replay_summary",
        lambda ticker=None: {
            "records": 51,
            "annual_records": 10,
            "quarterly_records": 41,
            "first_year": 2016,
            "last_year": 2026,
            "mean_abs_revenue_error_pct": 26.21,
            "mean_abs_margin_error_pp": 2.38,
        } if ticker == "LIGHT.AS" else {},
    )

    summary = eodhd_client._historical_replay_evidence_summary("LIGHT.AS")

    assert summary["records"] == 51
    assert summary["annual_records"] == 10
    assert summary["quarterly_records"] == 41
    assert summary["source"] == "deployment-replay-summary"


def test_eodhd_read_only_build_skips_learning_writes(monkeypatch):
    monkeypatch.setattr(eodhd_client, "_fetch_price", lambda _code: {"close": 100.0})
    monkeypatch.setattr(eodhd_client, "_fetch_fundamentals", lambda _code: _mock_fundamentals())
    monkeypatch.setattr(eodhd_client, "_get_risk_free_rate", lambda: 4.0)
    monkeypatch.setattr(eodhd_client, "_PEERS_AVAILABLE", False)

    def _knowledge_payload() -> dict:
        return {
            "summary": "Weighted knowledge model active.",
            "global_learning": {"enabled": False, "scope": None, "cohort_size": 0, "sector_span": 0, "confidence": 0.0},
            "assumption_weights": {"revenue_growth_near": {"source": "Knowledge model: test"}},
            "layered_learning": {"uncertainty": {"scenario_width_multiplier": 1.0}},
            "explainability": {
                "headline": "Read-only dashboard build.",
                "company_memory": {"weight_pct": 60, "history_window_years": 5, "review_due": False},
                "sector_memory": {"weight_pct": 25, "records": 5, "confidence": 0.6},
                "cohort_memory": {"weight_pct": 15, "records": 5, "confidence": 0.55},
                "global_brain": {"enabled": False, "scope": None, "cohort_size": 0, "sector_span": 0, "confidence": 0.0},
                "analog_evidence": {"enabled": False, "match_count": 0, "confidence": 0.0, "pattern_score": 0.0},
                "relationship_graph": {"enabled": False, "node_count": 0, "edge_count": 0, "sector_span": 0, "confidence": 0.0},
                "forecast_layers": [],
                "data_gaps": [],
            },
            "memory_hierarchy": {"relational": {"score": 0}, "procedural": {"score": 50}},
            "calibration_cohort_size": 5,
            "scenario_width_multiplier": 1.0,
            "revenue_growth_near": 7.0,
            "terminal_growth": 2.5,
            "ebit_margin_target": 12.0,
            "beta": 1.0,
            "wacc": 9.0,
            "tax_rate_pct": 20.0,
            "da_pct": 3.0,
            "capex_pct": 3.5,
            "sbc_pct": 0.5,
            "dso": 35.0,
            "dio": 50.0,
            "dpo": 40.0,
            "confidence_model": {
                "valuation_confidence": {"score": 0.57, "score_100": 57, "expected_error_pct": {"p50": 10.0, "p75": 14.0, "p90": 18.0}},
                "assumption_confidence": {"score": 0.61, "score_100": 61, "expected_error_index": 4.2},
                "ranking_signal": 0.57,
                "dashboard_breakdown": {
                    "total": 57,
                    "grade": "C",
                    "label": "Guarded Confidence",
                    "color": "amber",
                    "dcf_suitable": True,
                    "suitability_note": "",
                    "warnings": [],
                    "dimensions": [],
                },
            },
        }

    monkeypatch.setattr(knowledge_model_module, "refine_live_assumptions", lambda **_: _knowledge_payload())
    monkeypatch.setattr(eodhd_client, "_global_universe_summary", lambda _store: {"enabled": True, "tracked_symbols": 12})
    monkeypatch.setattr(eodhd_client, "_top_learned_peer_edges", lambda *_args, **_kwargs: [])
    monkeypatch.setattr(
        eodhd_client,
        "_register_global_universe_symbols",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(AssertionError("read-only build should not register symbols")),
    )
    monkeypatch.setattr(
        eodhd_client,
        "_record_peer_learning_signals",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(AssertionError("read-only build should not record peer signals")),
    )
    monkeypatch.setattr(
        eodhd_client,
        "_auto_bootstrap_current_ticker",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(AssertionError("read-only build should not bootstrap")),
    )
    monkeypatch.setattr(
        eodhd_client,
        "_backfill_learning_actuals",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(AssertionError("read-only build should not backfill")),
    )
    monkeypatch.setattr(
        eodhd_client,
        "_run_learning_maintenance",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(AssertionError("read-only build should not run maintenance")),
    )
    monkeypatch.setattr(
        eodhd_client,
        "_persist_learning_snapshot",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(AssertionError("read-only build should not persist snapshots")),
    )

    data = eodhd_client.build_dashboard_data("TEST", mutate_learning=False)

    assert data["knowledge_model"]["learning_bootstrap"]["reason"] == "mutate_learning disabled"
    assert data["knowledge_model"]["learning_maintenance"]["reason"] == "mutate_learning disabled"
    assert data["knowledge_model"]["learning_persistence"]["reason"] == "mutate_learning disabled"
    assert data["knowledge_model"]["global_universe"]["tracked_symbols"] == 12
    assert data["knowledge_model"]["model_accuracy"]["enabled"] is True
    assert data["knowledge_model"]["explainability"]["model_accuracy"]["enabled"] is True
    assert data["model_view"]["enabled"] is True
    assert data["model_view"]["probabilities"]["base"] > 0
    assert data["model_view"]["recommendation"] == data["learned_recommendation"]


def test_eodhd_high_margin_scenarios_stay_ordered(monkeypatch):
    monkeypatch.setattr(eodhd_client, "_fetch_price", lambda _code: {"close": 100.0})
    monkeypatch.setattr(eodhd_client, "_fetch_fundamentals", lambda _code: _mock_high_margin_fundamentals())
    monkeypatch.setattr(eodhd_client, "_get_risk_free_rate", lambda: 4.0)
    monkeypatch.setattr(eodhd_client, "_PEERS_AVAILABLE", False)

    data = eodhd_client.build_dashboard_data("HIMG")
    scenarios = data["scenarios"]

    assert scenarios["bull"]["margin_target"] >= scenarios["base"]["margin_target"]
    assert scenarios["bull"]["iv"] > scenarios["base"]["iv"] > scenarios["bear"]["iv"]


def test_normalize_requested_ticker_maps_manual_suffixes_and_exchange_hints():
    assert eodhd_client.normalize_requested_ticker("005930.KS") == "005930.KO"
    assert eodhd_client.normalize_requested_ticker("BHP.AX") == "BHP.AU"
    assert eodhd_client.normalize_requested_ticker("SAP.DE") == "SAP.XETRA"
    assert eodhd_client.normalize_requested_ticker("VOD.L") == "VOD.LSE"
    assert eodhd_client.normalize_requested_ticker("BRK.B") == "BRK-B.US"
    assert eodhd_client.normalize_requested_ticker("BHP", exchange="LSE") == "BHP.LSE"


def test_valuate_route_applies_exchange_hint_to_manual_ticker():
    from webapp.app import app as webapp_app

    webapp_app.config.update(TESTING=True, SECRET_KEY="test-secret")

    with webapp_app.test_client() as client:
        response = client.post(
            "/valuate",
            data={"ticker": "BHP", "exchange": "LSE", "currency": "USD", "years": "10"},
        )

    assert response.status_code == 302
    assert response.headers["Location"].endswith("/loading/BHP.LSE")


def test_valuate_route_resolves_company_name_before_normalizing(monkeypatch):
    import webapp.app as webapp_module

    webapp_module.app.config.update(TESTING=True, SECRET_KEY="test-secret")
    monkeypatch.setattr(
        webapp_module,
        "resolve_search_input",
        lambda value, exchange="auto": "005930.KO" if value == "Samsung Electronics Co Ltd" else None,
    )

    with webapp_module.app.test_client() as client:
        response = client.post(
            "/valuate",
            data={"ticker": "Samsung Electronics Co Ltd", "exchange": "auto", "currency": "USD", "years": "10"},
        )

    assert response.status_code == 302
    assert response.headers["Location"].endswith("/loading/005930.KO")


def test_api_ticker_search_returns_results(monkeypatch):
    import webapp.app as webapp_module

    webapp_module.app.config.update(TESTING=True, SECRET_KEY="test-secret")
    discovery_calls: list[bool] = []

    monkeypatch.setattr(
        webapp_module,
        "search_tickers",
        lambda query, limit=12, exchange="auto": [
            {
                "ticker": "005930.KO",
                "code": "005930",
                "name": "Samsung Electronics Co Ltd",
                "exchange": "KO",
                "country": "Korea",
            }
        ],
    )
    monkeypatch.setattr(webapp_module, "_safe_discovery_store", lambda: discovery_calls.append(True) or None)

    with webapp_module.app.test_client() as client:
        response = client.get("/api/ticker-search?q=samsung&exchange=auto&limit=12")

    assert response.status_code == 200
    payload = response.get_json()
    assert payload["results"][0]["ticker"] == "005930.KO"
    assert discovery_calls == []


def test_api_ticker_search_does_not_persist_external_learning_state(monkeypatch):
    import webapp.app as webapp_module

    webapp_module.app.config.update(TESTING=True, SECRET_KEY="test-secret")
    persisted: list[bool] = []
    discovery_calls: list[bool] = []

    monkeypatch.setattr(
        webapp_module,
        "search_tickers",
        lambda query, limit=12, exchange="auto": [{"ticker": "NKE", "name": "Nike Inc", "exchange": "US"}],
    )
    monkeypatch.setattr(webapp_module, "_safe_discovery_store", lambda: discovery_calls.append(True) or None)
    monkeypatch.setattr(webapp_module, "_persist_external_learning_state", lambda force=False: persisted.append(force) or {"ok": True})

    with webapp_module.app.test_client() as client:
        response = client.get("/api/ticker-search?q=nike")

    assert response.status_code == 200
    assert discovery_calls == []
    assert persisted == []


def test_valuate_route_defaults_to_since_ipo_years():
    import webapp.app as webapp_module

    webapp_module.app.config.update(TESTING=True, SECRET_KEY="test-secret")

    with webapp_module.app.test_client() as client:
        response = client.post(
            "/valuate",
            data={"ticker": "NKE", "exchange": "auto", "currency": "USD"},
        )
        with client.session_transaction() as flask_session:
            years = flask_session["years"]

    assert response.status_code == 302
    assert years == "since_ipo"


def test_watchlist_and_manual_compare_routes_persist_discovery_state(tmp_path, monkeypatch):
    import webapp.app as webapp_module
    from auto_valuation.learning.discovery import DiscoveryStore
    from auto_valuation.learning.universe import SymbolUniverseStore

    webapp_module.app.config.update(TESTING=True, SECRET_KEY="test-secret")
    universe = SymbolUniverseStore(tmp_path / "symbol-universe.db")
    discovery = DiscoveryStore(tmp_path / "discovery.db", universe_store=universe)
    monkeypatch.setattr(webapp_module, "_safe_discovery_store", lambda: discovery)

    with webapp_module.app.test_client() as client:
        add_watch = client.post(
            "/api/watchlist",
            json={
                "ticker": "005930.KO",
                "company_name": "Samsung Electronics Co Ltd",
                "exchange": "KO",
                "sector": "Technology",
                "industry": "Consumer Electronics",
            },
        )
        compare = client.post(
            "/api/manual-compare",
            json={
                "subject": {
                    "ticker": "005930.KO",
                    "company_name": "Samsung Electronics Co Ltd",
                    "exchange": "KO",
                    "sector": "Technology",
                    "industry": "Consumer Electronics",
                },
                "peer": {
                    "ticker": "AAPL",
                    "company_name": "Apple Inc",
                    "exchange": "US",
                    "sector": "Technology",
                    "industry": "Consumer Electronics",
                },
            },
        )
        watchlist = client.get("/api/watchlist")
        recent = client.get("/api/manual-compare?subject=005930.KO")

    assert add_watch.status_code == 200
    assert compare.status_code == 200
    assert watchlist.get_json()["items"][0]["ticker"] == "005930.KO"
    assert recent.get_json()["items"][0]["ticker"] == "AAPL"
    relationship = discovery.get_peer_relationship("005930.KO", "AAPL")
    assert relationship is not None
    assert relationship["manual_compare_hits"] == 1


def test_manual_compare_route_is_idempotent_for_device_event_id(tmp_path, monkeypatch):
    import webapp.app as webapp_module
    from auto_valuation.learning.discovery import DiscoveryStore
    from auto_valuation.learning.universe import SymbolUniverseStore

    webapp_module.app.config.update(TESTING=True, SECRET_KEY="test-secret")
    universe = SymbolUniverseStore(tmp_path / "symbol-universe.db")
    discovery = DiscoveryStore(tmp_path / "discovery.db", universe_store=universe)
    monkeypatch.setattr(webapp_module, "_safe_discovery_store", lambda: discovery)

    payload = {
        "event_id": "device:005930.KO:AAPL",
        "subject": {
            "ticker": "005930.KO",
            "company_name": "Samsung Electronics Co Ltd",
            "exchange": "KO",
            "sector": "Technology",
            "industry": "Consumer Electronics",
        },
        "peer": {
            "ticker": "AAPL",
            "company_name": "Apple Inc",
            "exchange": "US",
            "sector": "Technology",
            "industry": "Consumer Electronics",
        },
    }

    with webapp_module.app.test_client() as client:
        first = client.post("/api/manual-compare", json=payload)
        second = client.post("/api/manual-compare", json=payload)

    relationship = discovery.get_peer_relationship("005930.KO", "AAPL")
    recent = discovery.list_manual_compares(subject_ticker="005930.KO", limit=5)

    assert first.status_code == 200
    assert second.status_code == 200
    assert relationship is not None
    assert relationship["manual_compare_hits"] == 1
    assert recent[0]["event_id"] == "device:005930.KO:AAPL"


def test_index_and_dashboard_render_discovery_controls(monkeypatch):
    import webapp.app as webapp_module

    webapp_module.app.config.update(TESTING=True, SECRET_KEY="test-secret")
    monkeypatch.setattr(webapp_module, "get_dashboard_data", lambda ticker, **kwargs: copy.deepcopy(_sample_dashboard_data()))

    with webapp_module.app.test_client() as client:
        index_response = client.get("/")
        dashboard_response = client.get("/dashboard/NKE")

    assert index_response.status_code == 200
    assert dashboard_response.status_code == 200
    index_html = index_response.get_data(as_text=True)
    dashboard_html = dashboard_response.get_data(as_text=True)
    assert "Global Brain Watchlist" in index_html
    assert "Track In Global Brain" in dashboard_html
    assert "Compare + Learn" in dashboard_html
    assert "Manual Compare Feed" in dashboard_html
    assert "Brain Fit" in dashboard_html


def test_dashboard_peer_actions_use_peer_metadata(monkeypatch):
    import webapp.app as webapp_module

    data = copy.deepcopy(_sample_dashboard_data())
    data["exchange"] = "NYSE"
    data["sector"] = "Consumer Cyclical"
    data["industry"] = "Footwear & Accessories"
    data["peers"] = [
        {
            "ticker": "RMS.PA",
            "name": "Hermes Intl",
            "exchange": "PA",
            "sector": "Consumer Cyclical",
            "industry": "Luxury Goods",
            "canonical_industry": "Luxury Goods",
            "industry_family": "luxury-fashion",
            "same_industry_cluster": False,
            "same_industry_family": True,
            "peer_rank": 0,
            "peer_learning_score": 8.4,
            "base_peer_learning_score": 6.9,
            "pair_strength_score": 1.5,
            "pair_auto_peer_hits": 2,
            "pair_manual_compare_hits": 1,
            "subject": False,
            "ev_ebitda": 19.2,
            "ev_ebit": 23.4,
            "ev_rev": 6.1,
            "pe": 28.5,
            "p_fcf": 32.1,
        }
    ]
    data["peer_median"] = {}

    webapp_module.app.config.update(TESTING=True, SECRET_KEY="test-secret")
    monkeypatch.setattr(webapp_module, "get_dashboard_data", lambda ticker, **kwargs: copy.deepcopy(data))

    with webapp_module.app.test_client() as client:
        response = client.get("/dashboard/NKE")

    html_text = html.unescape(response.get_data(as_text=True))
    assert response.status_code == 200
    assert "Brain Fit" in html_text
    assert '"ticker": "RMS.PA"' in html_text
    assert '"exchange": "PA"' in html_text
    assert '"industry": "Luxury Goods"' in html_text
    assert '"peer_learning_score": 8.4' in html_text


def test_index_renders_company_search_autocomplete():
    import webapp.app as webapp_module

    webapp_module.app.config.update(TESTING=True, SECRET_KEY="test-secret")

    with webapp_module.app.test_client() as client:
        response = client.get("/")

    assert response.status_code == 200
    html = response.get_data(as_text=True)
    assert "Company Name Or Ticker" in html
    assert "Search by company name or ticker across all available exchanges" in html
    assert "selectedTicker" in html
    assert "/api/ticker-search" in html
    assert '<option value="since_ipo" selected>Since IPO</option>' in html
    assert '<option value="10">10 years</option>' in html
    assert '<option value="15">15 years</option>' in html


def test_loading_page_preloads_dashboard_once_without_api_duplicate():
    import webapp.app as webapp_module

    webapp_module.app.config.update(TESTING=True, SECRET_KEY="test-secret")

    with webapp_module.app.test_client() as client:
        response = client.get("/loading/NKE")

    html = response.get_data(as_text=True)
    assert response.status_code == 200
    assert "/api/snapshot" not in html
    assert "/api/dashboard" not in html
    assert "fetch(dashboardUrl" in html
    assert "document.write(dashboardHtml)" in html


def test_safe_dashboard_data_passes_requested_history_window(monkeypatch):
    import webapp.app as webapp_module

    webapp_module.app.config.update(TESTING=True, SECRET_KEY="test-secret")
    seen_kwargs: dict[str, object] = {}

    def _fake_dashboard_data(_ticker, **kwargs):
        seen_kwargs.update(kwargs)
        return {"ticker": "NKE", "historical": {"years": []}}

    monkeypatch.setattr(webapp_module, "get_dashboard_data", _fake_dashboard_data)

    with webapp_module.app.test_request_context("/"):
        webapp_module.session["years"] = "15"
        webapp_module._safe_dashboard_data("NKE", mutate_learning=False)

    assert seen_kwargs["historical_years"] == 15


def test_eodhd_cross_currency_quotes_use_quote_fx_and_usd_reporting(monkeypatch):
    monkeypatch.setattr(eodhd_client, "_fetch_price", lambda _code: {"close": 2911.0})
    monkeypatch.setattr(eodhd_client, "_fetch_fundamentals", lambda _code: _mock_cross_currency_fundamentals())
    monkeypatch.setattr(eodhd_client, "_get_risk_free_rate", lambda: 4.0)
    monkeypatch.setattr(eodhd_client, "_PEERS_AVAILABLE", False)

    def _fx(ccy: str) -> float:
        return {"USD": 1.0, "GBX": 0.0126}.get(ccy, 1.0)

    monkeypatch.setattr(eodhd_client, "_get_fx_rate", _fx)
    gbx_data = eodhd_client.build_dashboard_data("BHP.LSE")

    monkeypatch.setattr(eodhd_client, "_fetch_price", lambda _code: {"close": 36.6786})
    usd_fundamentals = _mock_cross_currency_fundamentals()
    usd_fundamentals["General"]["CurrencyCode"] = "USD"
    monkeypatch.setattr(eodhd_client, "_fetch_fundamentals", lambda _code: usd_fundamentals)
    usd_data = eodhd_client.build_dashboard_data("BHP.US")

    assert gbx_data["currency"] == "USD"
    assert gbx_data["quote_currency"] == "GBX"
    assert gbx_data["reporting_currency"] == "USD"
    assert gbx_data["price"] == pytest.approx(36.68, abs=0.01)
    assert gbx_data["intrinsic_value"] == pytest.approx(usd_data["intrinsic_value"], abs=0.01)


def test_api_dashboard_exposes_learning_explainability_alias(monkeypatch):
    import webapp.app as webapp_module

    webapp_module.app.config.update(TESTING=True, SECRET_KEY="test-secret")
    expected = _sample_dashboard_data()
    expected["knowledge_model"] = {
        "summary": "Weighted knowledge model active.",
        "explainability": {
            "headline": "Core forecast mix: 62% company memory, 24% sector prior, and 14% realised cohort learning.",
            "company_memory": {"weight_pct": 62},
        },
    }

    monkeypatch.setattr(webapp_module, "get_dashboard_data", lambda ticker, **kwargs: copy.deepcopy(expected))

    with webapp_module.app.test_client() as client:
        response = client.get("/api/dashboard/NKE")

    assert response.status_code == 200
    payload = response.get_json()
    assert payload["learning_explainability"]["company_memory"]["weight_pct"] == 62
    assert payload["knowledge_model_summary"] == "Weighted knowledge model active."
    assert payload["intrinsic_value_per_share"] == expected["intrinsic_value"]
    assert payload["model_confidence_score"] == expected["confidence_score"]
    assert payload["historical_years"] == expected["historical"]["years"]


def test_dashboard_and_api_fall_back_to_demo_when_dashboard_data_raises(monkeypatch):
    import webapp.app as webapp_module

    webapp_module.app.config.update(TESTING=True, SECRET_KEY="test-secret")
    monkeypatch.setattr(
        webapp_module,
        "get_dashboard_data",
        lambda _ticker, **kwargs: (_ for _ in ()).throw(RuntimeError("upstream boom")),
    )

    with webapp_module.app.test_client() as client:
        dashboard_response = client.get("/dashboard/NKE")
        api_response = client.get("/api/dashboard/NKE")

    assert dashboard_response.status_code == 200
    assert api_response.status_code == 200
    payload = api_response.get_json()
    assert payload["is_demo"] is True
    assert payload["data_source"] == "demo-fallback"
    assert "temporarily unavailable" in payload["demo_note"]
    assert payload["runtime_warning"] == "upstream boom"


def test_dashboard_renders_everything_knows_model_panel(monkeypatch):
    import webapp.app as webapp_module

    webapp_module.app.config.update(TESTING=True, SECRET_KEY="test-secret")
    data = _sample_dashboard_data()
    data["knowledge_model"] = {
        "summary": "Weighted knowledge model active.",
        "global_learning": {
            "enabled": True,
            "scope": "regime",
            "cohort_size": 12,
            "sector_span": 5,
            "confidence": 0.67,
        },
        "learning_backfill": {
            "updated_records": 1,
            "matured_records": 2,
        },
        "learning_maintenance": {
            "ran": True,
            "annual_postmortems_created": 3,
            "quinquennial_reports_created": 1,
            "scanned_tickers": 5,
        },
        "learning_persistence": {
            "enabled": True,
            "persisted": True,
            "horizon_year": 2031,
        },
        "relationship_graph": {
            "enabled": True,
            "confidence": 0.74,
            "node_count": 4,
            "edge_count": 5,
            "sector_span": 2,
            "global_universe": {
                "enabled": True,
                "tracked_symbols": 18,
                "sector_span": 5,
                "exchange_span": 4,
                "bootstrapped_symbols": 9,
                "calibration_priority_candidates": [
                    {"ticker": "GRAPH", "mode": "ticker", "direct_samples": 3, "cohort_samples": 0},
                ],
            },
            "candidate_pool_size": 8,
            "analog_pool_size": 6,
            "realized_candidate_count": 2,
            "role_counts": {"analog": 1, "realized_peer": 1},
            "summary": "Relationship graph linked 3 connected symbols across 2 sectors with moderate confidence.",
            "central_nodes": [
                {"ticker": "GRAPH", "role": "analog", "score": 0.92, "similarity": 0.9},
            ],
            "pathways": [
                {"ticker": "GRAPH", "score": 0.92, "rationale": "Analog pathway reinforcing margin durability.", "impact": "Growth +0.3pp, margin +0.2pp"},
            ],
            "learned_peer_edges": [
                {
                    "ticker": "RMS.PA",
                    "company_name": "Hermes Intl",
                    "industry": "Luxury Goods",
                    "canonical_industry": "Luxury Goods",
                    "industry_family": "luxury-fashion",
                    "pair_strength_score": 3.6,
                    "pair_strength_score_raw": 5.2,
                    "pair_decay_multiplier": 0.69,
                    "pair_age_days": 24.0,
                    "pair_auto_peer_hits": 3,
                    "pair_manual_compare_hits": 1,
                    "pair_last_source": "manual-compare",
                    "last_seen_at": "2026-05-03T08:17:04+00:00",
                }
            ],
            "visualization": {
                "width": 560,
                "height": 340,
                "view_box": "0 0 560 340",
                "legend": [
                    {"role": "subject", "label": "Subject ticker", "fill": "#0f172a", "stroke": "#38bdf8", "text": "#f8fafc"},
                    {"role": "analog", "label": "Analog memory", "fill": "#115e59", "stroke": "#5eead4", "text": "#ecfeff"},
                    {"role": "realized-peer", "label": "Realized peer", "fill": "#9a3412", "stroke": "#fdba74", "text": "#fff7ed"},
                ],
                "nodes": [
                    {"ticker": "NKE", "role": "subject", "x": 280, "y": 170, "radius": 26, "short_label": "NKE", "label_y": 212, "fill": "#0f172a", "stroke": "#38bdf8", "text": "#f8fafc", "score": 1.0},
                    {"ticker": "GRAPH", "role": "analog", "x": 178, "y": 88, "radius": 21, "short_label": "GRAPH", "label_y": 58, "fill": "#115e59", "stroke": "#5eead4", "text": "#ecfeff", "score": 0.92},
                    {"ticker": "PAIR", "role": "realized-peer", "x": 394, "y": 254, "radius": 18, "short_label": "PAIR", "label_y": 280, "fill": "#9a3412", "stroke": "#fdba74", "text": "#fff7ed", "score": 0.76},
                ],
                "edges": [
                    {"source": "NKE", "target": "GRAPH", "relationship": "analog-fingerprint", "weight": 0.91, "x1": 280, "y1": 170, "x2": 178, "y2": 88, "stroke_width": 4.4, "opacity": 0.79},
                    {"source": "NKE", "target": "PAIR", "relationship": "realized-spillover", "weight": 0.72, "x1": 280, "y1": 170, "x2": 394, "y2": 254, "stroke_width": 3.9, "opacity": 0.67},
                ],
            },
            "connected_tickers": ["GRAPH", "PAIR"],
        },
        "explainability": {
            "headline": "Core forecast mix: 58% company memory, 24% sector prior, and 18% realised cohort learning.",
            "company_memory": {
                "weight_pct": 58,
                "history_window_years": 5,
                "review_due": False,
            },
            "global_universe": {
                "enabled": True,
                "tracked_symbols": 18,
                "sector_span": 5,
                "exchange_span": 4,
                "bootstrapped_symbols": 9,
                "background_target_symbols": 1000,
                "background_seed_prefix_per_cycle": 8,
                "background_seed_pool_size": 298,
                "background_runner": {
                    "last_run_at": "2026-05-03T10:15:00+00:00",
                    "tracked_symbols": 146,
                    "seed_pool_size": 298,
                    "seed_target_symbols": 1000,
                    "requested_exchanges": ["US", "LSE"],
                    "fetched_exchanges": ["US"],
                    "exchange_discovered_symbols": 3,
                    "exchange_enrolled_symbols": 3,
                    "requested_tickers": ["GRAPH", "PAIR", "RMS.PA"],
                },
                "priority_candidates": ["GRAPH", "PAIR"],
                "calibration_priority_candidates": [
                    {"ticker": "GRAPH", "mode": "ticker", "direct_samples": 3, "cohort_samples": 0},
                ],
            },
            "learning_bootstrap": {
                "enabled": True,
                "ran": True,
                "reason": None,
                "replay_predictions_created": 4,
                "realized_outcomes_created": 3,
            },
            "cohort_memory": {
                "records": 8,
                "confidence": 0.62,
                "weight_pct": 18,
            },
            "global_brain": {
                "enabled": True,
                "scope": "regime",
                "cohort_size": 12,
                "sector_span": 5,
                "confidence": 0.67,
            },
            "realized_evidence": {
                "matured_records": 2,
                "updated_records": 1,
            },
            "maintenance": {
                "ran": True,
                "annual_postmortems_created": 3,
                "quinquennial_reports_created": 1,
                "scanned_tickers": 5,
            },
            "current_snapshot": {
                "enabled": True,
                "persisted": True,
                "horizon_year": 2031,
            },
            "analog_evidence": {
                "pattern_label": "Capital Light Transition",
                "pattern_score": 0.88,
                "confidence": 0.81,
                "match_count": 2,
                "archetypes": ["Adobe 2015-2019", "Salesforce 2008-2014"],
                "top_matches": [
                    {
                        "ticker": "ANLG",
                        "sector": "Consumer Discretionary",
                        "industry": "Internet Retail",
                        "score": 0.94,
                        "similarity": 0.92,
                        "regime_similarity": 0.89,
                        "why_it_matters": "Gross Margin 61.0% vs 60.5%, FCF Conversion 0.28 vs 0.27",
                        "evidence": [
                            {"label": "Gross Margin", "similarity": 0.96},
                            {"label": "FCF Conversion", "similarity": 0.93},
                        ],
                    }
                ],
                "note": "Pattern analog Capital Light Transition is active with score 0.88.",
            },
            "relationship_graph": {
                "enabled": True,
                "confidence": 0.74,
                "node_count": 4,
                "edge_count": 5,
                "sector_span": 2,
                "candidate_pool_size": 8,
                "analog_pool_size": 6,
                "realized_candidate_count": 2,
                "role_counts": {"analog": 1, "realized_peer": 1},
                "summary": "Relationship graph linked 3 connected symbols across 2 sectors with moderate confidence.",
                "central_nodes": [
                    {"ticker": "GRAPH", "role": "analog", "score": 0.92, "similarity": 0.9},
                ],
                "pathways": [
                    {"ticker": "GRAPH", "score": 0.92, "rationale": "Analog pathway reinforcing margin durability.", "impact": "Growth +0.3pp, margin +0.2pp"},
                ],
                "learned_peer_edges": [
                    {
                        "ticker": "RMS.PA",
                        "company_name": "Hermes Intl",
                        "industry": "Luxury Goods",
                        "canonical_industry": "Luxury Goods",
                        "industry_family": "luxury-fashion",
                        "pair_strength_score": 3.6,
                        "pair_strength_score_raw": 5.2,
                        "pair_decay_multiplier": 0.69,
                        "pair_age_days": 24.0,
                        "pair_auto_peer_hits": 3,
                        "pair_manual_compare_hits": 1,
                        "pair_last_source": "manual-compare",
                        "last_seen_at": "2026-05-03T08:17:04+00:00",
                    }
                ],
                "visualization": {
                    "width": 560,
                    "height": 340,
                    "view_box": "0 0 560 340",
                    "legend": [
                        {"role": "subject", "label": "Subject ticker", "fill": "#0f172a", "stroke": "#38bdf8", "text": "#f8fafc"},
                        {"role": "analog", "label": "Analog memory", "fill": "#115e59", "stroke": "#5eead4", "text": "#ecfeff"},
                        {"role": "realized-peer", "label": "Realized peer", "fill": "#9a3412", "stroke": "#fdba74", "text": "#fff7ed"},
                    ],
                    "nodes": [
                        {"ticker": "NKE", "role": "subject", "x": 280, "y": 170, "radius": 26, "short_label": "NKE", "label_y": 212, "fill": "#0f172a", "stroke": "#38bdf8", "text": "#f8fafc", "score": 1.0},
                        {"ticker": "GRAPH", "role": "analog", "x": 178, "y": 88, "radius": 21, "short_label": "GRAPH", "label_y": 58, "fill": "#115e59", "stroke": "#5eead4", "text": "#ecfeff", "score": 0.92},
                        {"ticker": "PAIR", "role": "realized-peer", "x": 394, "y": 254, "radius": 18, "short_label": "PAIR", "label_y": 280, "fill": "#9a3412", "stroke": "#fdba74", "text": "#fff7ed", "score": 0.76},
                    ],
                    "edges": [
                        {"source": "NKE", "target": "GRAPH", "relationship": "analog-fingerprint", "weight": 0.91, "x1": 280, "y1": 170, "x2": 178, "y2": 88, "stroke_width": 4.4, "opacity": 0.79},
                        {"source": "NKE", "target": "PAIR", "relationship": "realized-spillover", "weight": 0.72, "x1": 280, "y1": 170, "x2": 394, "y2": 254, "stroke_width": 3.9, "opacity": 0.67},
                    ],
                },
                "connected_tickers": ["GRAPH", "PAIR"],
            },
            "forecast_layers": [
                {
                    "driver": "Revenue Growth",
                    "final_value": 6.1,
                    "unit": "%",
                    "weights": {"company_history": 58, "sector_prior": 24, "learned_cohort": 18},
                    "overlays": [{"label": "Global brain", "impact": -0.4, "unit": "pp"}],
                    "source": "Knowledge model: 58% company 5y history, 24% sector prior, 18% learned cohort",
                    "warn": None,
                }
            ],
            "confidence_decomposition": {
                "summary": "Shared-brain confidence explains how much of the forecast is supported by history, realised learning, analogs, and cross-symbol evidence.",
                "dominant_risk": "Discount-rate sensitivity",
                "assumption_confidence": {"score": 0.64, "score_100": 64, "label": "moderate", "expected_error_index": 4.3},
                "valuation_confidence": {
                    "score": 0.57,
                    "score_100": 57,
                    "label": "guarded",
                    "expected_error_pct": {"p50": 10.8, "p75": 13.9, "p90": 18.6},
                },
                "components": [
                    {"label": "Company memory", "score": 58, "detail": "Five years of operating history are carrying most of the load."},
                    {"label": "Global brain", "score": 67, "detail": "Cross-symbol overlays are active across five sectors."},
                    {"label": "Relational memory", "score": 74, "detail": "Relationship graph links GRAPH and PAIR into the current symbol brain."},
                ],
            },
            "memory_hierarchy": {
                "summary": "Memory is split across episodic ticker history, semantic sector/cohort knowledge, relational cross-symbol links, and procedural confidence controls.",
                "layers": [
                    {"label": "Episodic Memory", "score": 68, "status": "moderate", "note": "Ticker history remains the main episodic anchor."},
                    {"label": "Semantic Memory", "score": 61, "status": "guarded", "note": "Sector and cohort memory are contributing reusable priors."},
                    {"label": "Relational Memory", "score": 74, "status": "moderate", "note": "Relationship graph links GRAPH and PAIR into the live symbol brain."},
                    {"label": "Procedural Memory", "score": 70, "status": "moderate", "note": "Confidence scoring is widening ranges when evidence conflicts."},
                ],
            },
            "data_gaps": [
                {"title": "Analog evidence is weak", "detail": "Only one named analog archetype is active so far.", "severity": "amber"},
            ],
            "limitations_note": "When evidence is thin, the model falls back toward company history and sector priors.",
        },
    }

    monkeypatch.setattr(webapp_module, "get_dashboard_data", lambda ticker, **kwargs: copy.deepcopy(data))

    with webapp_module.app.test_client() as client:
        response = client.get("/dashboard/NKE")

    assert response.status_code == 200
    html = response.get_data(as_text=True)
    assert "Everything Knows Model" in html
    assert "Company Memory" in html
    assert "Global Brain" in html
    assert "Global Universe" in html
    assert "Calibration queue is prioritizing" in html
    assert "Analog Evidence" in html
    assert "Top live analogs" in html
    assert "ANLG" in html
    assert "Relationship Graph" in html
    assert 'id="relationship-graph-network"' in html
    assert "Live network map" in html
    assert "Strongest learned peer edges" in html
    assert "Background learning is now rotating through 298 locally seeded equities" in html
    assert "Latest autonomous sweep: US, LSE" in html
    assert "Background Queue" in html
    assert "Exchange sweep" in html
    assert "Bootstrap queue" in html
    assert "RMS.PA" in html
    assert "Analog memory" in html
    assert "Realized peer" in html
    assert "Memory Hierarchy" in html
    assert "GRAPH" in html
    assert "Assumption Confidence" in html
    assert "Expected Valuation Error" in html
    assert "Remaining Data Gaps" in html


def test_top_learned_peer_edges_flattens_relationship_audit_payload(tmp_path):
    universe = SymbolUniverseStore(tmp_path / "symbol-universe.db")
    discovery = DiscoveryStore(tmp_path / "discovery.db", universe_store=universe)
    discovery.record_manual_compare(
        {
            "ticker": "CDI.PA",
            "company_name": "Christian Dior",
            "exchange": "PA",
            "sector": "Consumer Cyclical",
            "industry": "Luxury Goods",
        },
        [
            {
                "ticker": "RMS.PA",
                "company_name": "Hermes Intl",
                "exchange": "PA",
                "sector": "Consumer Cyclical",
                "industry": "Luxury Goods",
                "canonical_industry": "Luxury Goods",
                "industry_family": "luxury-fashion",
                "peer_learning_score": 7.1,
                "base_peer_learning_score": 5.4,
            }
        ],
    )

    edges = eodhd_client._top_learned_peer_edges(discovery, "CDI.PA", limit=3)

    assert edges[0]["ticker"] == "RMS.PA"
    assert edges[0]["pair_manual_compare_hits"] == 1
    assert edges[0]["pair_strength_score_raw"] >= edges[0]["pair_strength_score"]
    assert edges[0]["pair_last_source"] == "manual-compare"


def test_register_global_universe_symbols_tracks_current_and_related_tickers(tmp_path):
    store = SymbolUniverseStore(tmp_path / "symbol-universe.db")
    knowledge_model = {
        "analogs": {
            "items": [
                {
                    "ticker": "GRAPH",
                    "sector": "Industrials",
                    "industry": "Machinery",
                    "score": 0.91,
                    "similarity": 0.88,
                }
            ]
        },
        "relationship_graph": {
            "nodes": [
                {"ticker": "TEST", "sector": "Industrials", "industry": "Manufacturing", "role": "subject", "score": 1.0},
                {"ticker": "PAIR", "sector": "Technology", "industry": "Software", "role": "realized-peer", "score": 0.82, "similarity": 0.8},
            ]
        },
    }
    peer_items = [
        {
            "ticker": "LUX",
            "company_name": "Luxury Peer",
            "exchange": "PA",
            "sector": "Consumer Cyclical",
            "industry": "Luxury Goods",
            "peer_learning_score": 3.2,
            "industry_similarity": 1.0,
        }
    ]

    eodhd_client._register_global_universe_symbols(
        store,
        ticker="TEST",
        company_name="Test Co",
        exchange="NYSE",
        country="United States",
        sector="Industrials",
        industry="Manufacturing",
        knowledge_model=knowledge_model,
        peer_items=peer_items,
    )

    current = store.get_symbol("TEST")
    analog = store.get_symbol("GRAPH")
    related = store.get_symbol("PAIR")
    peer = store.get_symbol("LUX")
    summary = eodhd_client._global_universe_summary(store)

    assert current is not None
    assert current["valuation_hits"] == 1
    assert current["fundamentals_cached"] is True
    assert analog is not None
    assert related is not None
    assert peer is not None
    assert peer["metadata"]["peer_candidate_hits"] == 1
    assert peer["metadata"]["canonical_industry"] == "Luxury Goods"
    assert summary["tracked_symbols"] >= 4


def test_global_universe_summary_includes_background_runner_state(tmp_path, monkeypatch):
    store = SymbolUniverseStore(tmp_path / "symbol-universe.db")
    store.upsert_symbol("TEST", company_name="Test Co", exchange="NYSE", source="dashboard-live")
    monkeypatch.setattr(
        background_runner_module,
        "read_background_runner_state",
        lambda state_path=None: {
            "last_run_at": "2026-05-03T10:15:00+00:00",
            "requested_exchanges": ["US", "LSE"],
            "requested_tickers": ["TEST", "PAIR"],
            "exchange_discovered_symbols": 4,
        },
    )

    summary = eodhd_client._global_universe_summary(store)

    assert summary["background_runner"]["requested_exchanges"] == ["US", "LSE"]
    assert summary["background_runner"]["requested_tickers"] == ["TEST", "PAIR"]


def test_auto_bootstrap_current_ticker_skips_when_recent(tmp_path, monkeypatch):
    store = SymbolUniverseStore(tmp_path / "symbol-universe.db")
    store.upsert_symbol(
        "TEST",
        company_name="Test Co",
        exchange="NYSE",
        sector="Industrials",
        industry="Manufacturing",
        source="dashboard-live",
        valued=True,
        fundamentals_cached=True,
        bootstrapped=True,
        bootstrap_status="realized",
    )
    monkeypatch.setattr(eodhd_client, "_live_learning_feedback_enabled", lambda: True)

    result = eodhd_client._auto_bootstrap_current_ticker("TEST", _mock_fundamentals(), store)

    assert result["enabled"] is True
    assert result["ran"] is False
    assert result["reason"] == "fresh"
