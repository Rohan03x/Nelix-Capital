"""
tests/test_e2e.py — Integration / end-to-end pipeline tests (Phase 11).

These tests wire the sub-modules together without touching any live network.
All fixture data uses pre-cleaned (post-standardise_field_names) field names
so they can be fed directly to model/forecast functions.

Pipeline under test (no main.py — avoids its API inconsistencies):
  cleaner  →  bridge  →  build_wacc  →  run_dcf
                          ↓
             sensitivity (wacc_growth, scenarios)
                          ↓
             output (format_summary, write_json, write_excel)

No FMP / FRED / yfinance calls are made.
"""

from __future__ import annotations

import json
import math
from pathlib import Path

import openpyxl
import pytest

# ── Pipeline imports ──────────────────────────────────────────────────────────
from auto_valuation.data.bridge import compute_net_debt
from auto_valuation.data.cleaner import (
    standardise_field_names,
    unit_normalize,
    deduplicate_financial_data,
    normalize_one_time_items,
)
from auto_valuation.forecast.dcf import DCFResult, run_dcf
from auto_valuation.assumptions.wacc import build_wacc
from auto_valuation.model.income_statement import normalise_tax_rate
from auto_valuation.model.dilution import compute_fully_diluted_shares, compute_price_per_share
from auto_valuation.output.report import (
    format_valuation_summary,
    print_valuation_summary,
    write_json_output,
)
from auto_valuation.output.excel import write_excel_output
from auto_valuation.sensitivity.analysis import (
    wacc_growth_sensitivity,
    run_scenario_analysis,
    scenario_summary_table,
)
from auto_valuation.validation.checks import run_all_data_checks


# ─────────────────────────────────────────────────────────────────────────────
# Shared fixture factory
# ─────────────────────────────────────────────────────────────────────────────

def _make_income_stmt(year: int, revenue: float) -> dict:
    """Return a pre-cleaned income statement dict for a given fiscal year."""
    ebit       = revenue * 0.16
    da         = revenue * 0.05
    ebt        = ebit - revenue * 0.01        # simple interest charge
    tax        = ebt * 0.21
    return {
        "date":                      f"{year}-12-31",
        "period":                    "FY",
        "revenue":                   revenue,
        "cost_of_revenue":           revenue * 0.55,
        "gross_profit":              revenue * 0.45,
        "operating_expenses":        revenue * 0.29,
        "ebit":                      ebit,
        "ebit_normalized":           ebit,
        "da":                        da,           # field name validation checks for
        "depreciation_amortization": da,
        "income_before_tax":         ebt,
        "income_tax_expense":        tax,
        "net_income":                ebt - tax,
        "rd_expense":                revenue * 0.04,
    }


def _make_balance_sheet(year: int, revenue: float) -> dict:
    """Return a pre-cleaned balance sheet dict."""
    return {
        "date":                 f"{year}-12-31",
        "period":               "FY",
        "cash":                 revenue * 0.08,
        "accounts_receivable":  revenue * 0.12,   # ~44 days
        "inventory":            revenue * 0.10,   # ~36 days (vs COGS)
        "total_current_assets": revenue * 0.35,
        "property_plant_equipment": revenue * 0.40,
        "total_assets":         revenue * 0.90,
        "accounts_payable":     revenue * 0.09,   # ~33 days
        "short_term_debt":      revenue * 0.04,
        "long_term_debt":       revenue * 0.20,
        "total_debt":           revenue * 0.24,
        "total_liabilities":    revenue * 0.55,
        "shareholders_equity":  revenue * 0.35,
        "total_equity":         revenue * 0.35,   # field name validation checks for
    }


def _make_cash_flow(year: int, revenue: float) -> dict:
    """Return a pre-cleaned cash flow statement dict."""
    return {
        "date":                     f"{year}-12-31",
        "period":                   "FY",
        "operating_cash_flow":      revenue * 0.14,
        "cfo":                      revenue * 0.14,   # field name validation checks for
        "capital_expenditures":     -(revenue * 0.05),   # negative = outflow
        "capex":                    -(revenue * 0.05),   # field name validation checks for
        "free_cash_flow":           revenue * 0.09,
        "depreciation_amortization": revenue * 0.05,
    }


def _make_fixture_statements(n_years: int = 5) -> tuple[list[dict], list[dict], list[dict]]:
    """
    Build n_years of annual statements for a fictional company 'ACME'.
    Most recent year first (FMP convention).
    Revenue grows ~8% per year going back from 2024.
    """
    base = 60_000.0   # USD millions, most recent year
    income_stmts  = []
    balance_sheets = []
    cash_flows    = []
    for i in range(n_years):
        year = 2024 - i
        rev  = base * (1 / 1.08) ** i
        income_stmts.append(_make_income_stmt(year, rev))
        balance_sheets.append(_make_balance_sheet(year, rev))
        cash_flows.append(_make_cash_flow(year, rev))
    return income_stmts, balance_sheets, cash_flows


def _base_dcf_kwargs(
    income_stmts: list[dict],
    balance_sheets: list[dict],
    cash_flows: list[dict],
) -> dict:
    return dict(
        ticker="ACME",
        scenario="base",
        income_stmts=income_stmts,
        cash_flows=cash_flows,
        balance_sheets=balance_sheets,
        wacc=0.09,
        terminal_growth=0.03,
        near_term_growth=0.08,
        target_ebit_margin=0.18,
        forecast_years=10,
    )


# ─────────────────────────────────────────────────────────────────────────────
# 1 — Data layer → DCF integration
# ─────────────────────────────────────────────────────────────────────────────

class TestDataToDcfPipeline:

    @pytest.fixture
    def stmts(self):
        return _make_fixture_statements()

    def test_compute_net_debt_returns_float(self, stmts):
        _, balance_sheets, _ = stmts
        result = compute_net_debt(balance_sheets[0])
        assert isinstance(result, float)

    def test_net_debt_positive_with_leveraged_firm(self, stmts):
        _, balance_sheets, _ = stmts
        nd = compute_net_debt(balance_sheets[0])
        # long_term_debt (0.20*rev) + st_debt (0.04*rev) >> cash (0.08*rev)
        assert nd > 0

    def test_run_dcf_produces_dcf_result(self, stmts):
        income_stmts, balance_sheets, cash_flows = stmts
        result = run_dcf(**_base_dcf_kwargs(income_stmts, balance_sheets, cash_flows))
        assert isinstance(result, DCFResult)

    def test_run_dcf_positive_enterprise_value(self, stmts):
        income_stmts, balance_sheets, cash_flows = stmts
        result = run_dcf(**_base_dcf_kwargs(income_stmts, balance_sheets, cash_flows))
        assert result.enterprise_value > 0

    def test_run_dcf_ev_components_sum(self, stmts):
        """PV(UFCFs) + PV(TV) == EV."""
        income_stmts, balance_sheets, cash_flows = stmts
        r = run_dcf(**_base_dcf_kwargs(income_stmts, balance_sheets, cash_flows))
        assert r.enterprise_value == pytest.approx(r.pv_ufcfs + r.pv_terminal_value,
                                                   rel=1e-4)

    def test_run_dcf_10_forecast_years(self, stmts):
        income_stmts, balance_sheets, cash_flows = stmts
        r = run_dcf(**_base_dcf_kwargs(income_stmts, balance_sheets, cash_flows))
        assert len(r.forecast_years_data) == 10

    def test_run_dcf_forecast_years_ascending_revenue(self, stmts):
        income_stmts, balance_sheets, cash_flows = stmts
        r = run_dcf(**_base_dcf_kwargs(income_stmts, balance_sheets, cash_flows))
        revenues = [fy.revenue for fy in r.forecast_years_data]
        # Revenue should grow (near-term growth = 8%)
        assert revenues[-1] > revenues[0]

    def test_run_dcf_tv_pct_between_zero_and_one(self, stmts):
        income_stmts, balance_sheets, cash_flows = stmts
        r = run_dcf(**_base_dcf_kwargs(income_stmts, balance_sheets, cash_flows))
        assert 0.0 < r.tv_pct_of_ev < 1.0

    def test_normalise_tax_rate_used_in_pipeline(self, stmts):
        income_stmts, _, _ = stmts
        tax = normalise_tax_rate(income_stmts)
        assert 0.10 < tax < 0.45

    def test_wacc_dcf_integration(self):
        """build_wacc output feeds directly into run_dcf with correct result."""
        income_stmts, balance_sheets, cash_flows = _make_fixture_statements()
        latest_bs = balance_sheets[0]
        rev = income_stmts[0]["revenue"]

        wacc_dict = build_wacc(
            market_cap=rev * 3.5,          # ~3.5× revenue market cap
            total_debt=rev * 0.24,
            preferred_stock=0.0,
            basic_shares_mm=500.0,
            current_price=150.0,
            risk_free_rate=0.045,
            equity_risk_premium=0.055,
            beta=1.05,
            pre_tax_cost_of_debt=0.05,
            tax_rate=0.21,
        )
        wacc = wacc_dict["wacc"]
        assert 0.05 < wacc < 0.15, f"WACC out of bounds: {wacc}"

        r = run_dcf(
            ticker="ACME",
            scenario="base",
            income_stmts=income_stmts,
            cash_flows=cash_flows,
            balance_sheets=balance_sheets,
            wacc=wacc,
            terminal_growth=0.03,
            near_term_growth=0.08,
            target_ebit_margin=0.18,
        )
        assert r.enterprise_value > 0


# ─────────────────────────────────────────────────────────────────────────────
# 2 — Cleaner → bridge → DCF (raw-field integration)
# ─────────────────────────────────────────────────────────────────────────────

class TestCleanerToDcfPipeline:

    def test_standardise_then_run_dcf(self):
        """Simulate raw FMP data being cleaned before feeding into DCF."""
        raw_income = [
            {
                "date": f"{2024 - i}-12-31",
                "revenue": 60_000 * (1 / 1.08) ** i,
                "costOfRevenue": 60_000 * 0.55 * (1 / 1.08) ** i,
                "grossProfit": 60_000 * 0.45 * (1 / 1.08) ** i,
                "operatingIncome": 60_000 * 0.16 * (1 / 1.08) ** i,
                "depreciationAndAmortization": 60_000 * 0.05 * (1 / 1.08) ** i,
                "incomeBeforeTax": 60_000 * 0.15 * (1 / 1.08) ** i,
                "incomeTaxExpense": 60_000 * 0.15 * 0.21 * (1 / 1.08) ** i,
                "netIncome": 60_000 * 0.15 * 0.79 * (1 / 1.08) ** i,
                "researchAndDevelopmentExpenses": 60_000 * 0.04 * (1 / 1.08) ** i,
            }
            for i in range(5)
        ]
        raw_balance = [
            {
                "date": f"{2024 - i}-12-31",
                "cashAndCashEquivalents": 60_000 * 0.08 * (1 / 1.08) ** i,
                "netReceivables": 60_000 * 0.12 * (1 / 1.08) ** i,
                "inventory": 60_000 * 0.10 * (1 / 1.08) ** i,
                "longTermDebt": 60_000 * 0.20 * (1 / 1.08) ** i,
                "shortTermDebt": 60_000 * 0.04 * (1 / 1.08) ** i,
                "accountsPayables": 60_000 * 0.09 * (1 / 1.08) ** i,
            }
            for i in range(5)
        ]
        raw_cashflow = [
            {
                "date": f"{2024 - i}-12-31",
                "capitalExpenditure": -(60_000 * 0.05 * (1 / 1.08) ** i),
                "depreciationAndAmortization": 60_000 * 0.05 * (1 / 1.08) ** i,
                "operatingCashFlow": 60_000 * 0.14 * (1 / 1.08) ** i,
            }
            for i in range(5)
        ]

        # Clean — both functions take and return list[dict]
        profile = {"currency": "USD", "reportedCurrency": "USD"}
        income_stmts  = standardise_field_names(unit_normalize(raw_income, profile))
        balance_sheets = standardise_field_names(unit_normalize(raw_balance, profile))
        cash_flows    = standardise_field_names(unit_normalize(raw_cashflow, profile))

        income_stmts  = deduplicate_financial_data(income_stmts)
        balance_sheets = deduplicate_financial_data(balance_sheets)
        cash_flows    = deduplicate_financial_data(cash_flows)

        r = run_dcf(
            ticker="ACME",
            scenario="base",
            income_stmts=income_stmts,
            cash_flows=cash_flows,
            balance_sheets=balance_sheets,
            wacc=0.09,
            terminal_growth=0.03,
            near_term_growth=0.08,
            target_ebit_margin=0.18,
        )
        assert r.enterprise_value > 0

    def test_normalize_one_time_items_then_dcf(self):
        income_stmts, balance_sheets, cash_flows = _make_fixture_statements()
        cleaned = normalize_one_time_items(income_stmts)  # takes full list, not per-row
        r = run_dcf(
            ticker="ACME",
            scenario="base",
            income_stmts=cleaned,
            cash_flows=cash_flows,
            balance_sheets=balance_sheets,
            wacc=0.09,
            terminal_growth=0.03,
            near_term_growth=0.08,
            target_ebit_margin=0.18,
        )
        assert isinstance(r, DCFResult)
        assert r.enterprise_value > 0


# ─────────────────────────────────────────────────────────────────────────────
# 3 — DCF → Sensitivity integration
# ─────────────────────────────────────────────────────────────────────────────

class TestDcfToSensitivity:

    @pytest.fixture
    def base_kwargs(self):
        income_stmts, balance_sheets, cash_flows = _make_fixture_statements()
        return _base_dcf_kwargs(income_stmts, balance_sheets, cash_flows)

    def test_wacc_growth_sensitivity_returns_dict(self, base_kwargs):
        result = wacc_growth_sensitivity(
            base_dcf_kwargs=base_kwargs,
            wacc_range=[0.08, 0.09, 0.10],
            growth_range=[0.025, 0.030, 0.035],
            net_debt=10_000,
            shares_mm=500,
        )
        assert isinstance(result, dict)

    def test_wacc_growth_sensitivity_has_required_keys(self, base_kwargs):
        result = wacc_growth_sensitivity(
            base_dcf_kwargs=base_kwargs,
            wacc_range=[0.09, 0.10],
            growth_range=[0.025, 0.030],
        )
        for k in ("wacc_range", "growth_range", "ev_table", "price_table"):
            assert k in result, f"Missing key: {k}"

    def test_ev_table_populated(self, base_kwargs):
        result = wacc_growth_sensitivity(
            base_dcf_kwargs=base_kwargs,
            wacc_range=[0.08, 0.09, 0.10],
            growth_range=[0.025, 0.030, 0.035],
        )
        assert len(result["ev_table"]) > 0

    def test_higher_wacc_lower_ev(self, base_kwargs):
        """Economic check: higher WACC → lower EV (all else equal)."""
        result = wacc_growth_sensitivity(
            base_dcf_kwargs=base_kwargs,
            wacc_range=[0.08, 0.12],
            growth_range=[0.030],
        )
        ev_table = result["ev_table"]
        ev_low_wacc  = ev_table.get((0.08, 0.03)) or ev_table.get((0.0800, 0.0300))
        ev_high_wacc = ev_table.get((0.12, 0.03)) or ev_table.get((0.1200, 0.0300))
        if ev_low_wacc and ev_high_wacc:
            assert ev_low_wacc > ev_high_wacc

    def test_scenario_analysis_returns_three_scenarios(self, base_kwargs):
        results = run_scenario_analysis(base_kwargs)
        assert set(results.keys()) == {"bull", "base", "bear"}

    def test_scenario_results_are_dcf_results(self, base_kwargs):
        results = run_scenario_analysis(base_kwargs)
        for name, r in results.items():
            assert isinstance(r, DCFResult), f"Scenario {name!r} is not a DCFResult"

    def test_bull_ev_greater_than_bear_ev(self, base_kwargs):
        results = run_scenario_analysis(base_kwargs)
        assert results["bull"].enterprise_value > results["bear"].enterprise_value

    def test_scenario_summary_table_shape(self, base_kwargs):
        results = run_scenario_analysis(base_kwargs)
        table = scenario_summary_table(results, net_debt=10_000, shares_mm=500)
        assert len(table) == 3
        for row in table:
            assert "enterprise_value" in row
            assert "price_per_share"  in row
            assert "wacc"             in row


# ─────────────────────────────────────────────────────────────────────────────
# 4 — DCF → Output (JSON + Excel) integration
# ─────────────────────────────────────────────────────────────────────────────

class TestDcfToOutputPipeline:

    @pytest.fixture
    def dcf_and_meta(self):
        income_stmts, balance_sheets, cash_flows = _make_fixture_statements()
        r = run_dcf(**_base_dcf_kwargs(income_stmts, balance_sheets, cash_flows))
        nd = compute_net_debt(balance_sheets[0])
        return r, nd

    def test_format_summary_from_real_dcf(self, dcf_and_meta):
        dcf_result, net_debt = dcf_and_meta
        summary = format_valuation_summary(
            "ACME", dcf_result, net_debt=net_debt, shares_mm=500.0
        )
        assert isinstance(summary, dict)
        assert summary["enterprise_value"] == pytest.approx(dcf_result.enterprise_value)

    def test_equity_value_correct(self, dcf_and_meta):
        dcf_result, net_debt = dcf_and_meta
        summary = format_valuation_summary(
            "ACME", dcf_result, net_debt=net_debt, shares_mm=500.0
        )
        assert summary["equity_value"] == pytest.approx(dcf_result.enterprise_value - net_debt)

    def test_intrinsic_price_positive(self, dcf_and_meta):
        dcf_result, net_debt = dcf_and_meta
        summary = format_valuation_summary(
            "ACME", dcf_result, net_debt=net_debt, shares_mm=500.0
        )
        assert summary["intrinsic_price"] > 0

    def test_print_summary_no_crash(self, dcf_and_meta, capsys):
        dcf_result, net_debt = dcf_and_meta
        summary = format_valuation_summary(
            "ACME", dcf_result, net_debt=net_debt, shares_mm=500.0,
            current_price=120.0,
        )
        print_valuation_summary(summary)

    def test_write_json_creates_valid_file(self, dcf_and_meta, tmp_path):
        dcf_result, net_debt = dcf_and_meta
        summary = format_valuation_summary(
            "ACME", dcf_result, net_debt=net_debt, shares_mm=500.0
        )
        out = str(tmp_path / "result.json")
        write_json_output(summary, out)
        with open(out) as fh:
            data = json.load(fh)
        assert data["ticker"] == "ACME"
        assert data["enterprise_value"] == pytest.approx(dcf_result.enterprise_value)

    def test_write_excel_creates_valid_file(self, dcf_and_meta, tmp_path):
        dcf_result, net_debt = dcf_and_meta
        out = str(tmp_path / "result.xlsx")
        write_excel_output(
            out, "ACME", dcf_result,
            net_debt=net_debt, shares_mm=500.0, current_price=120.0,
        )
        wb = openpyxl.load_workbook(out, data_only=True)
        assert "Summary" in wb.sheetnames
        assert "DCF"     in wb.sheetnames
        wb.close()

    def test_full_pipeline_json_roundtrip(self, tmp_path):
        """Complete pipeline: statements → DCF → JSON → reload → verify EV."""
        income_stmts, balance_sheets, cash_flows = _make_fixture_statements()
        dcf_result = run_dcf(**_base_dcf_kwargs(income_stmts, balance_sheets, cash_flows))
        net_debt   = compute_net_debt(balance_sheets[0])

        # Scenario analysis
        base_kw = _base_dcf_kwargs(income_stmts, balance_sheets, cash_flows)
        scenarios = run_scenario_analysis(base_kw)
        scen_table = scenario_summary_table(scenarios, net_debt=net_debt, shares_mm=500)

        # Sensitivity
        sens = wacc_growth_sensitivity(
            base_dcf_kwargs=base_kw,
            wacc_range=[0.08, 0.09, 0.10],
            growth_range=[0.025, 0.030, 0.035],
            net_debt=net_debt,
            shares_mm=500,
        )

        # Format + write JSON
        summary = format_valuation_summary(
            "ACME", dcf_result,
            net_debt=net_debt, shares_mm=500,
            current_price=150.0,
            scenario_table=scen_table,
        )
        json_path = str(tmp_path / "pipeline_result.json")
        write_json_output(summary, json_path)

        # Reload and verify
        with open(json_path) as fh:
            loaded = json.load(fh)
        assert loaded["ticker"] == "ACME"
        assert loaded["enterprise_value"] == pytest.approx(dcf_result.enterprise_value, rel=1e-4)
        assert loaded["equity_value"] == pytest.approx(dcf_result.enterprise_value - net_debt, rel=1e-4)
        assert "scenarios" in loaded
        assert len(loaded["scenarios"]) == 3

    def test_full_pipeline_excel_all_sheets(self, tmp_path):
        """Complete pipeline → Excel with sensitivity, scenarios, assumptions."""
        income_stmts, balance_sheets, cash_flows = _make_fixture_statements()
        dcf_result = run_dcf(**_base_dcf_kwargs(income_stmts, balance_sheets, cash_flows))
        net_debt   = compute_net_debt(balance_sheets[0])

        base_kw = _base_dcf_kwargs(income_stmts, balance_sheets, cash_flows)
        scenarios  = run_scenario_analysis(base_kw)
        scen_table = scenario_summary_table(scenarios, net_debt=net_debt, shares_mm=500)
        sens = wacc_growth_sensitivity(
            base_dcf_kwargs=base_kw,
            wacc_range=[0.08, 0.09, 0.10],
            growth_range=[0.025, 0.030, 0.035],
        )
        assumptions = {
            "ticker": "ACME", "wacc": 0.09, "terminal_growth": 0.03,
            "near_term_growth": 0.08, "target_ebit_margin": 0.18,
        }

        out = str(tmp_path / "full_pipeline.xlsx")
        write_excel_output(
            out, "ACME", dcf_result,
            net_debt=net_debt, shares_mm=500,
            current_price=150.0,
            scenario_table=scen_table,
            sensitivity_wacc_g=sens,
            assumptions=assumptions,
        )

        wb = openpyxl.load_workbook(out, data_only=True)
        assert "Summary"     in wb.sheetnames
        assert "DCF"         in wb.sheetnames
        assert "Sensitivity" in wb.sheetnames
        assert "Assumptions" in wb.sheetnames
        wb.close()


# ─────────────────────────────────────────────────────────────────────────────
# 5 — Validation integration
# ─────────────────────────────────────────────────────────────────────────────

class TestValidationIntegration:

    def test_run_all_data_checks_returns_list(self):
        income_stmts, balance_sheets, cash_flows = _make_fixture_statements()
        results = run_all_data_checks(income_stmts, balance_sheets, cash_flows)
        assert isinstance(results, list)
        assert len(results) > 0

    def test_clean_data_passes_validation(self):
        income_stmts, balance_sheets, cash_flows = _make_fixture_statements()
        results = run_all_data_checks(income_stmts, balance_sheets, cash_flows)
        # Fixture data is well-formed — errors should be zero or minimal
        errors = [r for r in results if not r.is_ok() and r.severity == "error"]
        assert len(errors) == 0, f"Validation errors: {[r.message for r in errors]}"

    def test_diluted_shares_computation(self):
        dil = compute_fully_diluted_shares(
            basic_shares_mm=490.0,
            options_outstanding_mm=15.0,
            options_avg_strike=80.0,
            current_price=120.0,
        )
        assert "fully_diluted_mm" in dil
        assert dil["fully_diluted_mm"] >= 490.0

    def test_price_per_share_computation(self):
        equity = 80_000.0   # USD mm
        shares = 500.0      # mm
        price  = compute_price_per_share(equity, shares)
        assert price == pytest.approx(160.0)


# ─────────────────────────────────────────────────────────────────────────────
# 6 — Economic / financial sanity checks
# ─────────────────────────────────────────────────────────────────────────────

class TestEconomicSanity:
    """Verify that DCF outputs respect fundamental valuation principles."""

    @pytest.fixture
    def base_result(self):
        income_stmts, balance_sheets, cash_flows = _make_fixture_statements()
        return run_dcf(**_base_dcf_kwargs(income_stmts, balance_sheets, cash_flows))

    def test_higher_growth_increases_ev(self):
        income_stmts, balance_sheets, cash_flows = _make_fixture_statements()
        base = _base_dcf_kwargs(income_stmts, balance_sheets, cash_flows)

        r_low  = run_dcf(**{**base, "near_term_growth": 0.04})
        r_high = run_dcf(**{**base, "near_term_growth": 0.14})
        assert r_high.enterprise_value > r_low.enterprise_value

    def test_higher_wacc_decreases_ev(self):
        income_stmts, balance_sheets, cash_flows = _make_fixture_statements()
        base = _base_dcf_kwargs(income_stmts, balance_sheets, cash_flows)

        r_low_wacc  = run_dcf(**{**base, "wacc": 0.07})
        r_high_wacc = run_dcf(**{**base, "wacc": 0.13})
        assert r_low_wacc.enterprise_value > r_high_wacc.enterprise_value

    def test_higher_margin_increases_ev(self):
        income_stmts, balance_sheets, cash_flows = _make_fixture_statements()
        base = _base_dcf_kwargs(income_stmts, balance_sheets, cash_flows)

        r_low  = run_dcf(**{**base, "target_ebit_margin": 0.10})
        r_high = run_dcf(**{**base, "target_ebit_margin": 0.25})
        assert r_high.enterprise_value > r_low.enterprise_value

    def test_terminal_growth_below_wacc_required(self, base_result):
        """Terminal growth must be < WACC for GGM to be valid."""
        assert base_result.terminal_growth < base_result.wacc

    def test_pv_ufcf_each_year_positive(self, base_result):
        """All discounted forecast FCFs should be positive (positive UFCF)."""
        for fy in base_result.forecast_years_data:
            assert fy.pv_ufcf > 0, f"Year {fy.year}: negative PV(UFCF)"

    def test_discount_factors_decreasing(self, base_result):
        """Discount factors must decrease monotonically with time."""
        dfs = [fy.discount_factor for fy in base_result.forecast_years_data]
        for i in range(len(dfs) - 1):
            assert dfs[i] > dfs[i + 1], f"Non-decreasing factor at year {i+1}"
