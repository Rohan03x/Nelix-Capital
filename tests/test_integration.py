"""
tests/test_integration.py — End-to-end integration tests using mock data.

No live API calls. All external I/O is mocked via pytest monkeypatching.

Tests:
  - Full DCF pipeline with clean synthetic NKE data
  - Sensitivity table generation
  - Report formatting + JSON serialisation
  - Excel workbook written and readable
  - Scenario analysis produces ordered EVs
  - Data validation returns results for valid data
  - TTM computation end-to-end
"""

from __future__ import annotations

import json
import os
import tempfile
from pathlib import Path

import pytest

from auto_valuation.data.cleaner import standardise_field_names, unit_normalize
from auto_valuation.data.fiscal_year import compute_ttm
from auto_valuation.data.bridge import compute_net_debt, compute_equity_value
from auto_valuation.forecast.dcf import run_dcf, DCFResult
from auto_valuation.assumptions.wacc import build_wacc
from auto_valuation.assumptions.growth import build_growth_assumptions
from auto_valuation.model.income_statement import normalise_tax_rate
from auto_valuation.model.dilution import compute_fully_diluted_shares, compute_price_per_share
from auto_valuation.sensitivity.analysis import (
    wacc_growth_sensitivity,
    run_scenario_analysis,
    scenario_summary_table,
)
from auto_valuation.data.transactions import (
    compute_transaction_multiples,
    compute_transaction_comps_result,
)
from auto_valuation.data.comps import build_football_field
from auto_valuation.output.report import format_valuation_summary, write_json_output
from auto_valuation.output.excel import write_excel_output
from auto_valuation.validation.checks import run_all_data_checks
from auto_valuation.utils.error import DataQualityError

_PROFILE = {"currency": "USD"}


def _prep(records: list[dict]) -> list[dict]:
    """Standardise a list of FMP-style statement dicts to canonical names."""
    return standardise_field_names(unit_normalize(records, _PROFILE))


# ─────────────────────────────────────────────────────────────────────────────
# TTM computation
# ─────────────────────────────────────────────────────────────────────────────

class TestTTM:
    def test_revenue_sums_four_quarters(self):
        # compute_ttm handles both camelCase and canonical field names.
        q_is = [
            {"date": "2023-02-28", "revenue": 12500, "operatingIncome": 1500,
             "depreciationAndAmortization": 130, "incomeTaxExpense": 310,
             "netIncome": 1300, "period": "Q3"},
            {"date": "2022-11-30", "revenue": 13000, "operatingIncome": 1600,
             "depreciationAndAmortization": 130, "incomeTaxExpense": 320,
             "netIncome": 1350, "period": "Q2"},
            {"date": "2022-08-31", "revenue": 12800, "operatingIncome": 1550,
             "depreciationAndAmortization": 125, "incomeTaxExpense": 300,
             "netIncome": 1280, "period": "Q1"},
            {"date": "2022-05-31", "revenue": 11917, "operatingIncome": 1551,
             "depreciationAndAmortization": 130, "incomeTaxExpense": 281,
             "netIncome": 1140, "period": "Q4"},
        ]
        q_cf = [{"date": "2023-02-28", "capitalExpenditure": -220,
                  "stockBasedCompensation": 120, "operatingCashFlow": 1100,
                  "period": "Q3"}]
        q_bs = [{"date": "2023-02-28", "cash": 9000,
                  "long_term_debt": 8000, "total_equity": 9500,
                  "period": "Q3"}]
        ttm = compute_ttm(q_is, q_cf, q_bs)
        expected_rev = 12500 + 13000 + 12800 + 11917
        assert abs(ttm.get("revenue", 0) - expected_rev) < 1.0

    def test_balance_sheet_uses_latest(self):
        q_bs = [
            {"date": "2023-02-28", "cash": 9000, "long_term_debt": 8000,
             "total_equity": 9500, "period": "Q3"},
            {"date": "2022-11-30", "cash": 7000, "long_term_debt": 9000,
             "total_equity": 8500, "period": "Q2"},
        ]
        ttm = compute_ttm([], [], q_bs)
        assert isinstance(ttm, dict)
        # Most-recent-first: cash should be from first (latest) BS record
        assert ttm.get("cash") == 9000


# ─────────────────────────────────────────────────────────────────────────────
# Data validation
# ─────────────────────────────────────────────────────────────────────────────

class TestValidation:
    @pytest.fixture
    def std_data(self, fake_income_statement, fake_balance_sheet, fake_cash_flow):
        return (
            _prep(fake_income_statement),
            _prep(fake_balance_sheet),
            _prep(fake_cash_flow),
        )

    def test_run_all_data_checks_returns_list(self, std_data):
        is_std, bs_std, cf_std = std_data
        results = run_all_data_checks(is_std, bs_std, cf_std)
        assert isinstance(results, list)
        assert len(results) >= 1

    def test_results_have_name_and_status(self, std_data):
        is_std, bs_std, cf_std = std_data
        results = run_all_data_checks(is_std, bs_std, cf_std)
        for r in results:
            assert hasattr(r, "name")
            assert hasattr(r, "status")
            assert r.status in ("PASS", "WARN", "FAIL")

    def test_empty_data_raises(self):
        with pytest.raises(DataQualityError):
            run_all_data_checks([], [], [])


# ─────────────────────────────────────────────────────────────────────────────
# Full pipeline — DCF with real model objects
# ─────────────────────────────────────────────────────────────────────────────

class TestFullPipeline:
    @pytest.fixture(autouse=True)
    def _setup(self, fake_income_statement, fake_balance_sheet, fake_cash_flow):
        self.income_stmts  = _prep(fake_income_statement)
        self.balance_sheets = _prep(fake_balance_sheet)
        self.cash_flows    = _prep(fake_cash_flow)
        self.latest_bs     = self.balance_sheets[0]
        self.net_debt      = compute_net_debt(self.latest_bs)

    def test_net_debt_computed(self):
        # NKE synthetic: LT debt 8925, cash 9403 → net cash ≈ −478
        assert self.net_debt < 500

    def test_wacc_in_plausible_range(self):
        d = build_wacc(
            market_cap=151_000,
            total_debt=8_925,
            preferred_stock=0.0,
            basic_shares_mm=1_500,
            current_price=100.0,
            risk_free_rate=0.043,
            equity_risk_premium=0.055,
            beta=0.87,
            pre_tax_cost_of_debt=0.04,
            tax_rate=0.21,
        )
        assert 0.06 < d["wacc"] < 0.15

    def test_tax_rate_normalised(self):
        t = normalise_tax_rate(self.income_stmts, statutory_rate=0.21, years=3)
        assert 0.05 <= t <= 0.35

    def test_dcf_ev_positive(self):
        res = run_dcf(
            ticker="NKE",
            scenario="base",
            income_stmts=self.income_stmts,
            cash_flows=self.cash_flows,
            balance_sheets=self.balance_sheets,
            wacc=0.088,
            terminal_growth=0.025,
            near_term_growth=0.08,
            target_ebit_margin=0.135,
            forecast_years=5,
            tax_rate_override=0.21,
        )
        assert res.enterprise_value > 0

    def test_equity_bridge(self):
        ev = 111_825
        eq = compute_equity_value(ev, self.latest_bs)
        assert isinstance(eq, float)

    def test_price_per_share(self):
        ev    = 111_825
        eq    = ev - self.net_debt
        dil   = compute_fully_diluted_shares(basic_shares_mm=1500, current_price=100)
        price = compute_price_per_share(eq, dil["fully_diluted_mm"])
        assert price > 0


# ─────────────────────────────────────────────────────────────────────────────
# Sensitivity
# ─────────────────────────────────────────────────────────────────────────────

class TestSensitivity:
    @pytest.fixture(autouse=True)
    def _setup(self, fake_income_statement, fake_balance_sheet, fake_cash_flow):
        self.kwargs = dict(
            ticker="NKE",
            scenario="base",
            income_stmts=_prep(fake_income_statement),
            cash_flows=_prep(fake_cash_flow),
            balance_sheets=_prep(fake_balance_sheet),
            wacc=0.088,
            terminal_growth=0.025,
            near_term_growth=0.08,
            target_ebit_margin=0.135,
            forecast_years=5,
            tax_rate_override=0.21,
        )

    def test_wacc_growth_table_shape(self):
        sens = wacc_growth_sensitivity(
            self.kwargs,
            wacc_range=[0.08, 0.09, 0.10],
            growth_range=[0.020, 0.025],
        )
        # Up to 3×2 = 6 valid combos (some may be skipped if WACC ≤ g)
        assert len(sens["ev_table"]) >= 4

    def test_lower_wacc_higher_ev(self):
        sens = wacc_growth_sensitivity(
            self.kwargs,
            wacc_range=[0.08, 0.12],
            growth_range=[0.025],
        )
        keys = list(sens["ev_table"].keys())
        if len(keys) >= 2:
            evs = sorted([(k[0], v) for k, v in sens["ev_table"].items()], key=lambda x: x[0])
            assert evs[0][1] > evs[-1][1]  # lower WACC → higher EV

    def test_scenario_analysis_ordered(self):
        results = run_scenario_analysis(self.kwargs)
        assert "base" in results
        if "bull" in results and "bear" in results:
            assert results["bull"].enterprise_value > results["base"].enterprise_value > results["bear"].enterprise_value

    def test_scenario_summary_table_keys(self):
        results = run_scenario_analysis(self.kwargs, scenarios=["base"])
        table = scenario_summary_table(results, net_debt=0, shares_mm=1500)
        assert len(table) == 1
        row = table[0]
        assert "enterprise_value" in row
        assert "price_per_share"  in row


# ─────────────────────────────────────────────────────────────────────────────
# Transactions + football field
# ─────────────────────────────────────────────────────────────────────────────

class TestTransactionIntegration:
    @pytest.fixture
    def deals(self):
        return [
            {"target": "Acme", "ev_mm": 12000, "ebitda_mm": 600, "revenue_mm": 3000},
            {"target": "Beta", "ev_mm": 9000,  "ebitda_mm": 500, "revenue_mm": 2500},
        ]

    def test_pipeline(self, deals):
        mults  = compute_transaction_multiples(deals)
        result = compute_transaction_comps_result(700, 3500, mults)
        assert result.get("blended_ev_range", {}).get("low", 0) > 0

    def test_football_field_from_txn(self, deals):
        mults  = compute_transaction_multiples(deals)
        result = compute_transaction_comps_result(700, 3500, mults)
        lo = result["blended_ev_range"]["low"]
        hi = result["blended_ev_range"]["high"]
        ff = build_football_field(80000, 120000, lo, hi,
                                   transactions_ev_low=lo, transactions_ev_high=hi,
                                   net_debt=2000, shares_mm=1500)
        assert len(ff) >= 3


# ─────────────────────────────────────────────────────────────────────────────
# Output: report + Excel
# ─────────────────────────────────────────────────────────────────────────────

class TestOutputIntegration:
    @pytest.fixture(autouse=True)
    def _setup(self, fake_income_statement, fake_balance_sheet, fake_cash_flow):
        self.dcf = run_dcf(
            ticker="NKE",
            scenario="base",
            income_stmts=_prep(fake_income_statement),
            cash_flows=_prep(fake_cash_flow),
            balance_sheets=_prep(fake_balance_sheet),
            wacc=0.088,
            terminal_growth=0.025,
            near_term_growth=0.08,
            target_ebit_margin=0.135,
            forecast_years=5,
            tax_rate_override=0.21,
        )
        self.net_debt  = 478.0
        self.shares_mm = 1515.0

    def test_format_summary_keys(self):
        summary = format_valuation_summary(
            ticker="NKE",
            dcf_result=self.dcf,
            net_debt=self.net_debt,
            shares_mm=self.shares_mm,
            current_price=96.42,
        )
        for key in ("ticker", "enterprise_value", "equity_value",
                    "intrinsic_price", "wacc", "terminal_growth"):
            assert key in summary

    def test_json_serialisable(self):
        summary = format_valuation_summary("NKE", self.dcf, self.net_debt, self.shares_mm)
        with tempfile.TemporaryDirectory() as td:
            path = os.path.join(td, "out.json")
            write_json_output(summary, path)
            assert os.path.exists(path)
            with open(path) as fh:
                data = json.load(fh)
            assert data["ticker"] == "NKE"

    def test_excel_created(self):
        with tempfile.TemporaryDirectory() as td:
            path = os.path.join(td, "NKE_test.xlsx")
            out  = write_excel_output(
                output_path=path,
                ticker="NKE",
                dcf_result=self.dcf,
                net_debt=self.net_debt,
                shares_mm=self.shares_mm,
                current_price=96.42,
                assumptions={"wacc": 0.088},
            )
            assert os.path.exists(out)
            assert os.path.getsize(out) > 1000

    def test_excel_has_expected_sheets(self):
        import openpyxl
        with tempfile.TemporaryDirectory() as td:
            path = os.path.join(td, "NKE_test.xlsx")
            write_excel_output(
                output_path=path,
                ticker="NKE",
                dcf_result=self.dcf,
                net_debt=self.net_debt,
                shares_mm=self.shares_mm,
                assumptions={"wacc": 0.088},
            )
            wb = openpyxl.load_workbook(path)
            assert "Summary" in wb.sheetnames
            assert "DCF"     in wb.sheetnames
