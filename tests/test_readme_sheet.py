"""Tests for output/readme_sheet.py."""
import pytest
from datetime import date

openpyxl = pytest.importorskip("openpyxl")

from openpyxl import Workbook
from auto_valuation.output.readme_sheet import write_readme_sheet

_VALUATION = {
    "enterprise_value_mm": 2_800_000,
    "equity_value_mm": 2_700_000,
    "implied_share_price": 173.5,
    "current_price": 150.0,
    "upside_pct": 0.157,
    "wacc": 0.09,
    "terminal_g": 0.025,
    "tv_pct_of_ev": 0.72,
    "net_debt_mm": 100_000,
    "diluted_shares_mm": 15_552,
}


class TestWriteReadmeSheet:
    def test_creates_readme_sheet(self):
        wb = Workbook()
        write_readme_sheet(wb, "AAPL", "Apple Inc.", _VALUATION, [])
        assert "README" in wb.sheetnames

    def test_readme_is_first_sheet(self):
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        wb.create_sheet("Model")
        write_readme_sheet(wb, "AAPL", "Apple Inc.", _VALUATION, [])
        assert wb.sheetnames[0] == "README"

    def test_no_crash_with_warnings(self):
        wb = Workbook()
        warnings = ["Capex unusually high", "WACC near terminal g"]
        write_readme_sheet(wb, "AAPL", "Apple Inc.", _VALUATION, [], warnings=warnings)
        assert "README" in wb.sheetnames

    def test_no_crash_with_market_wacc(self):
        wb = Workbook()
        write_readme_sheet(
            wb, "AAPL", "Apple Inc.", _VALUATION, [],
            market_implied_wacc=0.085,
            run_date=date(2024, 1, 15),
        )
        assert "README" in wb.sheetnames

    def test_replaces_existing_sheet(self):
        wb = Workbook()
        wb.create_sheet("README")
        write_readme_sheet(wb, "AAPL", "Apple Inc.", _VALUATION, [])
        assert wb.sheetnames.count("README") == 1

    def test_data_written(self):
        wb = Workbook()
        write_readme_sheet(wb, "AAPL", "Apple Inc.", _VALUATION, [])
        ws = wb["README"]
        assert ws.max_row >= 2
