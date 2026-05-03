"""Tests for output/assumptions_sheet.py."""
import pytest

openpyxl = pytest.importorskip("openpyxl")

from openpyxl import Workbook
from auto_valuation.output.assumptions_sheet import write_assumptions_sheet

_ASSUMPTIONS = {
    "rev_cagr_y1_3": 0.06,
    "rev_cagr_y4_7": 0.04,
    "ebit_margin_y1": 0.295,
    "ebit_margin_y5": 0.31,
    "da_pct_rev": 0.03,
    "capex_pct_rev": 0.028,
    "sbc_pct_rev": 0.027,
    "tax_rate": 0.21,
    "nowc_pct_rev": -0.046,
    "wacc": 0.09,
    "terminal_g": 0.025,
}

_OVERRIDES = {
    "wacc": 0.095,
}


class TestWriteAssumptionsSheet:
    def test_creates_assumptions_sheet(self):
        wb = Workbook()
        write_assumptions_sheet(wb, "AAPL", _ASSUMPTIONS)
        assert "Assumptions" in wb.sheetnames

    def test_with_overrides(self):
        wb = Workbook()
        write_assumptions_sheet(wb, "AAPL", _ASSUMPTIONS, overrides=_OVERRIDES)
        assert "Assumptions" in wb.sheetnames

    def test_replaces_existing_sheet(self):
        wb = Workbook()
        wb.create_sheet("Assumptions")
        write_assumptions_sheet(wb, "AAPL", _ASSUMPTIONS)
        assert wb.sheetnames.count("Assumptions") == 1

    def test_no_crash_empty_assumptions(self):
        wb = Workbook()
        write_assumptions_sheet(wb, "AAPL", {})

    def test_data_written(self):
        wb = Workbook()
        write_assumptions_sheet(wb, "AAPL", _ASSUMPTIONS)
        ws = wb["Assumptions"]
        assert ws.max_row >= 2
