"""Tests for output/formatting.py."""
import pytest

openpyxl = pytest.importorskip("openpyxl")

from openpyxl import Workbook
from auto_valuation.output.formatting import (
    apply_tab_colors,
    apply_freeze_panes,
    set_print_area_and_page_setup,
    apply_conditional_formatting,
    protect_formula_sheets,
    apply_workbook_formatting,
)

_SHEETS = ["README", "Model", "Assumptions", "Comps", "Macro",
           "Raw IS", "Raw BS", "Raw CF"]


def _make_wb() -> Workbook:
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    for name in _SHEETS:
        wb.create_sheet(name)
    return wb


class TestApplyTabColors:
    def test_no_crash(self):
        wb = _make_wb()
        apply_tab_colors(wb)

    def test_empty_workbook(self):
        wb = Workbook()
        apply_tab_colors(wb)


class TestApplyFreezePanes:
    def test_no_crash(self):
        wb = _make_wb()
        apply_freeze_panes(wb)


class TestSetPrintArea:
    def test_no_crash(self):
        wb = _make_wb()
        set_print_area_and_page_setup(wb)


class TestApplyConditionalFormatting:
    def test_no_crash(self):
        wb = _make_wb()
        apply_conditional_formatting(wb)


class TestProtectFormulaSheets:
    def test_no_crash(self):
        wb = _make_wb()
        protect_formula_sheets(wb, password="")


class TestApplyWorkbookFormatting:
    def test_no_crash_no_protect(self):
        wb = _make_wb()
        apply_workbook_formatting(wb, protect=False)

    def test_with_protect(self):
        wb = _make_wb()
        apply_workbook_formatting(wb, protect=True)
