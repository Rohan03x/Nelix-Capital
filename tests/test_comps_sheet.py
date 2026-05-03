"""Tests for output/comps_sheet.py."""
import pytest

openpyxl = pytest.importorskip("openpyxl")

from openpyxl import Workbook
from auto_valuation.output.comps_sheet import write_comps_sheet

_PEERS = [
    {"ticker": "MSFT", "company_name": "Microsoft",
     "ev_mm": 3_000_000, "ebitda_ntm": 120_000, "ebit_ntm": 100_000,
     "revenue_ntm": 240_000, "net_income_ntm": 80_000,
     "ev_ebitda": 25.0, "ev_ebit": 30.0, "ev_revenue": 12.5,
     "pe": 37.5, "fcf_yield": 0.025, "roe": 0.42, "roic": 0.28},
    {"ticker": "GOOGL", "company_name": "Alphabet",
     "ev_mm": 1_800_000, "ebitda_ntm": 90_000, "ebit_ntm": 75_000,
     "revenue_ntm": 310_000, "net_income_ntm": 60_000,
     "ev_ebitda": 20.0, "ev_ebit": 24.0, "ev_revenue": 5.8,
     "pe": 30.0, "fcf_yield": 0.030, "roe": 0.22, "roic": 0.18},
]

_SUBJECT = {
    "ticker": "AAPL", "company_name": "Apple",
    "ev_mm": 2_800_000, "ebitda_ntm": 130_000, "ebit_ntm": 114_000,
    "revenue_ntm": 400_000, "net_income_ntm": 97_000,
    "ev_ebitda": 21.5, "ev_ebit": 24.6, "ev_revenue": 7.0,
    "pe": 29.0, "fcf_yield": 0.034, "roe": 1.60, "roic": 0.50,
}


class TestWriteCompsSheet:
    def test_creates_comps_sheet(self):
        wb = Workbook()
        write_comps_sheet(wb, "AAPL", _PEERS, subject=_SUBJECT)
        assert "Comps" in wb.sheetnames

    def test_no_crash_empty_peers(self):
        wb = Workbook()
        write_comps_sheet(wb, "AAPL", [])
        assert "Comps" in wb.sheetnames

    def test_proforma_flagged_no_crash(self):
        wb = Workbook()
        write_comps_sheet(wb, "AAPL", _PEERS, proforma_flagged=["MSFT"])
        assert "Comps" in wb.sheetnames

    def test_replaces_existing_sheet(self):
        wb = Workbook()
        wb.create_sheet("Comps")
        write_comps_sheet(wb, "AAPL", _PEERS)
        assert wb.sheetnames.count("Comps") == 1

    def test_data_written(self):
        wb = Workbook()
        write_comps_sheet(wb, "AAPL", _PEERS)
        ws = wb["Comps"]
        assert ws.max_row >= 2
