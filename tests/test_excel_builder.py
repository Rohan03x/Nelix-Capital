"""Tests for output/excel_builder.py."""
import pytest
import tempfile
import os
from datetime import date
from pathlib import Path

openpyxl = pytest.importorskip("openpyxl")

from auto_valuation.output.excel_builder import build_workbook


def _make_forecast_year(yr):
    class FY:
        pass
    fy = FY()
    fy.year = yr
    fy.revenue = 400_000 + yr * 10_000
    fy.ebit = 120_000
    fy.ebit_margin = 0.30
    fy.nopat = 94_800
    fy.da = 12_000
    fy.sbc = 10_000
    fy.capex = 11_000
    fy.delta_nowc = -500
    fy.ufcf = 95_300
    fy.ocf = 107_000
    fy.net_income = 97_000
    fy.tax_rate = 0.21
    fy.ibd = 100_000
    fy.interest_expense = 5_000
    return fy


_FORECAST = [_make_forecast_year(yr) for yr in range(1, 8)]

_HIST_DATA = {
    "income": [
        {"date": "2023-12-31", "revenue": 383_285, "ebit": 114_301,
         "netIncome": 96_995, "da": 11_519, "sbc": 10_833,
         "capex": 10_959, "nowc": -17_000},
    ],
    "balance": [
        {"date": "2023-12-31", "totalAssets": 352_583, "totalLiabilities": 290_437,
         "totalEquity": 62_146, "cash": 29_965},
    ],
    "cashflow": [
        {"date": "2023-12-31", "operatingCashFlow": 110_543,
         "capitalExpenditure": -10_959, "netIncome": 96_995},
    ],
}

_VALUATION = {
    "enterprise_value_mm": 2_800_000,
    "equity_value_mm": 2_700_000,
    "implied_share_price": 173.5,
    "price_per_share": 173.5,
    "current_price": 150.0,
    "upside_pct": 0.157,
    "implied_upside_pct": 0.157,
    "wacc": 0.09,
    "terminal_g": 0.025,
    "tv_pct_of_ev": 0.72,
    "net_debt_mm": 100_000,
    "diluted_shares_mm": 15_552,
    "market_implied_wacc": 0.085,
}

_ASSUMPTIONS = {
    "rev_cagr_y1_3": 0.06,
    "ebit_margin_y5": 0.31,
    "wacc": 0.09,
    "terminal_g": 0.025,
}

_MACRO = {
    "risk_free_rate": 0.043,
    "risk_free_series": "GS10",
    "risk_free_currency": "USD",
    "equity_risk_premium": 0.056,
    "erp_source": "Damodaran",
    "country_risk_premium": 0.0,
    "country_name": "United States",
    "size_premium": 0.0,
    "beta_value": 1.29,
    "fx_rate": 1.0,
    "reporting_currency": "USD",
}


class TestBuildWorkbook:
    def test_returns_path(self, tmp_path):
        result = build_workbook(
            ticker="AAPL",
            company_name="Apple Inc.",
            valuation=_VALUATION,
            historical_data=_HIST_DATA,
            forecast_years=_FORECAST,
            assumptions=_ASSUMPTIONS,
            output_dir=str(tmp_path),
            run_date=date(2024, 1, 15),
        )
        assert isinstance(result, Path)

    def test_file_created(self, tmp_path):
        result = build_workbook(
            ticker="AAPL",
            company_name="Apple Inc.",
            valuation=_VALUATION,
            historical_data=_HIST_DATA,
            forecast_years=_FORECAST,
            assumptions=_ASSUMPTIONS,
            output_dir=str(tmp_path),
            run_date=date(2024, 1, 15),
        )
        assert result.exists()
        assert result.suffix == ".xlsx"

    def test_workbook_has_expected_sheets(self, tmp_path):
        from openpyxl import load_workbook
        result = build_workbook(
            ticker="AAPL",
            company_name="Apple Inc.",
            valuation=_VALUATION,
            historical_data=_HIST_DATA,
            forecast_years=_FORECAST,
            assumptions=_ASSUMPTIONS,
            macro_data=_MACRO,
            output_dir=str(tmp_path),
            run_date=date(2024, 1, 15),
        )
        wb = load_workbook(result)
        assert "README" in wb.sheetnames
        assert "Model" in wb.sheetnames
        assert "Assumptions" in wb.sheetnames

    def test_with_peers(self, tmp_path):
        peers = [
            {"ticker": "MSFT", "company_name": "Microsoft",
             "ev_mm": 3_000_000, "ebitda_ntm": 120_000, "ebit_ntm": 100_000,
             "revenue_ntm": 240_000, "net_income_ntm": 80_000,
             "ev_ebitda": 25.0, "ev_ebit": 30.0, "ev_revenue": 12.5,
             "pe": 37.5, "fcf_yield": 0.025, "roe": 0.42, "roic": 0.28},
        ]
        result = build_workbook(
            ticker="AAPL",
            company_name="Apple Inc.",
            valuation=_VALUATION,
            historical_data=_HIST_DATA,
            forecast_years=_FORECAST,
            assumptions=_ASSUMPTIONS,
            peers=peers,
            output_dir=str(tmp_path),
            run_date=date(2024, 1, 15),
        )
        from openpyxl import load_workbook
        wb = load_workbook(result)
        assert "Comps" in wb.sheetnames
