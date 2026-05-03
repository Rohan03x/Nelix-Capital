"""
webapp/data/excel_export.py
────────────────────────────
Build a complete, formatted Excel workbook from a dashboard data dict.
Uses openpyxl only. Returns raw bytes suitable for Flask send_file().

Sheets:
  1. Cover      — key metrics, recommendation, valuation summary
  2. DCF Model  — 7-year forecast schedule + EV bridge
  3. Sensitivity — WACC × g intrinsic-value grid
  4. Scenarios   — bear / base / bull comparison
  5. Historical  — raw historical financials
  6. Assumptions — model driver table
  7. Comps       — peer comparable table (if available)
"""

from __future__ import annotations

import io
import logging
from datetime import date
from typing import Any

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False
    logger.warning("openpyxl not installed — Excel export unavailable.")


# ─── Colour palette ───────────────────────────────────────────────────────────
_BG_DARK     = "0D1117"   # header background
_BG_MID      = "161B22"   # section header
_BG_LIGHT    = "1C2128"   # alternating row
_BG_WHITE    = "FFFFFF"
_FG_WHITE    = "FFFFFF"
_FG_GREY     = "8B949E"
_GREEN_BG    = "0D4F2A"
_GREEN_FG    = "3FB950"
_AMBER_BG    = "4D3800"
_AMBER_FG    = "E3B341"
_RED_BG      = "4D1A1A"
_RED_FG      = "F85149"
_ORANGE      = "F3811D"
_ACCENT      = "1F6FEB"   # blue accent


# ─── Style helpers ────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold: bool = False, color: str = _FG_WHITE, size: int = 10,
          italic: bool = False) -> Font:
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")


def _border_bottom() -> Border:
    thin = Side(style="thin", color="303038")
    return Border(bottom=thin)


def _align(h: str = "left", v: str = "center", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _write_header_row(ws, row: int, cols: list[str],
                      bg: str = _BG_MID, fg: str = _FG_GREY) -> None:
    for c, label in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.font      = _font(bold=True, color=fg, size=9)
        cell.fill      = _fill(bg)
        cell.alignment = _align("center")


def _rec_colors(rec_class: str) -> tuple[str, str]:
    """Return (bg, fg) hex pair for a recommendation class."""
    if rec_class == "green":
        return _GREEN_BG, _GREEN_FG
    if rec_class == "red":
        return _RED_BG, _RED_FG
    return _AMBER_BG, _AMBER_FG


def _set_col_widths(ws, widths: dict[int, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _money(v: Any) -> str:
    """Format a $M value as '$X,XXXm'."""
    try:
        return f"${float(v):,.0f}M"
    except Exception:
        return str(v)


def _pct(v: Any, dp: int = 1) -> str:
    try:
        return f"{float(v):.{dp}f}%"
    except Exception:
        return str(v)


def _num(v: Any, dp: int = 2) -> str:
    try:
        return f"{float(v):,.{dp}f}"
    except Exception:
        return str(v)


# ─── Sheet builders ───────────────────────────────────────────────────────────

def _write_cover(ws, d: dict) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = _ORANGE

    rec_class = d.get("recommendation_class", "amber")
    rec_bg, rec_fg = _rec_colors(rec_class)

    # ── Title block ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value     = f"{d.get('company_name', d.get('ticker', 'N/A'))}  ({d.get('ticker', '')})"
    t.font      = _font(bold=True, color=_FG_WHITE, size=16)
    t.fill      = _fill(_BG_DARK)
    t.alignment = _align("left")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:F2")
    sub = ws["A2"]
    sub.value     = (f"{d.get('exchange', '')} · {d.get('sector', '')} · "
                     f"{d.get('industry', '')} · as of {d.get('price_date', str(date.today()))}")
    sub.font      = _font(color=_FG_GREY, size=9, italic=True)
    sub.fill      = _fill(_BG_DARK)
    sub.alignment = _align("left")

    # ── Recommendation pill (row 4) ──────────────────────────────────────────
    ws.merge_cells("A4:C4")
    rp = ws["A4"]
    rp.value     = f"⬤  {d.get('recommendation', 'N/A').upper()}"
    rp.font      = Font(bold=True, color=rec_fg, size=20, name="Calibri")
    rp.fill      = _fill(rec_bg)
    rp.alignment = _align("center")
    ws.row_dimensions[4].height = 36

    ws.merge_cells("D4:F4")
    up_cell = ws["D4"]
    upside = d.get("upside_pct", 0)
    up_cell.value     = f"{'+' if upside >= 0 else ''}{upside:.1f}% upside"
    up_cell.font      = Font(bold=True, color=rec_fg, size=14, name="Calibri")
    up_cell.fill      = _fill(rec_bg)
    up_cell.alignment = _align("center")

    # ── Key metrics table (rows 6-16) ────────────────────────────────────────
    metrics = [
        ("METRIC",                    "VALUE",                  "METRIC",                  "VALUE"),
        ("Current Price",             f"${d.get('price', 0):.2f}",  "Intrinsic Value",    f"${d.get('intrinsic_value', 0):.2f}"),
        ("Market Cap",                _money(d.get("market_cap", 0)), "Enterprise Value",  _money(d.get("enterprise_value", 0))),
        ("WACC",                      _pct(d.get("wacc", 0)),        "Terminal Growth",    _pct(d.get("terminal_growth", 0))),
        ("Cost of Equity",            _pct(d.get("cost_of_equity", 0)), "Cost of Debt (pre-tax)", _pct(d.get("cost_of_debt_pre", 0))),
        ("Beta",                      f"{d.get('beta', 1):.2f}×",    "Risk-Free Rate",     _pct(d.get("risk_free_rate", 0))),
        ("Net Debt",                  _money(d.get("net_debt", 0)),  "Diluted Shares",     f"{d.get('diluted_shares', 0):.1f}M"),
        ("PV UFCFs",                  _money(d.get("pv_ufcfs", 0)),  "PV Terminal Value",  _money(d.get("pv_terminal", 0))),
        ("TV % of EV",                _pct(d.get("tv_pct", 0)),      "EBIT Margin (base)", _pct(d.get("ebit_margin_base", 0))),
        ("EBIT Margin (target Y7)",   _pct(d.get("ebit_margin_target", 0)), "Rev Growth (near)", _pct(d.get("revenue_growth_near", 0))),
        ("Confidence Score",          f"{d.get('confidence_score', 0)}/100", "Data Source", d.get("data_freshness", "Demo")),
        ("52-Wk Low",                 f"${d.get('fifty_two_week_low', 0):.2f}", "52-Wk High", f"${d.get('fifty_two_week_high', 0):.2f}"),
        ("Analyst Low",               f"${d.get('analyst_low', 0):.2f}", "Analyst High",  f"${d.get('analyst_high', 0):.2f}"),
    ]

    for r_off, row_data in enumerate(metrics):
        r = 6 + r_off
        is_hdr = r_off == 0
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            if is_hdr:
                cell.font      = _font(bold=True, color=_FG_GREY, size=9)
                cell.fill      = _fill(_BG_MID)
                cell.alignment = _align("center")
            elif c_idx in (1, 3):
                cell.font      = _font(color=_FG_GREY, size=10)
                cell.fill      = _fill(_BG_LIGHT if r % 2 == 0 else _BG_DARK)
                cell.alignment = _align("left")
            else:
                cell.font      = _font(bold=True, color=_FG_WHITE, size=10)
                cell.fill      = _fill(_BG_LIGHT if r % 2 == 0 else _BG_DARK)
                cell.alignment = _align("right")

    # ── Description (row 21) ─────────────────────────────────────────────────
    ws.merge_cells("A21:F21")
    desc = ws["A21"]
    desc.value     = d.get("description", "")
    desc.font      = _font(color=_FG_GREY, size=9, italic=True)
    desc.fill      = _fill(_BG_DARK)
    desc.alignment = _align("left", wrap=True)
    ws.row_dimensions[21].height = 48

    _set_col_widths(ws, {1: 28, 2: 18, 3: 28, 4: 18, 5: 4, 6: 4})


def _write_dcf_model(ws, d: dict) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = _ACCENT

    # Title
    ws.merge_cells("A1:K1")
    t = ws["A1"]
    t.value     = f"DCF Model — {d.get('company_name', d.get('ticker', ''))} ({d.get('ticker', '')})"
    t.font      = _font(bold=True, color=_FG_WHITE, size=14)
    t.fill      = _fill(_BG_DARK)
    t.alignment = _align("left")
    ws.row_dimensions[1].height = 28

    # Forecast header
    headers = ["Year", "n", "Revenue ($M)", "EBIT Margin", "EBIT ($M)",
               "NOPAT ($M)", "D&A ($M)", "SBC ($M)", "CapEx ($M)", "ΔNWC ($M)",
               "UFCF ($M)", "Disc. Factor", "PV UFCF ($M)"]
    _write_header_row(ws, 3, headers)

    forecast = d.get("forecast") or []
    for r_off, row in enumerate(forecast):
        r   = 4 + r_off
        alt = _BG_LIGHT if r_off % 2 == 0 else _BG_DARK
        vals = [
            row.get("year", ""),
            row.get("n", r_off + 1),
            row.get("revenue", 0),
            f"{row.get('ebit_m', 0):.1f}%",
            row.get("ebit", 0),
            row.get("nopat", 0),
            row.get("da", 0),
            row.get("sbc", 0),
            row.get("capex", 0),
            row.get("d_nowc", 0),
            row.get("ufcf", 0),
            f"{row.get('df', 0):.4f}",
            row.get("pv", 0),
        ]
        for c_idx, val in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            cell.fill      = _fill(alt)
            cell.alignment = _align("right" if c_idx > 1 else "left")
            bold_col = c_idx in (11, 13)  # UFCF and PV
            cell.font = _font(bold=bold_col, size=10)

    # EV Bridge
    bridge_row = 4 + len(forecast) + 2
    bridge_data = [
        ("EV BRIDGE",                     "",           "", "VALUE"),
        ("PV of UFCFs (Years 1–7)",        "",           "", _money(d.get("pv_ufcfs", 0))),
        ("PV of Terminal Value",           "",           "", _money(d.get("pv_terminal", 0))),
        ("Enterprise Value",               "",           "", _money(d.get("enterprise_value", 0))),
        ("(–) Net Debt",                   "",           "", _money(d.get("net_debt", 0))),
        ("Equity Value",                   "",           "", _money(d.get("equity_value", 0))),
        ("÷ Diluted Shares",               "",           "", f"{d.get('diluted_shares', 0):.1f}M"),
        ("Intrinsic Value per Share",      "",           "", f"${d.get('intrinsic_value', 0):.2f}"),
        ("Terminal Value % of EV",         "",           "", _pct(d.get("tv_pct", 0))),
    ]
    for r_off, (label, _, __, val) in enumerate(bridge_data):
        r   = bridge_row + r_off
        is_hdr = r_off == 0
        for c_idx, content in enumerate([label, val], start=1):
            col = 1 if c_idx == 1 else 4
            cell = ws.cell(row=r, column=col, value=content)
            cell.fill = _fill(_BG_MID if is_hdr else (_BG_LIGHT if r_off % 2 == 0 else _BG_DARK))
            is_total = label in ("Enterprise Value", "Equity Value", "Intrinsic Value per Share")
            cell.font = _font(bold=is_total or is_hdr,
                              color=_FG_GREY if is_hdr else (_ORANGE if is_total else _FG_WHITE))
            cell.alignment = _align("right" if c_idx == 2 else "left")

    _set_col_widths(ws, {1: 12, 2: 4, 3: 16, 4: 14, 5: 14, 6: 14,
                         7: 12, 8: 12, 9: 12, 10: 12, 11: 14, 12: 14, 13: 14})


def _write_sensitivity(ws, d: dict) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = _AMBER_FG

    sens = d.get("sensitivity") or {}
    wacc_labels = sens.get("wacc_labels", [])
    g_labels    = sens.get("g_labels",    [])
    iv_grid     = sens.get("iv_grid",     [])
    base_w_idx  = sens.get("base_wacc_idx", 2)
    base_g_idx  = sens.get("base_g_idx",   2)

    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value     = f"Sensitivity: WACC vs Terminal Growth — {d.get('ticker', '')}"
    t.font      = _font(bold=True, color=_FG_WHITE, size=13)
    t.fill      = _fill(_BG_DARK)
    t.alignment = _align("left")
    ws.row_dimensions[1].height = 24

    price = d.get("price", 0)

    # Header row: blank + WACC labels
    header_row = 3
    ws.cell(row=header_row, column=1, value="g ↓ / WACC →").font = _font(bold=True, color=_FG_GREY, size=9)
    ws.cell(row=header_row, column=1).fill = _fill(_BG_MID)
    ws.cell(row=header_row, column=1).alignment = _align("center")

    for c_idx, label in enumerate(wacc_labels, start=2):
        cell = ws.cell(row=header_row, column=c_idx, value=label)
        cell.font      = _font(bold=True, color=_FG_GREY, size=9)
        cell.fill      = _fill(_BG_MID)
        cell.alignment = _align("center")

    # Data rows
    for r_off, (g_label, row) in enumerate(zip(g_labels, iv_grid)):
        r = header_row + 1 + r_off
        g_cell = ws.cell(row=r, column=1, value=g_label)
        g_cell.font      = _font(bold=True, color=_FG_GREY, size=9)
        g_cell.fill      = _fill(_BG_LIGHT)
        g_cell.alignment = _align("center")

        for c_off, iv in enumerate(row):
            c   = 2 + c_off
            val = f"${iv:.1f}" if iv is not None else "N/A"
            cell = ws.cell(row=r, column=c, value=val)

            is_base = (r_off == base_g_idx and c_off == base_w_idx)
            is_above = iv is not None and price > 0 and iv > price * 1.15
            is_below = iv is not None and price > 0 and iv < price * 0.85

            if is_base:
                bg, fg = _ORANGE, _BG_DARK
            elif is_above:
                bg, fg = _GREEN_BG, _GREEN_FG
            elif is_below:
                bg, fg = _RED_BG, _RED_FG
            else:
                bg, fg = _BG_DARK, _FG_WHITE

            cell.font      = _font(bold=is_base, color=fg, size=10)
            cell.fill      = _fill(bg)
            cell.alignment = _align("center")

    # Legend
    leg_row = header_row + len(g_labels) + 3
    for label, bg, fg in [
        ("■ Base case", _ORANGE, _BG_DARK),
        ("■ >15% upside", _GREEN_BG, _GREEN_FG),
        ("■ >15% downside", _RED_BG, _RED_FG),
    ]:
        cell = ws.cell(row=leg_row, column=1, value=label)
        cell.font = _font(color=fg, size=8)
        cell.fill = _fill(bg)
        leg_row += 1

    _set_col_widths(ws, {i: 14 for i in range(1, 8)})


def _write_scenarios(ws, d: dict) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = _GREEN_FG

    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value     = f"Scenarios — {d.get('company_name', d.get('ticker', ''))} ({d.get('ticker', '')})"
    t.font      = _font(bold=True, color=_FG_WHITE, size=13)
    t.fill      = _fill(_BG_DARK)
    t.alignment = _align("left")
    ws.row_dimensions[1].height = 24

    scenarios = d.get("scenarios") or {}
    price     = d.get("price", 0)
    hdr = ["",  "Bear Case", "Base Case", "Bull Case"]
    _write_header_row(ws, 3, hdr, _BG_MID)

    scenario_keys = ["bear", "base", "bull"]
    rows = [
        ("Scenario",         lambda s: s.get("label", "")),
        ("WACC",             lambda s: _pct(s.get("wacc", 0))),
        ("Terminal Growth",  lambda s: _pct(s.get("g", 0))),
        ("Margin Target Y7", lambda s: _pct(s.get("margin_target", 0))),
        ("Rev Growth",       lambda s: _pct(s.get("rev_growth", 0))),
        ("Intrinsic Value",  lambda s: f"${s.get('iv', 0):.2f}"),
        ("Upside / Downside",lambda s: f"{'+' if s.get('upside', 0) >= 0 else ''}{s.get('upside', 0):.1f}%"),
        ("Enterprise Value", lambda s: _money(s.get("ev", 0))),
        ("Recommendation",   lambda s: s.get("recommendation", "")),
    ]

    for r_off, (label, fn) in enumerate(rows):
        r = 4 + r_off
        alt = _BG_LIGHT if r_off % 2 == 0 else _BG_DARK
        ws.cell(row=r, column=1, value=label).font = _font(color=_FG_GREY, size=10)
        ws.cell(row=r, column=1).fill      = _fill(alt)
        ws.cell(row=r, column=1).alignment = _align("left")

        for c_off, key in enumerate(scenario_keys):
            sc = scenarios.get(key) or {}
            c  = 2 + c_off
            val = fn(sc)
            cell = ws.cell(row=r, column=c, value=val)
            if label == "Recommendation":
                bg_c = _GREEN_BG if sc.get("recommendation") == "Undervalued" else \
                       (_RED_BG if sc.get("recommendation") == "Overvalued" else _AMBER_BG)
                fg_c = _GREEN_FG if sc.get("recommendation") == "Undervalued" else \
                       (_RED_FG if sc.get("recommendation") == "Overvalued" else _AMBER_FG)
                cell.fill = _fill(bg_c)
                cell.font = _font(bold=True, color=fg_c, size=10)
            else:
                cell.fill = _fill(alt)
                cell.font = _font(bold=(label == "Intrinsic Value"), size=10)
            cell.alignment = _align("center")

    # Narrative
    narr_row = 4 + len(rows) + 2
    for c_off, key in enumerate(scenario_keys):
        sc = scenarios.get(key) or {}
        narr = sc.get("narrative", "")
        if narr:
            cell = ws.cell(row=narr_row, column=2 + c_off, value=narr)
            cell.font      = _font(color=_FG_GREY, size=8, italic=True)
            cell.fill      = _fill(_BG_DARK)
            cell.alignment = _align("left", wrap=True)
            ws.row_dimensions[narr_row].height = 60

    _set_col_widths(ws, {1: 24, 2: 22, 3: 22, 4: 22, 5: 4})


def _write_historical(ws, d: dict) -> None:
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:K1")
    t = ws["A1"]
    t.value     = f"Historical Financials — {d.get('company_name', d.get('ticker', ''))} ($M)"
    t.font      = _font(bold=True, color=_FG_WHITE, size=13)
    t.fill      = _fill(_BG_DARK)
    t.alignment = _align("left")
    ws.row_dimensions[1].height = 24

    h    = d.get("historical") or {}
    yrs  = h.get("years", [])
    cols = ["Metric"] + [str(y) for y in yrs]
    _write_header_row(ws, 3, cols)

    rows = [
        ("Revenue ($M)",       "revenue"),
        ("Gross Profit ($M)",  None),      # computed = revenue * gross_margin/100
        ("Gross Margin (%)",   "gross_margin"),
        ("EBIT Margin (%)",    "ebit_margin"),
        ("Net Income ($M)",    "net_income"),
        ("FCF ($M)",           "fcf"),
        ("CapEx ($M)",         "capex"),
        ("Total Debt ($M)",    "debt"),
        ("ROIC (%)",           "roic"),
        ("Shares Out. (M)",    "shares"),
    ]

    for r_off, (label, key) in enumerate(rows):
        r   = 4 + r_off
        alt = _BG_LIGHT if r_off % 2 == 0 else _BG_DARK
        cell = ws.cell(row=r, column=1, value=label)
        cell.font      = _font(color=_FG_GREY, size=10)
        cell.fill      = _fill(alt)
        cell.alignment = _align("left")

        if key is None:
            # Gross profit = revenue × gross_margin / 100
            revs = h.get("revenue", [])
            gms  = h.get("gross_margin", [])
            vals = [round(r_ * gm / 100) if gm else 0 for r_, gm in zip(revs, gms)]
        else:
            vals = h.get(key, [])

        for c_off, val in enumerate(vals):
            c    = 2 + c_off
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill      = _fill(alt)
            cell.font      = _font(size=10)
            cell.alignment = _align("right")

    _set_col_widths(ws, {1: 22, **{i: 14 for i in range(2, 12)}})


def _write_assumptions(ws, d: dict) -> None:
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value     = f"Model Assumptions — {d.get('ticker', '')}"
    t.font      = _font(bold=True, color=_FG_WHITE, size=13)
    t.fill      = _fill(_BG_DARK)
    t.alignment = _align("left")
    ws.row_dimensions[1].height = 24

    _write_header_row(ws, 3, ["Driver", "Auto Value", "Active Value", "Unit", "Source"])

    assumptions = d.get("assumptions") or []
    for r_off, row in enumerate(assumptions):
        r   = 4 + r_off
        alt = _BG_LIGHT if r_off % 2 == 0 else _BG_DARK
        vals = [
            row.get("driver", ""),
            row.get("auto",   ""),
            row.get("active", ""),
            row.get("unit",   ""),
            row.get("source", ""),
        ]
        for c_idx, val in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            cell.fill      = _fill(alt)
            cell.font      = _font(size=10)
            cell.alignment = _align("left")

    _set_col_widths(ws, {1: 36, 2: 14, 3: 14, 4: 8, 5: 50})


def _write_comps(ws, d: dict) -> None:
    ws.sheet_view.showGridLines = False

    peers = d.get("peers") or []
    if not peers:
        ws.cell(row=1, column=1, value="No comparable data available — fetch live FMP data for peers.")
        return

    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value     = f"Comparable Company Analysis — {d.get('ticker', '')} peer group"
    t.font      = _font(bold=True, color=_FG_WHITE, size=13)
    t.fill      = _fill(_BG_DARK)
    t.alignment = _align("left")
    ws.row_dimensions[1].height = 24

    _write_header_row(ws, 3, ["Company", "Ticker", "EV/EBITDA", "EV/EBIT",
                               "EV/Revenue", "P/E", "P/FCF", "Subject"])

    for r_off, peer in enumerate(peers):
        r   = 4 + r_off
        is_subj = peer.get("subject", False)
        alt = _ACCENT if is_subj else (_BG_LIGHT if r_off % 2 == 0 else _BG_DARK)
        vals = [
            peer.get("name", ""),
            peer.get("ticker", ""),
            f"{peer.get('ev_ebitda', 0):.1f}×",
            f"{peer.get('ev_ebit', 0):.1f}×",
            f"{peer.get('ev_rev', 0):.2f}×",
            f"{peer.get('pe', 0):.1f}×",
            f"{peer.get('p_fcf', 0):.1f}×",
            "★" if is_subj else "",
        ]
        for c_idx, val in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            cell.fill      = _fill(alt)
            cell.font      = _font(bold=is_subj, size=10)
            cell.alignment = _align("center" if c_idx > 2 else "left")

    # Peer median
    pm = d.get("peer_median") or {}
    if pm:
        med_row = 4 + len(peers)
        med_vals = ["Peer Median", "—",
                    f"{pm.get('ev_ebitda', 0):.1f}×",
                    f"{pm.get('ev_ebit', 0):.1f}×",
                    f"{pm.get('ev_rev', 0):.2f}×",
                    f"{pm.get('pe', 0):.1f}×",
                    f"{pm.get('p_fcf', 0):.1f}×", ""]
        for c_idx, val in enumerate(med_vals, start=1):
            cell = ws.cell(row=med_row, column=c_idx, value=val)
            cell.fill      = _fill(_BG_MID)
            cell.font      = _font(bold=True, color=_FG_GREY, size=10)
            cell.alignment = _align("center" if c_idx > 2 else "left")

    _set_col_widths(ws, {1: 20, 2: 8, 3: 12, 4: 12, 5: 12, 6: 10, 7: 10, 8: 10})


# ─── Main entry point ─────────────────────────────────────────────────────────

def build_excel_bytes(data: dict) -> bytes:
    """Build the Excel workbook and return raw bytes.

    Raises RuntimeError if openpyxl is not installed.
    """
    if not _HAS_OPENPYXL:
        raise RuntimeError("openpyxl is required for Excel export. Install it with: pip install openpyxl")

    wb = Workbook()

    # Remove default sheet
    default = wb.active
    wb.remove(default)

    # Create sheets in order
    cover_ws  = wb.create_sheet("Cover")
    dcf_ws    = wb.create_sheet("DCF Model")
    sens_ws   = wb.create_sheet("Sensitivity")
    scen_ws   = wb.create_sheet("Scenarios")
    hist_ws   = wb.create_sheet("Historical")
    asmp_ws   = wb.create_sheet("Assumptions")
    comps_ws  = wb.create_sheet("Comps")

    _write_cover(cover_ws,       data)
    _write_dcf_model(dcf_ws,     data)
    _write_sensitivity(sens_ws,  data)
    _write_scenarios(scen_ws,    data)
    _write_historical(hist_ws,   data)
    _write_assumptions(asmp_ws,  data)
    _write_comps(comps_ws,       data)

    # Apply dark theme to all sheet tabs / backgrounds
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF", ""):
                    cell.fill = _fill(_BG_DARK)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
