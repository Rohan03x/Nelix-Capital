"""
webapp/data/excel_export.py  —  Nike-style single-sheet DCF workbook
─────────────────────────────────────────────────────────────────────
Generates a single-sheet Excel workbook that mirrors the structure of
NIKE Valuation Strategy.xlsx:

  Sheet: "valuation"
  Col A      : row labels
  Cols B–K   : historical data  (10 years)
  Cols L–R   : forecast data    (7 years)

  Row layout (matching Nike row numbering):
    1–7    header (dates, period index, hist/fcst labels)
    8–39   DCF summary (UFCF schedule, WACC, TV, equity value, price)
    42–56  comparable companies
    60–74  sensitivity table (WACC × terminal growth)
   300–395  Income Statement
   415–469  Cash Flow Statement
   480–541  Balance Sheet

Public entry-point:
  build_excel_bytes(data: dict) -> bytes
"""

from __future__ import annotations

import io
import logging
from datetime import date, datetime
from typing import List, Optional

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False
    logger.warning("openpyxl not installed -- Excel export unavailable.")

# ---- Layout constants --------------------------------------------------------
_LABEL_COL = 1           # column A
_N_HIST    = 10          # ten years of history  -> cols B-K (indices 2-11)
_N_FCST    = 7           # seven forecast years  -> cols L-R (indices 12-18)
_HIST_COLS = list(range(2, 2 + _N_HIST))
_FCST_COLS = list(range(2 + _N_HIST, 2 + _N_HIST + _N_FCST))

# ---- Colour palette ----------------------------------------------------------
_NAVY       = "1F4E79"
_BOLD_BLUE  = "003366"
_HIST_BG    = "D6E4F0"
_FCST_BG    = "E2EFDA"
_LIGHT_GREY = "F2F2F2"
_WHITE      = "FFFFFF"

# ---- Helpers -----------------------------------------------------------------

def _col(n: int) -> str:
    return get_column_letter(n)


def _v(lst, idx, default=None):
    """Safe list access; returns default for out-of-range / None."""
    if lst is None:
        return default
    try:
        val = lst[idx]
        return default if val is None else val
    except (IndexError, TypeError):
        return default


def _s(val, default=0):
    """Return val if not None, else default."""
    return val if val is not None else default


# ---- Public entry-point ------------------------------------------------------

def build_excel_bytes(data: dict) -> bytes:
    """Build and return the complete Excel workbook as raw bytes."""
    if not _HAS_OPENPYXL:
        raise RuntimeError("openpyxl is not installed -- cannot export Excel.")

    wb = Workbook()
    ws = wb.active
    ws.title = "valuation"

    _build_sheet(ws, data)
    _apply_formatting(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---- Sheet builder -----------------------------------------------------------

def _build_sheet(ws, data: dict) -> None:
    h      = data.get("historical", {}) or {}
    fc     = data.get("forecast",   []) or []
    pm     = data.get("peer_median") or {}
    peers  = data.get("peers", []) or []
    sens   = data.get("sensitivity") or {}

    # -- historical series (all in $M) ----------------------------------------
    rev_h    = h.get("revenue",       [])
    gp_h     = h.get("gross_profit",  [])
    ebit_h   = h.get("ebit",          [])
    ni_h     = h.get("net_income",    [])
    da_h     = h.get("da",            [])
    capex_h  = h.get("capex",         [])
    op_cf_h  = h.get("op_cf",         [])
    sbc_h    = h.get("sbc",           [])
    shares_h = h.get("shares",        [])
    cash_h   = h.get("cash",          [])
    debt_h   = h.get("debt",          [])
    equity_h = h.get("equity",        [])
    ta_h     = h.get("total_assets",  [])
    pretax_h = h.get("pretax_income", [])
    tax_h    = h.get("tax",           [])
    fcf_h    = h.get("fcf",           [])
    years_h  = h.get("years",         [])

    # -- derived historical series --------------------------------------------
    cogs_h: List = []
    for i in range(_N_HIST):
        r, g = _v(rev_h, i), _v(gp_h, i)
        cogs_h.append((r - g) if (r is not None and g is not None) else None)

    sga_h: List = []
    for i in range(_N_HIST):
        g, e, d = _v(gp_h, i), _v(ebit_h, i), _v(da_h, i)
        sga_h.append((_s(g) - _s(e) - _s(d)) if g is not None else None)

    int_exp_h: List = []
    for i in range(_N_HIST):
        pt, e = _v(pretax_h, i), _v(ebit_h, i)
        int_exp_h.append((_s(pt) - _s(e)) if pt is not None else None)

    # -- top-level scalars -----------------------------------------------------
    ticker     = data.get("ticker", "")
    company    = data.get("company_name", ticker)
    price      = _s(data.get("current_price") or data.get("price", 0))
    iv         = _s(data.get("intrinsic_value", 0))
    wacc       = _s(data.get("wacc", 0))
    terminal_g = _s(data.get("terminal_growth", 0))
    tax_rate   = _s(data.get("tax_rate", 0))
    cod        = _s(data.get("cost_of_debt_pre") or data.get("cost_of_debt", 0))
    net_debt   = _s(data.get("net_debt", 0))
    shares_out = _s(data.get("diluted_shares") or data.get("shares_outstanding", 0))
    ev         = _s(data.get("enterprise_value", 0))
    pv_ufcfs   = _s(data.get("pv_ufcfs", 0))
    pv_tv      = _s(data.get("pv_terminal", 0))
    upside_pct = _s(data.get("upside_pct", 0))
    ebitda_ltm = _s(data.get("ebitda_ltm", 0))
    dso        = _s(data.get("dso", 30))
    dio        = _s(data.get("dio", 30))
    dpo        = _s(data.get("dpo", 60))
    tv_pct     = _s(data.get("tv_pct", 0))

    # ratio drivers from last historical year
    last_rev   = _s(_v(rev_h,  -1), 1)
    last_gm    = _s(_v(gp_h,   -1)) / last_rev if last_rev else 0
    last_sga_p = _s(_v(sga_h,  -1)) / last_rev if last_rev else 0
    last_ta    = _s(_v(ta_h,   -1))
    last_debt  = _s(_v(debt_h, -1))
    dt_ratio   = last_debt / last_ta if last_ta else 0

    # -- forecast series -------------------------------------------------------
    def _fv(key: str) -> list:
        return [f.get(key) for f in fc]

    fcst_rev    = _fv("revenue")
    fcst_ebit   = _fv("ebit")
    fcst_nopat  = _fv("nopat")
    fcst_da     = _fv("da")
    fcst_sbc    = _fv("sbc")
    fcst_capex  = _fv("capex")
    fcst_dnowc  = _fv("d_nowc")
    fcst_ufcf   = _fv("ufcf")
    fcst_df     = _fv("df")
    fcst_pv     = _fv("pv")
    fcst_ebit_m = _fv("ebit_m")
    fcst_years  = _fv("year")

    # forecast IS
    tax_dec   = tax_rate / 100
    fcst_gp   = [_s(_v(fcst_rev, j)) * last_gm  for j in range(_N_FCST)]
    fcst_cogs = [-(_s(_v(fcst_rev, j)) * (1 - last_gm)) for j in range(_N_FCST)]
    fcst_sga  = [_s(_v(fcst_rev, j)) * last_sga_p for j in range(_N_FCST)]
    fcst_opex = [_s(_v(fcst_sga, j)) + _s(_v(fcst_da, j)) for j in range(_N_FCST)]
    fcst_opinc= [_s(_v(fcst_gp, j)) - _s(_v(fcst_opex, j)) for j in range(_N_FCST)]
    fcst_ebt  = [_s(_v(fcst_opinc, j)) for j in range(_N_FCST)]
    fcst_tax  = [_s(_v(fcst_ebt, j)) * tax_dec  for j in range(_N_FCST)]
    fcst_ni   = [_s(_v(fcst_ebt, j)) - _s(_v(fcst_tax, j)) for j in range(_N_FCST)]
    fcst_ebitda = [_s(_v(fcst_ebit, j)) + _s(_v(fcst_da, j)) for j in range(_N_FCST)]

    # forecast CF
    fcst_op_cf = [_s(_v(fcst_ni, j)) + _s(_v(fcst_da, j)) for j in range(_N_FCST)]
    fcst_cap_cf= [-_s(_v(fcst_capex, j)) for j in range(_N_FCST)]
    fcst_net_c = [_s(_v(fcst_op_cf, j)) + _s(_v(fcst_cap_cf, j)) for j in range(_N_FCST)]

    # forecast BS
    fcst_ta_bs: List[float] = [
        _s(_v(fcst_rev, j)) / last_rev * last_ta for j in range(_N_FCST)
    ]
    fcst_recv  = [_s(_v(fcst_rev, j)) / 365 * dso for j in range(_N_FCST)]
    fcst_inv   = [abs(_s(_v(fcst_cogs, j))) / 365 * dio for j in range(_N_FCST)]
    fcst_cash: List[float] = []
    _prior = _s(_v(cash_h, -1))
    for j in range(_N_FCST):
        _prior += _s(_v(fcst_net_c, j))
        fcst_cash.append(_prior)
    fcst_ca   = [_s(_v(fcst_cash, j)) + _s(_v(fcst_recv, j)) + _s(_v(fcst_inv, j))
                 for j in range(_N_FCST)]
    fcst_ap   = [abs(_s(_v(fcst_cogs, j))) / 365 * dpo for j in range(_N_FCST)]
    fcst_ltd  = [last_debt] * _N_FCST
    fcst_tl   = [_s(_v(fcst_ap, j)) + _s(_v(fcst_ltd, j)) for j in range(_N_FCST)]
    fcst_eq   = [_s(_v(fcst_ta_bs, j)) - _s(_v(fcst_tl, j)) for j in range(_N_FCST)]

    # ==========================================================================
    # HEADER  (rows 1-5)
    # ==========================================================================

    # Row 1: year dates
    for i, yr in enumerate(years_h):
        try:
            ws.cell(1, _HIST_COLS[i]).value = datetime(int(yr), 12, 31)
            ws.cell(1, _HIST_COLS[i]).number_format = "YYYY"
        except (ValueError, TypeError):
            ws.cell(1, _HIST_COLS[i]).value = yr

    for j, yr_str in enumerate(fcst_years):
        try:
            yr_num = int(str(yr_str).replace("FY", ""))
            ws.cell(1, _FCST_COLS[j]).value = datetime(yr_num, 12, 31)
            ws.cell(1, _FCST_COLS[j]).number_format = "YYYY"
        except (ValueError, TypeError, AttributeError):
            ws.cell(1, _FCST_COLS[j]).value = yr_str

    # Row 2: period index  (-9...0  then  1...7)
    ws.cell(2, _LABEL_COL).value = "Period"
    for i in range(_N_HIST):
        ws.cell(2, _HIST_COLS[i]).value = -((_N_HIST - 1) - i)
    for j in range(_N_FCST):
        ws.cell(2, _FCST_COLS[j]).value = j + 1

    # Row 3: "Historical" / "Forecast" section labels
    ws.cell(3, _HIST_COLS[0]).value = "Historical"
    ws.cell(3, _FCST_COLS[0]).value = "Forecast"

    # Row 4: company
    ws.cell(4, _LABEL_COL).value = "Company"
    ws.cell(4, _HIST_COLS[-1]).value = f"{ticker}  --  {company}"

    # Row 5: units
    ws.cell(5, _LABEL_COL).value = "Units"
    ws.cell(5, _HIST_COLS[0]).value = "All financials in $M (USD millions) unless noted"

    # ==========================================================================
    # DCF SUMMARY  (rows 8-39)
    # Key scalars live in col R (last forecast col = index 18), matching Nike.
    # ==========================================================================
    R = _FCST_COLS[-1]   # column R
    K = _HIST_COLS[-1]   # column K

    ws.cell(8,  _LABEL_COL).value = "-- DCF Summary --"

    ws.cell(9,  _LABEL_COL).value = "Company"
    ws.cell(9,  K).value = f"{ticker} -- {company}"

    ws.cell(10, _LABEL_COL).value = "Valuation Date"
    ws.cell(10, K).value = date.today()

    ws.cell(11, _LABEL_COL).value = "Current Market Price ($/share)"
    ws.cell(11, K).value = price
    ws.cell(11, K).number_format = "$#,##0.00"

    # Row 14-24: UFCF schedule
    ws.cell(14, _LABEL_COL).value = "Forecast Year"
    for j, f in enumerate(fc):
        ws.cell(14, _FCST_COLS[j]).value = f.get("n", j + 1)

    ws.cell(15, _LABEL_COL).value = "Terminal Growth Rate (g)"
    ws.cell(15, R).value = terminal_g / 100
    ws.cell(15, R).number_format = "0.0%"

    ws.cell(16, _LABEL_COL).value = "Revenue ($M)"
    for i in range(_N_HIST):
        ws.cell(16, _HIST_COLS[i]).value = _v(rev_h, i)
    for j in range(_N_FCST):
        ws.cell(16, _FCST_COLS[j]).value = _v(fcst_rev, j)

    ws.cell(17, _LABEL_COL).value = "EBIT ($M)"
    for i in range(_N_HIST):
        ws.cell(17, _HIST_COLS[i]).value = _v(ebit_h, i)
    for j in range(_N_FCST):
        ws.cell(17, _FCST_COLS[j]).value = _v(fcst_ebit, j)

    ws.cell(18, _LABEL_COL).value = "NOPAT ($M)"
    for j in range(_N_FCST):
        ws.cell(18, _FCST_COLS[j]).value = _v(fcst_nopat, j)

    ws.cell(19, _LABEL_COL).value = "  (-) CapEx ($M)"
    for j in range(_N_FCST):
        ws.cell(19, _FCST_COLS[j]).value = -_s(_v(fcst_capex, j))

    ws.cell(20, _LABEL_COL).value = "  (+) D&A ($M)"
    for j in range(_N_FCST):
        ws.cell(20, _FCST_COLS[j]).value = _v(fcst_da, j)

    ws.cell(21, _LABEL_COL).value = "  (+/-) Delta NWC ($M)"
    for j in range(_N_FCST):
        ws.cell(21, _FCST_COLS[j]).value = _v(fcst_dnowc, j)

    ws.cell(22, _LABEL_COL).value = "Unlevered Free Cash Flow ($M)"
    for i in range(_N_HIST):
        ws.cell(22, _HIST_COLS[i]).value = _v(fcf_h, i)
    for j in range(_N_FCST):
        ws.cell(22, _FCST_COLS[j]).value = _v(fcst_ufcf, j)

    ws.cell(23, _LABEL_COL).value = "Discount Factor"
    for j in range(_N_FCST):
        ws.cell(23, _FCST_COLS[j]).value = _v(fcst_df, j)
        ws.cell(23, _FCST_COLS[j]).number_format = "0.0000"

    ws.cell(24, _LABEL_COL).value = "PV of UFCF ($M)"
    for j in range(_N_FCST):
        ws.cell(24, _FCST_COLS[j]).value = _v(fcst_pv, j)

    # Key assumption scalars (Nike R25, R26, R27, R28)
    ws.cell(25, _LABEL_COL).value = "Tax Rate"
    ws.cell(25, R).value = tax_rate / 100
    ws.cell(25, R).number_format = "0.0%"

    ws.cell(26, _LABEL_COL).value = "Cost of Debt (Pre-Tax)"
    ws.cell(26, R).value = cod / 100
    ws.cell(26, R).number_format = "0.0%"

    ws.cell(27, _LABEL_COL).value = "LT Debt / Total Assets"
    ws.cell(27, R).value = dt_ratio
    ws.cell(27, R).number_format = "0.0%"

    ws.cell(28, _LABEL_COL).value = "WACC"
    ws.cell(28, R).value = wacc / 100
    ws.cell(28, R).number_format = "0.0%"

    # Equity bridge (rows 29-39)
    ws.cell(29, _LABEL_COL).value = "Cash & Equivalents ($M)"
    ws.cell(29, K).value = _v(cash_h, -1)

    ws.cell(30, _LABEL_COL).value = "Total Debt ($M)"
    ws.cell(30, K).value = _v(debt_h, -1)

    ws.cell(31, _LABEL_COL).value = "Net Debt ($M)"
    ws.cell(31, K).value = net_debt

    ws.cell(32, _LABEL_COL).value = "Sum PV(UFCF) ($M)"
    ws.cell(32, K).value = pv_ufcfs

    ws.cell(33, _LABEL_COL).value = "PV of Terminal Value ($M)"
    ws.cell(33, K).value = pv_tv

    ws.cell(34, _LABEL_COL).value = "Diluted Shares Outstanding (M)"
    ws.cell(34, K).value = shares_out

    ws.cell(35, _LABEL_COL).value = "DCF Intrinsic Value ($/share)"
    ws.cell(35, K).value = iv
    ws.cell(35, K).number_format = "$#,##0.00"

    ws.cell(36, _LABEL_COL).value = "Market Price ($/share)"
    ws.cell(36, K).value = price
    ws.cell(36, K).number_format = "$#,##0.00"

    ws.cell(37, _LABEL_COL).value = "Upside / (Downside)"
    ws.cell(37, K).value = upside_pct / 100
    ws.cell(37, K).number_format = "0.0%"

    ws.cell(38, _LABEL_COL).value = "Enterprise Value ($M)"
    ws.cell(38, K).value = ev

    ws.cell(39, _LABEL_COL).value = "Terminal Value as % of EV"
    ws.cell(39, K).value = tv_pct / 100
    ws.cell(39, K).number_format = "0.0%"

    # ==========================================================================
    # COMPARABLE COMPANIES  (rows 42-56)
    # ==========================================================================
    ws.cell(42, _LABEL_COL).value = "-- Comparable Companies --"
    comp_hdrs = ["Company", "Ticker", "Rev ($M)", "EBITDA ($M)", "EBIT ($M)",
                 "P/E", "EV/EBITDA", "EV/EBIT", "EV/Rev", "P/FCF"]
    for ci, h_txt in enumerate(comp_hdrs):
        ws.cell(43, ci + 1).value = h_txt
        ws.cell(43, ci + 1).font = Font(bold=True, name="Calibri", size=10)

    for pi, peer in enumerate(peers[:10]):
        r = 44 + pi
        ws.cell(r, 1).value  = peer.get("name", "")
        ws.cell(r, 2).value  = peer.get("ticker", "")
        ws.cell(r, 3).value  = peer.get("revenue")
        ws.cell(r, 4).value  = peer.get("ebitda")
        ws.cell(r, 5).value  = peer.get("ebit")
        ws.cell(r, 6).value  = peer.get("pe")
        ws.cell(r, 7).value  = peer.get("ev_ebitda")
        ws.cell(r, 8).value  = peer.get("ev_ebit")
        ws.cell(r, 9).value  = peer.get("ev_rev")
        ws.cell(r, 10).value = peer.get("p_fcf")

    ws.cell(54, _LABEL_COL).value = "Peer Median"
    ws.cell(54, _LABEL_COL).font  = Font(bold=True, name="Calibri", size=10)
    ws.cell(54, 6).value  = pm.get("pe")
    ws.cell(54, 7).value  = pm.get("ev_ebitda")
    ws.cell(54, 8).value  = pm.get("ev_ebit")
    ws.cell(54, 9).value  = pm.get("ev_rev")
    ws.cell(54, 10).value = pm.get("p_fcf")

    ws.cell(55, _LABEL_COL).value = f"{ticker} (subject)"
    ws.cell(55, _LABEL_COL).font  = Font(bold=True, name="Calibri", size=10)
    ws.cell(55, 3).value = last_rev
    ws.cell(55, 4).value = ebitda_ltm
    ws.cell(55, 5).value = _v(ebit_h, -1)
    ws.cell(55, 7).value = round(ev / ebitda_ltm, 2) if ebitda_ltm else None
    ws.cell(55, 9).value = round(ev / last_rev,   2) if last_rev   else None

    # ==========================================================================
    # SENSITIVITY TABLE  (rows 60-74)
    # ==========================================================================
    ws.cell(60, _LABEL_COL).value = "-- Sensitivity: Intrinsic Value per Share --"
    ws.cell(61, _LABEL_COL).value = "WACC \\ Terminal g ->"

    g_labels    = sens.get("g_labels",    [])
    wacc_labels = sens.get("wacc_labels", [])
    iv_grid     = sens.get("iv_grid",     [])
    base_g_idx   = _s(sens.get("base_g_idx",   2))
    base_wacc_idx= _s(sens.get("base_wacc_idx", 2))

    for ci, gl in enumerate(g_labels):
        ws.cell(61, 2 + ci).value = gl
        ws.cell(61, 2 + ci).font  = Font(bold=True, name="Calibri", size=10)

    for ri, wl in enumerate(wacc_labels):
        row = 62 + ri
        ws.cell(row, _LABEL_COL).value = wl
        ws.cell(row, _LABEL_COL).font  = Font(bold=True, name="Calibri", size=10)
        grid_row = iv_grid[ri] if ri < len(iv_grid) else []
        for ci, iv_val in enumerate(grid_row):
            c = ws.cell(row, 2 + ci)
            c.value = iv_val
            c.number_format = "$#,##0.00"
            if ri == base_wacc_idx and ci == base_g_idx:
                c.fill = PatternFill("solid", fgColor="FFC000")
                c.font = Font(bold=True, name="Calibri", size=10)

    ws.cell(68, _LABEL_COL).value = (
        f"Base case:  WACC = {wacc:.1f}%   |   Terminal g = {terminal_g:.1f}%"
        f"   |   IV = ${iv:.2f}"
    )

    # ==========================================================================
    # INCOME STATEMENT  (rows 300-395)
    # Row 338 = Total Revenue  (key reference row for all IS forecast formulas)
    # Row 344 = D&A            (sourced from CF section)
    # Row 361 = Net Income     (key reference for CF/BS forecast formulas)
    # Row 380 = EBIT           (used in Comps section)
    # ==========================================================================

    def _wr(row: int, label: str, hist_vals, fcst_vals=None,
            fmt: Optional[str] = None, bold: bool = False) -> None:
        ws.cell(row, _LABEL_COL).value = label
        if bold:
            ws.cell(row, _LABEL_COL).font = Font(bold=True, name="Calibri", size=10)
        for i in range(_N_HIST):
            v = _v(hist_vals, i) if hist_vals else None
            if v is not None:
                c = ws.cell(row, _HIST_COLS[i])
                c.value = v
                if fmt:
                    c.number_format = fmt
        if fcst_vals:
            for j in range(_N_FCST):
                v = _v(fcst_vals, j)
                if v is not None:
                    c = ws.cell(row, _FCST_COLS[j])
                    c.value = v
                    if fmt:
                        c.number_format = fmt

    ws.cell(300, _LABEL_COL).value = "INCOME STATEMENT ($M)"
    ws.cell(337, _LABEL_COL).value = "Revenue"

    _wr(338, "Total Revenue",        rev_h,  fcst_rev,  bold=True)
    _wr(339, "  Cost of Goods Sold", [-_s(_v(cogs_h, i)) for i in range(_N_HIST)], fcst_cogs)
    _wr(340, "  Gross Profit",       gp_h,   fcst_gp,   bold=True)

    _wr(342, "  SG&A excl. D&A",     sga_h,  fcst_sga)
    _wr(344, "  D&A Total",          da_h,   fcst_da)
    _wr(345, "  Total Operating Expenses",
             [_s(_v(sga_h, i)) + _s(_v(da_h, i)) for i in range(_N_HIST)],
             fcst_opex)
    _wr(346, "Operating Income (EBIT)",
             [_s(_v(gp_h, i)) - _s(_v(sga_h, i)) - _s(_v(da_h, i))
              for i in range(_N_HIST)],
             fcst_opinc, bold=True)

    _wr(348, "  Interest Expense",
             [-_s(_v(int_exp_h, i)) for i in range(_N_HIST)])
    _wr(353, "EBT (Earnings Before Tax)",   pretax_h, fcst_ebt)
    _wr(355, "EBT incl. Unusual Items",     pretax_h, fcst_ebt)
    _wr(359, "  Income Tax Expense",        tax_h,    fcst_tax)
    _wr(361, "Net Income to Company",       ni_h,     fcst_ni,   bold=True)
    _wr(362, "Net Income",                  ni_h,     fcst_ni)

    ws.cell(371, _LABEL_COL).value = "Diluted Shares Outstanding (M)"
    for i in range(_N_HIST):
        ws.cell(371, _HIST_COLS[i]).value = _v(shares_h, i)

    hist_ebitda = [_s(_v(ebit_h, i)) + _s(_v(da_h, i)) for i in range(_N_HIST)]
    _wr(378, "EBITDA",    hist_ebitda, fcst_ebitda)
    _wr(380, "EBIT",      ebit_h, [_v(fcst_ebit, j) for j in range(_N_FCST)], bold=True)
    _wr(381, "  EBIT Margin (%)",
             [_s(_v(ebit_h, i)) / _s(_v(rev_h, i), 1) * 100
              if _v(rev_h, i) else None for i in range(_N_HIST)],
             fcst_ebit_m, fmt="0.0%")

    # SBC line
    _wr(382, "  Stock-Based Compensation", sbc_h, fcst_sbc)

    # ==========================================================================
    # CASH FLOW STATEMENT  (rows 415-469)
    # Row 437 = D&A Total  (source for IS row 344)
    # Row 445 = Cash from Operations
    # Row 469 = Net Change in Cash
    # ==========================================================================
    ws.cell(415, _LABEL_COL).value = "CASH FLOW STATEMENT ($M)"
    ws.cell(416, _LABEL_COL).value = "Operating Activities"

    _wr(434, "Net Income",                    ni_h,    fcst_ni)
    _wr(435, "  Depreciation & Amortisation", da_h,    fcst_da)
    _wr(437, "  D&A Total",                   da_h,    fcst_da)
    _wr(439, "  Stock-Based Compensation",    sbc_h,   fcst_sbc)
    _wr(440, "  Change in Net Working Capital",
             [_s(_v(op_cf_h, i)) - _s(_v(ni_h, i)) - _s(_v(da_h, i))
              for i in range(_N_HIST)],
             fcst_dnowc)
    _wr(445, "Cash from Operations",          op_cf_h, fcst_op_cf, bold=True)

    ws.cell(446, _LABEL_COL).value = "Investing Activities"
    cap_neg_h = [-_s(_v(capex_h, i)) for i in range(_N_HIST)]
    _wr(447, "  Capital Expenditure",         cap_neg_h, fcst_cap_cf)
    _wr(451, "Cash from Investing",           cap_neg_h, fcst_cap_cf, bold=True)

    ws.cell(452, _LABEL_COL).value = "Financing Activities"
    _wr(459, "  Issuance of Common Stock",    [0] * _N_HIST)
    _wr(460, "  Repurchase of Common Stock",  [0] * _N_HIST)
    _wr(464, "  Dividends Paid",              [0] * _N_HIST)
    _wr(466, "Cash from Financing",           [0] * _N_HIST, [0] * _N_FCST, bold=True)
    _wr(468, "  FX / Other Adjustment",       [0] * _N_HIST, [0] * _N_FCST)

    hist_net_cash = [_s(_v(op_cf_h, i)) - _s(_v(capex_h, i)) for i in range(_N_HIST)]
    _wr(469, "Net Change in Cash",            hist_net_cash, fcst_net_c, bold=True)

    # ==========================================================================
    # BALANCE SHEET  (rows 480-541)
    # Row 500 = Cash & Equivalents
    # Row 518 = Total Assets      (key reference for LT Debt forecast)
    # Row 528 = Long-term Debt    (IBD — used for net-debt / interest expense)
    # Row 539 = Total Common Equity
    # Row 541 = Total Liabilities and Equity
    # ==========================================================================
    ws.cell(480, _LABEL_COL).value = "BALANCE SHEET ($M)"
    ws.cell(481, _LABEL_COL).value = "Assets ($M)"

    hist_recv_bs = [_s(_v(rev_h, i)) / 365 * dso for i in range(_N_HIST)]
    hist_inv_bs  = [_s(_v(cogs_h, i)) / 365 * dio
                    if _v(cogs_h, i) is not None else 0
                    for i in range(_N_HIST)]
    hist_ca_bs   = [_s(_v(cash_h, i)) + hist_recv_bs[i] + hist_inv_bs[i]
                    for i in range(_N_HIST)]
    hist_ppe_bs  = [_s(_v(ta_h, i)) - _s(_v(cash_h, i)) - hist_recv_bs[i]
                    for i in range(_N_HIST)]

    _wr(500, "Cash and Equivalents",          cash_h,       fcst_cash,   bold=True)
    _wr(503, "Cash & Short-term Investments", cash_h,       fcst_cash)
    _wr(504, "Accounts Receivable",           hist_recv_bs, fcst_recv)
    _wr(506, "Total Receivables",             hist_recv_bs, fcst_recv)
    _wr(507, "Inventory",                     hist_inv_bs,  fcst_inv)
    _wr(510, "Total Current Assets",          hist_ca_bs,   fcst_ca,     bold=True)
    _wr(513, "Net PP&E",                      hist_ppe_bs,
             [_s(_v(fcst_ta_bs, j)) - _s(_v(fcst_cash, j)) - _s(_v(fcst_recv, j))
              for j in range(_N_FCST)])
    _wr(518, "Total Assets",                  ta_h,         fcst_ta_bs,  bold=True)

    ws.cell(519, _LABEL_COL).value = "Liabilities ($M)"
    hist_ap_bs = [_s(_v(cogs_h, i)) / 365 * dpo
                  if _v(cogs_h, i) is not None else 0
                  for i in range(_N_HIST)]
    hist_tl_bs = [hist_ap_bs[i] + _s(_v(debt_h, i)) for i in range(_N_HIST)]

    _wr(520, "Accounts Payable",              hist_ap_bs, fcst_ap)
    _wr(527, "Total Current Liabilities",     hist_ap_bs, fcst_ap,     bold=True)
    _wr(528, "Long-term Debt",                debt_h,     fcst_ltd)
    _wr(533, "Total Liabilities",             hist_tl_bs, fcst_tl,     bold=True)

    ws.cell(534, _LABEL_COL).value = "Equity ($M)"
    _wr(537, "Retained Earnings",             equity_h)
    _wr(539, "Total Common Equity",           equity_h, fcst_eq,       bold=True)
    _wr(540, "Total Equity",                  equity_h, fcst_eq)
    _wr(541, "Total Liabilities and Equity",  ta_h,     fcst_ta_bs,   bold=True)


# ---- Formatting pass ---------------------------------------------------------

def _apply_formatting(ws) -> None:
    """Column widths, fills, fonts, number formats."""

    # Column widths
    ws.column_dimensions["A"].width = 36
    for col_idx in range(2, 2 + _N_HIST + _N_FCST):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15

    # Fills / fonts
    NAVY_FILL  = PatternFill("solid", fgColor=_NAVY)
    HIST_FILL  = PatternFill("solid", fgColor=_HIST_BG)
    FCST_FILL  = PatternFill("solid", fgColor=_FCST_BG)
    GREY_FILL  = PatternFill("solid", fgColor=_LIGHT_GREY)
    HDATA_FILL = PatternFill("solid", fgColor="EBF4FB")
    FDATA_FILL = PatternFill("solid", fgColor="EBF5EB")
    WHITE_BOLD = Font(bold=True, color=_WHITE,     name="Calibri", size=10)
    NAVY_BOLD  = Font(bold=True, color=_BOLD_BLUE, name="Calibri", size=10)

    # Rows to treat as section dividers (full-width navy fill)
    _SECTION_ROWS = {300, 415, 480}
    # Rows containing financial figures (for number-format + colour)
    _FIN_ROWS = (
        set(range(14, 40))
        | set(range(300, 396))
        | set(range(415, 470))
        | set(range(480, 542))
    )

    _HIST_COLS_SET = set(_HIST_COLS)
    _FCST_COLS_SET = set(_FCST_COLS)

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None and cell.row not in _SECTION_ROWS:
                continue
            r, c = cell.row, cell.column

            # Section-header rows (col A only)
            if c == _LABEL_COL and r in _SECTION_ROWS:
                cell.font = WHITE_BOLD
                cell.fill = NAVY_FILL
                continue

            # Separator labels ("-- ... --")
            if c == _LABEL_COL and isinstance(cell.value, str) and "--" in cell.value:
                cell.font = NAVY_BOLD
                cell.fill = GREY_FILL
                continue

            # Row 1: date headers
            if r == 1 and cell.value is not None:
                cell.font = Font(bold=True, name="Calibri", size=10)
                if c in _HIST_COLS_SET:
                    cell.fill = HIST_FILL
                elif c in _FCST_COLS_SET:
                    cell.fill = FCST_FILL
                continue

            # Row 3: "Historical" / "Forecast" labels
            if r == 3 and cell.value is not None:
                if c in _HIST_COLS_SET:
                    cell.fill = HIST_FILL
                    cell.font = Font(bold=True, color=_BOLD_BLUE, name="Calibri", size=10)
                elif c in _FCST_COLS_SET:
                    cell.fill = FCST_FILL
                    cell.font = Font(bold=True, color="1A5C2A", name="Calibri", size=10)
                continue

            # Financial data cells
            if r in _FIN_ROWS and isinstance(cell.value, (int, float)):
                if cell.number_format not in ("$#,##0.00", "0.0%", "0.0000",
                                               "YYYY", "0.00%", "0.0%"):
                    cell.number_format = "#,##0" if abs(cell.value) >= 10 else "#,##0.0"
                if c in _HIST_COLS_SET:
                    cell.fill = HDATA_FILL
                elif c in _FCST_COLS_SET:
                    cell.fill = FDATA_FILL

    # Freeze at B4 so dates/labels stay visible while scrolling
    ws.freeze_panes = "B4"

