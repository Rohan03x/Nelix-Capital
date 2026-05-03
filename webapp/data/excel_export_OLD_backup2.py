"""
webapp/data/excel_export.py  (professional formula-linked model)
─────────────────────────────────────────────────────────────────
Generates a multi-sheet, formula-linked Excel DCF workbook from a
dashboard data dict.  Inspired by professional sell-side models.

Sheet architecture (20 sheets):
  1.  Readme        — how to use the model
  2.  Cover         — summary dashboard
  3.  Assumptions   — all input cells (editable, B3-B30)
  4.  Raw_IS        — raw income statement values (from yfinance)
  5.  Raw_BS        — raw balance sheet values
  6.  Raw_CF        — raw cash-flow statement values
  7.  WACC_Calc     — CAPM / WACC build-up (formula-linked to Assumptions)
  8.  IS_Forecast   — revenue & margin forecast (formula-linked)
  9.  CF_Forecast   — cash-flow & UFCF forecast (formula-linked)
 10.  DCF_Calc      — discount factors, PV, terminal value (formula-linked)
 11.  Historical    — historical ratios & trend analysis
 12.  Sensitivity   — WACC × g sensitivity table (values from model)
 13.  Scenarios     — bear / base / bull comparison
 14.  Comps         — peer comparable analysis
 15.  Peer_Data     — raw peer data table
 16.  Football_Field— valuation bridge chart data
 17.  Data_Quality  — source and data integrity flags
 18.  Source_Log    — raw data pull log
 19.  Rough_Work    — scratch calculations
 20.  (auto-named)  — formula legend

Public entry-point:
  build_excel_bytes(data: dict) → bytes
"""

from __future__ import annotations

import io
import logging
from datetime import date
from typing import Any

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False
    logger.warning("openpyxl not installed — Excel export unavailable.")


# ─── Colour palette (dark Bloomberg-style) ───────────────────────────────────
_BG_DARK    = "0D1117"
_BG_MID     = "161B22"
_BG_LIGHT   = "1C2128"
_BG_WHITE   = "FFFFFF"
_BG_INPUT   = "F0F4FF"   # light blue — input/assumption cells
_BG_FORMULA = "F5FFF5"   # light green — formula cells
_BG_HEADER  = "003366"   # dark navy
_BG_SUBHDR  = "1F4E79"   # medium navy
_BG_GREEN   = "E2EFDA"
_BG_RED     = "FCE4D6"
_BG_AMBER   = "FFF2CC"
_FG_WHITE   = "FFFFFF"
_FG_NAVY    = "003366"
_FG_GREY    = "595959"
_FG_GREEN   = "375623"
_FG_RED     = "9C0006"
_FG_AMBER   = "7D5A00"
_ACCENT     = "2E75B6"


# ─── Style helpers ────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> "PatternFill":
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=_FG_NAVY, size=10, italic=False, name="Calibri"):
    return Font(bold=bold, color=color, size=size, italic=italic, name=name)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _thin_border():
    t = Side(style="thin", color="BFBFBF")
    return Border(left=t, right=t, top=t, bottom=t)

def _bottom_border():
    t = Side(style="thin", color="BFBFBF")
    return Border(bottom=t)

def _set_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

def _header_cell(ws, row, col, value, bg=_BG_HEADER, fg=_FG_WHITE, size=10, bold=True, h="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = _font(bold=bold, color=fg, size=size)
    c.fill      = _fill(bg)
    c.alignment = _align(h, "center")
    return c

def _input_cell(ws, row, col, value):
    """Blue-tinted editable input cell."""
    c = ws.cell(row=row, column=col, value=value)
    c.font      = _font(bold=True, color=_FG_NAVY, size=10)
    c.fill      = _fill(_BG_INPUT)
    c.alignment = _align("right")
    c.border    = _thin_border()
    return c

def _formula_cell(ws, row, col, formula):
    """Green-tinted formula cell."""
    c = ws.cell(row=row, column=col, value=formula)
    c.font      = _font(color=_FG_GREEN, size=10)
    c.fill      = _fill(_BG_FORMULA)
    c.alignment = _align("right")
    return c

def _label_cell(ws, row, col, value, bold=False, indent=0):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = _font(bold=bold, color=_FG_NAVY, size=10)
    c.alignment = _align("left", "center")
    if indent:
        c.alignment = Alignment(horizontal="left", vertical="center", indent=indent)
    return c

def _value_cell(ws, row, col, value, fmt=None, color=_FG_NAVY):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = _font(color=color, size=10)
    c.alignment = _align("right")
    if fmt:
        c.number_format = fmt
    return c

def _section_row(ws, row, label, ncols=20, bg=_BG_SUBHDR, fg=_FG_WHITE):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=label)
    c.font      = _font(bold=True, color=fg, size=9)
    c.fill      = _fill(bg)
    c.alignment = _align("left", "center")
    return c


# ─── Number formats ───────────────────────────────────────────────────────────
_FMT_MONEY  = '#,##0'          # whole-number $M
_FMT_MONEY1 = '#,##0.0'
_FMT_PCT    = '0.0%'
_FMT_PCT1   = '0.0%'
_FMT_MULT   = '0.0"×"'
_FMT_MULT2  = '0.00"×"'
_FMT_PRICE  = '$#,##0.00'
_FMT_PCT_RAW= '0.0'            # already a percentage, no symbol


# ─── Column-letter helpers ────────────────────────────────────────────────────

def _col_letter(n: int) -> str:
    """1-indexed column → letter. 1→A, 2→B … 26→Z, 27→AA …"""
    return get_column_letter(n)


# ─── Safe value helper ────────────────────────────────────────────────────────

def _v(d: dict, *keys, default=0):
    for k in keys:
        if k in d:
            v = d[k]
            try:
                return float(v) if v is not None else default
            except Exception:
                return default
    return default

def _sv(d: dict, *keys, default=""):
    for k in keys:
        if k in d:
            return str(d[k]) if d[k] is not None else default
    return default


###############################################################################
# SHEET 1: Readme
###############################################################################

def _build_readme(wb, d: dict):
    ws = wb.create_sheet("Readme")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 70

    _header_cell(ws, 1, 1, "", bg=_BG_HEADER)
    ws.merge_cells("B1:C1")
    c = ws.cell(row=1, column=2, value="📊  DCF Valuation Model — How to Use")
    c.font = _font(bold=True, color=_FG_WHITE, size=14)
    c.fill = _fill(_BG_HEADER)
    c.alignment = _align("left", "center")
    ws.row_dimensions[1].height = 30

    rows = [
        (3,  "Purpose",       "Discounted Cash Flow valuation model — 7-year forecast horizon, mid-year discounting, TV = UFCF_n / (WACC - g)."),
        (4,  "Data Source",   f"Yahoo Finance (yfinance). Pulled on {date.today().isoformat()}. Raw data is in Raw_IS / Raw_BS / Raw_CF sheets."),
        (5,  "Input Cells",   "All editable model inputs are on the Assumptions sheet (blue cells, column B). Change values there — model sheets update automatically."),
        (6,  "Formula Sheets","WACC_Calc, IS_Forecast, CF_Forecast and DCF_Calc all use Excel formulas referencing Assumptions + Raw sheets. Do NOT paste values over formula cells."),
        (7,  "Sensitivity",   "Sensitivity sheet shows IV across WACC (rows) and terminal-g (cols). Values are computed by the Python engine — re-export to refresh."),
        (8,  "Comps",         "Comps sheet shows peer multiples from Yahoo Finance. Peer data cached for 24 hours."),
        (9,  "Scenarios",     "Scenarios sheet shows Bear / Base / Bull IV estimates."),
        (10, "Colour Code",   "Blue cells = user inputs. Green cells = Excel formulas. Grey cells = labels. White cells = model output."),
        (11, "Units",         "All monetary values in $M (millions USD) unless stated otherwise."),
        (12, "DCF Convention","Terminal value uses perpetuity-growth method: TV = UFCF_terminal / (WACC - g). Mid-year discounting applied: PV factor = 1 / (1+WACC)^(n-0.5)."),
        (13, "Key Sheets",    "Assumptions → WACC_Calc → IS_Forecast → CF_Forecast → DCF_Calc → Cover (summary)"),
        (15, "DISCLAIMER",    "This model is for educational and research purposes only. It does not constitute investment advice."),
    ]

    for r, label, text in rows:
        _label_cell(ws, r, 2, label, bold=True)
        c = ws.cell(row=r, column=3, value=text)
        c.font      = _font(color=_FG_GREY, size=9)
        c.alignment = _align("left", "center", wrap=True)
        ws.row_dimensions[r].height = 22

    ws.freeze_panes = "B2"


###############################################################################
# SHEET 2: Cover
###############################################################################

def _build_cover(wb, d: dict):
    ws = wb.create_sheet("Cover")
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 40

    _set_col_widths(ws, {1: 3, 2: 28, 3: 18, 4: 18, 5: 18, 6: 18, 7: 18})

    ticker = _sv(d, "ticker", default="N/A")
    name   = _sv(d, "company_name", default=ticker)
    price  = _v(d, "price")
    iv     = _v(d, "intrinsic_value")
    upside = _v(d, "upside_pct")
    rec    = _sv(d, "recommendation", default="HOLD")
    wacc   = _v(d, "wacc")
    g      = _v(d, "terminal_growth")
    tv_pct = _v(d, "tv_pct")
    mktcap = _v(d, "market_cap_m")
    ev     = _v(d, "enterprise_value_m")
    nd     = _v(d, "net_debt")
    shares = _v(d, "diluted_shares")
    sector = _sv(d, "sector")
    industry = _sv(d, "industry")
    conf   = _v(d, "confidence_score", default=0)

    # Title bar
    ws.merge_cells("B1:G1")
    c = ws.cell(row=1, column=2, value=f"  {ticker}  —  {name}  |  DCF Valuation")
    c.font = _font(bold=True, color=_FG_WHITE, size=16, name="Calibri")
    c.fill = _fill(_BG_HEADER)
    c.alignment = _align("left", "center")

    # Sub-header
    ws.merge_cells("B2:G2")
    c = ws.cell(row=2, column=2, value=f"  {sector}  ·  {industry}  ·  As of {date.today().isoformat()}")
    c.font = _font(color=_FG_WHITE, size=10)
    c.fill = _fill(_BG_SUBHDR)
    c.alignment = _align("left", "center")

    # KPI block headers
    _section_row(ws, 4, "  KEY METRICS", 6)

    kpis = [
        ("Market Price",      f"${price:,.2f}"),
        ("Intrinsic Value",   f"${iv:,.2f}"),
        ("Upside / Downside", f"{upside:+.1f}%"),
        ("Recommendation",    rec),
        ("WACC",              f"{wacc:.1f}%"),
        ("Terminal Growth",   f"{g:.1f}%"),
        ("TV % of EV",        f"{tv_pct:.1f}%"),
        ("Confidence",        f"{int(conf)}/100"),
        ("Market Cap ($M)",   f"${mktcap:,.0f}M"),
        ("Enterprise Value",  f"${ev:,.0f}M"),
        ("Net Debt ($M)",     f"${nd:,.0f}M"),
        ("Diluted Shares",    f"{shares:,.1f}M"),
    ]

    for i, (label, val) in enumerate(kpis):
        r = 5 + (i // 3)
        c_col = 2 + (i % 3) * 2
        ws.cell(row=r, column=c_col, value=label).font   = _font(bold=True, size=9, color=_FG_GREY)
        ws.cell(row=r, column=c_col+1, value=val).font   = _font(bold=True, size=11, color=_FG_NAVY)
        ws.cell(row=r, column=c_col+1).alignment         = _align("left")
        ws.row_dimensions[r].height = 20

    # Assumptions summary
    r = 11
    _section_row(ws, r, "  MODEL ASSUMPTIONS", 6)
    assump = d.get("assumptions", [])
    for i, a in enumerate(assump[:18]):
        r_a = 12 + (i // 2)
        c_col = 2 + (i % 2) * 3
        ws.cell(row=r_a, column=c_col, value=a.get("driver", "")).font = _font(size=9, color=_FG_GREY)
        ws.cell(row=r_a, column=c_col+1, value=f"{a.get('active', '')} {a.get('unit', '')}").font = _font(bold=True, size=9, color=_FG_NAVY)
        ws.cell(row=r_a, column=c_col+2, value=a.get("source", "")).font = _font(size=8, color=_FG_GREY, italic=True)
        ws.row_dimensions[r_a].height = 16

    # Scenarios
    r = 24
    _section_row(ws, r, "  SCENARIO ANALYSIS", 6)
    headers = ["", "Bear", "Base", "Bull"]
    for ci, h in enumerate(headers):
        _header_cell(ws, r+1, 2+ci, h, bg=_BG_SUBHDR)
    scen = d.get("scenarios", {})
    for ri, (label, key) in enumerate([("Revenue CAGR", "revenue_cagr"), ("EBIT Margin", "ebit_margin"), ("WACC", "wacc"), ("Intrinsic Value", "intrinsic_value")]):
        for ci, s in enumerate(["bear", "base", "bull"]):
            sv = scen.get(s, {})
            val = sv.get(key, "-")
            ws.cell(row=r+2+ri, column=2+ci+1, value=val)
        ws.cell(row=r+2+ri, column=2, value=label).font = _font(size=9, color=_FG_NAVY)
        ws.row_dimensions[r+2+ri].height = 16


###############################################################################
# SHEET 3: Assumptions (INPUT SHEET — user-editable blue cells)
###############################################################################

# Fixed cell map — Assumptions!B3-B30 are referenced by formula sheets
_ASSUMP_MAP = {
    "revenue_growth_near": 3,   # B3
    "revenue_growth_far":  4,   # B4
    "ebit_margin_base":    5,   # B5
    "ebit_margin_target":  6,   # B6
    "tax_rate":            7,   # B7
    "da_pct":              8,   # B8
    "capex_pct":           9,   # B9
    "sbc_pct":             10,  # B10
    "dso":                 11,  # B11
    "dio":                 12,  # B12
    "dpo":                 13,  # B13
    "buyback_yield":       14,  # B14
    "dividend_yield":      15,  # B15
    "wacc":                16,  # B16
    "terminal_growth":     17,  # B17
    "cost_of_equity":      18,  # B18
    "cost_of_debt_pretax": 19,  # B19
    "beta":                20,  # B20
    "risk_free_rate":      21,  # B21
    "equity_risk_premium": 22,  # B22
    "equity_weight":       23,  # B23
    "debt_weight":         24,  # B24
    "diluted_shares":      25,  # B25
    "net_debt":            26,  # B26
    "minority_interest":   27,  # B27
    "preferred_equity":    28,  # B28
    "forecast_years":      29,  # B29
}

def _build_assumptions(wb, d: dict):
    ws = wb.create_sheet("Assumptions")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {1: 3, 2: 32, 3: 16, 4: 14, 5: 30})

    _header_cell(ws, 1, 2, "MODEL ASSUMPTIONS — EDITABLE INPUTS", bg=_BG_HEADER, fg=_FG_WHITE, size=12)
    ws.merge_cells("B1:E1")
    ws.row_dimensions[1].height = 28

    _header_cell(ws, 2, 2, "Driver", bg=_BG_SUBHDR, size=9)
    _header_cell(ws, 2, 3, "Value", bg=_BG_SUBHDR, size=9)
    _header_cell(ws, 2, 4, "Unit", bg=_BG_SUBHDR, size=9)
    _header_cell(ws, 2, 5, "Source / Notes", bg=_BG_SUBHDR, size=9)

    assump_rows = [
        # (B row, label, value, unit, source, section_before)
        ("REVENUE & MARGIN FORECASTS", None, None, None, None),
        ("Revenue Growth — Near Term (Yrs 1-3)", d.get("revenue_growth_near", 5.0), "%", "Mean of LTM CAGR + analyst consensus", "revenue_growth_near"),
        ("Revenue Growth — Long Term (Yrs 4-7)", d.get("revenue_growth_far",  3.0), "%", "Blend toward terminal rate", "revenue_growth_far"),
        ("EBIT Margin — Base (current LTM)",      d.get("ebit_margin_base", 10.0),  "%", "LTM EBIT / Revenue", "ebit_margin_base"),
        ("EBIT Margin — Target (Year 7)",          d.get("ebit_margin_target", 12.0),"%", "Linear convergence from base", "ebit_margin_target"),
        ("Tax Rate (effective)",                   d.get("tax_rate", 21.0),          "%", "LTM Tax Provision / Pretax Income", "tax_rate"),
        ("CAPEX, D&A, WORKING CAPITAL", None, None, None, None),
        ("D&A % Revenue",                          d.get("da_pct", 5.0),             "%", "LTM D&A / Revenue", "da_pct"),
        ("CapEx % Revenue",                        d.get("capex_pct", 6.0),          "%", "LTM CapEx / Revenue", "capex_pct"),
        ("SBC % Revenue",                          d.get("sbc_pct", 2.0),            "%", "LTM SBC / Revenue", "sbc_pct"),
        ("DSO (Days Sales Outstanding)",           d.get("dso", 30.0),               "days", "LTM AR / (Rev/365)", "dso"),
        ("DIO (Days Inventory Outstanding)",       d.get("dio", 20.0),               "days", "LTM Inventory / (COGS/365)", "dio"),
        ("DPO (Days Payable Outstanding)",         d.get("dpo", 25.0),               "days", "LTM AP / (COGS/365)", "dpo"),
        ("Buyback Yield",                          d.get("buyback_yield", 1.0),       "%", "LTM Buybacks / Market Cap", "buyback_yield"),
        ("Dividend Yield",                         d.get("dividend_yield", 1.0),      "%", "Current Dividend / Price", "dividend_yield"),
        ("DISCOUNT RATE", None, None, None, None),
        ("WACC (weighted avg cost of capital)",    d.get("wacc", 9.0),               "%", "=WACC_Calc!B15 (computed)", "wacc"),
        ("Terminal Growth Rate",                   d.get("terminal_growth", 2.5),    "%", "Long-run nominal GDP growth", "terminal_growth"),
        ("Cost of Equity (CAPM)",                  d.get("cost_of_equity", 10.0),    "%", "Rf + β × ERP", "cost_of_equity"),
        ("Cost of Debt (pre-tax)",                 d.get("cost_of_debt_pretax", 5.0),"%", "Interest Expense / Avg Debt", "cost_of_debt_pretax"),
        ("Beta (levered)",                         d.get("beta", 1.0),               "×", "5-year monthly vs S&P 500", "beta"),
        ("Risk-Free Rate (10-yr Treasury)",        d.get("risk_free_rate", 4.2),     "%", "FRED DGS10", "risk_free_rate"),
        ("Equity Risk Premium",                    d.get("equity_risk_premium", 5.5),"%", "Damodaran ERP", "equity_risk_premium"),
        ("Equity Weight (% of capital)",           d.get("equity_weight_pct", 80.0), "%", "Market cap / EV", "equity_weight"),
        ("Debt Weight (% of capital)",             d.get("debt_weight_pct", 20.0),   "%", "Debt / EV", "debt_weight"),
        ("BRIDGE TO EQUITY VALUE", None, None, None, None),
        ("Diluted Shares Outstanding (M)",         d.get("diluted_shares", 100.0),   "M shares", "Current shares from yfinance", "diluted_shares"),
        ("Net Debt ($M)",                          d.get("net_debt", 0.0),           "$M", "Total Debt − Cash", "net_debt"),
        ("Minority Interest ($M)",                 d.get("minority_interest_m", 0.0),"$M", "Non-controlling interests", "minority_interest"),
        ("Preferred Equity ($M)",                  d.get("preferred_equity_m", 0.0), "$M", "Preferred stock at par", "preferred_equity"),
        ("Forecast Years",                         7,                                "years", "Hard-coded DCF horizon", "forecast_years"),
    ]

    current_b_row = 3  # tracks B-row index for formula references
    r = 3
    for item in assump_rows:
        if len(item) == 5 and item[1] is None:
            # Section header
            _section_row(ws, r, f"  {item[0]}", 4)
            r += 1
            continue
        label, val, unit, src, key = item
        _label_cell(ws, r, 2, label, bold=False)
        # value in B col (blue input)
        try:
            fval = float(val) if val is not None else 0.0
        except Exception:
            fval = 0.0
        _input_cell(ws, r, 3, round(fval, 3))
        _label_cell(ws, r, 4, unit or "")
        c_src = ws.cell(row=r, column=5, value=src or "")
        c_src.font      = _font(color=_FG_GREY, size=8, italic=True)
        c_src.alignment = _align("left", "center")
        ws.row_dimensions[r].height = 17
        r += 1

    ws.freeze_panes = "C3"

    # Named range legend
    r += 2
    _section_row(ws, r, "  CELL REFERENCE MAP (for formulas in model sheets)", 4)
    r += 1
    ws.cell(row=r, column=2, value="Cell").font = _font(bold=True, size=9, color=_FG_WHITE)
    ws.cell(row=r, column=2).fill = _fill(_BG_SUBHDR)
    ws.cell(row=r, column=3, value="Assumption").font = _font(bold=True, size=9, color=_FG_WHITE)
    ws.cell(row=r, column=3).fill = _fill(_BG_SUBHDR)
    r += 1
    b_row = 3
    for item in assump_rows:
        if item[1] is None:
            b_row += 1
            continue
        ws.cell(row=r, column=2, value=f"Assumptions!$C${b_row}").font = _font(size=9, color=_FG_GREY)
        ws.cell(row=r, column=3, value=item[0]).font = _font(size=9, color=_FG_NAVY)
        r += 1
        b_row += 1


###############################################################################
# HELPERS: figure out the C-row in Assumptions for a given key
###############################################################################

def _arow(key: str) -> int:
    """Return the Assumptions row (C column) for assumption key."""
    b_row = 3
    assump_rows_keys = [
        None,  # section "REVENUE & MARGIN FORECASTS"
        "revenue_growth_near", "revenue_growth_far",
        "ebit_margin_base", "ebit_margin_target", "tax_rate",
        None,  # section "CAPEX, D&A, WORKING CAPITAL"
        "da_pct", "capex_pct", "sbc_pct", "dso", "dio", "dpo",
        "buyback_yield", "dividend_yield",
        None,  # section "DISCOUNT RATE"
        "wacc", "terminal_growth", "cost_of_equity", "cost_of_debt_pretax",
        "beta", "risk_free_rate", "equity_risk_premium",
        "equity_weight", "debt_weight",
        None,  # section "BRIDGE TO EQUITY VALUE"
        "diluted_shares", "net_debt", "minority_interest",
        "preferred_equity", "forecast_years",
    ]
    for k in assump_rows_keys:
        if k is None:
            b_row += 1
        elif k == key:
            return b_row
        else:
            b_row += 1
    return 99  # fallback

# Pre-compute frequently used assumption rows
_AR_REV_NEAR    = _arow("revenue_growth_near")   # C3
_AR_REV_FAR     = _arow("revenue_growth_far")    # C4
_AR_EBIT_BASE   = _arow("ebit_margin_base")      # C5
_AR_EBIT_TGT    = _arow("ebit_margin_target")    # C6
_AR_TAX         = _arow("tax_rate")              # C7
_AR_DA          = _arow("da_pct")                # C9  (section adds 1)
_AR_CAPEX       = _arow("capex_pct")             # C10
_AR_SBC         = _arow("sbc_pct")              # C11
_AR_WACC        = _arow("wacc")
_AR_TGROWTH     = _arow("terminal_growth")
_AR_COE         = _arow("cost_of_equity")
_AR_COD         = _arow("cost_of_debt_pretax")
_AR_BETA        = _arow("beta")
_AR_RF          = _arow("risk_free_rate")
_AR_ERP         = _arow("equity_risk_premium")
_AR_EW          = _arow("equity_weight")
_AR_DW          = _arow("debt_weight")
_AR_SHARES      = _arow("diluted_shares")
_AR_ND          = _arow("net_debt")
_AR_MI          = _arow("minority_interest")
_AR_PE          = _arow("preferred_equity")


def _aref(key: str) -> str:
    """Return absolute Excel cell reference for Assumptions!$C$row."""
    return f"Assumptions!$C${_arow(key)}"


###############################################################################
# SHEET 4-6: Raw_IS / Raw_BS / Raw_CF  (VALUES ONLY — yfinance source data)
###############################################################################

# Raw_IS row map (rows are fixed so formula sheets can reference them)
_RAW_IS_ROWS = {
    "header":          1,
    "revenue":         2,
    "gross_profit":    3,
    "ebit":            4,
    "net_income":      5,
    "diluted_shares":  6,
    "tax_provision":   7,
    "pretax_income":   8,
    "gross_margin":    10,  # formula
    "ebit_margin":     11,  # formula
    "net_margin":      12,  # formula
}

# Raw_CF row map
_RAW_CF_ROWS = {
    "header":     1,
    "op_cf":      2,
    "capex":      3,
    "fcf":        4,
    "da":         5,
    "sbc":        6,
    "interest":   7,
    "buybacks":   8,
    "da_pct":     10,  # formula
    "capex_pct":  11,  # formula
    "fcf_margin": 12,  # formula
    "sbc_pct":    13,  # formula
}

# Raw_BS row map
_RAW_BS_ROWS = {
    "header":       1,
    "total_assets": 2,
    "total_debt":   3,
    "cash":         4,
    "net_debt":     5,  # formula
    "equity":       6,
    "roic":         8,  # formula
}


def _year_cols(years: list, start_col: int = 2) -> dict:
    """Map year → column index."""
    return {y: start_col + i for i, y in enumerate(sorted(years))}


def _build_raw_is(wb, d: dict):
    ws = wb.create_sheet("Raw_IS")
    ws.sheet_view.showGridLines = False

    hist   = d.get("historical", {})
    years  = hist.get("years", [])
    ycols  = _year_cols(years)

    _set_col_widths(ws, {1: 28})
    for col in range(2, 2 + len(years) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    _header_cell(ws, 1, 1, "INCOME STATEMENT — RAW DATA ($M)", bg=_BG_HEADER, fg=_FG_WHITE, size=11, h="left")
    for y, col in ycols.items():
        _header_cell(ws, 1, col, str(y), bg=_BG_SUBHDR)

    rows_def = [
        (_RAW_IS_ROWS["revenue"],        "Total Revenue ($M)",     "revenue"),
        (_RAW_IS_ROWS["gross_profit"],    "Gross Profit ($M)",      "gross_profit"),
        (_RAW_IS_ROWS["ebit"],            "EBIT ($M)",              "ebit"),
        (_RAW_IS_ROWS["net_income"],      "Net Income ($M)",        "net_income"),
        (_RAW_IS_ROWS["diluted_shares"],  "Diluted Shares (M)",     "shares"),
        (_RAW_IS_ROWS["tax_provision"],   "Tax Provision ($M)",     "tax"),
        (_RAW_IS_ROWS["pretax_income"],   "Pretax Income ($M)",     "pretax"),
    ]

    arrays = {
        "revenue":      hist.get("revenue", []),
        "gross_profit": hist.get("gross_profit", []),
        "ebit":         hist.get("ebit", []),
        "net_income":   hist.get("net_income", []),
        "shares":       hist.get("shares", []),
        "tax":          hist.get("tax", []),
        "pretax":       hist.get("pretax_income", []),
    }

    for row_idx, label, arr_key in rows_def:
        ws.cell(row=row_idx, column=1, value=label).font = _font(size=9, color=_FG_NAVY)
        arr = arrays.get(arr_key, [])
        for i, (y, col) in enumerate(ycols.items()):
            val = arr[i] if i < len(arr) else 0
            _value_cell(ws, row_idx, col, val, _FMT_MONEY)

    # Derived ratio rows (formulas)
    ws.cell(row=9, column=1, value="— RATIOS —").font = _font(bold=True, size=8, color=_FG_GREY)
    margin_defs = [
        (_RAW_IS_ROWS["gross_margin"],  "Gross Margin %",    _RAW_IS_ROWS["gross_profit"]),
        (_RAW_IS_ROWS["ebit_margin"],   "EBIT Margin %",     _RAW_IS_ROWS["ebit"]),
        (_RAW_IS_ROWS["net_margin"],    "Net Margin %",      _RAW_IS_ROWS["net_income"]),
    ]
    for row_idx, label, num_row in margin_defs:
        ws.cell(row=row_idx, column=1, value=label).font = _font(size=9, color=_FG_GREY)
        for y, col in ycols.items():
            rev_col = get_column_letter(col)
            rev_row = _RAW_IS_ROWS["revenue"]
            formula = f"=IF({rev_col}{rev_row}<>0,{rev_col}{row_idx-1 if row_idx>_RAW_IS_ROWS['ebit'] else row_idx}/{rev_col}{rev_row},0)"
            # Simpler: just use the numerator / revenue
            formula = f"=IF({rev_col}{rev_row}<>0,{rev_col}{num_row}/{rev_col}{rev_row},0)"
            _formula_cell(ws, row_idx, col, formula).number_format = _FMT_PCT

    ws.freeze_panes = "B2"


def _build_raw_bs(wb, d: dict):
    ws = wb.create_sheet("Raw_BS")
    ws.sheet_view.showGridLines = False

    hist  = d.get("historical", {})
    years = hist.get("years", [])
    ycols = _year_cols(years)

    _set_col_widths(ws, {1: 28})
    for col in range(2, 2 + len(years) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    _header_cell(ws, 1, 1, "BALANCE SHEET — RAW DATA ($M)", bg=_BG_HEADER, fg=_FG_WHITE, size=11, h="left")
    for y, col in ycols.items():
        _header_cell(ws, 1, col, str(y), bg=_BG_SUBHDR)

    bs_defs = [
        (_RAW_BS_ROWS["total_assets"], "Total Assets ($M)",   hist.get("total_assets", [])),
        (_RAW_BS_ROWS["total_debt"],   "Total Debt ($M)",     hist.get("total_debt", [])),
        (_RAW_BS_ROWS["cash"],         "Cash & Equivalents",  hist.get("cash", [])),
        (_RAW_BS_ROWS["equity"],       "Total Equity ($M)",   hist.get("equity", [])),
    ]
    for row_idx, label, arr in bs_defs:
        ws.cell(row=row_idx, column=1, value=label).font = _font(size=9, color=_FG_NAVY)
        for i, (y, col) in enumerate(ycols.items()):
            _value_cell(ws, row_idx, col, arr[i] if i < len(arr) else 0, _FMT_MONEY)

    # Net debt formula row
    ws.cell(row=_RAW_BS_ROWS["net_debt"], column=1, value="Net Debt ($M)").font = _font(size=9, color=_FG_GREY)
    for y, col in ycols.items():
        cc = get_column_letter(col)
        _formula_cell(ws, _RAW_BS_ROWS["net_debt"], col,
            f"={cc}{_RAW_BS_ROWS['total_debt']}-{cc}{_RAW_BS_ROWS['cash']}").number_format = _FMT_MONEY

    ws.freeze_panes = "B2"


def _build_raw_cf(wb, d: dict):
    ws = wb.create_sheet("Raw_CF")
    ws.sheet_view.showGridLines = False

    hist  = d.get("historical", {})
    years = hist.get("years", [])
    ycols = _year_cols(years)

    _set_col_widths(ws, {1: 28})
    for col in range(2, 2 + len(years) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    _header_cell(ws, 1, 1, "CASH FLOW STATEMENT — RAW DATA ($M)", bg=_BG_HEADER, fg=_FG_WHITE, size=11, h="left")
    for y, col in ycols.items():
        _header_cell(ws, 1, col, str(y), bg=_BG_SUBHDR)

    cf_defs = [
        (_RAW_CF_ROWS["op_cf"],   "Operating Cash Flow",      hist.get("op_cf", [])),
        (_RAW_CF_ROWS["capex"],   "Capital Expenditure",       hist.get("capex", [])),
        (_RAW_CF_ROWS["fcf"],     "Free Cash Flow",            hist.get("fcf", [])),
        (_RAW_CF_ROWS["da"],      "D&A",                       hist.get("da", [])),
        (_RAW_CF_ROWS["sbc"],     "Stock-Based Compensation",  hist.get("sbc", [])),
        (_RAW_CF_ROWS["buybacks"],"Share Buybacks",            hist.get("buybacks", [])),
    ]
    for row_idx, label, arr in cf_defs:
        ws.cell(row=row_idx, column=1, value=label).font = _font(size=9, color=_FG_NAVY)
        for i, (y, col) in enumerate(ycols.items()):
            _value_cell(ws, row_idx, col, arr[i] if i < len(arr) else 0, _FMT_MONEY)

    # Ratio formula rows (need to reference Raw_IS)
    ws.cell(row=9, column=1, value="— % OF REVENUE —").font = _font(bold=True, size=8, color=_FG_GREY)
    pct_defs = [
        (_RAW_CF_ROWS["da_pct"],    "D&A % Revenue",     _RAW_CF_ROWS["da"]),
        (_RAW_CF_ROWS["capex_pct"], "CapEx % Revenue",   _RAW_CF_ROWS["capex"]),
        (_RAW_CF_ROWS["fcf_margin"],"FCF Margin %",      _RAW_CF_ROWS["fcf"]),
        (_RAW_CF_ROWS["sbc_pct"],   "SBC % Revenue",     _RAW_CF_ROWS["sbc"]),
    ]
    n_hist_years = len(years)
    for row_idx, label, num_row in pct_defs:
        ws.cell(row=row_idx, column=1, value=label).font = _font(size=9, color=_FG_GREY)
        for i, (y, col) in enumerate(ycols.items()):
            cc = get_column_letter(col)
            rev_row = _RAW_IS_ROWS["revenue"]
            formula = f"=IF(Raw_IS!{cc}{rev_row}<>0,{cc}{num_row}/Raw_IS!{cc}{rev_row},0)"
            _formula_cell(ws, row_idx, col, formula).number_format = _FMT_PCT

    ws.freeze_panes = "B2"


###############################################################################
# SHEET 7: WACC_Calc  (formula-linked to Assumptions!$C$...)
###############################################################################

def _build_wacc_calc(wb, d: dict):
    ws = wb.create_sheet("WACC_Calc")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {1: 3, 2: 36, 3: 18, 4: 28})

    _header_cell(ws, 1, 2, "WACC BUILD-UP (CAPM METHOD)", bg=_BG_HEADER, fg=_FG_WHITE, size=12, h="left")
    ws.merge_cells("B1:D1")
    ws.row_dimensions[1].height = 28

    def _row(r, label, formula_or_val, note="", is_section=False, is_result=False):
        if is_section:
            _section_row(ws, r, f"  {label}", 3)
            return
        ws.cell(row=r, column=2, value=label).font = _font(
            size=10, bold=is_result,
            color=_FG_NAVY if not is_result else _FG_WHITE
        )
        if is_result:
            ws.cell(row=r, column=2).fill = _fill(_BG_HEADER)
        c = ws.cell(row=r, column=3, value=formula_or_val)
        c.font = _font(bold=is_result,
                       color=_FG_GREEN if isinstance(formula_or_val, str) and formula_or_val.startswith("=") else _FG_NAVY,
                       size=10)
        if isinstance(formula_or_val, str) and formula_or_val.startswith("="):
            c.fill = _fill(_BG_FORMULA)
        else:
            c.fill = _fill(_BG_INPUT)
        c.number_format = _FMT_PCT_RAW
        c.alignment = _align("right")
        ws.cell(row=r, column=4, value=note).font = _font(color=_FG_GREY, size=9, italic=True)
        ws.row_dimensions[r].height = 18

    _row(2, "COST OF EQUITY", None, is_section=True)
    _row(3,  "Risk-Free Rate (Rf)",     f"={_aref('risk_free_rate')}",        "10-yr Treasury yield from FRED")
    _row(4,  "Beta (β, levered)",       f"={_aref('beta')}",                  "5-yr monthly regression vs S&P 500")
    _row(5,  "Equity Risk Premium (ERP)",f"={_aref('equity_risk_premium')}",  "Damodaran implied ERP")
    _row(6,  "Cost of Equity (Ke)",     f"=C3+C4*C5",                        "Ke = Rf + β × ERP  (CAPM)")

    _row(7, "COST OF DEBT", None, is_section=True)
    _row(8,  "Cost of Debt (pre-tax)",  f"={_aref('cost_of_debt_pretax')}",   "Interest Expense / Avg Debt")
    _row(9,  "Tax Rate",                f"={_aref('tax_rate')}/100",          "LTM effective rate")
    _row(10, "Cost of Debt (after-tax)",f"=C8*(1-C9)",                       "Kd × (1 − t)")

    _row(11, "CAPITAL STRUCTURE", None, is_section=True)
    _row(12, "Equity Weight (We)",      f"={_aref('equity_weight')}/100",    "Market cap / Enterprise Value")
    _row(13, "Debt Weight (Wd)",        f"={_aref('debt_weight')}/100",      "Debt / Enterprise Value")
    _row(14, "Check (We + Wd = 1)",     f"=C12+C13",                        "Should equal 1.0 (100%)")

    _row(15, "WACC", None, is_section=True)
    _row(16, "WACC",                    f"=C12*C6+C13*C10",                  "We×Ke + Wd×Kd×(1-t)", is_result=True)
    _row(17, "WACC (from Assumptions)", f"={_aref('wacc')}/100",            "Manual override / cross-check")
    _row(18, "Terminal Growth (g)",     f"={_aref('terminal_growth')}/100",  "From Assumptions")
    _row(19, "WACC − g spread",         f"=C17-C18",                        "Must be ≥ 0.5% for DCF to be valid")

    # Highlight the WACC row
    ws.cell(row=16, column=2).fill = _fill(_BG_HEADER)
    ws.cell(row=16, column=3).fill = _fill(_BG_FORMULA)

    ws.freeze_panes = "B2"


###############################################################################
# SHEET 8: IS_Forecast  (formula-linked)
###############################################################################

# IS_Forecast column map: A=label, B=Yr1, C=Yr2 ... H=Yr7
_FCY_LABEL_COL   = 1
_FCY_START_COL   = 2   # Year 1
_FCY_N_YEARS     = 7

# IS_Forecast row map
_ISF_ROWS = {
    "year_hdr":     1,
    "revenue":      2,
    "rev_growth":   3,
    "gross_profit": 4,
    "gp_margin":    5,
    "ebit":         6,
    "ebit_margin":  7,
    "da":           8,
    "ebitda":       9,
    "tax":          10,
    "nopat":        11,
    "sbc":          12,
    "nwc_change":   13,
}

def _build_is_forecast(wb, d: dict):
    ws = wb.create_sheet("IS_Forecast")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {1: 26, **{c: 14 for c in range(2, 2 + _FCY_N_YEARS + 1)}})

    _header_cell(ws, 1, 1, "INCOME STATEMENT FORECAST ($M)", bg=_BG_HEADER, fg=_FG_WHITE, size=11, h="left")
    base_year = max(d.get("historical", {}).get("years", [2023]))
    for i in range(_FCY_N_YEARS):
        yr = base_year + i + 1
        _header_cell(ws, 1, _FCY_START_COL + i, f"FY{yr}", bg=_BG_SUBHDR)

    # Revenue base = most recent actual (from Raw_IS)
    # We need to figure out the last column in Raw_IS
    hist_years = sorted(d.get("historical", {}).get("years", []))
    n_hist = len(hist_years)
    last_hist_col = get_column_letter(1 + n_hist)   # Raw_IS col for latest year

    labels = {
        _ISF_ROWS["revenue"]:      "Revenue ($M)",
        _ISF_ROWS["rev_growth"]:   "Revenue Growth (%)",
        _ISF_ROWS["gross_profit"]: "Gross Profit ($M)",
        _ISF_ROWS["gp_margin"]:    "Gross Margin (%)",
        _ISF_ROWS["ebit"]:         "EBIT ($M)",
        _ISF_ROWS["ebit_margin"]:  "EBIT Margin (%)",
        _ISF_ROWS["da"]:           "D&A ($M)",
        _ISF_ROWS["ebitda"]:       "EBITDA ($M)",
        _ISF_ROWS["tax"]:          "Tax (NOPAT adj.) ($M)",
        _ISF_ROWS["nopat"]:        "NOPAT ($M)",
        _ISF_ROWS["sbc"]:          "SBC ($M)",
        _ISF_ROWS["nwc_change"]:   "NWC Change ($M)",
    }
    for rr, lbl in labels.items():
        ws.cell(row=rr, column=1, value=lbl).font = _font(size=9, color=_FG_NAVY, bold=(rr in [2, 6, 11]))
        ws.row_dimensions[rr].height = 17

    # The growth blend: years 1-3 use near-term rate, 4-7 use far rate
    # Use Assumptions near rate for cols 2-4, far rate for cols 5-8
    for i in range(_FCY_N_YEARS):
        col = _FCY_START_COL + i
        cc  = get_column_letter(col)

        # Revenue
        if i == 0:
            # Year 1 off the last historical Raw_IS value
            rev_formula = f"=Raw_IS!{last_hist_col}{_RAW_IS_ROWS['revenue']}*(1+{_aref('revenue_growth_near')}/100)"
        elif i < 3:
            prev = get_column_letter(col - 1)
            rev_formula = f"={prev}{_ISF_ROWS['revenue']}*(1+{_aref('revenue_growth_near')}/100)"
        else:
            # Blend: linear taper from near to far over years 4-7
            prev = get_column_letter(col - 1)
            rev_formula = f"={prev}{_ISF_ROWS['revenue']}*(1+{_aref('revenue_growth_far')}/100)"
        _formula_cell(ws, _ISF_ROWS["revenue"], col, rev_formula).number_format = _FMT_MONEY

        # Rev growth label (just show the applicable rate)
        g_formula = f"={cc}{_ISF_ROWS['revenue']}/{get_column_letter(col-1) if i>0 else 'Raw_IS!'+last_hist_col}{_ISF_ROWS['revenue']}-1"
        _formula_cell(ws, _ISF_ROWS["rev_growth"], col, g_formula if i > 0 else f"={_aref('revenue_growth_near')}/100").number_format = _FMT_PCT

        # EBIT margin — linear interpolation from base to target over 7 years
        # margin_yr_i = base + (target - base) * i / (N-1)
        frac = i / (_FCY_N_YEARS - 1)
        ebit_m_formula = (
            f"={_aref('ebit_margin_base')}/100"
            f"+({_aref('ebit_margin_target')}/100-{_aref('ebit_margin_base')}/100)*{frac:.4f}"
        )
        _formula_cell(ws, _ISF_ROWS["ebit_margin"], col, ebit_m_formula).number_format = _FMT_PCT

        # EBIT
        _formula_cell(ws, _ISF_ROWS["ebit"], col, f"={cc}{_ISF_ROWS['revenue']}*{cc}{_ISF_ROWS['ebit_margin']}").number_format = _FMT_MONEY

        # Gross profit (proxy: EBIT + D&A, since we don't have detailed opex splits)
        _formula_cell(ws, _ISF_ROWS["gross_profit"], col,
            f"={cc}{_ISF_ROWS['ebit']}+{cc}{_ISF_ROWS['da']}").number_format = _FMT_MONEY
        _formula_cell(ws, _ISF_ROWS["gp_margin"], col,
            f"=IF({cc}{_ISF_ROWS['revenue']}<>0,{cc}{_ISF_ROWS['gross_profit']}/{cc}{_ISF_ROWS['revenue']},0)").number_format = _FMT_PCT

        # D&A
        _formula_cell(ws, _ISF_ROWS["da"], col,
            f"={cc}{_ISF_ROWS['revenue']}*{_aref('da_pct')}/100").number_format = _FMT_MONEY

        # EBITDA
        _formula_cell(ws, _ISF_ROWS["ebitda"], col,
            f"={cc}{_ISF_ROWS['ebit']}+{cc}{_ISF_ROWS['da']}").number_format = _FMT_MONEY

        # Tax
        _formula_cell(ws, _ISF_ROWS["tax"], col,
            f"={cc}{_ISF_ROWS['ebit']}*{_aref('tax_rate')}/100").number_format = _FMT_MONEY

        # NOPAT
        _formula_cell(ws, _ISF_ROWS["nopat"], col,
            f"={cc}{_ISF_ROWS['ebit']}-{cc}{_ISF_ROWS['tax']}").number_format = _FMT_MONEY

        # SBC
        _formula_cell(ws, _ISF_ROWS["sbc"], col,
            f"={cc}{_ISF_ROWS['revenue']}*{_aref('sbc_pct')}/100").number_format = _FMT_MONEY

        # NWC change (simple DSO/DIO/DPO approach, placeholder as 0 for simplicity)
        _formula_cell(ws, _ISF_ROWS["nwc_change"], col, "=0").number_format = _FMT_MONEY

    ws.freeze_panes = "B2"


###############################################################################
# SHEET 9: CF_Forecast  (UFCF computation, formula-linked)
###############################################################################

_CFF_ROWS = {
    "year_hdr":   1,
    "nopat":      2,   # =IS_Forecast!Bxx
    "da":         3,   # + D&A
    "sbc":        4,   # - SBC
    "capex":      5,   # - CapEx
    "nwc":        6,   # - ΔNWC
    "ufcf":       7,   # = UFCF
    "ufcf_margin":8,   # UFCF / Revenue
    "cum_pv":     9,   # placeholder
}

def _build_cf_forecast(wb, d: dict):
    ws = wb.create_sheet("CF_Forecast")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {1: 26, **{c: 14 for c in range(2, 2 + _FCY_N_YEARS + 1)}})

    base_year = max(d.get("historical", {}).get("years", [2023]))
    _header_cell(ws, 1, 1, "UNLEVERED FREE CASH FLOW FORECAST ($M)", bg=_BG_HEADER, fg=_FG_WHITE, size=11, h="left")
    for i in range(_FCY_N_YEARS):
        yr = base_year + i + 1
        _header_cell(ws, 1, _FCY_START_COL + i, f"FY{yr}", bg=_BG_SUBHDR)

    labels = {
        _CFF_ROWS["nopat"]:       "NOPAT ($M)",
        _CFF_ROWS["da"]:          "+ D&A ($M)",
        _CFF_ROWS["sbc"]:         "− SBC ($M)",
        _CFF_ROWS["capex"]:       "− CapEx ($M)",
        _CFF_ROWS["nwc"]:         "− ΔNWC ($M)",
        _CFF_ROWS["ufcf"]:        "= UFCF ($M)",
        _CFF_ROWS["ufcf_margin"]: "UFCF Margin (%)",
    }
    for rr, lbl in labels.items():
        ws.cell(row=rr, column=1, value=lbl).font = _font(size=9, color=_FG_NAVY,
                                                           bold=(rr == _CFF_ROWS["ufcf"]))
        ws.row_dimensions[rr].height = 17

    for i in range(_FCY_N_YEARS):
        col = _FCY_START_COL + i
        cc  = get_column_letter(col)

        _formula_cell(ws, _CFF_ROWS["nopat"], col,
            f"=IS_Forecast!{cc}{_ISF_ROWS['nopat']}").number_format = _FMT_MONEY
        _formula_cell(ws, _CFF_ROWS["da"],    col,
            f"=IS_Forecast!{cc}{_ISF_ROWS['da']}").number_format = _FMT_MONEY
        _formula_cell(ws, _CFF_ROWS["sbc"],   col,
            f"=IS_Forecast!{cc}{_ISF_ROWS['sbc']}").number_format = _FMT_MONEY
        _formula_cell(ws, _CFF_ROWS["capex"],  col,
            f"=IS_Forecast!{cc}{_ISF_ROWS['revenue']}*{_aref('capex_pct')}/100").number_format = _FMT_MONEY

        # NWC change (zero for now — referenced from IS_Forecast NWC row)
        _formula_cell(ws, _CFF_ROWS["nwc"],   col, "=0").number_format = _FMT_MONEY

        # UFCF = NOPAT + D&A - SBC - CapEx - ΔNWC
        _formula_cell(ws, _CFF_ROWS["ufcf"],  col,
            f"={cc}{_CFF_ROWS['nopat']}+{cc}{_CFF_ROWS['da']}"
            f"-{cc}{_CFF_ROWS['sbc']}-{cc}{_CFF_ROWS['capex']}-{cc}{_CFF_ROWS['nwc']}").number_format = _FMT_MONEY

        _formula_cell(ws, _CFF_ROWS["ufcf_margin"], col,
            f"=IF(IS_Forecast!{cc}{_ISF_ROWS['revenue']}<>0,"
            f"{cc}{_CFF_ROWS['ufcf']}/IS_Forecast!{cc}{_ISF_ROWS['revenue']},0)").number_format = _FMT_PCT

    ws.freeze_panes = "B2"


###############################################################################
# SHEET 10: DCF_Calc  (discount factors, PV sums, TV, bridge)
###############################################################################

_DCF_ROWS = {
    "year_hdr":   1,
    "year_n":     2,   # 1, 2, 3 … 7
    "ufcf":       3,   # =CF_Forecast!Bxx
    "disc_factor":4,   # =1/(1+WACC)^(n-0.5)
    "pv_ufcf":    5,   # =ufcf * disc_factor
    "cumulative": 6,   # running sum
    "terminal":   8,   # terminal value
    "pv_tv":      9,   # PV of TV
    "sum_pv":     11,  # sum of PV UFCFs
    "pv_tv_sum":  12,  # PV TV
    "ev":         13,  # Enterprise Value
    "nd":         14,  # Net debt
    "minority":   15,  # minority interest
    "preferred":  16,  # preferred equity
    "equity_val": 17,  # equity value
    "shares":     18,  # shares
    "iv":         19,  # intrinsic value per share
    "price":      20,  # current price
    "upside":     21,  # upside
    "wacc_used":  23,  # WACC used
    "g_used":     24,  # terminal g used
    "tv_pct_ev":  25,  # TV % of EV
}

def _build_dcf_calc(wb, d: dict):
    ws = wb.create_sheet("DCF_Calc")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {1: 30, **{c: 14 for c in range(2, 2 + _FCY_N_YEARS + 1)}, _FCY_N_YEARS + 2: 18})

    base_year = max(d.get("historical", {}).get("years", [2023]))
    _header_cell(ws, 1, 1, "DCF CALCULATION & EQUITY BRIDGE ($M)", bg=_BG_HEADER, fg=_FG_WHITE, size=11, h="left")
    for i in range(_FCY_N_YEARS):
        yr = base_year + i + 1
        _header_cell(ws, 1, _FCY_START_COL + i, f"FY{yr}", bg=_BG_SUBHDR)

    last_fc_col = get_column_letter(_FCY_START_COL + _FCY_N_YEARS - 1)

    labels = {
        _DCF_ROWS["year_n"]:     "Year (n)",
        _DCF_ROWS["ufcf"]:       "UFCF ($M)",
        _DCF_ROWS["disc_factor"]:"Discount Factor  [1/(1+WACC)^(n-0.5)]",
        _DCF_ROWS["pv_ufcf"]:   "PV of UFCF ($M)",
        _DCF_ROWS["cumulative"]: "Cumulative PV ($M)",
    }
    for rr, lbl in labels.items():
        ws.cell(row=rr, column=1, value=lbl).font = _font(size=9, color=_FG_NAVY,
                                                           bold=(rr == _DCF_ROWS["ufcf"]))
        ws.row_dimensions[rr].height = 17

    # WACC reference cell (use the Assumptions WACC)
    wacc_ref = f"{_aref('wacc')}/100"

    for i in range(_FCY_N_YEARS):
        col = _FCY_START_COL + i
        cc  = get_column_letter(col)
        n   = i + 1

        ws.cell(row=_DCF_ROWS["year_n"], column=col, value=n).font = _font(size=9, color=_FG_GREY)

        _formula_cell(ws, _DCF_ROWS["ufcf"], col,
            f"=CF_Forecast!{cc}{_CFF_ROWS['ufcf']}").number_format = _FMT_MONEY

        _formula_cell(ws, _DCF_ROWS["disc_factor"], col,
            f"=1/(1+{wacc_ref})^({n}-0.5)").number_format = "0.0000"

        _formula_cell(ws, _DCF_ROWS["pv_ufcf"], col,
            f"={cc}{_DCF_ROWS['ufcf']}*{cc}{_DCF_ROWS['disc_factor']}").number_format = _FMT_MONEY

        if i == 0:
            cum_formula = f"={cc}{_DCF_ROWS['pv_ufcf']}"
        else:
            prev = get_column_letter(col - 1)
            cum_formula = f"={prev}{_DCF_ROWS['cumulative']}+{cc}{_DCF_ROWS['pv_ufcf']}"
        _formula_cell(ws, _DCF_ROWS["cumulative"], col, cum_formula).number_format = _FMT_MONEY

    # Terminal value section
    r = _DCF_ROWS["terminal"]
    _section_row(ws, r - 1, "  TERMINAL VALUE", _FCY_N_YEARS + 1)
    ws.cell(row=r, column=1, value="Terminal UFCF ($M)").font = _font(size=9, color=_FG_NAVY)
    _formula_cell(ws, r, 2,
        f"=CF_Forecast!{last_fc_col}{_CFF_ROWS['ufcf']}").number_format = _FMT_MONEY

    ws.cell(row=_DCF_ROWS["pv_tv"] - 1, column=1, value="Terminal Value ($M)").font = _font(size=9, color=_FG_NAVY)
    # TV = terminal_UFCF / (WACC - g)   [Nike convention: last year UFCF / spread]
    tv_formula = f"=B{r}/({wacc_ref}-{_aref('terminal_growth')}/100)"
    _formula_cell(ws, _DCF_ROWS["pv_tv"] - 1, 2, tv_formula).number_format = _FMT_MONEY

    tv_disc_yr = _FCY_N_YEARS  # discount TV at year N
    ws.cell(row=_DCF_ROWS["pv_tv"], column=1, value="PV of Terminal Value ($M)").font = _font(size=9, color=_FG_NAVY)
    _formula_cell(ws, _DCF_ROWS["pv_tv"], 2,
        f"=B{_DCF_ROWS['pv_tv']-1}/(1+{wacc_ref})^{tv_disc_yr}").number_format = _FMT_MONEY

    # Equity bridge
    r2 = _DCF_ROWS["sum_pv"]
    _section_row(ws, r2 - 1, "  EQUITY VALUE BRIDGE", _FCY_N_YEARS + 1)

    sum_pv_cols = "+".join(
        get_column_letter(_FCY_START_COL + i) + str(_DCF_ROWS["pv_ufcf"])
        for i in range(_FCY_N_YEARS)
    )
    bridge_rows = [
        (r2,                         "Sum of PV UFCFs ($M)",         f"={sum_pv_cols}"),
        (_DCF_ROWS["pv_tv_sum"],     "+ PV Terminal Value ($M)",     f"=B{_DCF_ROWS['pv_tv']}"),
        (_DCF_ROWS["ev"],            "= Enterprise Value ($M)",      f"=B{r2}+B{_DCF_ROWS['pv_tv_sum']}"),
        (_DCF_ROWS["nd"],            "− Net Debt ($M)",              f"={_aref('net_debt')}"),
        (_DCF_ROWS["minority"],      "− Minority Interest ($M)",     f"={_aref('minority_interest')}"),
        (_DCF_ROWS["preferred"],     "− Preferred Equity ($M)",      f"={_aref('preferred_equity')}"),
        (_DCF_ROWS["equity_val"],    "= Equity Value ($M)",          f"=B{_DCF_ROWS['ev']}-B{_DCF_ROWS['nd']}-B{_DCF_ROWS['minority']}-B{_DCF_ROWS['preferred']}"),
        (_DCF_ROWS["shares"],        "Diluted Shares Outstanding (M)",f"={_aref('diluted_shares')}"),
        (_DCF_ROWS["iv"],            "Intrinsic Value per Share ($)", f"=IF(B{_DCF_ROWS['shares']}<>0,B{_DCF_ROWS['equity_val']}/B{_DCF_ROWS['shares']},0)"),
        (_DCF_ROWS["price"],         "Current Market Price ($)",     d.get("price", 0)),
        (_DCF_ROWS["upside"],        "Upside / Downside (%)",        f"=IF(B{_DCF_ROWS['price']}<>0,(B{_DCF_ROWS['iv']}-B{_DCF_ROWS['price']})/B{_DCF_ROWS['price']},0)"),
    ]
    for row_idx, label, formula_or_val in bridge_rows:
        is_result = row_idx in [_DCF_ROWS["ev"], _DCF_ROWS["equity_val"], _DCF_ROWS["iv"]]
        ws.cell(row=row_idx, column=1, value=label).font = _font(size=9, bold=is_result, color=_FG_NAVY)
        c = ws.cell(row=row_idx, column=2, value=formula_or_val)
        if isinstance(formula_or_val, str) and formula_or_val.startswith("="):
            c.fill = _fill(_BG_FORMULA)
            c.font = _font(color=_FG_GREEN, bold=is_result, size=10)
        else:
            c.fill = _fill(_BG_INPUT)
            c.font = _font(color=_FG_NAVY, bold=is_result, size=10)
        if row_idx == _DCF_ROWS["iv"]:
            c.number_format = _FMT_PRICE
        elif row_idx == _DCF_ROWS["upside"]:
            c.number_format = _FMT_PCT
        else:
            c.number_format = _FMT_MONEY
        c.alignment = _align("right")
        ws.row_dimensions[row_idx].height = 18
        if is_result:
            ws.cell(row=row_idx, column=1).fill = _fill(_BG_SUBHDR)
            ws.cell(row=row_idx, column=1).font = _font(bold=True, color=_FG_WHITE, size=9)

    # Key stat rows
    r3 = _DCF_ROWS["wacc_used"]
    _section_row(ws, r3 - 1, "  KEY PARAMETERS USED", _FCY_N_YEARS + 1)
    for rr, lbl, fml in [
        (r3,    "WACC used",                f"={_aref('wacc')}/100"),
        (r3+1,  "Terminal Growth Rate (g)", f"={_aref('terminal_growth')}/100"),
        (r3+2,  "TV % of Enterprise Value", f"=IF(B{_DCF_ROWS['ev']}<>0,B{_DCF_ROWS['pv_tv']}/B{_DCF_ROWS['ev']},0)"),
    ]:
        ws.cell(row=rr, column=1, value=lbl).font = _font(size=9, color=_FG_GREY)
        c = ws.cell(row=rr, column=2, value=fml)
        c.font = _font(color=_FG_GREEN, size=9)
        c.fill = _fill(_BG_FORMULA)
        c.number_format = _FMT_PCT if "%" in lbl or rr < r3+2 else _FMT_PCT
        c.alignment = _align("right")

    ws.freeze_panes = "B2"


###############################################################################
# SHEET 11: Historical  (trend analysis)
###############################################################################

def _build_historical(wb, d: dict):
    ws = wb.create_sheet("Historical")
    ws.sheet_view.showGridLines = False

    hist  = d.get("historical", {})
    years = hist.get("years", [])
    if not years:
        return

    ycols = _year_cols(years)
    _set_col_widths(ws, {1: 30, **{c: 14 for c in range(2, 2 + len(years) + 2)}})

    _header_cell(ws, 1, 1, "HISTORICAL FINANCIAL ANALYSIS", bg=_BG_HEADER, fg=_FG_WHITE, size=11, h="left")
    # Write year headers in row 1 alongside title (no merge — avoids MergedCell conflict)
    for y, col in ycols.items():
        _header_cell(ws, 1, col, str(y), bg=_BG_SUBHDR)
    for y, col in ycols.items():
        _header_cell(ws, 1, col, str(y), bg=_BG_SUBHDR)

    metrics = [
        ("INCOME STATEMENT ($M)", None, None),
        ("Revenue",              hist.get("revenue", []),        _FMT_MONEY),
        ("Gross Profit",         hist.get("gross_profit", []),   _FMT_MONEY),
        ("EBIT",                 hist.get("ebit", []),           _FMT_MONEY),
        ("Net Income",           hist.get("net_income", []),     _FMT_MONEY),
        ("MARGINS", None, None),
        ("Gross Margin",         hist.get("gross_margin", []),   _FMT_PCT_RAW),
        ("EBIT Margin",          hist.get("ebit_margin", []),    _FMT_PCT_RAW),
        ("Net Margin",           hist.get("net_margin", []),     _FMT_PCT_RAW),
        ("CASH FLOW ($M)", None, None),
        ("Operating CF",         hist.get("op_cf", []),          _FMT_MONEY),
        ("CapEx",                hist.get("capex", []),          _FMT_MONEY),
        ("Free Cash Flow",       hist.get("fcf", []),            _FMT_MONEY),
        ("D&A",                  hist.get("da", []),             _FMT_MONEY),
        ("SBC",                  hist.get("sbc", []),            _FMT_MONEY),
        ("RETURNS & LEVERAGE", None, None),
        ("ROIC (%)",             hist.get("roic", []),           _FMT_PCT_RAW),
        ("Total Debt",           hist.get("total_debt", []),     _FMT_MONEY),
    ]

    r = 2
    for item in metrics:
        label, arr, fmt = item
        if arr is None:
            _section_row(ws, r, f"  {label}", max(len(years) + 1, 6))
            r += 1
            continue
        ws.cell(row=r, column=1, value=label).font = _font(size=9, color=_FG_NAVY)
        for i, (y, col) in enumerate(ycols.items()):
            val = arr[i] if i < len(arr) else None
            if val is not None:
                c = ws.cell(row=r, column=col, value=val)
                c.font = _font(size=9, color=_FG_NAVY)
                c.number_format = fmt or _FMT_MONEY
                c.alignment = _align("right")
        r += 1

    ws.freeze_panes = "B2"


###############################################################################
# SHEET 12: Sensitivity  (value grid)
###############################################################################

def _build_sensitivity(wb, d: dict):
    ws = wb.create_sheet("Sensitivity")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {1: 14, **{c: 11 for c in range(2, 12)}})

    _header_cell(ws, 1, 1, "SENSITIVITY ANALYSIS — Intrinsic Value per Share ($)", bg=_BG_HEADER, fg=_FG_WHITE, size=11)
    ws.merge_cells("A1:K1")

    sens = d.get("sensitivity", {})
    wacc_vals = sens.get("wacc_range", [])
    g_vals    = sens.get("g_range", [])
    grid      = sens.get("grid", [])

    if not (wacc_vals and g_vals and grid):
        ws.cell(row=3, column=1, value="Sensitivity data unavailable.").font = _font(color=_FG_GREY, size=9)
        return

    # Headers
    ws.cell(row=2, column=1, value="WACC \\ Term.g").font = _font(bold=True, size=9, color=_FG_WHITE)
    ws.cell(row=2, column=1).fill = _fill(_BG_HEADER)
    ws.cell(row=2, column=1).alignment = _align("center")

    for j, g in enumerate(g_vals):
        _header_cell(ws, 2, 2 + j, f"{g:.1f}%", bg=_BG_SUBHDR)

    price = d.get("price", 0)
    for i, w in enumerate(wacc_vals):
        r = 3 + i
        ws.cell(row=r, column=1, value=f"{w:.1f}%").font = _font(bold=True, size=9, color=_FG_WHITE)
        ws.cell(row=r, column=1).fill = _fill(_BG_SUBHDR)
        ws.cell(row=r, column=1).alignment = _align("center")
        for j, g in enumerate(g_vals):
            iv = grid[i][j] if i < len(grid) and j < len(grid[i]) else 0
            c = ws.cell(row=r, column=2 + j, value=round(iv, 2))
            c.number_format = _FMT_PRICE
            c.alignment     = _align("center")
            upside = (iv - price) / price if price else 0
            if upside > 0.15:
                c.fill = _fill(_BG_GREEN); c.font = _font(color=_FG_GREEN, bold=True, size=9)
            elif upside < -0.15:
                c.fill = _fill(_BG_RED);   c.font = _font(color=_FG_RED,   bold=True, size=9)
            else:
                c.fill = _fill(_BG_AMBER); c.font = _font(color=_FG_AMBER, bold=True, size=9)

    # Legend
    r_l = 3 + len(wacc_vals) + 2
    ws.cell(row=r_l, column=1, value="Colour legend:").font = _font(bold=True, size=9, color=_FG_GREY)
    for col, (bg, fg, lbl) in enumerate([
        (_BG_GREEN, _FG_GREEN, ">+15% upside"),
        (_BG_AMBER, _FG_AMBER, "±15% (fair value)"),
        (_BG_RED,   _FG_RED,   ">15% downside"),
    ], start=2):
        c = ws.cell(row=r_l, column=col, value=lbl)
        c.fill = _fill(bg); c.font = _font(color=fg, size=9); c.alignment = _align("center")


###############################################################################
# SHEET 13: Scenarios
###############################################################################

def _build_scenarios(wb, d: dict):
    ws = wb.create_sheet("Scenarios")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {1: 3, 2: 30, 3: 18, 4: 18, 5: 18, 6: 28})

    _header_cell(ws, 1, 2, "SCENARIO ANALYSIS", bg=_BG_HEADER, fg=_FG_WHITE, size=12)
    ws.merge_cells("B1:F1")
    ws.row_dimensions[1].height = 28

    headers = ["Assumption / Metric", "Bear", "Base", "Bull", "Notes"]
    for ci, h in enumerate(headers):
        _header_cell(ws, 2, 2 + ci, h, bg=_BG_SUBHDR)

    scen = d.get("scenarios", {})
    rows = [
        ("Revenue CAGR",        "revenue_cagr",   "%"),
        ("EBIT Margin",         "ebit_margin",    "%"),
        ("WACC",                "wacc",           "%"),
        ("Terminal Growth",     "terminal_growth","%"),
        ("Intrinsic Value ($)", "intrinsic_value","$"),
        ("Upside / Downside",   "upside_pct",     "%"),
        ("Recommendation",      "recommendation",  ""),
    ]
    for ri, (label, key, unit) in enumerate(rows):
        r = 3 + ri
        ws.cell(row=r, column=2, value=label).font = _font(size=9, bold=(key == "intrinsic_value"), color=_FG_NAVY)
        for ci, sn in enumerate(["bear", "base", "bull"]):
            sv = scen.get(sn, {})
            val = sv.get(key, "-")
            c = ws.cell(row=r, column=3 + ci, value=val)
            c.alignment = _align("center")
            c.font = _font(size=9, color=_FG_NAVY)
            if key == "intrinsic_value":
                bg_map = {"bear": _BG_RED, "base": _BG_AMBER, "bull": _BG_GREEN}
                fg_map = {"bear": _FG_RED, "base": _FG_AMBER, "bull": _FG_GREEN}
                c.fill = _fill(bg_map[sn])
                c.font = _font(bold=True, color=fg_map[sn], size=9)
        ws.cell(row=r, column=6, value=unit).font = _font(size=8, color=_FG_GREY)
        ws.row_dimensions[r].height = 17


###############################################################################
# SHEET 14: Comps
###############################################################################

def _build_comps(wb, d: dict):
    ws = wb.create_sheet("Comps")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {1: 3, 2: 20, 3: 22, 4: 16, 5: 16, 6: 16, 7: 14, 8: 14, 9: 14, 10: 14, 11: 14})

    _header_cell(ws, 1, 2, "COMPARABLE COMPANY ANALYSIS", bg=_BG_HEADER, fg=_FG_WHITE, size=12)
    ws.merge_cells("B1:K1")
    ws.row_dimensions[1].height = 26

    headers = ["Ticker", "Company", "Mkt Cap ($M)", "EV ($M)", "Revenue ($M)", "EV/Rev", "EV/EBITDA", "EV/EBIT", "P/E", "P/FCF"]
    for ci, h in enumerate(headers):
        _header_cell(ws, 2, 2 + ci, h, bg=_BG_SUBHDR, size=9)

    peers = d.get("peers", [])
    if not peers:
        ws.merge_cells("B3:K3")
        ws.cell(row=3, column=2, value="No peer data available. Data is fetched live from Yahoo Finance (cached 24h). Re-export to refresh.").font = _font(color=_FG_GREY, size=9)
        return

    for ri, p in enumerate(peers):
        r = 3 + ri
        is_subj = p.get("subject", False)
        bg = _BG_SUBHDR if is_subj else (_BG_LIGHT if ri % 2 else _BG_WHITE)
        fg = _FG_WHITE if is_subj else _FG_NAVY

        def _cv(col, val, fmt=None):
            c = ws.cell(row=r, column=col, value=val)
            c.font      = _font(bold=is_subj, color=fg, size=9)
            c.fill      = _fill(bg)
            c.alignment = _align("center" if col > 3 else "left")
            if fmt:
                c.number_format = fmt
            return c

        _cv(2, p.get("ticker", ""))
        _cv(3, p.get("name", ""))
        _cv(4, p.get("market_cap", None), _FMT_MONEY)
        _cv(5, p.get("ev", None), _FMT_MONEY)
        _cv(6, p.get("revenue", None), _FMT_MONEY)
        _cv(7, p.get("ev_rev", None), _FMT_MULT2)
        _cv(8, p.get("ev_ebitda", None), _FMT_MULT)
        _cv(9, p.get("ev_ebit", None), _FMT_MULT)
        _cv(10, p.get("pe", None), _FMT_MULT)
        _cv(11, p.get("p_fcf", None), _FMT_MULT)
        ws.row_dimensions[r].height = 16

    # Median row
    r_med = 3 + len(peers) + 1
    pm = d.get("peer_median") or {}
    _header_cell(ws, r_med, 2, "PEER MEDIAN", bg=_BG_HEADER, size=9)
    ws.cell(row=r_med, column=3, value="").fill = _fill(_BG_HEADER)
    for col_idx, key, fmt in [
        (7,  "ev_rev",    _FMT_MULT2),
        (8,  "ev_ebitda", _FMT_MULT),
        (9,  "ev_ebit",   _FMT_MULT),
        (10, "pe",        _FMT_MULT),
        (11, "p_fcf",     _FMT_MULT),
    ]:
        val = pm.get(key)
        c = ws.cell(row=r_med, column=col_idx, value=val)
        c.font = _font(bold=True, color=_FG_WHITE, size=9)
        c.fill = _fill(_BG_HEADER)
        c.number_format = fmt
        c.alignment = _align("center")

    # Implied values section
    r_imp = r_med + 2
    _section_row(ws, r_imp, "  IMPLIED EQUITY VALUE FROM PEER MULTIPLES", 10)
    ticker = d.get("ticker", "")
    shares = d.get("diluted_shares", 1) or 1
    nd = d.get("net_debt", 0) or 0
    ebitda_ltm = d.get("ebitda_ltm", 0)
    rev_ltm    = d.get("revenue_base", 0)

    imp_rows = [
        ("EV/Revenue implied",   pm.get("ev_rev"),   rev_ltm),
        ("EV/EBITDA implied",    pm.get("ev_ebitda"), ebitda_ltm),
    ]
    r_imp2 = r_imp + 1
    headers2 = ["Multiple Applied", "LTM Metric ($M)", "Implied EV ($M)", "- Net Debt", "Implied Equity Val ($M)", "Implied IV / Share"]
    for ci, h in enumerate(headers2):
        _header_cell(ws, r_imp2, 2+ci, h, bg=_BG_SUBHDR, size=9)
    for i, (lbl, mult, ltm_val) in enumerate(imp_rows):
        r = r_imp2 + 1 + i
        ws.cell(row=r, column=2, value=lbl).font = _font(size=9, color=_FG_NAVY)
        ws.cell(row=r, column=3, value=mult).number_format = _FMT_MULT2
        ws.cell(row=r, column=4, value=ltm_val).number_format = _FMT_MONEY
        implied_ev = (mult * ltm_val) if (mult and ltm_val) else None
        ws.cell(row=r, column=5, value=implied_ev).number_format = _FMT_MONEY
        ws.cell(row=r, column=6, value=nd).number_format = _FMT_MONEY
        implied_iv = ((implied_ev - nd) / shares) if (implied_ev and shares) else None
        c = ws.cell(row=r, column=7, value=implied_iv)
        c.number_format = _FMT_PRICE
        c.font = _font(bold=True, color=_FG_NAVY, size=9)
        ws.row_dimensions[r].height = 16


###############################################################################
# SHEET 15: Peer_Data  (raw peer data table)
###############################################################################

def _build_peer_data(wb, d: dict):
    ws = wb.create_sheet("Peer_Data")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {1: 3, 2: 12, 3: 22, 4: 16, 5: 16, 6: 16, 7: 16, 8: 16, 9: 14, 10: 14, 11: 14})

    _header_cell(ws, 1, 2, "RAW PEER DATA (Yahoo Finance)", bg=_BG_HEADER, fg=_FG_WHITE, size=11)
    ws.merge_cells("B1:K1")

    headers = ["Ticker", "Company", "Mkt Cap ($M)", "EV ($M)", "Revenue ($M)", "EBITDA ($M)", "EBIT ($M)", "Net Income", "FCF ($M)", "EV/Rev", "EV/EBITDA"]
    for ci, h in enumerate(headers):
        _header_cell(ws, 2, 2 + ci, h, bg=_BG_SUBHDR, size=9)

    for ri, p in enumerate(d.get("peers", [])):
        r = 3 + ri
        vals = [p.get("ticker",""), p.get("name",""), p.get("market_cap"),
                p.get("ev"), p.get("revenue"), p.get("ebitda"),
                p.get("ebit"), p.get("net_income"), p.get("fcf"),
                p.get("ev_rev"), p.get("ev_ebitda")]
        fmts = ["","", _FMT_MONEY, _FMT_MONEY, _FMT_MONEY, _FMT_MONEY, _FMT_MONEY, _FMT_MONEY, _FMT_MONEY, _FMT_MULT2, _FMT_MULT]
        for ci, (val, fmt) in enumerate(zip(vals, fmts)):
            c = ws.cell(row=r, column=2+ci, value=val)
            c.font = _font(size=9, color=_FG_NAVY)
            c.alignment = _align("right" if ci > 1 else "left")
            if fmt:
                c.number_format = fmt
        ws.row_dimensions[r].height = 15


###############################################################################
# SHEET 16: Football_Field
###############################################################################

def _build_football_field(wb, d: dict):
    ws = wb.create_sheet("Football_Field")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {1: 3, 2: 30, 3: 14, 4: 14, 5: 14})

    _header_cell(ws, 1, 2, "FOOTBALL FIELD — VALUATION BRIDGE DATA", bg=_BG_HEADER, fg=_FG_WHITE, size=11)
    ws.merge_cells("B1:E1")

    headers = ["Method", "Low ($)", "Mid ($)", "High ($)"]
    for ci, h in enumerate(headers):
        _header_cell(ws, 2, 2+ci, h, bg=_BG_SUBHDR)

    price  = d.get("price", 0)
    iv     = d.get("intrinsic_value", 0)
    scen   = d.get("scenarios", {})
    bear_iv = scen.get("bear", {}).get("intrinsic_value", 0)
    bull_iv = scen.get("bull", {}).get("intrinsic_value", 0)

    pm = d.get("peer_median") or {}
    shares = d.get("diluted_shares", 1) or 1
    nd     = d.get("net_debt", 0) or 0
    rev_ltm   = d.get("revenue_base", 0) or 0
    ebitda_ltm= d.get("ebitda_ltm", 0) or 0

    def _peers_range(mult_low, mult_high, ltm):
        if not (mult_low and mult_high and ltm and shares):
            return None, None
        lo = (mult_low * ltm - nd) / shares
        hi = (mult_high * ltm - nd) / shares
        return round(lo, 2), round(hi, 2)

    ev_rev_lo, ev_rev_hi = _peers_range(
        pm.get("ev_rev_p25"), pm.get("ev_rev_p75"), rev_ltm)
    ev_ebitda_lo, ev_ebitda_hi = _peers_range(
        pm.get("ev_ebitda_p25"), pm.get("ev_ebitda_p75"), ebitda_ltm)

    ff_rows = [
        ("DCF — Scenario Range",       bear_iv, iv, bull_iv),
        ("DCF — Sensitivity Range",    d.get("sensitivity", {}).get("iv_min"), iv, d.get("sensitivity", {}).get("iv_max")),
        ("Peers EV/Revenue",           ev_rev_lo,    (((ev_rev_lo or 0)+(ev_rev_hi or 0))/2) or None, ev_rev_hi),
        ("Peers EV/EBITDA",            ev_ebitda_lo, (((ev_ebitda_lo or 0)+(ev_ebitda_hi or 0))/2) or None, ev_ebitda_hi),
        ("52-Week Range",              d.get("week52_low"), price, d.get("week52_high")),
        ("Analyst Target Range",       d.get("analyst_low"), d.get("analyst_consensus", {}).get("price_target"), d.get("analyst_high")),
        ("Current Market Price",       price, price, price),
    ]

    for ri, (method, lo, mid, hi) in enumerate(ff_rows):
        r = 3 + ri
        ws.cell(row=r, column=2, value=method).font = _font(size=9, color=_FG_NAVY)
        for ci, val in enumerate([lo, mid, hi]):
            c = ws.cell(row=r, column=3+ci, value=round(val, 2) if val else None)
            c.number_format = _FMT_PRICE
            c.font = _font(size=9, color=_FG_NAVY)
            c.alignment = _align("right")
        ws.row_dimensions[r].height = 16


###############################################################################
# SHEET 17: Data_Quality
###############################################################################

def _build_data_quality(wb, d: dict):
    ws = wb.create_sheet("Data_Quality")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {1: 3, 2: 30, 3: 20, 4: 50})

    _header_cell(ws, 1, 2, "DATA QUALITY & INTEGRITY FLAGS", bg=_BG_HEADER, fg=_FG_WHITE, size=11)
    ws.merge_cells("B1:D1")

    _header_cell(ws, 2, 2, "Flag", bg=_BG_SUBHDR, size=9)
    _header_cell(ws, 2, 3, "Status", bg=_BG_SUBHDR, size=9)
    _header_cell(ws, 2, 4, "Details", bg=_BG_SUBHDR, size=9)

    flags = d.get("flags", [])
    dq    = d.get("data_quality", {})
    r = 3
    for f in flags:
        status = f.get("status", "pass")
        bg = _BG_GREEN if status == "pass" else (_BG_RED if status == "fail" else _BG_AMBER)
        fg = _FG_GREEN if status == "pass" else (_FG_RED if status == "fail" else _FG_AMBER)
        ws.cell(row=r, column=2, value=f.get("name", "")).font = _font(size=9, color=_FG_NAVY)
        c_st = ws.cell(row=r, column=3, value=status.upper())
        c_st.fill = _fill(bg); c_st.font = _font(bold=True, color=fg, size=9); c_st.alignment = _align("center")
        c_msg = ws.cell(row=r, column=4, value=f.get("message", ""))
        c_msg.font = _font(size=9, color=_FG_GREY)
        ws.row_dimensions[r].height = 16
        r += 1

    r += 1
    _section_row(ws, r, "  DATA COVERAGE", 3)
    r += 1
    for label, val in [
        ("Annual years available", dq.get("annual_years", "?")),
        ("Quarterly periods",      dq.get("quarterly_periods", "?")),
        ("Has quarterly recon",    str(dq.get("has_quarterly_recon", False))),
        ("Reconstructed extra yrs",dq.get("reconstructed_years", 0)),
        ("Data source",            dq.get("source", "Yahoo Finance")),
        ("Export date",            date.today().isoformat()),
    ]:
        ws.cell(row=r, column=2, value=label).font = _font(size=9, color=_FG_GREY)
        ws.cell(row=r, column=3, value=val).font   = _font(size=9, color=_FG_NAVY)
        ws.row_dimensions[r].height = 15
        r += 1


###############################################################################
# SHEET 18: Source_Log
###############################################################################

def _build_source_log(wb, d: dict):
    ws = wb.create_sheet("Source_Log")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {1: 3, 2: 30, 3: 50, 4: 20})

    _header_cell(ws, 1, 2, "DATA SOURCE LOG", bg=_BG_HEADER, fg=_FG_WHITE, size=11)
    ws.merge_cells("B1:D1")

    _header_cell(ws, 2, 2, "Data Item",   bg=_BG_SUBHDR, size=9)
    _header_cell(ws, 2, 3, "Source",      bg=_BG_SUBHDR, size=9)
    _header_cell(ws, 2, 4, "Retrieved",   bg=_BG_SUBHDR, size=9)

    today = date.today().isoformat()
    rows = [
        ("Income Statement (annual)",   "Yahoo Finance — yfinance t.financials",     today),
        ("Balance Sheet (annual)",      "Yahoo Finance — yfinance t.balance_sheet",  today),
        ("Cash Flow (annual)",          "Yahoo Finance — yfinance t.cashflow",        today),
        ("Quarterly data",              "Yahoo Finance — yfinance t.quarterly_*",    today),
        ("Company Info / Price",        "Yahoo Finance — yfinance t.info / t.fast_info", today),
        ("Risk-Free Rate",              "FRED DGS10 (10-yr Treasury)",               today),
        ("Beta",                        "Yahoo Finance (5-yr monthly vs SPY)",       today),
        ("Peer Metrics",                "Yahoo Finance (peer_lists.py cache 24h)",   today),
        ("Analyst Consensus",           "Yahoo Finance — t.info analyst fields",     today),
    ]
    for ri, (item, src, retrieved) in enumerate(rows):
        r = 3 + ri
        ws.cell(row=r, column=2, value=item).font      = _font(size=9, color=_FG_NAVY)
        ws.cell(row=r, column=3, value=src).font       = _font(size=9, color=_FG_GREY)
        ws.cell(row=r, column=4, value=retrieved).font = _font(size=9, color=_FG_GREY)
        ws.row_dimensions[r].height = 15


###############################################################################
# SHEET 19: Rough_Work  (scratch calculations, assumption back-testing)
###############################################################################

def _build_rough_work(wb, d: dict):
    ws = wb.create_sheet("Rough_Work")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {1: 3, 2: 32, 3: 18, 4: 18, 5: 18, 6: 18, 7: 18})

    _header_cell(ws, 1, 2, "ROUGH WORK — ASSUMPTION BACK-TESTING & CHECKS", bg=_BG_HEADER, fg=_FG_WHITE, size=11)
    ws.merge_cells("B1:G1")

    hist  = d.get("historical", {})
    years = hist.get("years", [])
    revs  = hist.get("revenue", [])

    r = 3
    _section_row(ws, r, "  REVENUE GROWTH HISTORY", 6)
    r += 1
    ws.cell(row=r, column=2, value="Year").font = _font(bold=True, size=9, color=_FG_WHITE)
    ws.cell(row=r, column=2).fill = _fill(_BG_SUBHDR)
    ws.cell(row=r, column=3, value="Revenue ($M)").font = _font(bold=True, size=9, color=_FG_WHITE)
    ws.cell(row=r, column=3).fill = _fill(_BG_SUBHDR)
    ws.cell(row=r, column=4, value="YoY Growth").font = _font(bold=True, size=9, color=_FG_WHITE)
    ws.cell(row=r, column=4).fill = _fill(_BG_SUBHDR)
    r += 1
    for i, (y, rev) in enumerate(zip(years, revs)):
        ws.cell(row=r, column=2, value=y).font = _font(size=9, color=_FG_NAVY)
        ws.cell(row=r, column=3, value=rev).number_format = _FMT_MONEY
        ws.cell(row=r, column=3).font = _font(size=9, color=_FG_NAVY)
        if i > 0 and revs[i-1] and revs[i-1] != 0:
            growth = (rev - revs[i-1]) / revs[i-1]
            ws.cell(row=r, column=4, value=growth).number_format = _FMT_PCT
            ws.cell(row=r, column=4).font = _font(size=9, color=_FG_GREEN if growth > 0 else _FG_RED)
        r += 1

    r += 1
    _section_row(ws, r, "  WACC CROSS-CHECK", 6)
    r += 1
    wacc = d.get("wacc", 9.0)
    beta = d.get("beta", 1.0)
    rf   = d.get("risk_free_rate", 4.2)
    erp  = d.get("equity_risk_premium", 5.5)
    ke   = d.get("cost_of_equity", 10.0)
    kd   = d.get("cost_of_debt_pretax", 5.0)
    ew   = d.get("equity_weight_pct", 80.0)
    dw   = d.get("debt_weight_pct", 20.0)
    tr   = d.get("tax_rate", 21.0)

    checks = [
        ("Cost of Equity (Ke = Rf + β×ERP)",      f"{rf:.1f}% + {beta:.2f}×{erp:.1f}% = {ke:.2f}%"),
        ("Cost of Debt (after-tax)",                f"{kd:.1f}% × (1 − {tr:.0f}%) = {kd*(1-tr/100):.2f}%"),
        ("WACC = We×Ke + Wd×Kd(1-t)",              f"{ew:.0f}%×{ke:.2f}% + {dw:.0f}%×{kd*(1-tr/100):.2f}% ≈ {wacc:.2f}%"),
        ("WACC − g spread",                         f"{wacc:.1f}% − {d.get('terminal_growth', 2.5):.1f}% = {wacc - d.get('terminal_growth', 2.5):.1f}pp"),
    ]
    for label, val in checks:
        ws.cell(row=r, column=2, value=label).font = _font(size=9, color=_FG_NAVY)
        ws.cell(row=r, column=3, value=val).font   = _font(size=9, color=_FG_GREY)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        ws.row_dimensions[r].height = 16
        r += 1

    r += 1
    _section_row(ws, r, "  CONFIDENCE SCORING BREAKDOWN", 6)
    r += 1
    conf_dims = d.get("confidence_dimensions", [])
    if not conf_dims:
        conf_score = d.get("confidence_score", 0)
        ws.cell(row=r, column=2, value=f"Confidence Score: {conf_score}/100").font = _font(size=9, color=_FG_NAVY)
    else:
        for dim in conf_dims:
            ws.cell(row=r, column=2, value=dim.get("name", "")).font = _font(size=9, color=_FG_NAVY)
            ws.cell(row=r, column=3, value=f"{dim.get('score',0)}/{dim.get('max_score',0)}").font = _font(size=9, color=_FG_GREY)
            ws.cell(row=r, column=4, value=dim.get("note", "")).font  = _font(size=8, color=_FG_GREY, italic=True)
            r += 1


###############################################################################
# MASTER ENTRY POINT
###############################################################################

def build_excel_bytes(data: dict) -> bytes:
    """Build a professional formula-linked Excel workbook and return raw bytes."""
    if not _HAS_OPENPYXL:
        raise ImportError("openpyxl is required for Excel export.")

    wb = Workbook()
    # Remove the default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Make historical arrays easily accessible
    hist = data.get("historical", {})
    years = hist.get("years", [])

    # Build all sheets in order
    _build_readme(wb, data)
    _build_cover(wb, data)
    _build_assumptions(wb, data)
    _build_raw_is(wb, data)
    _build_raw_bs(wb, data)
    _build_raw_cf(wb, data)
    _build_wacc_calc(wb, data)
    _build_is_forecast(wb, data)
    _build_cf_forecast(wb, data)
    _build_dcf_calc(wb, data)
    _build_historical(wb, data)
    _build_sensitivity(wb, data)
    _build_scenarios(wb, data)
    _build_comps(wb, data)
    _build_peer_data(wb, data)
    _build_football_field(wb, data)
    _build_data_quality(wb, data)
    _build_source_log(wb, data)
    _build_rough_work(wb, data)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
