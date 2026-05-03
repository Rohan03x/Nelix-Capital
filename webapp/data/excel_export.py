"""
webapp/data/excel_export.py  —  Nike-style single-sheet DCF workbook
Generates a single-sheet Excel workbook mirroring Nike Valuation Strategy.xlsx.

  Sheet: "valuation"
  Col A   : row labels
  Cols B-K: historical data  (10 years)
  Cols L-R: forecast data    (7 years)

  Row layout:
    1-7    header
    8-39   DCF summary
    42-56  comparable companies
    60-74  sensitivity table
   190-295 Working Calculations
   300-412 Income Statement
   415-478 Cash Flow Statement
   480-590 Balance Sheet

Public entry-point:
  build_excel_bytes(data: dict) -> bytes
"""
from __future__ import annotations
import io
import logging
from datetime import date, datetime
from typing import List, Optional

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False
    logger.warning("openpyxl not installed -- Excel export unavailable.")

_LABEL_COL = 1
_N_HIST    = 10
_N_FCST    = 7
_HIST_COLS = list(range(2, 2 + _N_HIST))
_FCST_COLS = list(range(2 + _N_HIST, 2 + _N_HIST + _N_FCST))

# Last calendar day for each month (non-leap-year approx; fine for fiscal year labeling)
_MONTH_END_DAY = {1:31,2:28,3:31,4:30,5:31,6:30,7:31,8:31,9:30,10:31,11:30,12:31}
_MONTH_NAMES   = {
    "January":1,"February":2,"March":3,"April":4,"May":5,"June":6,
    "July":7,"August":8,"September":9,"October":10,"November":11,"December":12,
}

_NAVY       = "1F4E79"
_BOLD_BLUE  = "003366"
_HIST_BG    = "D6E4F0"
_FCST_BG    = "E2EFDA"
_LIGHT_GREY = "F2F2F2"
_WHITE      = "FFFFFF"


def _v(lst, idx, default=None):
    if lst is None:
        return default
    try:
        val = lst[idx]
        return default if val is None else val
    except (IndexError, TypeError):
        return default


def _s(val, default=0):
    return val if val is not None else default


def build_excel_bytes(data: dict) -> bytes:
    """Build and return the complete Excel workbook as raw bytes."""
    if not _HAS_OPENPYXL:
        raise RuntimeError("openpyxl is not installed -- cannot export Excel.")
    wb = Workbook()
    ws = wb.active
    ws.title = "valuation"
    _build_sheet(ws, data)
    _apply_formatting(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _build_sheet(ws, data: dict) -> None:
    h      = data.get("historical", {}) or {}
    fc     = data.get("forecast",   []) or []
    pm     = data.get("peer_median") or {}
    peers  = data.get("peers", []) or []
    sens   = data.get("sensitivity") or {}

    rev_h    = h.get("revenue",       [])
    gp_h     = h.get("gross_profit",  [])
    gm_h_pct = h.get("gross_margin",  [])
    ebit_h   = h.get("ebit",          [])
    ni_h     = h.get("net_income",    [])
    da_h     = h.get("da",            [])
    capex_h  = h.get("capex",         [])
    op_cf_h  = h.get("op_cf",         [])
    sbc_h    = h.get("sbc",           [])
    shares_h = h.get("shares",        [])
    cash_h   = h.get("cash",          [])
    debt_h   = h.get("debt",          [])
    equity_h = h.get("equity",        [])
    ta_h     = h.get("total_assets",  [])
    pretax_h = h.get("pretax_income", [])
    tax_h    = h.get("tax",           [])
    fcf_h    = h.get("fcf",           [])
    years_h  = h.get("years",         [])

    if not gp_h and gm_h_pct:
        gp_h = [
            (_s(_v(rev_h, i)) * _s(_v(gm_h_pct, i)) / 100) if _v(rev_h, i) is not None else None
            for i in range(_N_HIST)
        ]

    cogs_h: List = []
    for i in range(_N_HIST):
        r, g = _v(rev_h, i), _v(gp_h, i)
        cogs_h.append((r - g) if (r is not None and g is not None) else None)

    sga_h: List = []
    for i in range(_N_HIST):
        g, e, d = _v(gp_h, i), _v(ebit_h, i), _v(da_h, i)
        sga_h.append((_s(g) - _s(e) - _s(d)) if g is not None else None)

    # ── Extended historical data (from EODHD extended fields) ────────────
    int_exp_direct_h = h.get("interest_expense", [])
    int_inc_direct_h = h.get("interest_income", [])
    ar_direct_h      = h.get("accounts_receivable", [])
    inv_direct_h     = h.get("inventory_bs", [])
    ap_direct_h      = h.get("accounts_payable", [])
    ca_direct_h      = h.get("total_current_assets", [])
    cl_direct_h      = h.get("total_current_liabilities", [])
    ppe_direct_h     = h.get("net_ppe", [])
    gross_ppe_raw_h  = h.get("gross_ppe", [])
    accum_dep_raw_h  = h.get("accum_dep", [])
    goodwill_raw_h   = h.get("goodwill", [])
    intang_raw_h     = h.get("intangibles", [])
    re_direct_h      = h.get("retained_earnings", [])
    div_direct_h     = h.get("dividends_paid", [])
    buyback_direct_h = h.get("buybacks", [])
    stock_iss_h      = h.get("stock_issued", [])
    net_borr_h       = h.get("net_borrowings", [])
    dates_h          = h.get("dates", [])
    fy_end_month_str = data.get("fiscal_year_end_month", "December")
    _fy_month        = _MONTH_NAMES.get(fy_end_month_str, 12)

    # Availability flags for new data arrays
    _has_ar   = any(_v(ar_direct_h, i) for i in range(_N_HIST))
    _has_inv  = any(_v(inv_direct_h, i) for i in range(_N_HIST))
    _has_ap   = any(_v(ap_direct_h, i) for i in range(_N_HIST))
    _has_ca   = any(_v(ca_direct_h, i) for i in range(_N_HIST))
    _has_cl   = any(_v(cl_direct_h, i) for i in range(_N_HIST))
    _has_ppe  = any(_v(ppe_direct_h, i) for i in range(_N_HIST))
    _has_div  = any(_v(div_direct_h, i) for i in range(_N_HIST))
    _has_buy  = any(_v(buyback_direct_h, i) for i in range(_N_HIST))
    _has_iss  = any(_v(stock_iss_h, i) for i in range(_N_HIST))
    _has_borr = any(_v(net_borr_h, i) for i in range(_N_HIST))

    # Interest expense: use actual if available, else derive as |EBIT - EBT| (positive magnitude)
    int_exp_h: List = []
    for i in range(_N_HIST):
        actual = _v(int_exp_direct_h, i)
        if actual is not None and actual != 0:
            int_exp_h.append(abs(_s(actual)))
        else:
            pt, e = _v(pretax_h, i), _v(ebit_h, i)
            int_exp_h.append(abs(_s(e) - _s(pt)) if pt is not None else None)

    ticker     = data.get("ticker", "")
    company    = data.get("company_name", ticker)
    display_currency = str(data.get("display_currency") or data.get("currency") or "USD")
    display_currency_symbol = str(data.get("display_currency_symbol") or "$")
    model_currency = str(data.get("model_currency") or data.get("reporting_currency") or data.get("currency") or display_currency)
    exchange = str(data.get("exchange") or "")
    sector = str(data.get("sector") or "")
    industry = str(data.get("industry") or "")
    suitability_note = str(((data.get("confidence_breakdown") or {}).get("suitability_note") or "")).strip()
    price      = _s(data.get("current_price") or data.get("price", 0))
    iv         = _s(data.get("intrinsic_value", 0))
    wacc       = _s(data.get("wacc", 0))
    terminal_g = _s(data.get("terminal_growth", 0))
    tax_rate   = _s(data.get("tax_rate", 0))
    cod        = _s(data.get("cost_of_debt_pre") or data.get("cost_of_debt", 0))
    net_debt   = _s(data.get("net_debt", 0))
    shares_out = _s(data.get("diluted_shares") or data.get("shares_outstanding", 0))
    ev         = _s(data.get("enterprise_value", 0))
    pv_ufcfs   = _s(data.get("pv_ufcfs", 0))
    pv_tv      = _s(data.get("pv_terminal", 0))
    upside_pct = _s(data.get("upside_pct", 0))
    ebitda_ltm = _s(data.get("ebitda_ltm", 0))
    dso        = _s(data.get("dso", 30))
    dio        = _s(data.get("dio", 30))
    dpo        = _s(data.get("dpo", 60))
    tv_pct     = _s(data.get("tv_pct", 0))

    last_rev   = _s(_v(rev_h,  -1), 1)
    last_gm    = _s(_v(gp_h,   -1)) / last_rev if last_rev else 0
    last_ebit_margin = _s(_v(ebit_h, -1)) / last_rev if last_rev else 0
    last_sga_p = _s(_v(sga_h,  -1)) / last_rev if last_rev else 0
    last_ta    = _s(_v(ta_h,   -1))
    last_debt  = _s(_v(debt_h, -1))
    dt_ratio   = last_debt / last_ta if last_ta else 0
    buyback_yield = _s(data.get("buyback_yield", 0))
    dividend_yield = _s(data.get("dividend_yield", 0))

    share_shrink_rates = []
    for i in range(1, _N_HIST):
        prev_shares = _v(shares_h, i - 1)
        curr_shares = _v(shares_h, i)
        if prev_shares and curr_shares and prev_shares > 0:
            share_shrink_rates.append(max(0.0, min(0.05, (prev_shares - curr_shares) / prev_shares)))
    share_shrink_rate = 0.0
    if share_shrink_rates:
        trailing_rates = share_shrink_rates[-3:]
        share_shrink_rate = sum(trailing_rates) / len(trailing_rates)
    elif buyback_yield > 0:
        share_shrink_rate = min(0.05, buyback_yield / 100)

    def _fv(key):
        return [f.get(key) for f in fc]

    fcst_rev    = _fv("revenue")
    fcst_ebit   = _fv("ebit")
    fcst_nopat  = _fv("nopat")
    fcst_da     = _fv("da")
    fcst_sbc    = _fv("sbc")
    fcst_capex  = _fv("capex")
    fcst_dnowc  = _fv("d_nowc")
    fcst_ufcf   = _fv("ufcf")
    fcst_df     = _fv("df")
    fcst_pv     = _fv("pv")
    fcst_ebit_m = _fv("ebit_m")
    fcst_years  = _fv("year")

    fcst_shares = []
    _share_base = _s(_v(shares_h, -1), shares_out)
    for _ in range(_N_FCST):
        _share_base *= (1 - share_shrink_rate)
        fcst_shares.append(round(_share_base, 1))

    tax_dec   = tax_rate / 100
    fcst_gm_p = []
    for j in range(_N_FCST):
        ebit_margin_j = _s(_v(fcst_ebit_m, j)) / 100
        gm_lift = max(0.0, ebit_margin_j - last_ebit_margin) * 0.5
        fcst_gm_p.append(min(0.95, max(0.0, last_gm + gm_lift)))
    fcst_gp   = [_s(_v(fcst_rev, j)) * _s(_v(fcst_gm_p, j)) for j in range(_N_FCST)]
    fcst_cogs = [_s(_v(fcst_rev, j)) - _s(_v(fcst_gp, j)) for j in range(_N_FCST)]
    # Derive forecast SGA so that IS Operating Income = DCF EBIT (consistent model):
    #   SGA = GP - EBIT - D&A  (positive value; floor at 0 to avoid negative expense)
    fcst_sga  = [max(0.0, _s(_v(fcst_gp, j)) - _s(_v(fcst_ebit, j)) - _s(_v(fcst_da, j)))
                 for j in range(_N_FCST)]
    fcst_opex = [_s(_v(fcst_sga, j)) + _s(_v(fcst_da, j)) for j in range(_N_FCST)]
    fcst_opinc= [_s(_v(fcst_gp, j)) - _s(_v(fcst_opex, j)) for j in range(_N_FCST)]
    fcst_interest_exp = [max(0.0, last_debt * cod / 100)] * _N_FCST
    fcst_ebt  = [_s(_v(fcst_ebit, j)) - _s(_v(fcst_interest_exp, j)) for j in range(_N_FCST)]
    fcst_tax  = [max(_s(_v(fcst_ebt, j)), 0) * tax_dec for j in range(_N_FCST)]
    fcst_ni   = [_s(_v(fcst_ebt, j)) - _s(_v(fcst_tax, j)) for j in range(_N_FCST)]
    fcst_ebitda = [_s(_v(fcst_ebit, j)) + _s(_v(fcst_da, j)) for j in range(_N_FCST)]

    fcst_ta_bs: List[float] = [
        _s(_v(fcst_rev, j)) / last_rev * last_ta for j in range(_N_FCST)
    ]
    fcst_recv  = [_s(_v(fcst_rev, j)) / 365 * dso for j in range(_N_FCST)]
    fcst_inv   = [abs(_s(_v(fcst_cogs, j))) / 365 * dio for j in range(_N_FCST)]
    fcst_ap   = [abs(_s(_v(fcst_cogs, j))) / 365 * dpo for j in range(_N_FCST)]
    fcst_ltd  = [last_debt] * _N_FCST

    fcst_chg_ar: List[float] = []
    fcst_chg_inv: List[float] = []
    fcst_chg_ap: List[float] = []
    fcst_chg_other_oa: List[float] = []
    prev_recv_bs = _s(_v(ar_direct_h, -1)) if _has_ar else _s(_v(rev_h, -1)) / 365 * dso
    prev_inv_bs = _s(_v(inv_direct_h, -1)) if _has_inv else (_s(_v(cogs_h, -1)) / 365 * dio if _v(cogs_h, -1) is not None else 0)
    prev_ap_bs = _s(_v(ap_direct_h, -1)) if _has_ap else (_s(_v(cogs_h, -1)) / 365 * dpo if _v(cogs_h, -1) is not None else 0)
    for j in range(_N_FCST):
        recv_j = _s(_v(fcst_recv, j))
        inv_j = _s(_v(fcst_inv, j))
        ap_j = _s(_v(fcst_ap, j))
        fcst_chg_ar.append(-(recv_j - prev_recv_bs))
        fcst_chg_inv.append(-(inv_j - prev_inv_bs))
        fcst_chg_ap.append(ap_j - prev_ap_bs)
        fcst_chg_other_oa.append(
            -_s(_v(fcst_dnowc, j)) - fcst_chg_ar[-1] - fcst_chg_inv[-1] - fcst_chg_ap[-1]
        )
        prev_recv_bs, prev_inv_bs, prev_ap_bs = recv_j, inv_j, ap_j

    fcst_op_cf = [
        _s(_v(fcst_ni, j)) + _s(_v(fcst_da, j)) + _s(_v(fcst_sbc, j))
        + _s(_v(fcst_chg_ar, j)) + _s(_v(fcst_chg_inv, j)) + _s(_v(fcst_chg_ap, j))
        + _s(_v(fcst_chg_other_oa, j))
        for j in range(_N_FCST)
    ]
    fcst_cap_cf = [-_s(_v(fcst_capex, j)) for j in range(_N_FCST)]
    implied_div_cf = -(shares_out * price * dividend_yield / 100) if dividend_yield > 0 and price > 0 and shares_out > 0 else 0.0
    implied_buy_cf = -(shares_out * price * buyback_yield / 100) if buyback_yield > 0 and price > 0 and shares_out > 0 else 0.0
    fcst_div_cf = [(-_s(_v(div_direct_h, -1)) if _has_div else implied_div_cf)] * _N_FCST
    fcst_buy_cf = [(-_s(_v(buyback_direct_h, -1)) if _has_buy else implied_buy_cf)] * _N_FCST
    fcst_iss_cf = [(_s(_v(stock_iss_h, -1)) if _has_iss else 0.0)] * _N_FCST
    fcst_fin_cf = [
        _s(_v(fcst_iss_cf, j)) + _s(_v(fcst_buy_cf, j)) + _s(_v(fcst_div_cf, j))
        for j in range(_N_FCST)
    ]
    fcst_net_c = [
        _s(_v(fcst_op_cf, j)) + _s(_v(fcst_cap_cf, j)) + _s(_v(fcst_fin_cf, j))
        for j in range(_N_FCST)
    ]

    fcst_cash = []
    _prior = _s(_v(cash_h, -1))
    for j in range(_N_FCST):
        _prior = max(0.0, _prior + _s(_v(fcst_net_c, j)))
        fcst_cash.append(_prior)
    fcst_ca = [_s(_v(fcst_cash, j)) + _s(_v(fcst_recv, j)) + _s(_v(fcst_inv, j))
               for j in range(_N_FCST)]
    fcst_tl = [_s(_v(fcst_ap, j)) + _s(_v(fcst_ltd, j)) for j in range(_N_FCST)]
    fcst_eq = [_s(_v(fcst_ta_bs, j)) - _s(_v(fcst_tl, j)) for j in range(_N_FCST)]

    fcst_re = []
    _retained = _s(_v(re_direct_h, -1))
    for j in range(_N_FCST):
        _retained += _s(_v(fcst_ni, j)) + min(0.0, _s(_v(fcst_div_cf, j)))
        fcst_re.append(_retained)

    # Use actual EODHD BS data if available, fall back to turnover-ratio derived
    hist_recv_bs = ([_s(_v(ar_direct_h, i)) for i in range(_N_HIST)] if _has_ar
                    else [_s(_v(rev_h, i)) / 365 * dso for i in range(_N_HIST)])
    hist_inv_bs  = ([_s(_v(inv_direct_h, i)) for i in range(_N_HIST)] if _has_inv
                    else [_s(_v(cogs_h, i)) / 365 * dio
                          if _v(cogs_h, i) is not None else 0 for i in range(_N_HIST)])
    hist_ap_bs   = ([_s(_v(ap_direct_h, i)) for i in range(_N_HIST)] if _has_ap
                    else [_s(_v(cogs_h, i)) / 365 * dpo
                          if _v(cogs_h, i) is not None else 0 for i in range(_N_HIST)])
    # PPE: use actual net_ppe if available, else derive as residual (ta - cash - AR)
    hist_ppe_bs  = ([_s(_v(ppe_direct_h, i)) for i in range(_N_HIST)] if _has_ppe
                    else [max(0, _s(_v(ta_h, i)) - _s(_v(cash_h, i)) - hist_recv_bs[i])
                          for i in range(_N_HIST)])
    # Total current assets: use actual if available, else cash+AR+Inv
    hist_ca_bs   = ([_s(_v(ca_direct_h, i)) for i in range(_N_HIST)] if _has_ca
                    else [_s(_v(cash_h, i)) + hist_recv_bs[i] + hist_inv_bs[i]
                          for i in range(_N_HIST)])
    hist_tl_bs   = [_s(_v(ta_h, i)) - _s(_v(equity_h, i)) for i in range(_N_HIST)]
    _last_ppe_bs = hist_ppe_bs[-1] if hist_ppe_bs else 0
    _last_gw_bs = _s(_v(goodwill_raw_h, -1))
    _last_ia_bs = _s(_v(intang_raw_h, -1))
    fcst_ta_bs = [
        max(
            _s(_v(fcst_ta_bs, j)),
            _s(_v(fcst_cash, j))
            + _s(_v(fcst_recv, j))
            + _s(_v(fcst_inv, j))
            + _last_ppe_bs
            + _last_gw_bs
            + _last_ia_bs,
        )
        for j in range(_N_FCST)
    ]

    yr_labels = []
    for yr in years_h:
        try:
            yr_labels.append(f"{int(yr)} FY")
        except (ValueError, TypeError):
            yr_labels.append(str(yr) if yr else "")

    def _wr(row, label, hist_vals, fcst_vals=None, fmt=None, bold=False):
        ws.cell(row, _LABEL_COL).value = label
        if bold:
            ws.cell(row, _LABEL_COL).font = Font(bold=True, name="Calibri", size=10)
        for i in range(_N_HIST):
            v = _v(hist_vals, i) if hist_vals else None
            if v is not None:
                c = ws.cell(row, _HIST_COLS[i])
                c.value = v
                if fmt:
                    c.number_format = fmt
        if fcst_vals:
            for j in range(_N_FCST):
                v = _v(fcst_vals, j)
                if v is not None:
                    c = ws.cell(row, _FCST_COLS[j])
                    c.value = v
                    if fmt:
                        c.number_format = fmt

    def _wrf(row, label, formula, bold=False, fmt=None):
        """Write an Excel formula row — same formula pattern for every column.
        '@' in the formula string is replaced with the column letter for each column."""
        ws.cell(row, _LABEL_COL).value = label
        if bold:
            ws.cell(row, _LABEL_COL).font = Font(bold=True, name="Calibri", size=10)
        for col in _HIST_COLS + _FCST_COLS:
            cell = ws.cell(row, col)
            cell.value = formula.replace("@", get_column_letter(col))
            if fmt:
                cell.number_format = fmt

    # =========================================================
    # HEADER (rows 1-5)
    # =========================================================
    ws.cell(1, _LABEL_COL).value = "Date"
    for i, yr in enumerate(years_h):
        date_obj = _v(dates_h, i)
        if date_obj is not None:
            ws.cell(1, _HIST_COLS[i]).value = date_obj
            ws.cell(1, _HIST_COLS[i]).number_format = "YYYY"
        else:
            try:
                _m, _d = _fy_month, _MONTH_END_DAY.get(_fy_month, 31)
                ws.cell(1, _HIST_COLS[i]).value = datetime(int(yr), _m, _d)
                ws.cell(1, _HIST_COLS[i]).number_format = "YYYY"
            except (ValueError, TypeError):
                ws.cell(1, _HIST_COLS[i]).value = yr

    for j, yr_str in enumerate(fcst_years):
        try:
            yr_num = int(str(yr_str).replace("FY", ""))
            _m, _d = _fy_month, _MONTH_END_DAY.get(_fy_month, 31)
            ws.cell(1, _FCST_COLS[j]).value = datetime(yr_num, _m, _d)
            ws.cell(1, _FCST_COLS[j]).number_format = "YYYY"
        except (ValueError, TypeError, AttributeError):
            ws.cell(1, _FCST_COLS[j]).value = yr_str

    ws.cell(2, _LABEL_COL).value = "Time"
    for i in range(_N_HIST):
        ws.cell(2, _HIST_COLS[i]).value = -((_N_HIST - 1) - i)
    for j in range(_N_FCST):
        ws.cell(2, _FCST_COLS[j]).value = j + 1

    ws.cell(3, _LABEL_COL).value = "Data"
    for i in range(_N_HIST):
        ws.cell(3, _HIST_COLS[i]).value = "Historical"
    for j in range(_N_FCST):
        ws.cell(3, _FCST_COLS[j]).value = "Forecast"

    ws.cell(4, _LABEL_COL).value = "Company"
    ws.cell(4, _HIST_COLS[-1]).value = f"{ticker}  --  {company}"

    ws.cell(5, _LABEL_COL).value = "Units"
    ws.cell(5, _HIST_COLS[0]).value = f"All financials in {display_currency} M unless noted; model currency {model_currency}"
    ws.cell(6, _LABEL_COL).value = "Profile"
    ws.cell(6, _HIST_COLS[0]).value = " | ".join(part for part in (exchange, sector, industry) if part)
    ws.cell(7, _LABEL_COL).value = "DCF Suitability"
    ws.cell(7, _HIST_COLS[0]).value = suitability_note or "Use DCF alongside scenario analysis when business quality or cyclicality is uncertain."

    # =========================================================
    # DCF SUMMARY (rows 8-39)
    # =========================================================
    R = _FCST_COLS[-1]
    K = _HIST_COLS[-1]

    ws.cell(8,  _LABEL_COL).value = "-- DCF Summary --"
    ws.cell(9,  _LABEL_COL).value = "Company"
    ws.cell(9,  K).value = f"{ticker} -- {company}"
    ws.cell(10, _LABEL_COL).value = "Valuation Date"
    ws.cell(10, K).value = date.today()
    ws.cell(11, _LABEL_COL).value = f"Current Market Price ({display_currency}/share)"
    ws.cell(11, K).value = price
    ws.cell(11, K).number_format = f'"{display_currency_symbol}"#,##0.00'

    ws.cell(14, _LABEL_COL).value = "Forecast Year"
    for j, f in enumerate(fc):
        ws.cell(14, _FCST_COLS[j]).value = f.get("n", j + 1)

    ws.cell(15, _LABEL_COL).value = "Terminal Growth Rate (g)"
    ws.cell(15, R).value = terminal_g / 100
    ws.cell(15, R).number_format = "0.0%"

    ws.cell(16, _LABEL_COL).value = "Revenue ($M)"
    for i in range(_N_HIST):
        ws.cell(16, _HIST_COLS[i]).value = _v(rev_h, i)
    for j in range(_N_FCST):
        ws.cell(16, _FCST_COLS[j]).value = _v(fcst_rev, j)

    ws.cell(17, _LABEL_COL).value = "EBIT ($M)"
    for i in range(_N_HIST):
        ws.cell(17, _HIST_COLS[i]).value = _v(ebit_h, i)
    for j in range(_N_FCST):
        ws.cell(17, _FCST_COLS[j]).value = _v(fcst_ebit, j)

    ws.cell(18, _LABEL_COL).value = "NOPAT ($M)"
    for j in range(_N_FCST):
        ws.cell(18, _FCST_COLS[j]).value = _v(fcst_nopat, j)

    ws.cell(19, _LABEL_COL).value = "  (-) CapEx ($M)"
    for j in range(_N_FCST):
        ws.cell(19, _FCST_COLS[j]).value = -_s(_v(fcst_capex, j))

    ws.cell(20, _LABEL_COL).value = "  (+) D&A ($M)"
    for j in range(_N_FCST):
        ws.cell(20, _FCST_COLS[j]).value = _v(fcst_da, j)

    ws.cell(21, _LABEL_COL).value = "  (+/-) Delta NWC ($M)"
    for j in range(_N_FCST):
        ws.cell(21, _FCST_COLS[j]).value = _v(fcst_dnowc, j)

    ws.cell(22, _LABEL_COL).value = "Free Cash Flow (Levered Hist.) / UFCF (Fcst) ($M)"
    for i in range(_N_HIST):
        ws.cell(22, _HIST_COLS[i]).value = _v(fcf_h, i)
    for j in range(_N_FCST):
        ws.cell(22, _FCST_COLS[j]).value = _v(fcst_ufcf, j)

    ws.cell(23, _LABEL_COL).value = "Discount Factor"
    for j in range(_N_FCST):
        ws.cell(23, _FCST_COLS[j]).value = _v(fcst_df, j)
        ws.cell(23, _FCST_COLS[j]).number_format = "0.0000"

    ws.cell(24, _LABEL_COL).value = "PV of UFCF ($M)"
    for j in range(_N_FCST):
        c = get_column_letter(_FCST_COLS[j])
        ws.cell(24, _FCST_COLS[j]).value = f"=ROUND({c}22*{c}23,0)"

    ws.cell(25, _LABEL_COL).value = "Tax Rate"
    ws.cell(25, R).value = tax_rate / 100
    ws.cell(25, R).number_format = "0.0%"

    ws.cell(26, _LABEL_COL).value = "Cost of Debt (Pre-Tax)"
    ws.cell(26, R).value = cod / 100
    ws.cell(26, R).number_format = "0.0%"

    ws.cell(27, _LABEL_COL).value = "LT Debt / Total Assets"
    ws.cell(27, R).value = dt_ratio
    ws.cell(27, R).number_format = "0.0%"

    ws.cell(28, _LABEL_COL).value = "WACC"
    ws.cell(28, R).value = wacc / 100
    ws.cell(28, R).number_format = "0.0%"

    # K = last hist col letter, R = last fcst col letter
    _K = get_column_letter(K)
    _R = get_column_letter(R)

    ws.cell(29, _LABEL_COL).value = "Cash & Equivalents ($M)"
    ws.cell(29, K).value = f"={_K}500"                    # links to BS Cash row
    ws.cell(30, _LABEL_COL).value = "Total Debt ($M)"
    ws.cell(30, K).value = f"={_K}528"                    # links to BS LT Debt row
    ws.cell(31, _LABEL_COL).value = "Net Debt ($M)"
    ws.cell(31, K).value = f"={_K}30-{_K}29"             # = Total Debt - Cash
    ws.cell(32, _LABEL_COL).value = "Sum PV(UFCF) ($M)"
    _L = get_column_letter(_FCST_COLS[0])                  # first forecast col
    ws.cell(32, K).value = f"=SUM({_L}24:{_R}24)"         # sum all PV UFCF years
    ws.cell(33, _LABEL_COL).value = "PV of Terminal Value ($M)"
    # Gordon Growth Model: TV = UFCF_last*(1+g)/(WACC-g), discounted to Year 7
    # Wrapped in IFERROR in case WACC ≤ g (spread ≤ 0) to prevent #DIV/0!.
    ws.cell(33, K).value = (
        f"=IFERROR(ROUND({_R}22*(1+{_R}15)/({_R}28-{_R}15)/(1+{_R}28)^7,0),0)"
    )
    ws.cell(34, _LABEL_COL).value = "Diluted Shares Outstanding (M)"
    ws.cell(34, K).value = shares_out
    ws.cell(35, _LABEL_COL).value = f"DCF Intrinsic Value ({display_currency}/share)"
    ws.cell(35, K).value = f"=IFERROR(({_K}32+{_K}33-{_K}31)/{_K}34,0)"
    ws.cell(35, K).number_format = f'"{display_currency_symbol}"#,##0.00'
    ws.cell(36, _LABEL_COL).value = f"Market Price ({display_currency}/share)"
    ws.cell(36, K).value = price
    ws.cell(36, K).number_format = f'"{display_currency_symbol}"#,##0.00'
    ws.cell(37, _LABEL_COL).value = "Upside / (Downside)"
    ws.cell(37, K).value = f"=IFERROR({_K}35/{_K}36-1,0)"
    ws.cell(37, K).number_format = "0.0%"
    ws.cell(38, _LABEL_COL).value = "Enterprise Value ($M)"
    ws.cell(38, K).value = f"={_K}32+{_K}33"
    ws.cell(39, _LABEL_COL).value = "Terminal Value as % of EV"
    ws.cell(39, K).value = f"=IFERROR({_K}33/{_K}38,0)"
    ws.cell(39, K).number_format = "0.0%"

    # =========================================================
    # COMPARABLE COMPANIES (rows 42-56)
    # =========================================================
    ws.cell(42, _LABEL_COL).value = "-- Comparable Companies --"
    comp_hdrs = ["Company", "Ticker", "Rev ($M)", "EBITDA ($M)", "EBIT ($M)",
                 "P/E", "EV/EBITDA", "EV/EBIT", "EV/Rev", "P/FCF"]
    for ci, h_txt in enumerate(comp_hdrs):
        ws.cell(43, ci + 1).value = h_txt
        ws.cell(43, ci + 1).font = Font(bold=True, name="Calibri", size=10)

    for pi, peer in enumerate(peers[:10]):
        r = 44 + pi
        ws.cell(r, 1).value  = peer.get("name", "")
        ws.cell(r, 2).value  = peer.get("ticker", "")
        ws.cell(r, 3).value  = peer.get("revenue")
        ws.cell(r, 4).value  = peer.get("ebitda")
        ws.cell(r, 5).value  = peer.get("ebit")
        ws.cell(r, 6).value  = peer.get("pe")
        ws.cell(r, 7).value  = peer.get("ev_ebitda")
        ws.cell(r, 8).value  = peer.get("ev_ebit")
        ws.cell(r, 9).value  = peer.get("ev_rev")
        ws.cell(r, 10).value = peer.get("p_fcf")

    ws.cell(54, _LABEL_COL).value = "Peer Median"
    ws.cell(54, _LABEL_COL).font  = Font(bold=True, name="Calibri", size=10)
    ws.cell(54, 6).value  = pm.get("pe")
    ws.cell(54, 7).value  = pm.get("ev_ebitda")
    ws.cell(54, 8).value  = pm.get("ev_ebit")
    ws.cell(54, 9).value  = pm.get("ev_rev")
    ws.cell(54, 10).value = pm.get("p_fcf")

    ws.cell(55, _LABEL_COL).value = f"{ticker} (subject)"
    ws.cell(55, _LABEL_COL).font  = Font(bold=True, name="Calibri", size=10)
    ws.cell(55, 3).value = last_rev
    ws.cell(55, 4).value = ebitda_ltm
    ws.cell(55, 5).value = _v(ebit_h, -1)
    ws.cell(55, 7).value = round(ev / ebitda_ltm, 2) if ebitda_ltm else None
    ws.cell(55, 9).value = round(ev / last_rev,   2) if last_rev   else None

    # =========================================================
    # SENSITIVITY TABLE (rows 60-74)
    # =========================================================
    ws.cell(60, _LABEL_COL).value = "-- Sensitivity: Intrinsic Value per Share --"
    ws.cell(61, _LABEL_COL).value = "WACC \\ Terminal g ->"

    g_labels    = sens.get("g_labels",    [])
    wacc_labels = sens.get("wacc_labels", [])
    iv_grid     = sens.get("iv_grid",     [])
    base_g_idx   = _s(sens.get("base_g_idx",   2))
    base_wacc_idx= _s(sens.get("base_wacc_idx", 2))

    for ci, gl in enumerate(g_labels):
        ws.cell(61, 2 + ci).value = gl
        ws.cell(61, 2 + ci).font  = Font(bold=True, name="Calibri", size=10)

    for ri, wl in enumerate(wacc_labels):
        row = 62 + ri
        ws.cell(row, _LABEL_COL).value = wl
        ws.cell(row, _LABEL_COL).font  = Font(bold=True, name="Calibri", size=10)
        grid_row = iv_grid[ri] if ri < len(iv_grid) else []
        for ci, iv_val in enumerate(grid_row):
            c = ws.cell(row, 2 + ci)
            c.value = iv_val
            c.number_format = "$#,##0.00"
            if ri == base_wacc_idx and ci == base_g_idx:
                c.fill = PatternFill("solid", fgColor="FFC000")
                c.font = Font(bold=True, name="Calibri", size=10)

    ws.cell(68, _LABEL_COL).value = (
        f"Base case:  WACC = {wacc:.1f}%   |   Terminal g = {terminal_g:.1f}%"
        f"   |   IV = ${iv:.2f}"
    )

    # =========================================================
    # WORKING CALCULATIONS (rows 190-295)
    # =========================================================
    ws.cell(190, _LABEL_COL).value = "-- Working Calculations --"
    ws.cell(191, _LABEL_COL).value = f"{ticker} | Working Calculations"
    ws.cell(192, _LABEL_COL).value = "Internal reference rows (cross-checks)"
    ws.cell(193, _HIST_COLS[0]).value = "Historical"
    ws.cell(193, _FCST_COLS[0]).value = "Forecast"

    # BS aggregates — use actual EODHD data where available
    # CA excl cash: actual (totalCA - cash) if available, else AR+Inv
    if _has_ca:
        ca_excl_h = [max(0, _s(_v(ca_direct_h, i)) - _s(_v(cash_h, i))) for i in range(_N_HIST)]
    else:
        ca_excl_h = [hist_recv_bs[i] + hist_inv_bs[i] for i in range(_N_HIST)]
    ca_excl_f = [_s(_v(fcst_recv, j)) + _s(_v(fcst_inv, j)) for j in range(_N_FCST)]
    # CL excl IBD: actual total CL if available (includes IBD portions but much better than just AP)
    if _has_cl:
        cl_excl_h = [_s(_v(cl_direct_h, i)) for i in range(_N_HIST)]
    else:
        cl_excl_h = hist_ap_bs[:]
    cl_excl_f = fcst_ap[:]
    nca_excl_h = [0] * _N_HIST
    nca_excl_f = [0] * _N_FCST
    ppe_intang_h = [max(0, _s(_v(ta_h, i)) - _s(_v(cash_h, i)) - ca_excl_h[i])
                    for i in range(_N_HIST)]
    # PPE forecast: hold at last historical to prevent spurious decline to zero
    _last_ppe_wc = ppe_intang_h[-1] if ppe_intang_h else 0
    ppe_intang_f = [_last_ppe_wc] * _N_FCST
    fcst_display_ta = [
        _s(_v(fcst_cash, j)) + _s(_v(ca_excl_f, j)) + _s(_v(nca_excl_f, j)) + _s(_v(ppe_intang_f, j))
        for j in range(_N_FCST)
    ]
    fcst_ta_bs = [max(_s(_v(fcst_ta_bs, j)), fcst_display_ta[j]) for j in range(_N_FCST)]
    # IBD and book equity must be computed before NCL (NCL is residual of BS)
    ibd_h = [_s(_v(debt_h, i)) for i in range(_N_HIST)]
    ibd_f = fcst_ltd[:]
    be_h = [_s(_v(equity_h, i)) for i in range(_N_HIST)]
    # NCL excl IBD = residual: TA - CL_excl - IBD - OE  (balances the BS to zero)
    # Only reliable when we have actual current-liabilities data from EODHD
    if _has_cl:
        ncl_excl_h = [max(0, _s(_v(ta_h, i)) - cl_excl_h[i] - ibd_h[i] - be_h[i])
                      for i in range(_N_HIST)]
    else:
        ncl_excl_h = [0] * _N_HIST
    _last_ncl = ncl_excl_h[-1] if ncl_excl_h else 0
    ncl_excl_f = [_last_ncl] * _N_FCST
    fcst_tl = [_s(_v(fcst_ap, j)) + _s(_v(fcst_ltd, j)) + _s(_v(ncl_excl_f, j)) for j in range(_N_FCST)]
    fcst_eq = [_s(_v(fcst_ta_bs, j)) - _s(_v(fcst_tl, j)) for j in range(_N_FCST)]
    be_f = fcst_eq[:]
    cash_agg_h = [_s(_v(cash_h, i)) for i in range(_N_HIST)]
    cash_agg_f = fcst_cash[:]

    _wr(194, "BalanceSheetFundingDeficiency = A-L-OE",
        [_s(_v(ta_h, i)) - hist_tl_bs[i] - _s(_v(equity_h, i))
         for i in range(_N_HIST)])
    _wr(196, "CurrentAssetsExcludingCash", ca_excl_h, ca_excl_f)
    _wr(197, "CurrentLiabilitiesExcludingIBD", cl_excl_h, cl_excl_f)
    _wr(198, "NonCurrentAssetsExclPPEAndIntangibles", nca_excl_h, nca_excl_f)
    _wr(199, "PPEAndIntangibles", ppe_intang_h, ppe_intang_f)
    _wr(200, "NonCurrentLiabilitiesExcludingIBD", ncl_excl_h, ncl_excl_f)
    _wr(201, "InterestBearingDebt (IBD)", ibd_h, ibd_f)
    _wr(202, "BookEquity (OE)", be_h, be_f)
    _wr(203, "Cash", cash_agg_h, cash_agg_f)
    _wrf(204, "BalanceSheetCheck (should be zero)",
              "=@196-@197+@198+@199-@200-@201-@202+@203")

    # Delta rows (206-215)
    ca_all  = ca_excl_h  + ca_excl_f
    cl_all  = cl_excl_h  + cl_excl_f
    nca_all = nca_excl_h + nca_excl_f
    ppe_all = ppe_intang_h + ppe_intang_f
    ncl_all = ncl_excl_h + ncl_excl_f
    ibd_all = ibd_h + ibd_f
    oe_all  = be_h  + be_f
    c_all   = cash_agg_h + cash_agg_f

    delta_labels = {
        206: "\u0394CAExclCash",
        207: "\u0394CLExclIBD",
        208: "\u0394NCAExclPPE",
        209: "\u0394PPEAndIntangibles",
        210: "\u0394NCLExclIBD",
        211: "\u0394IBD",
        212: "\u0394OE",
        213: "\u0394Cash = CashNow - CashBefore",
        214: "\u0394Cash reconciled",
        215: "BSCheck2 (should be zero)",
    }
    for rn, lbl in delta_labels.items():
        ws.cell(rn, _LABEL_COL).value = lbl

    all_cols = _HIST_COLS[1:] + _FCST_COLS
    all_src_idx = list(range(1, _N_HIST + _N_FCST))
    series_map = {206: ca_all, 207: cl_all, 208: nca_all, 209: ppe_all,
                  210: ncl_all, 211: ibd_all, 212: oe_all, 213: c_all}
    for rn, series in series_map.items():
        for k, src_idx in enumerate(all_src_idx):
            ws.cell(rn, all_cols[k]).value = series[src_idx] - series[src_idx - 1]

    for k, src_idx in enumerate(all_src_idx):
        col = all_cols[k]
        dca  = ca_all[src_idx] - ca_all[src_idx-1]
        dcl  = cl_all[src_idx] - cl_all[src_idx-1]
        dnca = nca_all[src_idx] - nca_all[src_idx-1]
        dppe = ppe_all[src_idx] - ppe_all[src_idx-1]
        dncl = ncl_all[src_idx] - ncl_all[src_idx-1]
        dibd = ibd_all[src_idx] - ibd_all[src_idx-1]
        doe  = oe_all[src_idx]  - oe_all[src_idx-1]
        dc   = c_all[src_idx]   - c_all[src_idx-1]
        ws.cell(214, col).value = -(dca - dcl) - (dnca + dppe - dncl) + dibd + doe
        ws.cell(215, col).value = dca - dcl + dnca + dppe - dncl - dibd - doe + dc

    # EFCF (rows 217-224)
    efcf_labels = {
        217: "NetIncome",
        218: "Dividends",
        219: "Buybacks",
        220: "EquityRaisings",
        221: "\u0394Reserves",
        222: "\u0394OE estimated",
        223: "\u0394OE from BS",
        224: "EFCF = Dividends + Buybacks - EquityRaisings",
    }
    for rn, lbl in efcf_labels.items():
        ws.cell(rn, _LABEL_COL).value = lbl

    ni_all  = [_s(_v(ni_h, i)) for i in range(_N_HIST)] + fcst_ni
    # Use actual dividends/buybacks if available, else zero
    if _has_div:
        _last_div = _s(_v(div_direct_h, -1))
        div_all = [_s(_v(div_direct_h, i)) for i in range(_N_HIST)] + [_last_div] * _N_FCST
    else:
        div_all = [0] * (_N_HIST + _N_FCST)
    if _has_buy:
        _last_buy = _s(_v(buyback_direct_h, -1))
        buy_all = [_s(_v(buyback_direct_h, i)) for i in range(_N_HIST)] + [_last_buy] * _N_FCST
    else:
        buy_all = [0] * (_N_HIST + _N_FCST)
    if _has_iss:
        _last_iss = _s(_v(stock_iss_h, -1))
        eq_all  = [_s(_v(stock_iss_h, i)) for i in range(_N_HIST)] + [_last_iss] * _N_FCST
    else:
        eq_all = [0] * (_N_HIST + _N_FCST)

    for k, src_idx in enumerate(all_src_idx):
        col = all_cols[k]
        ws.cell(217, col).value = ni_all[src_idx]
        ws.cell(218, col).value = div_all[src_idx]
        ws.cell(219, col).value = buy_all[src_idx]
        ws.cell(220, col).value = eq_all[src_idx]
        ws.cell(221, col).value = 0
        ws.cell(222, col).value = (ni_all[src_idx] - div_all[src_idx]
                                   - buy_all[src_idx] + eq_all[src_idx])
        ws.cell(223, col).value = oe_all[src_idx] - oe_all[src_idx - 1]
        ws.cell(224, col).value = (div_all[src_idx] + buy_all[src_idx]
                                   - eq_all[src_idx])

    # Cash flow decomp (rows 226-237)
    cf_labels = {
        226: "\u0394IBD",
        227: "IntExp",
        229: "D&A",
        230: "\u0394NOWC",
        231: "CapEx = \u0394PPE + DA",
        233: "OperatingCashflows = NI + DA - \u0394NOWC",
        234: "InvestingCashflows = -CapEx",
        235: "FinancingCashflows = \u0394IBD - EFCF",
        236: "\u0394Cash estimated",
        237: "\u0394Cash from CF statement",
    }
    for rn, lbl in cf_labels.items():
        ws.cell(rn, _LABEL_COL).value = lbl

    ie_all  = [_s(_v(int_exp_h, i)) for i in range(_N_HIST)] + [_s(_v(fcst_interest_exp, j)) for j in range(_N_FCST)]
    da_all  = ([_s(_v(da_h, i)) for i in range(_N_HIST)]
               + [_s(_v(fcst_da, j)) for j in range(_N_FCST)])

    for k, src_idx in enumerate(all_src_idx):
        col = all_cols[k]
        dibd  = ibd_all[src_idx] - ibd_all[src_idx - 1]
        dca2  = ca_all[src_idx]  - ca_all[src_idx - 1]
        dcl2  = cl_all[src_idx]  - cl_all[src_idx - 1]
        dnca2 = nca_all[src_idx] - nca_all[src_idx - 1]
        dppe2 = ppe_all[src_idx] - ppe_all[src_idx - 1]
        dncl2 = ncl_all[src_idx] - ncl_all[src_idx - 1]
        da_v  = da_all[src_idx]
        ni_v  = ni_all[src_idx]
        ie_v  = ie_all[src_idx]
        efcf_v = div_all[src_idx] + buy_all[src_idx] - eq_all[src_idx]
        dnowc  = dca2 + dnca2 - dcl2 - dncl2
        capex_v= dppe2 + da_v
        ocf_v  = ni_v + da_v - dnowc
        icf_v  = -capex_v
        fcf2_v = dibd - efcf_v
        ws.cell(226, col).value = dibd
        ws.cell(227, col).value = ie_v
        ws.cell(229, col).value = da_v
        ws.cell(230, col).value = dnowc
        ws.cell(231, col).value = capex_v
        ws.cell(233, col).value = ocf_v
        ws.cell(234, col).value = icf_v
        ws.cell(235, col).value = fcf2_v
        ws.cell(236, col).value = ocf_v + icf_v + fcf2_v
        ws.cell(237, col).value = c_all[src_idx] - c_all[src_idx - 1]

    # Q9a FFCFwITS (rows 240-246)
    ws.cell(240, _LABEL_COL).value = "Question 9a:"
    ws.cell(241, _LABEL_COL).value = "FFCFwITS = DebtCF + EFCF"
    ws.cell(243, _LABEL_COL).value = "DebtCashFlowPaidToDebtHolders"
    ws.cell(244, _LABEL_COL).value = "FFCFwITS = DebtCashFlow + EFCF"
    ws.cell(245, _LABEL_COL).value = "Interest tax shield per year (ITS)"
    ws.cell(246, _LABEL_COL).value = "FFCFwITS = UFCF + ITS (fcst)"

    for k, src_idx in enumerate(all_src_idx):
        col = all_cols[k]
        efcf_v = div_all[src_idx] + buy_all[src_idx] - eq_all[src_idx]
        dibd   = ibd_all[src_idx] - ibd_all[src_idx - 1]
        ie_v   = ie_all[src_idx]
        ws.cell(243, col).value = dibd
        ws.cell(244, col).value = efcf_v + dibd
        ws.cell(245, col).value = ie_v * tax_dec

    for j in range(_N_FCST):
        col = _FCST_COLS[j]
        ufcf_v = _s(_v(fcst_ufcf, j))
        its_v  = ws.cell(245, col).value or 0
        ws.cell(246, col).value = ufcf_v + its_v

    # Q9b
    ws.cell(248, _LABEL_COL).value = "Question 9b:"
    ws.cell(249, _LABEL_COL).value = "EV with ITS adjustment"

    # Q10 CFADS (rows 259-261)
    ws.cell(259, _LABEL_COL).value = "Question 10:"
    ws.cell(260, _LABEL_COL).value = "CFADS (actual)"
    ws.cell(261, _LABEL_COL).value = "CFADS notional"

    for k, src_idx in enumerate(all_src_idx):
        col = all_cols[k]
        da_v   = da_all[src_idx]
        ni_v   = ni_all[src_idx]
        ie_v   = ie_all[src_idx]
        dca3   = ca_all[src_idx]  - ca_all[src_idx-1]
        dcl3   = cl_all[src_idx]  - cl_all[src_idx-1]
        dnca3  = nca_all[src_idx] - nca_all[src_idx-1]
        dncl3  = ncl_all[src_idx] - ncl_all[src_idx-1]
        dppe3  = ppe_all[src_idx] - ppe_all[src_idx-1]
        dnowc3 = dca3 + dnca3 - dcl3 - dncl3
        capex3 = dppe3 + da_v
        ws.cell(260, col).value = ni_v + da_v + ie_v * (1 - tax_dec) - dnowc3 - capex3
        ws.cell(261, col).value = ni_v + da_v + ie_v * (1 - tax_dec) - dnowc3 - da_v

    # Efficiency (rows 263-274)
    eff_labels = {
        263: "Fixed Assets", 264: "Revenue", 265: "COGS",
        266: "Receivables", 267: "Inventory", 268: "Payables",
        269: "Efficiency:",
    }
    for rn, lbl in eff_labels.items():
        ws.cell(rn, _LABEL_COL).value = lbl

    ppe_bs  = ppe_intang_h + ppe_intang_f
    rev_bs  = ([_s(_v(rev_h, i)) for i in range(_N_HIST)]
               + [_s(_v(fcst_rev, j)) for j in range(_N_FCST)])
    cogs_bs = ([_s(_v(cogs_h, i)) for i in range(_N_HIST)]
               + [abs(_s(_v(fcst_cogs, j))) for j in range(_N_FCST)])
    recv_bs = hist_recv_bs + fcst_recv
    inv_bs  = hist_inv_bs  + fcst_inv
    pay_bs  = hist_ap_bs   + fcst_ap

    all_period_cols = _HIST_COLS + _FCST_COLS
    for k in range(_N_HIST + _N_FCST):
        col = all_period_cols[k]
        ws.cell(263, col).value = ppe_bs[k]
        ws.cell(264, col).value = rev_bs[k]
        ws.cell(265, col).value = cogs_bs[k]
        ws.cell(266, col).value = recv_bs[k]
        ws.cell(267, col).value = inv_bs[k]
        ws.cell(268, col).value = pay_bs[k]

    # Efficiency ratios — write as Excel formulas referencing the data rows above
    ws.cell(270, _LABEL_COL).value = "Receivable Days"
    ws.cell(271, _LABEL_COL).value = "Inventory Days"
    ws.cell(272, _LABEL_COL).value = "Payable Days"
    ws.cell(273, _LABEL_COL).value = "Cash Conversion Cycle (days)"
    ws.cell(274, _LABEL_COL).value = "Fixed Asset Turnover (years)"
    for col in all_period_cols:
        c = get_column_letter(col)
        ws.cell(270, col).value = f"=IFERROR({c}266/{c}264*365,\"\")"
        ws.cell(271, col).value = f"=IFERROR({c}267/{c}265*365,\"\")"
        ws.cell(272, col).value = f"=IFERROR({c}268/{c}265*365,\"\")"
        ws.cell(273, col).value = f"={c}270+{c}271-{c}272"
        ws.cell(274, col).value = f"=IFERROR({c}264/{c}263,\"\")"

    # Liquidity / Leverage (rows 276-292)
    lev_labels = {
        276: "Liquidity", 277: "Current Ratio", 278: "Quick Ratio",
        280: "Book assets", 281: "Book equity", 282: "Book debt",
        283: "Leverage ratios", 284: "Debt to assets",
        285: "Debt to equity", 286: "Asset to equity",
        288: "EBITDA", 289: "Tax",
        290: "Interest coverage", 291: "EBITDA less Tax coverage",
        292: "CFADS notional coverage",
    }
    for rn, lbl in lev_labels.items():
        ws.cell(rn, _LABEL_COL).value = lbl

    # Write formula strings referencing the BS/IS rows already written above
    for col in all_period_cols:
        c = get_column_letter(col)
        # Liquidity
        ws.cell(277, col).value = f"=IFERROR({c}510/{c}527,\"\")"   # TCA / TCL
        ws.cell(278, col).value = f"=IFERROR(({c}510-{c}507)/{c}527,\"\")"  # (TCA-Inv)/TCL
        # Book values — reference BS totals
        ws.cell(280, col).value = f"={c}518"   # Total Assets
        ws.cell(281, col).value = f"={c}540"   # Total Equity
        ws.cell(282, col).value = f"={c}528"   # LT Debt
        # Leverage
        ws.cell(284, col).value = f"=IFERROR({c}282/{c}280,\"\")"   # Debt/Assets
        ws.cell(285, col).value = f"=IFERROR({c}282/{c}281,\"\")"   # Debt/Equity
        ws.cell(286, col).value = f"=IFERROR({c}280/{c}281,\"\")"   # Assets/Equity
        # Coverage — reference IS rows
        ws.cell(288, col).value = f"={c}380+{c}344"               # EBITDA = EBIT + D&A
        ws.cell(289, col).value = f"={c}359"                       # Tax
        ws.cell(290, col).value = f"=IFERROR({c}380/{c}227,\"\")"  # Interest coverage
        ws.cell(291, col).value = f"=IFERROR(({c}288-{c}289)/{c}227,\"\")"  # EBITDA-Tax cov.
        ws.cell(292, col).value = f"=IFERROR({c}261/{c}227,\"\")"  # CFADS cov.

    ws.cell(295, _LABEL_COL).value = (
        "Note: Working calc rows 190-295 mirror Nike VSTR template cross-checks."
    )

    # =========================================================
    # INCOME STATEMENT (rows 300-412)
    # =========================================================
    ws.cell(300, _LABEL_COL).value = "INCOME STATEMENT ($M)"
    ws.cell(301, _LABEL_COL).value = f"{company} | Income Statement"
    ws.cell(302, _LABEL_COL).value = f"{ticker}"
    ws.cell(304, _LABEL_COL).value = "Source: Recommended"
    ws.cell(305, _LABEL_COL).value = "Period Category: Custom"
    ws.cell(306, _LABEL_COL).value = "Period Type: Custom"
    ws.cell(307, _LABEL_COL).value = "Reporting Basis: Custom"
    ws.cell(308, _LABEL_COL).value = "Sort Order: Custom"
    ws.cell(309, _LABEL_COL).value = "Currency: U.S. Dollar (USD)"
    ws.cell(310, _LABEL_COL).value = "Magnitude: Millions (M)"

    ws.cell(311, _LABEL_COL).value = "Recommended: S&P Capital IQ - Standard"
    for i in range(_N_HIST):
        ws.cell(311, _HIST_COLS[i]).value = yr_labels[i] if i < len(yr_labels) else ""
    ws.cell(312, _LABEL_COL).value = "Period Ended"
    for i, yr in enumerate(years_h):
        date_obj = _v(dates_h, i)
        if date_obj is not None:
            ws.cell(312, _HIST_COLS[i]).value = date_obj
            ws.cell(312, _HIST_COLS[i]).number_format = "M/D/YYYY"
        else:
            try:
                _m2, _d2 = _fy_month, _MONTH_END_DAY.get(_fy_month, 31)
                ws.cell(312, _HIST_COLS[i]).value = datetime(int(yr), _m2, _d2)
                ws.cell(312, _HIST_COLS[i]).number_format = "M/D/YYYY"
            except (ValueError, TypeError):
                pass
    ws.cell(313, _LABEL_COL).value = "Currency"
    for i in range(_N_HIST):
        ws.cell(313, _HIST_COLS[i]).value = "USD"
    ws.cell(314, _LABEL_COL).value = "Units"
    for i in range(_N_HIST):
        ws.cell(314, _HIST_COLS[i]).value = "Millions"
    ws.cell(315, _LABEL_COL).value = "Revenues"

    ws.cell(316, _LABEL_COL).value = "Revenues"
    for i in range(_N_HIST):
        ws.cell(316, _HIST_COLS[i]).value = _v(rev_h, i)

    # Assumption driver rows 320-327
    ws.cell(320, _LABEL_COL).value = "growth rate revenue"
    for i in range(1, _N_HIST):
        r_c, r_p = _v(rev_h, i), _v(rev_h, i - 1)
        ws.cell(320, _HIST_COLS[i]).value = (r_c / r_p - 1) if r_p and r_c else None
    terminal_g_dec = terminal_g / 100
    last_hist_g = 0
    if _v(rev_h, -1) and _v(rev_h, -2):
        last_hist_g = _v(rev_h, -1) / _v(rev_h, -2) - 1
    for j in range(_N_FCST):
        # L3 fix: derive growth from actual forecast revenues so row 320 is
        # consistent with the DCF revenue schedule in rows 16 / 337.
        prev_r = _s(_v(fcst_rev, j - 1)) if j > 0 else _s(_v(rev_h, -1))
        curr_r = _s(_v(fcst_rev, j))
        gval = (curr_r / prev_r - 1) if prev_r and curr_r else terminal_g_dec
        ws.cell(320, _FCST_COLS[j]).value = gval

    ws.cell(321, _LABEL_COL).value = "COGS/Sales"
    for i in range(_N_HIST):
        r, cg = _v(rev_h, i), _v(cogs_h, i)
        ws.cell(321, _HIST_COLS[i]).value = (cg / r) if r and cg else None
    for j in range(_N_FCST):
        rev_j = _s(_v(fcst_rev, j))
        ws.cell(321, _FCST_COLS[j]).value = (_s(_v(fcst_cogs, j)) / rev_j) if rev_j else None

    ws.cell(322, _LABEL_COL).value = "effective corporate tax rate"
    for i in range(_N_HIST):
        t, pt = _v(tax_h, i), _v(pretax_h, i)
        ws.cell(322, _HIST_COLS[i]).value = (t / pt) if pt and t else None
    for j in range(_N_FCST):
        ws.cell(322, _FCST_COLS[j]).value = tax_dec

    ws.cell(323, _LABEL_COL).value = "book long term debt to asset ratio"
    for i in range(_N_HIST):
        d, ta = _v(debt_h, i), _v(ta_h, i)
        ws.cell(323, _HIST_COLS[i]).value = (d / ta) if ta and d else None
    for j in range(_N_FCST):
        ws.cell(323, _FCST_COLS[j]).value = dt_ratio

    ws.cell(324, _LABEL_COL).value = "dividend payout%"
    for i in range(_N_HIST):
        ws.cell(324, _HIST_COLS[i]).value = 0
    for j in range(_N_FCST):
        ws.cell(324, _FCST_COLS[j]).value = 0

    ws.cell(325, _LABEL_COL).value = "average growth rate"
    ws.cell(326, _LABEL_COL).value = "average tax rate"
    ws.cell(327, _LABEL_COL).value = "interest tax expense rate"
    for i in range(_N_HIST):
        ie_v = _v(int_exp_h, i)
        d_v  = _v(debt_h, i)
        ws.cell(327, _HIST_COLS[i]).value = (ie_v / d_v) if d_v and ie_v else None

    ws.cell(329, _LABEL_COL).value = "Recommended: S&P Capital IQ - Standard"
    for i in range(_N_HIST):
        ws.cell(329, _HIST_COLS[i]).value = yr_labels[i] if i < len(yr_labels) else ""
    ws.cell(330, _LABEL_COL).value = " "
    for i in range(_N_HIST):
        ws.cell(330, _HIST_COLS[i]).value = "Current/Restated"
    ws.cell(331, _LABEL_COL).value = "Period Ended"
    for i, yr in enumerate(years_h):
        date_obj = _v(dates_h, i)
        if date_obj is not None:
            ws.cell(331, _HIST_COLS[i]).value = date_obj
            ws.cell(331, _HIST_COLS[i]).number_format = "M/D/YYYY"
        else:
            try:
                _m2, _d2 = _fy_month, _MONTH_END_DAY.get(_fy_month, 31)
                ws.cell(331, _HIST_COLS[i]).value = datetime(int(yr), _m2, _d2)
                ws.cell(331, _HIST_COLS[i]).number_format = "M/D/YYYY"
            except (ValueError, TypeError):
                pass
    ws.cell(332, _LABEL_COL).value = "Financial Filing Date"
    ws.cell(333, _LABEL_COL).value = "Spot Exchange Rate"
    for i in range(_N_HIST):
        ws.cell(333, _HIST_COLS[i]).value = 1
    ws.cell(334, _LABEL_COL).value = "Average Exchange Rate"
    for i in range(_N_HIST):
        ws.cell(334, _HIST_COLS[i]).value = 1
    ws.cell(335, _LABEL_COL).value = " "
    ws.cell(336, _LABEL_COL).value = "($M)"

    # IS data rows
    ws.cell(337, _LABEL_COL).value = "Revenue"
    for i in range(_N_HIST):
        ws.cell(337, _HIST_COLS[i]).value = _v(rev_h, i)
    for j in range(_N_FCST):                                        # source of truth for revenue
        ws.cell(337, _FCST_COLS[j]).value = _v(fcst_rev, j)

    _wrf(338, "        Total Revenue",   "=@337",                   bold=True)
    _wr(339,  "Cost Of Goods Sold",
              [_s(_v(cogs_h, i)) for i in range(_N_HIST)], fcst_cogs)
    _wrf(340, "        Gross Profit",    "=@338-@339",              bold=True)
    _wr(342,  " SG&A excl. Depreciation", sga_h, fcst_sga)
    _wr(344,  "        Depreciation & Amort., Total", da_h, fcst_da)
    _wrf(345, "        Other Operating Exp., Total",  "=@342+@344")
    _wrf(346, "        Operating Income",             "=@340-@345", bold=True)

    ws.cell(347, _LABEL_COL).value = " "
    _wr(348, "Interest Expense",
             [-_s(_v(int_exp_h, i)) for i in range(_N_HIST)],
             [-_s(_v(fcst_interest_exp, j)) for j in range(_N_FCST)])
    _wr(349, "Interest and Invest. Income", [0] * _N_HIST, [0] * _N_FCST)

    # Other Non-Operating Income: plug = EBT - EBIT + IntExp (reconciles EBT to formula-chain)
    other_nonop_h = [_s(_v(pretax_h, i)) - _s(_v(ebit_h, i)) + _s(_v(int_exp_h, i))
                     for i in range(_N_HIST)]
    _wr(350, "        Net Interest Exp.",
             [-_s(_v(int_exp_h, i)) for i in range(_N_HIST)],
             [-_s(_v(fcst_interest_exp, j)) for j in range(_N_FCST)])
    _wr(351, "Currency Exchange Gains",   [0] * _N_HIST, [0] * _N_FCST)
    _wr(352, "Other Non-Operating Inc.",  other_nonop_h,  [0] * _N_FCST)
    _wrf(353, "        EBT Excl Unusual Items", "=@346+@350+@351+@352")
    _wr(354,  "Restructuring Charges",     [0] * _N_HIST, [0] * _N_FCST)
    _wrf(355, "        EBT Incl. Unusual Items", "=@353+@354",        bold=True)
    _wr(359,  "Income Tax Expense",        tax_h, fcst_tax)
    _wrf(360, "        Earnings from Cont. Ops.", "=@355-@359")
    _wrf(361, "        Net Income to Company",    "=@362",             bold=True)
    _wr(362,  "        Net Income",               ni_h, fcst_ni)      # source of truth
    _wrf(363, "        NI to Common Incl Extra Items", "=@362")
    _wrf(364, "        NI to Common Excl. Extra Items", "=@362")

    ws.cell(365, _LABEL_COL).value = "Per Share Items ($)"
    _wr(368, "Weighted Avg. Basic Shares Out. (M)", shares_h, fcst_shares)
    _wr(371, "Weighted Avg. Diluted Shares Out. (M)", shares_h, fcst_shares)
    _wrf(366, "Basic EPS",                         "=IFERROR(@362/@368,\"\")")
    _wrf(367, "Basic EPS Excl. Extra Items",       "=@366")
    _wrf(369, "Diluted EPS Incl. Extra Items",     "=IFERROR(@362/@371,\"\")")
    _wrf(370, "Diluted EPS Excl. Extra Items",     "=@369")

    ws.cell(377, _LABEL_COL).value = "Supplemental Items ($M)"
    _wr(379, "EBITA",    ebit_h,      fcst_ebit)
    _wr(380, "EBIT",     ebit_h,
             [_v(fcst_ebit, j) for j in range(_N_FCST)], bold=True)   # source of truth
    _wrf(378, "EBITDA",  "=@380+@344")
    _wrf(381, "EBITDAR", "=@378")
    _wr(382, "Effective Tax Rate (%)",
             [_s(_v(tax_h, i)) / _s(_v(pretax_h, i), 1) * 100
              if _v(pretax_h, i) else None for i in range(_N_HIST)])

    ws.cell(390, _LABEL_COL).value = "Supplemental Operating Expense Items ($M)"
    _wr(399, "        Stock-Based Comp., Total", sbc_h, fcst_sbc)

    ws.cell(400, _LABEL_COL).value = "CIQ Restatement Type Code"
    for i in range(_N_HIST):
        ws.cell(400, _HIST_COLS[i]).value = "NC"
    ws.cell(401, _LABEL_COL).value = "CIQ Calculation Type Code"
    for i in range(_N_HIST):
        ws.cell(401, _HIST_COLS[i]).value = "REP"

    # =========================================================
    # CASH FLOW STATEMENT (rows 415-478)
    # =========================================================
    ws.cell(415, _LABEL_COL).value = "CASH FLOW STATEMENT ($M)"
    ws.cell(416, _LABEL_COL).value = f"{company} | Cash Flow"
    ws.cell(418, _LABEL_COL).value = "Source: Recommended"
    ws.cell(419, _LABEL_COL).value = "Period Category: Custom"
    ws.cell(420, _LABEL_COL).value = "Period Type: Custom"
    ws.cell(421, _LABEL_COL).value = "Reporting Basis: Custom"
    ws.cell(422, _LABEL_COL).value = "Sort Order: Custom"
    ws.cell(423, _LABEL_COL).value = "Currency: U.S. Dollar (USD)"
    ws.cell(424, _LABEL_COL).value = "Magnitude: Millions (M)"

    ws.cell(426, _LABEL_COL).value = "Recommended: S&P Capital IQ - Standard"
    for i in range(_N_HIST):
        ws.cell(426, _HIST_COLS[i]).value = yr_labels[i] if i < len(yr_labels) else ""
    ws.cell(427, _LABEL_COL).value = " "
    for i in range(_N_HIST):
        ws.cell(427, _HIST_COLS[i]).value = "Current/Restated"
    ws.cell(428, _LABEL_COL).value = "Period Ended"
    for i, yr in enumerate(years_h):
        date_obj = _v(dates_h, i)
        if date_obj is not None:
            ws.cell(428, _HIST_COLS[i]).value = date_obj
            ws.cell(428, _HIST_COLS[i]).number_format = "M/D/YYYY"
        else:
            try:
                _m2, _d2 = _fy_month, _MONTH_END_DAY.get(_fy_month, 31)
                ws.cell(428, _HIST_COLS[i]).value = datetime(int(yr), _m2, _d2)
                ws.cell(428, _HIST_COLS[i]).number_format = "M/D/YYYY"
            except (ValueError, TypeError):
                pass
    ws.cell(429, _LABEL_COL).value = "Financial Filing Date"
    ws.cell(430, _LABEL_COL).value = "Spot Exchange Rate"
    for i in range(_N_HIST):
        ws.cell(430, _HIST_COLS[i]).value = 1
    ws.cell(431, _LABEL_COL).value = "Average Exchange Rate"
    for i in range(_N_HIST):
        ws.cell(431, _HIST_COLS[i]).value = 1
    ws.cell(432, _LABEL_COL).value = " "

    ws.cell(433, _LABEL_COL).value = "Operating Activities ($M)"
    _wr(434, "Net Income - CF",                  ni_h,    fcst_ni)
    _wr(435, "Depreciation & Amort.",            da_h,    fcst_da)
    _wr(436, "Amort. of Goodwill and Intangibles", [0] * _N_HIST, [0] * _N_FCST)
    _wrf(437, "        Depreciation & Amort., Total", "=@435+@436")
    _wr(438, "Asset Writedown & Restructuring",  [0] * _N_HIST, [0] * _N_FCST)
    _wr(439, "Stock-Based Compensation",         sbc_h,   fcst_sbc)

    wc_chg_h = [_s(_v(op_cf_h, i)) - _s(_v(ni_h, i)) - _s(_v(da_h, i))
                for i in range(_N_HIST)]
    _wr(440, "Other Operating Activities",       wc_chg_h, [0] * _N_FCST)
    # F1 fix: populate forecast NWC changes from forecast BS arrays
    _prev_recv = hist_recv_bs[-1] if hist_recv_bs else 0.0
    _prev_inv  = hist_inv_bs[-1]  if hist_inv_bs  else 0.0
    _prev_ap   = hist_ap_bs[-1]   if hist_ap_bs   else 0.0
    fcst_dar = []
    fcst_dinv = []
    fcst_dap  = []
    for _j_nwc in range(_N_FCST):
        _cur_recv = _s(_v(fcst_recv, _j_nwc))
        _cur_inv  = _s(_v(fcst_inv,  _j_nwc))
        _cur_ap   = _s(_v(fcst_ap,   _j_nwc))
        fcst_dar.append(-(_cur_recv - _prev_recv))   # AR increase → negative CF
        fcst_dinv.append(-(_cur_inv  - _prev_inv))   # Inv increase → negative CF
        fcst_dap.append(  _cur_ap   - _prev_ap)      # AP increase → positive CF
        _prev_recv, _prev_inv, _prev_ap = _cur_recv, _cur_inv, _cur_ap
    _wr(441, "Change in Acc. Receivable",
             [-(hist_recv_bs[i] - (hist_recv_bs[i-1] if i > 0 else hist_recv_bs[i]))
              for i in range(_N_HIST)], fcst_dar)
    _wr(442, "Change In Inventories",
             [-(hist_inv_bs[i] - (hist_inv_bs[i-1] if i > 0 else hist_inv_bs[i]))
              for i in range(_N_HIST)], fcst_dinv)
    _wr(443, "Change in Acc. Payable",
             [hist_ap_bs[i] - (hist_ap_bs[i-1] if i > 0 else hist_ap_bs[i])
              for i in range(_N_HIST)], fcst_dap)
    _wr(444, "Change in Other Net Operating Assets", [0] * _N_HIST, fcst_chg_other_oa)
    _wrf(445, "        Cash from Ops.",
              "=@434+@437+@438+@439+@440+@441+@442+@443+@444", bold=True)

    ws.cell(446, _LABEL_COL).value = "Investing Activity ($M)"
    cap_neg_h = [-_s(_v(capex_h, i)) for i in range(_N_HIST)]
    _wr(447, "Capital Expenditure",              cap_neg_h, fcst_cap_cf)
    _wr(448, "Sale of Property, Plant and Equipment", [0] * _N_HIST, [0] * _N_FCST)
    _wr(449, "Invest. in Marketable & Equity Securt.", [0] * _N_HIST, [0] * _N_FCST)
    _wr(450, "Other Investing Activities",       [0] * _N_HIST, [0] * _N_FCST)
    _wrf(451, "        Cash from Investing",     "=@447+@448+@449+@450",  bold=True)

    ws.cell(452, _LABEL_COL).value = "Financing Activity ($M)"
    # CF financing: use actual data if available
    _debt_iss_h  = [max(0.0, _s(_v(net_borr_h, i))) for i in range(_N_HIST)] if _has_borr else [0] * _N_HIST
    _debt_rep_h  = [min(0.0, _s(_v(net_borr_h, i))) for i in range(_N_HIST)] if _has_borr else [0] * _N_HIST
    _div_cf_h    = [-_s(_v(div_direct_h, i)) for i in range(_N_HIST)] if _has_div else [0] * _N_HIST
    _buy_cf_h    = [-_s(_v(buyback_direct_h, i)) for i in range(_N_HIST)] if _has_buy else [0] * _N_HIST
    _iss_cf_h    = [_s(_v(stock_iss_h, i)) for i in range(_N_HIST)] if _has_iss else [0] * _N_HIST
    _div_cf_f    = -_s(_v(div_direct_h, -1)) if _has_div else 0
    _buy_cf_f    = -_s(_v(buyback_direct_h, -1)) if _has_buy else 0
    _iss_cf_f    = _s(_v(stock_iss_h, -1)) if _has_iss else 0
    _wr(453, "Short-term Debt Issued",    [0] * _N_HIST, [0] * _N_FCST)
    _wr(454, "Long-term Debt Issued",     _debt_iss_h,   [0] * _N_FCST)
    _wrf(455, "        Total Debt Issued", "=@453+@454")
    _wr(456, "Short-term Debt Repaid",    [0] * _N_HIST, [0] * _N_FCST)
    _wr(457, "Long-term Debt Repaid",     _debt_rep_h,   [0] * _N_FCST)
    _wrf(458, "        Total Debt Repaid", "=@456+@457")
    _wr(459, "Issuance of Common Stock",   _iss_cf_h, fcst_iss_cf)
    _wr(460, "Repurchase of Common Stock", _buy_cf_h, fcst_buy_cf)
    _wr(461, "Common Dividends Paid",      _div_cf_h, fcst_div_cf)
    _wr(462, "equity raisings",            _iss_cf_h, fcst_iss_cf)
    _wrf(463, "Common and/or Pref. Dividends Paid", "=@461")
    _wrf(464, "        Total Dividends Paid",        "=@461",                bold=True)
    _wr(465, "Other Financing Activities", [0] * _N_HIST, [0] * _N_FCST)
    _wrf(466, "        Cash from Financing",
              "=@453+@454+@456+@457+@459+@460+@461+@465",                   bold=True)

    ws.cell(467, _LABEL_COL).value = "Other Cash Flow ($M)"
    _wr(468, "Foreign Exchange Rate Adj.", [0] * _N_HIST, [0] * _N_FCST)
    _wrf(469, "        Net Change in Cash", "=@445+@451+@466+@468",         bold=True)

    ws.cell(470, _LABEL_COL).value = "Supplemental Items ($M)"
    _wr(471, "Cash Interest Paid",
             [_s(_v(int_exp_h, i)) for i in range(_N_HIST)],
             fcst_interest_exp)   # actual interest paid
    _wr(472, "Cash Taxes Paid",           tax_h, fcst_tax)
    _wr(473, "Levered Free Cash Flow",    fcf_h)
    _wr(474, "Free Cash Flow (Levered Hist.) / UFCF (Fcst)",  fcf_h,        fcst_ufcf)
    _wr(475, "Change in Net Working Capital", wc_chg_h, fcst_dnowc)
    _wr(476, "Net Debt Issued",           [0] * _N_HIST, [0] * _N_FCST)
    ws.cell(477, _LABEL_COL).value = "CIQ Restatement Type Code"
    for i in range(_N_HIST):
        ws.cell(477, _HIST_COLS[i]).value = "NC"
    ws.cell(478, _LABEL_COL).value = "CIQ Calculation Type Code"
    for i in range(_N_HIST):
        ws.cell(478, _HIST_COLS[i]).value = "REP"

    # =========================================================
    # BALANCE SHEET (rows 480-590)
    # =========================================================
    ws.cell(480, _LABEL_COL).value = "BALANCE SHEET ($M)"
    ws.cell(481, _LABEL_COL).value = f"{company} | Balance Sheet"
    ws.cell(483, _LABEL_COL).value = "Source: Recommended"
    ws.cell(484, _LABEL_COL).value = "Period Category: Custom"
    ws.cell(485, _LABEL_COL).value = "Period Type: Custom"
    ws.cell(486, _LABEL_COL).value = "Reporting Basis: Custom"
    ws.cell(487, _LABEL_COL).value = "Sort Order: Custom"
    ws.cell(488, _LABEL_COL).value = "Currency: U.S. Dollar (USD)"
    ws.cell(489, _LABEL_COL).value = "Magnitude: Millions (M)"

    ws.cell(491, _LABEL_COL).value = "Recommended: S&P Capital IQ - Standard"
    for i in range(_N_HIST):
        ws.cell(491, _HIST_COLS[i]).value = yr_labels[i] if i < len(yr_labels) else ""
    ws.cell(492, _LABEL_COL).value = " "
    for i in range(_N_HIST):
        ws.cell(492, _HIST_COLS[i]).value = "Current/Restated"
    ws.cell(493, _LABEL_COL).value = "Period Ended"
    for i, yr in enumerate(years_h):
        date_obj = _v(dates_h, i)
        if date_obj is not None:
            ws.cell(493, _HIST_COLS[i]).value = date_obj
            ws.cell(493, _HIST_COLS[i]).number_format = "M/D/YYYY"
        else:
            try:
                _m2, _d2 = _fy_month, _MONTH_END_DAY.get(_fy_month, 31)
                ws.cell(493, _HIST_COLS[i]).value = datetime(int(yr), _m2, _d2)
                ws.cell(493, _HIST_COLS[i]).number_format = "M/D/YYYY"
            except (ValueError, TypeError):
                pass
    ws.cell(494, _LABEL_COL).value = "Financial Filing Date"
    ws.cell(495, _LABEL_COL).value = "Spot Exchange Rate"
    for i in range(_N_HIST):
        ws.cell(495, _HIST_COLS[i]).value = 1
    ws.cell(496, _LABEL_COL).value = "Average Exchange Rate"
    for i in range(_N_HIST):
        ws.cell(496, _HIST_COLS[i]).value = 1
    ws.cell(497, _LABEL_COL).value = "Reported Currency Code"
    for i in range(_N_HIST):
        ws.cell(497, _HIST_COLS[i]).value = "USD"
    ws.cell(498, _LABEL_COL).value = " "
    ws.cell(499, _LABEL_COL).value = "Assets ($M)"

    # ── Balance-sheet plug helpers ────────────────────────────────────────────
    # R509 Other Current Assets: plug so that R503+R506+R507+R508+R509 = TCA
    hist_oca_plug = ([max(0, _s(_v(ca_direct_h, i)) - _s(_v(cash_h, i))
                          - hist_recv_bs[i] - hist_inv_bs[i])
                      for i in range(_N_HIST)] if _has_ca else [0] * _N_HIST)
    fcst_oca_plug = [max(0, _s(_v(fcst_ca, j)) - _s(_v(fcst_cash, j))
                         - _s(_v(fcst_recv, j)) - _s(_v(fcst_inv, j)))
                     for j in range(_N_FCST)]

    # R517 Other Long-term Assets: plug so that BS totals balance
    _last_gw  = _s(_v(goodwill_raw_h, -1))
    _last_ia  = _s(_v(intang_raw_h, -1))
    _last_ppe_fcst = _last_ppe_bs
    hist_olt_plug = [max(0, _s(_v(ta_h, i)) - hist_ca_bs[i] - hist_ppe_bs[i]
                         - _s(_v(goodwill_raw_h, i)) - _s(_v(intang_raw_h, i)))
                     for i in range(_N_HIST)]
    fcst_olt_plug = [max(0, _s(_v(fcst_ta_bs, j)) - _s(_v(fcst_ca, j))
                         - _last_ppe_fcst - _last_gw - _last_ia)
                     for j in range(_N_FCST)]

    # R526 Other Current Liabilities: plug so that R520:R526 sum = Total CL
    hist_ocl_plug = ([max(0, _s(_v(cl_direct_h, i)) - hist_ap_bs[i])
                      for i in range(_N_HIST)] if _has_cl else [0] * _N_HIST)
    fcst_ocl_plug = [0] * _N_FCST  # forecast CL modelled as AP-only

    # R532 Other NCL: ncl_excl_h / ncl_excl_f (already computed residual)

    # R538 Comprehensive Inc. and Other: plug so that equity sub-items sum to total equity
    hist_oe_plug = [_s(_v(equity_h, i)) - _s(_v(re_direct_h, i))
                    for i in range(_N_HIST)]
    # Forecast retained earnings: rolling prior_RE + NI - dividends
    _re_prior_val = _s(_v(re_direct_h, -1))
    _last_div_val = _s(_v(div_direct_h, -1)) if _has_div else 0.0
    fcst_re = []
    for _j_re in range(_N_FCST):
        _re_prior_val = _re_prior_val + _s(_v(fcst_ni, _j_re)) - _last_div_val
        fcst_re.append(_re_prior_val)
    # "Comprehensive Inc. and Other" absorbs the residual so equity still balances
    fcst_oe_plug = [_s(_v(fcst_eq, j)) - fcst_re[j]  # residual plug
                    for j in range(_N_FCST)]

    # Gross PP&E: use actual if EODHD provides it, else fall back to net_ppe
    _has_gross_ppe = any(gross_ppe_raw_h)
    _has_accum_dep = any(accum_dep_raw_h)
    if _has_gross_ppe:
        hist_gross_ppe = [_s(_v(gross_ppe_raw_h, i)) for i in range(_N_HIST)]
        fcst_gross_ppe = [_s(_v(gross_ppe_raw_h, -1))] * _N_FCST
    else:
        hist_gross_ppe = hist_ppe_bs           # fall back to net PPE
        fcst_gross_ppe = [_last_ppe_fcst] * _N_FCST
    if _has_accum_dep:
        hist_accum_dep = [_s(_v(accum_dep_raw_h, i)) for i in range(_N_HIST)]  # negative
        fcst_accum_dep = [_s(_v(accum_dep_raw_h, -1))] * _N_FCST
    else:
        hist_accum_dep = [0] * _N_HIST
        fcst_accum_dep = [0] * _N_FCST

    # ── Assets ────────────────────────────────────────────────────────────────
    _wr(500, "Cash and Equivalents",           cash_h,       fcst_cash,   bold=True)
    _wr(501, "Short-term Investments",         [0] * _N_HIST, [0] * _N_FCST)
    _wr(502, "Trading Asset Securities",       [0] * _N_HIST, [0] * _N_FCST)
    _wrf(503, "        Cash & Short-term Investments", "=@500+@501+@502")
    _wr(504, "Accounts Receivable",            hist_recv_bs, fcst_recv)
    _wr(505, "Notes Receivable",               [0] * _N_HIST, [0] * _N_FCST)
    _wrf(506, "        Total Receivables",     "=@504+@505")
    _wr(507, "Inventory",                      hist_inv_bs,  fcst_inv)
    _wr(508, "Prepaid Exp.",                   [0] * _N_HIST, [0] * _N_FCST)
    _wr(509, "Other Current Assets",           hist_oca_plug, fcst_oca_plug)
    _wrf(510, "        Total Current Assets",  "=@503+@506+@507+@508+@509",  bold=True)
    _wr(511, "Gross Property, Plant & Equipment",
             hist_gross_ppe, fcst_gross_ppe)
    _wr(512, "Accumulated Depreciation",       hist_accum_dep, fcst_accum_dep)
    _wrf(513, "        Net Property, Plant & Equipment", "=@511+@512")
    _wr(514, "Goodwill",
             goodwill_raw_h if any(goodwill_raw_h) else [0] * _N_HIST,
             [_last_gw] * _N_FCST if _last_gw else [0] * _N_FCST)
    _wr(515, "Other Intangibles",
             intang_raw_h if any(intang_raw_h) else [0] * _N_HIST,
             [_last_ia] * _N_FCST if _last_ia else [0] * _N_FCST)
    _wr(516, "Deferred Tax Assets, LT",        [0] * _N_HIST, [0] * _N_FCST)
    _wr(517, "Other Long-term Assets",         hist_olt_plug, fcst_olt_plug)
    _wrf(518, "Total Assets",                  "=@510+@513+@514+@515+@516+@517", bold=True)

    # ── Liabilities ───────────────────────────────────────────────────────────
    ws.cell(519, _LABEL_COL).value = "Liabilities ($M)"
    _wr(520, "Accounts Payable",               hist_ap_bs,   fcst_ap)
    _wr(521, "Accrued Exp.",                   [0] * _N_HIST, [0] * _N_FCST)
    _wr(522, "Short-term Borrowings",          [0] * _N_HIST, [0] * _N_FCST)
    _wr(523, "Curr. Port. of Long-term Debt",  [0] * _N_HIST, [0] * _N_FCST)
    _wr(524, "Current Portion of Leases",      [0] * _N_HIST, [0] * _N_FCST)
    _wr(525, "Curr. Income Taxes Payable",     [0] * _N_HIST, [0] * _N_FCST)
    _wr(526, "Other Current Liabilities",      hist_ocl_plug, fcst_ocl_plug)
    _wrf(527, "        Total Current Liabilities",
              "=@520+@521+@522+@523+@524+@525+@526",                         bold=True)
    _wr(528, "Long-term Debt",                 debt_h,       fcst_ltd)
    _wr(529, "Long-term Leases",               [0] * _N_HIST, [0] * _N_FCST)
    _wr(530, "Pension & Other Post-Retire. Benefits", [0] * _N_HIST, [0] * _N_FCST)
    _wr(531, "Def. Tax Liability, Non-Curr.",  [0] * _N_HIST, [0] * _N_FCST)
    _wr(532, "Other Non-Current Liabilities",  ncl_excl_h,   ncl_excl_f)
    _wrf(533, "Total Liabilities",             "=@527+@528+@529+@530+@531+@532", bold=True)

    # ── Equity ────────────────────────────────────────────────────────────────
    ws.cell(534, _LABEL_COL).value = "Equity ($M)"
    _wr(535, "Common Stock",                   [0] * _N_HIST, [0] * _N_FCST)
    _wr(536, "Additional Paid In Capital",     [0] * _N_HIST, [0] * _N_FCST)
    _wr(537, "Retained Earnings",
             [_s(_v(re_direct_h, i)) for i in range(_N_HIST)], fcst_re)
    _wr(538, "Comprehensive Inc. and Other",   hist_oe_plug,  fcst_oe_plug)
    _wrf(539, "        Total Common Equity",   "=@535+@536+@537+@538",       bold=True)
    _wrf(540, "Total Equity",                  "=@539",                      bold=True)
    _wrf(541, "Total Liabilities And Equity",  "=@533+@540",                 bold=True)

    ws.cell(542, _LABEL_COL).value = "Supplemental Items ($M)"
    _wr(543, "ECS Total Shares Outstanding on Filing Date (M)",
             shares_h, fcst_shares)
    _wr(544, "ECS Total Common Shares Outstanding (M)",
             shares_h, fcst_shares)
    _wrf(545, "Book Value per Share",          "=IFERROR(@540/@543,\"\")")
    _wrf(546, "Tangible Book Value",           "=@540-@514-@515")
    _wrf(547, "Tangible Book Value per Share", "=IFERROR(@546/@543,\"\")")
    _wrf(548, "Total Debt",                    "=@528")
    _wrf(549, "Net Debt",                      "=@528-@500")

    ws.cell(561, _LABEL_COL).value = "CIQ Restatement Type Code"
    for i in range(_N_HIST):
        ws.cell(561, _HIST_COLS[i]).value = "NC"
    ws.cell(562, _LABEL_COL).value = "CIQ Calculation Type Code"
    for i in range(_N_HIST):
        ws.cell(562, _HIST_COLS[i]).value = "REP"


def _apply_formatting(ws) -> None:
    ws.column_dimensions["A"].width = 36
    for col_idx in range(2, 2 + _N_HIST + _N_FCST):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15

    NAVY_FILL  = PatternFill("solid", fgColor=_NAVY)
    HIST_FILL  = PatternFill("solid", fgColor=_HIST_BG)
    FCST_FILL  = PatternFill("solid", fgColor=_FCST_BG)
    GREY_FILL  = PatternFill("solid", fgColor=_LIGHT_GREY)
    HDATA_FILL = PatternFill("solid", fgColor="EBF4FB")
    FDATA_FILL = PatternFill("solid", fgColor="EBF5EB")
    WHITE_BOLD = Font(bold=True, color=_WHITE,     name="Calibri", size=10)
    NAVY_BOLD  = Font(bold=True, color=_BOLD_BLUE, name="Calibri", size=10)

    _SECTION_ROWS = {300, 415, 480}
    _FIN_ROWS = (
        set(range(14, 40))
        | set(range(190, 296))
        | set(range(300, 413))
        | set(range(415, 479))
        | set(range(480, 591))
    )
    _HIST_COLS_SET = set(_HIST_COLS)
    _FCST_COLS_SET = set(_FCST_COLS)

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None and cell.row not in _SECTION_ROWS:
                continue
            r, c = cell.row, cell.column
            if c == _LABEL_COL and r in _SECTION_ROWS:
                cell.font = WHITE_BOLD
                cell.fill = NAVY_FILL
                continue
            if c == _LABEL_COL and isinstance(cell.value, str) and "--" in cell.value:
                cell.font = NAVY_BOLD
                cell.fill = GREY_FILL
                continue
            if r == 1 and cell.value is not None:
                cell.font = Font(bold=True, name="Calibri", size=10)
                if c in _HIST_COLS_SET:
                    cell.fill = HIST_FILL
                elif c in _FCST_COLS_SET:
                    cell.fill = FCST_FILL
                continue
            if r == 3 and cell.value is not None:
                if c in _HIST_COLS_SET:
                    cell.fill = HIST_FILL
                    cell.font = Font(bold=True, color=_BOLD_BLUE, name="Calibri", size=10)
                elif c in _FCST_COLS_SET:
                    cell.fill = FCST_FILL
                    cell.font = Font(bold=True, color="1A5C2A", name="Calibri", size=10)
                continue
            if r in _FIN_ROWS and isinstance(cell.value, (int, float)):
                if cell.number_format not in ("$#,##0.00", "0.0%", "0.0000",
                                               "YYYY", "0.00%", "0.0%"):
                    cell.number_format = "#,##0" if abs(cell.value) >= 10 else "#,##0.0"
                if c in _HIST_COLS_SET:
                    cell.fill = HDATA_FILL
                elif c in _FCST_COLS_SET:
                    cell.fill = FDATA_FILL

    ws.freeze_panes = "B4"
