"""
full_compare.py — Comprehensive 1:1 comparison ORIG vs GEN for NIKE
─────────────────────────────────────────────────────────────────────
ORIG = "NIKE Valuation strategy.xlsx"  (USD thousands, valuatione sheet)
GEN  = http://127.0.0.1:5000/api/export/NKE  (USD millions, valuation sheet)

Strategy:
  1. Dump every non-empty cell in both sheets
  2. Align columns by fiscal year header
  3. Compare all numeric rows (ORIG /1000 → millions)
  4. Report label mismatches, missing rows, numeric diffs
"""
import io, sys, urllib.request
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

ORIG_PATH = r"NIKE Valuation strategy.xlsx"
API_URL   = "http://127.0.0.1:5000/api/export/NKE"
MAX_ROW   = 600
MAX_COL   = 22   # A..V

# ══════════════════════════════════════════════════════════════════════════════
# 1. LOAD
# ══════════════════════════════════════════════════════════════════════════════
print("Loading ORIG…")
orig_wb = load_workbook(ORIG_PATH, data_only=True)
print(f"  ORIG sheets: {orig_wb.sheetnames}")

print("Fetching GEN from API…")
with urllib.request.urlopen(API_URL, timeout=60) as r:
    gen_bytes = r.read()
# Save for inspection
with open("generated_nke.xlsx", "wb") as f:
    f.write(gen_bytes)
gen_wb = load_workbook(io.BytesIO(gen_bytes), data_only=True)
print(f"  GEN  sheets: {gen_wb.sheetnames}")

orig_ws = orig_wb["valuatione"]
gen_ws  = gen_wb["valuation"]

# ══════════════════════════════════════════════════════════════════════════════
# 2. COLLECT ALL NON-EMPTY CELLS
# ══════════════════════════════════════════════════════════════════════════════
def collect(ws, max_row=MAX_ROW, max_col=MAX_COL):
    data = {}
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(r, c).value
            if v is not None:
                data[(r, c)] = v
    return data

orig_data = collect(orig_ws)
gen_data  = collect(gen_ws)

# ══════════════════════════════════════════════════════════════════════════════
# 3. HELPERS
# ══════════════════════════════════════════════════════════════════════════════
def num(v):
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        try:
            return float(v.replace(",","").replace("$","").replace("%","").strip())
        except Exception:
            pass
    return None

def sv(v):
    return str(v).strip() if v is not None else ""

def pct_diff(a, b):
    if a is None or b is None:
        return None
    if abs(b) < 0.01:
        return None if abs(a) < 0.01 else float("inf")
    return abs(a - b) / abs(b) * 100

# ══════════════════════════════════════════════════════════════════════════════
# 4. DETECT YEAR HEADERS IN EACH SHEET → build col→year mapping
# ══════════════════════════════════════════════════════════════════════════════
def extract_year(v):
    """Return integer year from a cell value if it looks like a year."""
    if isinstance(v, datetime):
        return v.year
    if isinstance(v, (int, float)):
        y = int(v)
        if 2000 <= y <= 2045:
            return y
    if isinstance(v, str):
        s = v.strip()
        # "FY2024", "2024", "FY24"
        for prefix in ("FY", "fy", ""):
            try:
                y = int(s.replace(prefix, ""))
                if 2000 <= y <= 2045:
                    return y
            except Exception:
                pass
    return None

# Scan rows 1-10 for year headers in both sheets
def find_year_map(data, search_rows=range(1, 12)):
    """Return {col_index: year} for all columns that have year-like values."""
    year_map = {}
    for r in search_rows:
        row_years = {}
        for c in range(2, MAX_COL + 1):
            y = extract_year(data.get((r, c)))
            if y:
                row_years[c] = y
        if len(row_years) >= 5:   # found a row with many years = the header row
            return row_years, r
    return {}, None

orig_year_map, orig_yr_row = find_year_map(orig_data)
gen_year_map,  gen_yr_row  = find_year_map(gen_data)

print(f"\nORIG year-header row: {orig_yr_row}")
print(f"  col→year: { {get_column_letter(c):y for c,y in sorted(orig_year_map.items())} }")
print(f"\nGEN  year-header row: {gen_yr_row}")
print(f"  col→year: { {get_column_letter(c):y for c,y in sorted(gen_year_map.items())} }")

# Reverse: year → col for each sheet
orig_yr2col = {y: c for c, y in orig_year_map.items()}
gen_yr2col  = {y: c for c, y in gen_year_map.items()}

# Common years
common_years = sorted(set(orig_yr2col.keys()) & set(gen_yr2col.keys()))
print(f"\nCommon fiscal years for comparison: {common_years}")

# ══════════════════════════════════════════════════════════════════════════════
# 5. LABEL SCAN — build row→label mapping for both sheets
# ══════════════════════════════════════════════════════════════════════════════
def build_label_map(data):
    """Return {label_text: [rows]} — col A labels plus col G (ORIG DCF section)."""
    lmap = {}
    for r in range(1, MAX_ROW + 1):
        lbl = sv(data.get((r, 1)))   # col A
        if not lbl:
            lbl = sv(data.get((r, 7)))   # col G for ORIG DCF labels
        if lbl:
            lmap.setdefault(lbl.lower(), []).append(r)
    return lmap

orig_labels = build_label_map(orig_data)
gen_labels  = build_label_map(gen_data)

# ══════════════════════════════════════════════════════════════════════════════
# 6. SECTION A: Full dump of ORIG
# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "═"*120)
print("SECTION A: ORIG 'valuatione' — every non-empty cell (rows 1-600, cols A-V)")
print("═"*120)
for row in range(1, MAX_ROW + 1):
    parts = []
    for col in range(1, MAX_COL + 1):
        v = orig_data.get((row, col))
        if v is not None:
            parts.append(f"{get_column_letter(col)}={repr(v)}")
    if parts:
        print(f"R{row:4d}: " + "  ".join(parts))

# ══════════════════════════════════════════════════════════════════════════════
# 7. SECTION B: Full dump of GEN
# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "═"*120)
print("SECTION B: GEN 'valuation' — every non-empty cell (rows 1-600, cols A-V)")
print("═"*120)
for row in range(1, MAX_ROW + 1):
    parts = []
    for col in range(1, MAX_COL + 1):
        v = gen_data.get((row, col))
        if v is not None:
            parts.append(f"{get_column_letter(col)}={repr(v)}")
    if parts:
        print(f"R{row:4d}: " + "  ".join(parts))

# ══════════════════════════════════════════════════════════════════════════════
# 8. SECTION C: Row-by-row label comparison (col A)
# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "═"*120)
print("SECTION C: Col-A label comparison (both sheets, rows 1-600)")
print("═"*120)
print(f"{'Row':<6} {'ORIG ColA':<60} {'GEN ColA':<60} Match?")
print("-"*140)
for row in range(1, MAX_ROW + 1):
    o = sv(orig_data.get((row, 1)))
    g = sv(gen_data.get((row, 1)))
    if o or g:
        ok = (o.lower() == g.lower())
        flag = "" if ok else "  <<<< DIFF"
        print(f"R{row:<5} {repr(o):<60} {repr(g):<60} {'OK' if ok else 'DIFF'}{flag}")

# ══════════════════════════════════════════════════════════════════════════════
# 9. SECTION D: Year-aligned numeric comparison
#    For each common year, compare every row that has a number in ORIG
#    ORIG is in THOUSANDS → divide by 1000 for millions comparison
# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "═"*120)
print("SECTION D: Year-aligned numeric comparison (ORIG/1000 = M vs GEN in M)")
print(f"  Comparing years: {common_years}")
print(f"  Threshold: >2% diff flagged")
print("═"*120)
print(f"{'Row':>5} {'Label':<45} {'Year':>6} {'ORIG(M)':>14} {'GEN(M)':>14} {'Diff%':>9}  Status")
print("-"*120)

all_diffs = []
all_matches = 0

for year in common_years:
    oc = orig_yr2col[year]
    gc = gen_yr2col[year]
    for row in range(1, MAX_ROW + 1):
        ov = orig_data.get((row, oc))
        gv = gen_data.get((row, gc))
        on = num(ov)
        gn = num(gv)
        if on is None and gn is None:
            continue
        on_m = on / 1000.0 if on is not None else None
        gn_m = gn
        diff = pct_diff(on_m, gn_m)
        # Get label: try col A in orig, then col G
        lbl = sv(orig_data.get((row, 1)) or orig_data.get((row, 7)) or gen_data.get((row, 1)) or "")
        if diff is None:
            status = "ONE_SIDE"
            all_diffs.append((row, year, on_m, gn_m, float("inf"), lbl))
        elif diff > 2.0:
            status = f"DIFF {diff:.1f}%"
            all_diffs.append((row, year, on_m, gn_m, diff, lbl))
        else:
            status = "OK"
            all_matches += 1
        o_s = f"{on_m:>14.3f}" if on_m is not None else f"{'<missing>':>14}"
        g_s = f"{gn_m:>14.3f}" if gn_m is not None else f"{'<missing>':>14}"
        d_s = f"{diff:>9.1f}" if (diff is not None and diff != float("inf")) else f"{'inf':>9}"
        print(f"R{row:4d}  {lbl[:44]:<45} {year:>6} {o_s} {g_s} {d_s}  {status}")

print(f"\n  >>> Total year-aligned mismatches (>2%): {len(all_diffs)}")
print(f"  >>> Year-aligned exact matches (<2%): {all_matches}")

# ══════════════════════════════════════════════════════════════════════════════
# 10. SECTION E: Summary — only the DIFF rows, grouped by label
# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "═"*120)
print("SECTION E: DIFF SUMMARY — rows with >2% mismatch grouped by label")
print("═"*120)

from collections import defaultdict
by_label = defaultdict(list)
for row, year, on_m, gn_m, diff, lbl in all_diffs:
    by_label[lbl].append((row, year, on_m, gn_m, diff))

for lbl, entries in sorted(by_label.items(), key=lambda x: x[1][0][0]):
    print(f"\nRow {entries[0][0]:3d}  Label: {repr(lbl)}")
    for row, year, on_m, gn_m, diff in entries:
        o_s = f"{on_m:.3f}" if on_m is not None else "<missing>"
        g_s = f"{gn_m:.3f}" if gn_m is not None else "<missing>"
        d_s = f"{diff:.1f}%" if diff != float("inf") else "inf"
        print(f"         FY{year}: ORIG={o_s:>12}M  GEN={g_s:>12}M  diff={d_s}")

# ══════════════════════════════════════════════════════════════════════════════
# 11. SECTION F: DCF section deep comparison (ORIG col G labels, rows 1-76)
# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "═"*120)
print("SECTION F: ORIG DCF section — col G labels (rows 1-76)")
print("═"*120)
for row in range(1, 77):
    g_lbl = orig_data.get((row, 7))
    if g_lbl:
        parts = [f"colG={repr(g_lbl)}"]
        for col in range(8, MAX_COL + 1):
            v = orig_data.get((row, col))
            if v is not None:
                parts.append(f"{get_column_letter(col)}={repr(v)}")
        print(f"R{row:3d}: " + "  ".join(parts))

# Compare GEN rows 1-76 (where GEN has its DCF)
print("\n--- GEN rows 1-76 (DCF area) ---")
for row in range(1, 77):
    parts = []
    for col in range(1, MAX_COL + 1):
        v = gen_data.get((row, col))
        if v is not None:
            parts.append(f"{get_column_letter(col)}={repr(v)}")
    if parts:
        print(f"R{row:3d}: " + "  ".join(parts))

# ══════════════════════════════════════════════════════════════════════════════
# 12. SECTION G: GEN rows missing from ORIG and vice versa (by label)
# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "═"*120)
print("SECTION G: Labels in ORIG but NOT in GEN")
print("═"*120)
for lbl, rows in sorted(orig_labels.items()):
    if lbl and lbl not in gen_labels:
        print(f"  MISSING in GEN: {repr(lbl)}  (ORIG rows: {rows})")

print("\n" + "═"*120)
print("SECTION G2: Labels in GEN but NOT in ORIG")
print("═"*120)
for lbl, rows in sorted(gen_labels.items()):
    if lbl and lbl not in orig_labels:
        print(f"  EXTRA in GEN: {repr(lbl)}  (GEN rows: {rows})")

# ══════════════════════════════════════════════════════════════════════════════
# 13. SECTION H: Raw column comparison (no year alignment, col-by-col)
#     Shows ALL rows where numbers differ between same columns
# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "═"*120)
print("SECTION H: Raw col-by-col numeric diff (same col index, ORIG/1000 vs GEN)")
print("  Only showing rows/cols where both have numeric values and diff > 5%")
print("═"*120)
print(f"{'Row':>5} {'Col':<4} {'ORIG(M)':>14} {'GEN(M)':>14} {'Diff%':>9}  Label")
print("-"*110)

raw_mismatches = []
for row in range(1, MAX_ROW + 1):
    for col in range(2, MAX_COL + 1):
        ov = orig_data.get((row, col))
        gv = gen_data.get((row, col))
        on = num(ov)
        gn = num(gv)
        if on is None or gn is None:
            continue
        on_m = on / 1000.0
        gn_m = gn
        diff = pct_diff(on_m, gn_m)
        if diff is not None and diff > 5.0:
            lbl = sv(orig_data.get((row, 1)) or orig_data.get((row, 7)) or gen_data.get((row, 1)) or "")
            raw_mismatches.append((row, col, on_m, gn_m, diff, lbl))

for row, col, on_m, gn_m, diff, lbl in raw_mismatches[:300]:
    print(f"R{row:4d}  {get_column_letter(col):<4} {on_m:>14.3f} {gn_m:>14.3f} {diff:>9.1f}  {lbl[:60]}")

print(f"\n  >>> Raw col-by-col mismatches (>5%): {len(raw_mismatches)}")

# ══════════════════════════════════════════════════════════════════════════════
# 14. SECTION I: Key metrics spot check (important rows by label match)
# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "═"*120)
print("SECTION I: Key metrics spot-check — most recent common year")
print("═"*120)

KEY_LABELS = [
    "revenue", "net revenue", "sales", "net sales",
    "cost of goods sold", "gross profit", "gross margin",
    "ebit", "operating income", "operating profit",
    "ebitda",
    "net income", "net earnings",
    "earnings per share", "eps",
    "depreciation", "depreciation & amortization", "d&a",
    "capital expenditures", "capex",
    "free cash flow", "unlevered free cash flow", "ufcf",
    "wacc", "terminal growth",
    "total assets", "total liabilities", "total equity",
    "cash", "long-term debt", "total debt",
    "dividends", "buybacks",
    "shares outstanding",
    "current ratio", "debt/equity", "roe", "roa", "roic",
    "interest expense",
]

if common_years:
    last_year = common_years[-1]
    oc = orig_yr2col[last_year]
    gc = gen_yr2col[last_year]
    print(f"  Comparing FY{last_year} (ORIG col {get_column_letter(oc)}, GEN col {get_column_letter(gc)})")
    print(f"{'Label':<45} {'ORIG(M)':>14} {'GEN(M)':>14} {'Diff%':>9}  Status")
    print("-"*100)
    for kl in KEY_LABELS:
        o_rows = orig_labels.get(kl, [])
        g_rows = gen_labels.get(kl, [])
        if not o_rows and not g_rows:
            continue
        # try best match
        for or_ in o_rows:
            ov = num(orig_data.get((or_, oc)))
            if ov is None:
                continue
            on_m = ov / 1000.0
            # find corresponding GEN row
            best_gr = None
            for gr in g_rows:
                if gen_data.get((gr, gc)) is not None:
                    best_gr = gr
                    break
            if best_gr is None:
                print(f"  {kl:<44} {on_m:>14.3f} {'<missing>':>14} {'inf':>9}  MISSING in GEN")
                continue
            gn_m = num(gen_data.get((best_gr, gc)))
            if gn_m is None:
                print(f"  {kl:<44} {on_m:>14.3f} {'<none>':>14} {'inf':>9}  NO NUMERIC IN GEN")
                continue
            diff = pct_diff(on_m, gn_m)
            d_s = f"{diff:.1f}%" if diff is not None and diff != float("inf") else "inf"
            status = "OK" if (diff is not None and diff <= 2.0) else "DIFF"
            print(f"  {kl:<44} {on_m:>14.3f} {gn_m:>14.3f} {d_s:>9}  {status}  OR={or_} GR={best_gr}")
            break

print("\n" + "═"*120)
print("DONE. Output saved to generated_nke.xlsx for manual inspection.")
print("═"*120)
