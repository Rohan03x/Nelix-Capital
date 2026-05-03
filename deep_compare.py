"""
deep_compare.py  —  Full structural analysis of ORIG vs GEN
Uses utf-8 output, handles encoding cleanly.
"""
import io, sys, urllib.request
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

ORIG_PATH = r"NIKE Valuation strategy.xlsx"
API_URL   = "http://127.0.0.1:5000/api/export/NKE"

# ─── load ───────────────────────────────────────────────────────────────────

orig_wb = load_workbook(ORIG_PATH, data_only=True)
print(f"ORIG sheets: {orig_wb.sheetnames}")

with urllib.request.urlopen(API_URL, timeout=30) as r:
    gen_bytes = r.read()
gen_wb = load_workbook(io.BytesIO(gen_bytes), data_only=True)
print(f"GEN  sheets: {gen_wb.sheetnames}")

orig_ws = orig_wb["valuatione"]
gen_ws  = gen_wb["valuation"]

MAX_ROW  = 600
MAX_COL  = 19   # A(1) through S(19) to capture col S too

# ─── collect all non-empty cells ────────────────────────────────────────────

def collect(ws, max_row=MAX_ROW, max_col=MAX_COL):
    data = {}
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(r, c).value
            if v is not None:
                data[(r, c)] = v
    return data

orig_data = collect(orig_ws)
gen_data  = collect(gen_ws)

# ─── helpers ─────────────────────────────────────────────────────────────────

def numeric(v):
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        try:
            return float(v.replace(",","").replace("$","").replace("%",""))
        except Exception:
            pass
    return None

def pct(a, b):
    if a is None or b is None:
        return None
    if abs(b) < 0.001:
        return None if abs(a) < 0.001 else float("inf")
    return abs(a - b) / abs(b) * 100

def sv(v):
    return str(v).strip() if v is not None else ""

# ─── SECTION 1: Full dump of ORIG ────────────────────────────────────────────

print("\n" + "="*90)
print("SECTION 1: ORIG 'valuatione' sheet — full non-empty cells (rows 1-600, cols A-S)")
print("="*90)

for row in range(1, MAX_ROW + 1):
    parts = []
    for col in range(1, MAX_COL + 1):
        v = orig_data.get((row, col))
        if v is not None:
            parts.append(f"{get_column_letter(col)}={repr(v)}")
    if parts:
        print(f"R{row:4d}: " + "  ".join(parts))

# ─── SECTION 2: Full dump of GEN ─────────────────────────────────────────────

print("\n" + "="*90)
print("SECTION 2: GEN 'valuation' sheet — full non-empty cells (rows 1-600, cols A-S)")
print("="*90)

for row in range(1, MAX_ROW + 1):
    parts = []
    for col in range(1, MAX_COL + 1):
        v = gen_data.get((row, col))
        if v is not None:
            parts.append(f"{get_column_letter(col)}={repr(v)}")
    if parts:
        print(f"R{row:4d}: " + "  ".join(parts))

# ─── SECTION 3: Col-A (label) comparison ─────────────────────────────────────

print("\n" + "="*90)
print("SECTION 3: Col-A label comparison (rows 1-600)")
print("="*90)
print(f"{'Row':<6} {'ORIG label':<55} {'GEN label':<55} {'Match?'}")
print("-"*170)

for row in range(1, MAX_ROW + 1):
    o = sv(orig_data.get((row, 1)))
    g = sv(gen_data.get((row, 1)))
    if o or g:
        match = "OK" if o.lower() == g.lower() else ("DIFF" if (o or g) else "")
        flag  = "  <<<< LABEL DIFF" if match == "DIFF" else ""
        print(f"R{row:<5} {repr(o):<55} {repr(g):<55} {match}{flag}")

# ─── SECTION 4: Numeric comparison ───────────────────────────────────────────
# ORIG is in THOUSANDS, GEN is in MILLIONS → scale ORIG by /1000 before compare

print("\n" + "="*90)
print("SECTION 4: Numeric comparison (ORIG in thousands /1000 vs GEN in millions)")
print("  Threshold: >2% difference reported")
print("="*90)
print(f"{'Row':>5} {'Col':<4} {'ORIG(M)':>14} {'GEN(M)':>14} {'Diff%':>8}  Label")
print("-"*110)

mismatches = []
matches    = 0

# map the column from ORIG to GEN
# ORIG: B=FY2015...K=FY2024 (hist), L=FY2025...R=FY2031 (fcst)
# GEN:  B=empty*6, then G-K=FY2022-FY2025 (hist), L-R=FY2026-2032 (fcst)
# The GEN only has ~4 years of historical data. We compare against ORIG K column (FY2024) = GEN col K

# For now compare col by col (raw) to catch structure issues
for row in range(1, MAX_ROW + 1):
    for col in range(2, MAX_COL + 1):
        ov = orig_data.get((row, col))
        gv = gen_data.get((row, col))
        on = numeric(ov)
        gn = numeric(gv)
        if on is None and gn is None:
            continue
        # scale ORIG from thousands to millions
        if on is not None:
            on_m = on / 1000.0
        else:
            on_m = None
        gn_m = gn
        diff = pct(on_m, gn_m)
        if diff is None:
            if on_m is not None or gn_m is not None:
                lbl = sv(orig_data.get((row,1)) or gen_data.get((row,1)))
                mismatches.append((row, col, on_m, gn_m, float("inf"), lbl))
        elif diff > 2.0:
            lbl = sv(orig_data.get((row,1)) or gen_data.get((row,1)))
            mismatches.append((row, col, on_m, gn_m, diff, lbl))
        else:
            matches += 1

for row, col, on_m, gn_m, diff, lbl in mismatches[:200]:
    col_ltr = get_column_letter(col)
    o_str = f"{on_m:>14.2f}" if on_m is not None else f"{'<missing>':>14}"
    g_str = f"{gn_m:>14.2f}" if gn_m is not None else f"{'<missing>':>14}"
    d_str = f"{diff:>8.1f}" if diff != float("inf") else f"{'inf':>8}"
    print(f"R{row:4d}  {col_ltr:<4} {o_str} {g_str} {d_str}  {lbl[:60]}")

print(f"\nTotal mismatches (>2%): {len(mismatches)},  exact matches: {matches}")

# ─── SECTION 5: Key DCF rows comparison ─────────────────────────────────────

print("\n" + "="*90)
print("SECTION 5: DCF key rows — ORIG col K (FY2024 current) vs GEN col K")
print("  (ORIG in thousands → /1000 = millions)")
print("="*90)

DCF_ROWS = {
    # ORIG row → label
    16: "Sales / Revenue",
    17: "Net income",
    18: "D&A",
    19: "CapEx",
    20: "NOWC change",
    21: "UFCF",
    22: "Discount factor",
    23: "PV(UFCF)",
    25: "WACC",
    26: "Terminal growth",
}
for r, lbl in DCF_ROWS.items():
    ok = orig_data.get((r, 11))   # col K
    gk = gen_data.get((r, 11))
    on_m = numeric(ok) / 1000 if numeric(ok) is not None else None
    gn_m = numeric(gk)
    diff = pct(on_m, gn_m)
    d_str = f"{diff:.1f}%" if diff is not None and diff != float("inf") else ("inf" if diff == float("inf") else "N/A")
    print(f"R{r:<4} {lbl:<30} ORIG(K→M)={on_m!r:<16} GEN={gn_m!r:<16} diff={d_str}")

# ─── SECTION 6: ORIG rows where label is in col G (not col A) ────────────────

print("\n" + "="*90)
print("SECTION 6: Rows where ORIG label is in col G (DCF model rows)")
print("="*90)
for row in range(1, MAX_ROW+1):
    g_lbl = orig_data.get((row, 7))   # col G
    if isinstance(g_lbl, str) and len(g_lbl.strip()) > 0:
        # show row data
        parts = [f"colG_label={repr(g_lbl)}"]
        for col in range(8, MAX_COL+1):
            v = orig_data.get((row, col))
            if v is not None:
                parts.append(f"{get_column_letter(col)}={repr(v)}")
        print(f"R{row:4d}: " + "  ".join(parts))

print("\nDONE.")
